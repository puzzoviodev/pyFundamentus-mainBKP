import openpyxl
import re
import time
import numpy as np
import pandas as pd
from unidecode import unidecode
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
import requests
# teste silvio 3e
import warnings
from openpyxl.styles import numbers

#from teste01 import metrica

fillvermelho= PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") # Vermelho

fillverde= PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid") # Verde

fillamarelo =  PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # Amarelo


fillazul = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid") # Azul


#filltitulo =   PatternFill(start_color="#002060", end_color="#002060", fill_type="solid")

filltitulo = PatternFill(start_color="002060", end_color="002060", fill_type="solid")  # Azul escuro

font_branca = Font(color="FFFFFF")  # Branco

TITLES = [
    'Identificação', 'Resumo Financeiro', 'Cotações', 'Informações Básicas',
    'Oscilações', 'Indicadores de Valuation', 'Indicadores de Rentabilidade',
    'Indicadores de Endividamento', 'Balanço Patrimonial', 'Demonstrativo de Resultados'
]

linha2 = 1
metricasts= ""

MetricasStatus = {
    'Dívida Líquida/EBIT': {
        'pessimo': {'min': 4, 'max': float('inf')},
        'ruim': {'min': 3, 'max': 4},
        'moderado': {'min': 1.5, 'max': 3},
        'bom': {'min': 0, 'max': 1.5},
        'otimo': {'min': float('-inf'), 'max': 0},
        'descricao': 'Relação entre dívida líquida e lucro operacional antes de juros e impostos. Valores altos indicam maior alavancagem e risco financeiro.',
        'agrupador': 'Endividamento',
        'descrpessimo': 'Endividamento crítico, leva mais de 4 anos de EBIT para quitar a dívida líquida, indicando alavancagem extrema e risco financeiro elevado. Comum em empresas em crise (ex.: OIBR3) ou setores cíclicos em baixa (ex.: construção, CSNA3). Em 2025, com juros de 10-12%, o custo da dívida pressiona ainda mais a geração de caixa, aumentando a vulnerabilidade a choques econômicos, como desaceleração ou aumento de taxas. Empresas nessa faixa frequentemente enfrentam dificuldades para cobrir juros, com risco de default ou necessidade de reestruturação. Investidores devem evitar, salvo sinais claros de recuperação, como redução agressiva de dívida, venda de ativos (ex.: Petrobrás em reestruturações passadas) ou aumento robusto do EBIT via eficiência operacional. Analise a cobertura de juros (EBIT/juros) e a tendência do EBIT nos últimos 3-5 anos para avaliar sustentabilidade. Setores intensivos em capital (ex.: infraestrutura) podem justificar alavancagem, mas exigem fluxo de caixa estável. Empresas com histórico de lucros voláteis ou em setores expostos a commodities (ex.: mineração) são particularmente arriscadas.',
        'descrruim': 'Endividamento elevado, leva de 3 a 4 anos de EBIT para quitar a dívida líquida. Risco significativo, especialmente em setores cíclicos (ex.: mineração, VALE3). Em 2025, com juros altos, a capacidade de pagamento é pressionada. Cautela é necessária; analise estabilidade do EBIT e plano de desalavancagem.',
        'descrmoderado': 'Endividamento moderado, leva de 1,5 a 3 anos de EBIT para quitar a dívida líquida. Aceitável em setores estáveis (ex.: utilities, ENGI11). Em 2025, monitorar fluxo de caixa livre e riscos macroeconômicos para garantir sustentabilidade.',
        'descrbom': 'Endividamento baixo, dívida quitada em até 1,5 anos de EBIT. Sinal de saúde financeira sólida, comum em setores maduros (ex.: bens de consumo, ITUB4). Em 2025, atrativo para investidores conservadores, mas verificar reinvestimento.',
        'descotimo': 'Dívida líquida negativa, caixa excede dívidas. Situação financeira excepcional, comum em tecnologia (ex.: MGLU3 em períodos de alta liquidez). Em 2025, avaliar alocação de caixa para evitar ineficiências.'
    },
    'Dívida Líquida/EBITDA': {
        'pessimo': {'min': 3.5, 'max': float('inf')},
        'ruim': {'min': 2.5, 'max': 3.5},
        'moderado': {'min': 1, 'max': 2.5},
        'bom': {'min': 0, 'max': 1},
        'otimo': {'min': float('-inf'), 'max': 0},
        'descricao': 'Relação entre dívida líquida e EBITDA. Valores altos indicam alavancagem elevada.',
        'agrupador': 'Endividamento',
        'descrpessimo': 'Endividamento crítico, acima de 3,5x o EBITDA. Alto risco, comum em crise (ex.: OIBR3). Em 2025, com juros de 10-12%, risco de default é elevado. Evitar, salvo recuperação clara.',
        'descrruim': 'Endividamento elevado, 2,5-3,5x o EBITDA. Risco em setores cíclicos (ex.: CSNA3). Cautela, verificar plano de redução de dívida.',
        'descrmoderado': 'Endividamento moderado, 1-2,5x o EBITDA. Aceitável em setores estáveis (ex.: EGIE3). Monitorar fluxo de caixa.',
        'descrbom': 'Endividamento baixo, até 1x o EBITDA. Sólido, comum em bens de consumo (ex.: ABEV3). Atraente para conservadores.',
        'descotimo': 'Dívida líquida negativa, caixa supera dívidas. Excelente, comum em tecnologia (ex.: WEGR3). Avaliar uso do caixa.'
    },
    'Dívida Líquida/Patrimônio Líquido': {
        'pessimo': {'min': 1, 'max': float('inf')},
        'ruim': {'min': 0.7, 'max': 1},
        'moderado': {'min': 0.3, 'max': 0.7},
        'bom': {'min': 0, 'max': 0.3},
        'otimo': {'min': float('-inf'), 'max': 0},
        'descricao': 'Relação entre dívida líquida e patrimônio líquido. Indica nível de alavancagem em relação ao capital próprio.',
        'agrupador': 'Endividamento',
        'descrpessimo': 'Alavancagem crítica, dívida excede PL. Alto risco (ex.: OIBR3). Em 2025, risco de insolvência é elevado. Evitar.',
        'descrruim': 'Alavancagem elevada, 70-100% do PL. Risco moderado (ex.: CSNA3). Cautela, verificar solvência.',
        'descrmoderado': 'Alavancagem moderada, 30-70% do PL. Aceitável (ex.: ENGI11). Monitorar ROE e fluxo de caixa.',
        'descrbom': 'Baixa alavancagem, até 30% do PL. Sólido (ex.: ITUB4). Atraente para conservadores.',
        'descotimo': 'Caixa excede dívida, situação robusta. Comum em tecnologia (ex.: MGLU3). Avaliar alocação de capital.'
    },
    'Patrimônio/Ativos': {
        'pessimo': {'min': 0, 'max': 0.2},
        'ruim': {'min': 0.2, 'max': 0.3},
        'moderado': {'min': 0.3, 'max': 0.5},
        'bom': {'min': 0.5, 'max': 0.7},
        'otimo': {'min': 0.7, 'max': float('inf')},
        'descricao': 'Proporção do patrimônio em relação aos ativos totais. Indica robustez financeira.',
        'agrupador': 'Endividamento',
        'descrpessimo': 'Estrutura frágil, patrimônio <20% dos ativos (ex.: OIBR3). Alto risco. Evitar.',
        'descrruim': 'Estrutura limitada, 20-30% dos ativos. Risco moderado (ex.: CSNA3). Cautela.',
        'descrmoderado': 'Estrutura equilibrada, 30-50% dos ativos (ex.: ENGI11). Seguro com boa gestão.',
        'descrbom': 'Estrutura forte, 50-70% dos ativos (ex.: ABEV3). Atraente para conservadores.',
        'descotimo': 'Estrutura robusta, >70% dos ativos (ex.: WEGR3). Excelente, verificar alocação.'
    },
    'PL/Ativos': {
        'pessimo': {'min': 0, 'max': 0.2},
        'ruim': {'min': 0.2, 'max': 0.3},
        'moderado': {'min': 0.3, 'max': 0.5},
        'bom': {'min': 0.5, 'max': 0.7},
        'otimo': {'min': 0.7, 'max': float('inf')},
        'descricao': 'Relação entre patrimônio líquido e ativos totais. Indica solidez financeira.',
        'agrupador': 'Endividamento',
        'descrpessimo': 'Patrimônio muito baixo, <20% dos ativos (ex.: OIBR3). Evitar.',
        'descrruim': 'Patrimônio limitado, 20-30% (ex.: CSNA3). Cautela.',
        'descrmoderado': 'Patrimônio equilibrado, 30-50% (ex.: ENGI11). Seguro.',
        'descrbom': 'Patrimônio forte, 50-70% (ex.: ABEV3). Atraente.',
        'descotimo': 'Patrimônio robusto, >70% (ex.: WEGR3). Excelente.'
    },
    'Dividend Yield (DY)': {
        'pessimo': {'min': 0, 'max': 0},
        'ruim': {'min': 0.1, 'max': 2},
        'moderado': {'min': 2, 'max': 4},
        'bom': {'min': 4, 'max': 6},
        'otimo': {'min': 6, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Renda',
        'descrpessimo': 'Sem dividendos, foco em reinvestimento ou prejuízo (ex.: startups). Evitar para renda.',
        'descrruim': 'Dividendos baixos, 0,1-2% (ex.: MGLU3). Pouco atrativo para renda.',
        'descrmoderado': 'Dividendos moderados, 2-4% (ex.: ENGI11). Equilíbrio entre renda e reinvestimento.',
        'descrbom': 'Dividendos atrativos, 4-6% (ex.: ITUB4). Ideal para renda.',
        'descotimo': 'Dividendos altos, >6% (ex.: TAEE11). Verificar sustentabilidade.'
    },
    'EV/EBIT': {
        'pessimo': {'min': 20, 'max': float('inf')},
        'ruim': {'min': 15, 'max': 20},
        'moderado': {'min': 10, 'max': 15},
        'bom': {'min': 5, 'max': 10},
        'otimo': {'min': 0, 'max': 5},
        'descricao': 'Relação entre valor da empresa e EBIT. Indica valuation.',
        'agrupador': 'Valuation',
        'descrpessimo': 'Valuation elevado, >20x (ex.: startups tech). Evitar, salvo crescimento excepcional.',
        'descrruim': 'Valuation alto, 15-20x (ex.: MGLU3). Cautela, verificar crescimento.',
        'descrmoderado': 'Valuation razoável, 10-15x (ex.: ABEV3). Justo com EBIT estável.',
        'descrbom': 'Valuation atrativo, 5-10x (ex.: ITUB4). Ideal para valor.',
        'descotimo': 'Valuation muito atrativo, <5x (ex.: VALE3 em recuperação). Oportunidade.'
    },
    'EV/EBITDA': {
        'pessimo': {'min': 15, 'max': float('inf')},
        'ruim': {'min': 10, 'max': 15},
        'moderado': {'min': 7, 'max': 10},
        'bom': {'min': 4, 'max': 7},
        'otimo': {'min': 0, 'max': 4},
        'descricao': 'Relação entre valor da empresa e EBITDA. Indica valuation.',
        'agrupador': 'Valuation',
        'descrpessimo': 'Valuation elevado, >15x (ex.: startups tech). Evitar.',
        'descrruim': 'Valuation alto, 10-15x (ex.: MGLU3). Cautela.',
        'descrmoderado': 'Valuation razoável, 7-10x (ex.: ABEV3). Justo.',
        'descrbom': 'Valuation atrativo, 4-7x (ex.: ITUB4). Ideal.',
        'descotimo': 'Valuation muito atrativo, <4x (ex.: VALE3). Oportunidade.'
    },
    'P/ATIVO': {
        'pessimo': {'min': 2, 'max': float('inf')},
        'ruim': {'min': 1.5, 'max': 2},
        'moderado': {'min': 1, 'max': 1.5},
        'bom': {'min': 0.5, 'max': 1},
        'otimo': {'min': 0, 'max': 0.5},
        'descricao': 'Relação entre preço e ativos totais. Indica valuation.',
        'agrupador': 'Valuation',
        'descrpessimo': 'Valuation elevado, >2x (ex.: MGLU3). Evitar.',
        'descrruim': 'Valuation alto, 1,5-2x (ex.: startups). Cautela.',
        'descrmoderado': 'Valuation razoável, 1-1,5x (ex.: ABEV3). Justo.',
        'descrbom': 'Valuation atrativo, 0,5-1x (ex.: ITUB4). Ideal.',
        'descotimo': 'Valuation muito atrativo, <0,5x (ex.: VALE3). Oportunidade.'
    },
    'P/Ativo Circulante Líquido': {
        'pessimo': {'min': 3, 'max': float('inf')},
        'ruim': {'min': 2, 'max': 3},
        'moderado': {'min': 1, 'max': 2},
        'bom': {'min': 0.5, 'max': 1},
        'otimo': {'min': 0, 'max': 0.5},
        'descricao': 'Relação entre preço e ativo circulante líquido. Indica valuation e liquidez.',
        'agrupador': 'Valuation',
        'descrpessimo': 'Valuation elevado, >3x (ex.: startups). Evitar.',
        'descrruim': 'Valuation alto, 2-3x (ex.: MGLU3). Cautela.',
        'descrmoderado': 'Valuation razoável, 1-2x (ex.: ABEV3). Justo.',
        'descrbom': 'Valuation atrativo, 0,5-1x (ex.: ITUB4). Ideal.',
        'descotimo': 'Valuation muito atrativo, <0,5x (ex.: VALE3). Oportunidade.'
    },
    'P/Capital de Giro': {
        'pessimo': {'min': 5, 'max': float('inf')},
        'ruim': {'min': 3, 'max': 5},
        'moderado': {'min': 1.5, 'max': 3},
        'bom': {'min': 0.5, 'max': 1.5},
        'otimo': {'min': 0, 'max': 0.5},
        'descricao': 'Relação entre preço e capital de giro. Indica valuation e eficiência.',
        'agrupador': 'Valuation',
        'descrpessimo': 'Valuation elevado, >5x (ex.: startups). Evitar.',
        'descrruim': 'Valuation alto, 3-5x (ex.: MGLU3). Cautela.',
        'descrmoderado': 'Valuation razoável, 1,5-3x (ex.: ABEV3). Justo.',
        'descrbom': 'Valuation atrativo, 0,5-1,5x (ex.: ITUB4). Ideal.',
        'descotimo': 'Valuation muito atrativo, <0,5x (ex.: VALE3). Oportunidade.'
    },
    'P/EBIT': {
        'pessimo': {'min': 25, 'max': float('inf')},
        'ruim': {'min': 20, 'max': 25},
        'moderado': {'min': 15, 'max': 20},
        'bom': {'min': 10, 'max': 15},
        'otimo': {'min': 0, 'max': 10},
        'descricao': 'Relação entre preço e EBIT. Indica valuation operacional.',
        'agrupador': 'Valuation',
        'descrpessimo': 'Valuation elevado, >25x (ex.: startups). Evitar.',
        'descrruim': 'Valuation alto, 20-25x (ex.: MGLU3). Cautela.',
        'descrmoderado': 'Valuation razoável, 15-20x (ex.: ABEV3). Justo.',
        'descrbom': 'Valuation atrativo, 10-15x (ex.: ITUB4). Ideal.',
        'descotimo': 'Valuation muito atrativo, <10x (ex.: VALE3). Oportunidade.'
    },
    'P/EBITDA': {
        'pessimo': {'min': 15, 'max': float('inf')},
        'ruim': {'min': 10, 'max': 15},
        'moderado': {'min': 7, 'max': 10},
        'bom': {'min': 4, 'max': 7},
        'otimo': {'min': 0, 'max': 4},
        'descricao': 'Relação entre preço e EBITDA. Indica valuation operacional.',
        'agrupador': 'Valuation',
        'descrpessimo': 'Valuation elevado, >15x (ex.: startups). Evitar.',
        'descrruim': 'Valuation alto, 10-15x (ex.: MGLU3). Cautela.',
        'descrmoderado': 'Valuation razoável, 7-10x (ex.: ABEV3). Justo.',
        'descrbom': 'Valuation atrativo, 4-7x (ex.: ITUB4). Ideal.',
        'descotimo': 'Valuation muito atrativo, <4x (ex.: VALE3). Oportunidade.'
    },
    'P/L': {
        'pessimo': {'min': 30, 'max': float('inf')},
        'ruim': {'min': 20, 'max': 30},
        'moderado': {'min': 15, 'max': 20},
        'bom': {'min': 10, 'max': 15},
        'otimo': {'min': 0, 'max': 10},
        'descricao': 'Relação entre preço e lucro líquido. Indica valuation.',
        'agrupador': 'Valuation',
        'descrpessimo': 'Valuation elevado, >30x (ex.: startups). Evitar.',
        'descrruim': 'Valuation alto, 20-30x (ex.: MGLU3). Cautela.',
        'descrmoderado': 'Valuation razoável, 15-20x (ex.: ABEV3). Justo.',
        'descrbom': 'Valuation atrativo, 10-15x (ex.: ITUB4). Ideal.',
        'descotimo': 'Valuation muito atrativo, <10x (ex.: VALE3). Oportunidade.'
    },
    'P/VPA': {
        'pessimo': {'min': 3, 'max': float('inf')},
        'ruim': {'min': 2, 'max': 3},
        'moderado': {'min': 1.5, 'max': 2},
        'bom': {'min': 1, 'max': 1.5},
        'otimo': {'min': 0, 'max': 1},
        'descricao': 'Relação entre preço e valor patrimonial por ação. Indica valuation.',
        'agrupador': 'Valuation',
        'descrpessimo': 'Valuation elevado, >3x (ex.: startups). Evitar.',
        'descrruim': 'Valuation alto, 2-3x (ex.: MGLU3). Cautela.',
        'descrmoderado': 'Valuation razoável, 1,5-2x (ex.: ABEV3). Justo.',
        'descrbom': 'Valuation atrativo, 1-1,5x (ex.: ITUB4). Ideal.',
        'descotimo': 'Valuation muito atrativo, <1x (ex.: VALE3). Oportunidade.'
    },
    'PSR': {
        'pessimo': {'min': 3, 'max': float('inf')},
        'ruim': {'min': 2, 'max': 3},
        'moderado': {'min': 1, 'max': 2},
        'bom': {'min': 0.5, 'max': 1},
        'otimo': {'min': 0, 'max': 0.5},
        'descricao': 'Relação entre preço e receita. Indica valuation em relação às vendas.',
        'agrupador': 'Valuation',
        'descrpessimo': 'Valuation elevado, >3x (ex.: startups). Evitar.',
        'descrruim': 'Valuation alto, 2-3x (ex.: MGLU3). Cautela.',
        'descrmoderado': 'Valuation razoável, 1-2x (ex.: ABEV3). Justo.',
        'descrbom': 'Valuation atrativo, 0,5-1x (ex.: ITUB4). Ideal.',
        'descotimo': 'Valuation muito atrativo, <0,5x (ex.: VALE3). Oportunidade.'
    },
    'Giro do Ativo': {
        'pessimo': {'min': 0, 'max': 0.5},
        'ruim': {'min': 0.5, 'max': 1},
        'moderado': {'min': 1, 'max': 1.5},
        'bom': {'min': 1.5, 'max': 2},
        'otimo': {'min': 2, 'max': float('inf')},
        'descricao': 'Eficiência na geração de receita por ativo. Indica produtividade.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'Baixa eficiência, <0,5x (ex.: OIBR3). Evitar.',
        'descrruim': 'Eficiência limitada, 0,5-1x (ex.: CSNA3). Cautela.',
        'descrmoderado': 'Eficiência razoável, 1-1,5x (ex.: ABEV3). Justo.',
        'descrbom': 'Alta eficiência, 1,5-2x (ex.: ITUB4). Ideal.',
        'descotimo': 'Eficiência excepcional, >2x (ex.: WEGR3). Atraente.'
    },
    'Liquidez Corrente': {
        'pessimo': {'min': 0, 'max': 1},
        'ruim': {'min': 1, 'max': 1.2},
        'moderado': {'min': 1.2, 'max': 1.5},
        'bom': {'min': 1.5, 'max': 2},
        'otimo': {'min': 2, 'max': float('inf')},
        'descricao': 'Capacidade de cobrir passivos de curto prazo com ativos circulantes.',
        'agrupador': 'Liquidez',
        'descrpessimo': 'Baixa liquidez, <1 (ex.: OIBR3). Evitar.',
        'descrruim': 'Liquidez limitada, 1-1,2 (ex.: CSNA3). Cautela.',
        'descrmoderado': 'Liquidez razoável, 1,2-1,5 (ex.: ABEV3). Justo.',
        'descrbom': 'Boa liquidez, 1,5-2 (ex.: ITUB4). Ideal.',
        'descotimo': 'Liquidez excepcional, >2 (ex.: WEGR3). Atraente.'
    },
    'LPA': {
        'pessimo': {'min': float('-inf'), 'max': 0},
        'ruim': {'min': 0, 'max': 0.5},
        'moderado': {'min': 0.5, 'max': 1},
        'bom': {'min': 1, 'max': 2},
        'otimo': {'min': 2, 'max': float('inf')},
        'descricao': 'Lucro por ação. Indica rentabilidade por ação.',
        'agrupador': 'Rentabilidade',
        'descrpessimo': 'Prejuízo por ação (ex.: OIBR3). Evitar.',
        'descrruim': 'Lucro baixo, 0-0,5 (ex.: CSNA3). Cautela.',
        'descrmoderado': 'Lucro razoável, 0,5-1 (ex.: ABEV3). Justo.',
        'descrbom': 'Lucro sólido, 1-2 (ex.: ITUB4). Ideal.',
        'descotimo': 'Lucro excepcional, >2 (ex.: WEGR3). Atraente.'
    },
    'Margem Bruta': {
        'pessimo': {'min': 0, 'max': 10},
        'ruim': {'min': 10, 'max': 20},
        'moderado': {'min': 20, 'max': 30},
        'bom': {'min': 30, 'max': 50},
        'otimo': {'min': 50, 'max': float('inf')},
        'descricao': 'Percentual de lucro bruto sobre receita. Indica eficiência operacional.',
        'agrupador': 'Rentabilidade',
        'descrpessimo': 'Margem muito baixa, <10% (ex.: OIBR3). Evitar.',
        'descrruim': 'Margem limitada, 10-20% (ex.: CSNA3). Cautela.',
        'descrmoderado': 'Margem razoável, 20-30% (ex.: ABEV3). Justo.',
        'descrbom': 'Margem sólida, 30-50% (ex.: ITUB4). Ideal.',
        'descotimo': 'Margem excepcional, >50% (ex.: WEGR3). Atraente.'
    },
    'Margem EBIT': {
        'pessimo': {'min': 0, 'max': 5},
        'ruim': {'min': 5, 'max': 10},
        'moderado': {'min': 10, 'max': 15},
        'bom': {'min': 15, 'max': 25},
        'otimo': {'min': 25, 'max': float('inf')},
        'descricao': 'Percentual de lucro operacional sobre receita. Indica eficiência.',
        'agrupador': 'Rentabilidade',
        'descrpessimo': 'Margem muito baixa, <5% (ex.: OIBR3). Evitar.',
        'descrruim': 'Margem limitada, 5-10% (ex.: CSNA3). Cautela.',
        'descrmoderado': 'Margem razoável, 10-15% (ex.: ABEV3). Justo.',
        'descrbom': 'Margem sólida, 15-25% (ex.: ITUB4). Ideal.',
        'descotimo': 'Margem excepcional, >25% (ex.: WEGR3). Atraente.'
    },
    'Margem EBITDA': {
        'pessimo': {'min': 0, 'max': 10},
        'ruim': {'min': 10, 'max': 15},
        'moderado': {'min': 15, 'max': 20},
        'bom': {'min': 20, 'max': 30},
        'otimo': {'min': 30, 'max': float('inf')},
        'descricao': 'Percentual de EBITDA sobre receita. Indica geração de caixa operacional.',
        'agrupador': 'Rentabilidade',
        'descrpessimo': 'Margem muito baixa, <10% (ex.: OIBR3). Evitar.',
        'descrruim': 'Margem limitada, 10-15% (ex.: CSNA3). Cautela.',
        'descrmoderado': 'Margem razoável, 15-20% (ex.: ABEV3). Justo.',
        'descrbom': 'Margem sólida, 20-30% (ex.: ITUB4). Ideal.',
        'descotimo': 'Margem excepcional, >30% (ex.: WEGR3). Atraente.'
    },
    'Margem Líquida': {
        'pessimo': {'min': float('-inf'), 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Percentual de lucro líquido sobre receita. Indica rentabilidade final.',
        'agrupador': 'Rentabilidade',
        'descrpessimo': 'Prejuízo líquido, <0% (ex.: OIBR3). Evitar.',
        'descrruim': 'Margem baixa, 0-5% (ex.: CSNA3). Cautela.',
        'descrmoderado': 'Margem razoável, 5-10% (ex.: ABEV3). Justo.',
        'descrbom': 'Margem sólida, 10-20% (ex.: ITUB4). Ideal.',
        'descotimo': 'Margem excepcional, >20% (ex.: WEGR3). Atraente.'
    },
    'ROA': {
        'pessimo': {'min': float('-inf'), 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 15},
        'otimo': {'min': 15, 'max': float('inf')},
        'descricao': 'Retorno sobre ativos. Indica eficiência na utilização dos ativos.',
        'agrupador': 'Rentabilidade',
        'descrpessimo': 'Retorno negativo, <0% (ex.: OIBR3). Evitar.',
        'descrruim': 'Retorno baixo, 0-5% (ex.: CSNA3). Cautela.',
        'descrmoderado': 'Retorno razoável, 5-10% (ex.: ABEV3). Justo.',
        'descrbom': 'Retorno sólido, 10-15% (ex.: ITUB4). Ideal.',
        'descotimo': 'Retorno excepcional, >15% (ex.: WEGR3). Atraente.'
    },
    'ROE': {
        'pessimo': {'min': float('-inf'), 'max': 0},
        'ruim': {'min': 0, 'max': 10},
        'moderado': {'min': 10, 'max': 15},
        'bom': {'min': 15, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Retorno sobre patrimônio líquido. Indica rentabilidade para acionistas.',
        'agrupador': 'Rentabilidade',
        'descrpessimo': 'Retorno negativo, <0% (ex.: OIBR3). Evitar.',
        'descrruim': 'Retorno baixo, 0-10% (ex.: CSNA3). Cautela.',
        'descrmoderado': 'Retorno razoável, 10-15% (ex.: ABEV3). Justo.',
        'descrbom': 'Retorno sólido, 15-20% (ex.: ITUB4). Ideal.',
        'descotimo': 'Retorno excepcional, >20% (ex.: WEGR3). Atraente.'
    },
    'CAGR Lucros 5 anos': {
        'pessimo': {'min': float('-inf'), 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Crescimento anual composto dos lucros em 5 anos. Indica tendência de crescimento.',
        'agrupador': 'Rentabilidade',
        'descrpessimo': 'Crescimento negativo, <0% (ex.: OIBR3). Evitar.',
        'descrruim': 'Crescimento baixo, 0-5% (ex.: CSNA3). Cautela.',
        'descrmoderado': 'Crescimento razoável, 5-10% (ex.: ABEV3). Justo.',
        'descrbom': 'Crescimento sólido, 10-20% (ex.: ITUB4). Ideal.',
        'descotimo': 'Crescimento excepcional, >20% (ex.: WEGR3). Atraente.'
    },
    'ROIC': {
        'pessimo': {'min': float('-inf'), 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 15},
        'otimo': {'min': 15, 'max': float('inf')},
        'descricao': 'Retorno sobre capital investido. Indica eficiência do capital.',
        'agrupador': 'Rentabilidade',
        'descrpessimo': 'Retorno negativo, <0% (ex.: OIBR3). Evitar.',
        'descrruim': 'Retorno baixo, 0-5% (ex.: CSNA3). Cautela.',
        'descrmoderado': 'Retorno razoável, 5-10% (ex.: ABEV3). Justo.',
        'descrbom': 'Retorno sólido, 10-15% (ex.: ITUB4). Ideal.',
        'descotimo': 'Retorno excepcional, >15% (ex.: WEGR3). Atraente.'
    },
    'VPA': {
        'pessimo': {'min': float('-inf'), 'max': 0},
        'ruim': {'min': 0, 'max': 1},
        'moderado': {'min': 1, 'max': 5},
        'bom': {'min': 5, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Valor patrimonial por ação. Indica valor contábil por ação.',
        'agrupador': 'Rentabilidade',
        'descrpessimo': 'Valor patrimonial negativo (ex.: OIBR3). Evitar.',
        'descrruim': 'Valor patrimonial baixo, 0-1 (ex.: CSNA3). Cautela.',
        'descrmoderado': 'Valor patrimonial razoável, 1-5 (ex.: ABEV3). Justo.',
        'descrbom': 'Valor patrimonial sólido, 5-10 (ex.: ITUB4). Ideal.',
        'descotimo': 'Valor patrimonial excepcional, >10 (ex.: WEGR3). Atraente.'
    }
}


wbsaida = openpyxl.Workbook()


# define selenium webdriver options
options = webdriver.ChromeOptions()

# create selenium webdriver instance
driver = webdriver.Chrome(options=options)

def categorizar_valor(metrica, valor):
    try:
        if metrica not in MetricasStatus:
            return 'Métrica não reconhecida'
        valor2 = float(valor)
        #print("categoriza_valor", valor2)
        for categoria, limites in MetricasStatus[metrica].items():

            if categoria in ['descricao', 'agrupador','descrpessimo','descrruim','descrmoderado','descrbom','descotimo']:
                continue
            if limites['min'] <= valor2 < limites['max']:
                return categoria
        return 'Valor fora do alcance definido'
    except Exception as e:
        print("Debug - limites:", limites)
        print(f"Erro inesperado categorizar: {e}")
        print(f"Erro metrica: {e}")
        print("categorizar_valor - Erro" , metrica)
        print("categoriza_valor", valor2)
        print('categoriza_valor' ,stock)
    finally:
       # print("categorizar_valor - OK2")
       pass

def criaPlanilhaIndRentabilidade(wbsaida):
    wbsaida.create_sheet('IndiRentabilidade')
    IndiRentabilidade = wbsaida['IndiRentabilidade']
    IndiRentabilidade.append(
        ['Agrupador', 'Fonte', 'ATIVO', 'Indicador', 'Valor', 'Referencia', 'Pessimo','Ruim', 'Moderado', 'Bom', 'Otimo', 'Descrição'])

    for cell in IndiRentabilidade[1]:  # Apenas o cabeçalho
        cell.fill = filltitulo
        cell.font = font_branca

def tratamento(indicador):
    indicador2 = indicador

    try:
        if indicador2 in ["-", "--","--%"]:
            indicador2 = 0
        elif indicador2 is None or is_null_zero_or_spaces(indicador2):
            indicador2 = 0
        else:
            if indicador2 == "":  # Verificação adicionada para string vazia
                indicador2 = 0
            else:
                indicador2 = float(indicador2.strip('%')) / 100
        return indicador2

    except Exception as e:
        print(f"Erro inesperado tratamento : {e}", "metrica  ", metricasts, " indicador  ", indicador, " stock  ",stock)
        # print(metrica)  # Certifique-se de que metrica está definida
        # print(indicadortratado)  # Certifique-se de que indicadortratado está definida
        #print('tratamneto - erro', stock, "   ", metrica)


    finally:
       # print('tratamneto OK')
       pass
def tratamento3(indicador):
    indicador2 = indicador

    try:
        if indicador2 in ["-", "--","--%"]:
            indicador2 = 0
        elif indicador2 is None or is_null_zero_or_spaces(indicador2):
            indicador2 = 0
        else:
            if indicador2 == "":  # Verificação adicionada para string vazia
                indicador2 = 0
            else:
                indicador2 = float(indicador2)
        return indicador2

    except Exception as e:
        print(f"Erro inesperado tratamento 3 : {e}", " metrica  ", metricasts, " indicador  ", indicador, " stock  ",stock)
        # print(metrica)  # Certifique-se de que metrica está definida
        # print(indicadortratado)  # Certifique-se de que indicadortratado está definida
        #print('tratamneto3 - erro', stock, "   ", metrica)


    finally:
        #print('tratamneto3 OK')
        pass
# Certifique-se de que as variáveis `metrica` e `indicadortratado` estão definidas corretamente no contexto onde a função é chamada.

def tratamento2(indicador):
    indicador2 = indicador

    try:
        if indicador2 in ["-", "--","--%"]:
            indicador2 = 0
        elif indicador2 is None or is_null_zero_or_spaces(indicador2):
            indicador2 = 0
        elif indicador2 == "":  # Verificação adicionada para string vazia
            indicador2 = 0
        else:
            indicador2 = float(indicador2)
        return indicador2


        return indicador2
    except Exception as e:
      # print(f"Erro inesperado tratamento2: {e}", " metrica  ", metrica, " indicador  ", indicador ," stock ", stock)
        # print(metrica)  # Certifique-se de que metrica está definida
        # print(indicadortratado)  # Certifique-se de que indicadortratado está definida
        print('tratamneto2 - erro', stock)


    finally:
        #print('tratamneto2 OK', indicador)
        pass
def gravaIndiEficiênciaoStaus(wsIndiRentabilidade, dict_stocks, stock):
    # fontes ['StatusInvest', 'Fundamentus']



    global linha2
    global metricasts
    #linha2 = 1
    try:
        #print(dict_stocks)
        for metrica, detalhes in MetricasStatus.items():
    #        print(f'Métrica: {metrica}')
            linha2 += 1
            metricasts = metrica
            if metrica in ['Giro ativos', 'Div. liquida/PL','Div. liquida/EBITDA','Div. liquida/EBIT','PL/Ativos',
                           'Passivos/Ativos','Liq. corrente','P/L','PEG Ratio','P/VP','EV/EBITDA','EV/EBIT',
                            'P/EBITDA','P/EBIT','VPA','P/Ativo','LPA',
                            'P/SR','P/Ativo Circ. Liq.']:
             #   print('IF tratamneto2 ',metrica )

                indicadortratado = tratamento2(dict_stocks[stock].get(metrica))
                valor_pl = indicadortratado
                categoria_pl = categorizar_valor(metrica, (valor_pl))
              #  print("categoria " + categoria_pl)
               # print("categoria tratamento2 " + categoria_pl)
            elif metrica in ['Valor atual','LIQUIDEZ MEDIA DIARIA','Patrimonio liquido',
                             'Ativos','Ativo circulante','Divida bruta','Disponibilidade',
                             'Divida liquida','Valor de mercado','Valor de firma']:
               # print('IF tratamneto3 ', metrica)
               # print("categoria " + categoria_pl)
                indicadortratado = tratamento3(dict_stocks[stock].get(metrica))
                valor_pl = indicadortratado
                categoria_pl = categorizar_valor(metrica, (valor_pl))
               # print("categoria tratamento3" + categoria_pl)
            else:
               # print('IF tratamneto ', metrica)
                indicadortratado = tratamento(dict_stocks[stock].get(metrica))
                valor_pl = indicadortratado
                categoria_pl = categorizar_valor(metrica,(valor_pl * 100))
                #print("categoria tratamento" + categoria_pl)
                # Certifique-se de que 'ROE' é o valor correto para a métrica
   #         print(f'O índice P/L {valor_pl} é categorizado como: {categoria_pl}')
   #         print(f"  Agrupador: {detalhes['agrupador']}")


            wsIndiRentabilidade.cell(row=linha2, column=1, value=detalhes['agrupador'])
            wsIndiRentabilidade.cell(row=linha2, column=2, value='StausInvest')
            wsIndiRentabilidade.cell(row=linha2, column=3, value=stock)
           # print('celula - Indicador', metrica)
            wsIndiRentabilidade.cell(row=linha2, column=4, value=metrica)
            if metrica in ['Giro ativos', 'Div. liquida/PL','Div. liquida/EBITDA','Div. liquida/EBITDA',
                           'Div. liquida/EBIT','PL/Ativos','Passivos/Ativos','Liq. corrente',
                           'P/L','PEG Ratio','P/VP','EV/EBITDA','EV/EBIT',
                            'P/EBITDA','P/EBIT','VPA','P/Ativo','LPA',
                            'P/SR','P/Ativo Circ. Liq.']:
               # print('IF da celula - Indicador', metrica )
               # print('IF da celula - valor', valor_pl)
                wsIndiRentabilidade.cell(row=linha2, column=5, value=valor_pl).number_format = numbers.FORMAT_NUMBER_00
            elif  metrica in ['Valor atual','LIQUIDEZ MEDIA DIARIA','Patrimonio liquido',
                             'Ativos','Ativo circulante','Divida bruta','Disponibilidade',
                             'Divida liquida','Valor de mercado','Valor de firma']:
                  wsIndiRentabilidade.cell(row=linha2, column=5, value=valor_pl).number_format = 'R$ #,##0.00'
            else:
                wsIndiRentabilidade.cell(row=linha2, column=5, value=valor_pl).number_format = numbers.FORMAT_PERCENTAGE_00

            wsIndiRentabilidade.cell(row=linha2, column=7,
                             value=f"Mínimo = {detalhes['pessimo']['min']}, Máximo = {detalhes['pessimo']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=8,
                                     value=f"Mínimo = {detalhes['ruim']['min']}, Máximo = {detalhes['ruim']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=9,
                                     value=f"Mínimo = {detalhes['moderado']['min']}, Máximo = {detalhes['moderado']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=10,
                                     value=f"Mínimo = {detalhes['bom']['min']}, Máximo = {detalhes['bom']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=11,
                                     value=f"Mínimo = {detalhes['otimo']['min']}, Máximo = {detalhes['otimo']['max']}")
            #print(detalhes)

            #if metrica == 'Div. liquida/EBITDA':
            if categoria_pl == 'pessimo':
                wsIndiRentabilidade.cell(row=linha2, column=6, value=categoria_pl).fill = fillvermelho
                wsIndiRentabilidade.cell(row=linha2, column=11, value=f"{detalhes['descrpessimo']}")
            if categoria_pl == 'ruim':
                wsIndiRentabilidade.cell(row=linha2, column=6, value=categoria_pl).fill = fillvermelho
                wsIndiRentabilidade.cell(row=linha2, column=11, value=f"{detalhes['descrruim']}")
            if  categoria_pl == 'moderado':
                wsIndiRentabilidade.cell(row=linha2, column=6, value=categoria_pl).fill =fillamarelo
                wsIndiRentabilidade.cell(row=linha2, column=11, value=f"{detalhes['descrmoderado']}")
            if  categoria_pl == 'bom':
                wsIndiRentabilidade.cell(row=linha2, column=6, value=categoria_pl).fill = fillverde
                wsIndiRentabilidade.cell(row=linha2, column=11, value=f"{detalhes['descrbom']}")
            if categoria_pl == 'otimo':
                wsIndiRentabilidade.cell(row=linha2, column=6, value=categoria_pl).fill =fillazul
                wsIndiRentabilidade.cell(row=linha2, column=11, value=f"{detalhes['descotimo']}")


    except Exception as e:
        print(f"Erro inesperado grava planilha2: {e}")
        print(metrica)
        print(indicadortratado)

        print('gravaIndiEficiênciaoStaus - erro' ,  stock,"    ", metrica)
    finally:
        print('gravaIndiEficiênciaoStaus  OK''', stock)

def is_null_zero_or_spaces(variable):
    # Verifica se a variável é None
    if variable is None:
        return True
    # Verifica se a variável é zero (0)
    elif variable == 0:
        return True
    # Verifica se a variável é uma string e contém apenas espaços
    elif isinstance(variable, str) and variable.strip() == '':
        return True
    elif variable == '-%':
        return True
    else:
        return False


def get_stock_soup(stock):
    ''' Get raw html from a stock '''

    # access the stock urlww
    driver.get(f'https://statusinvest.com.br/acoes/{stock}')

    # get html from stock
    html = driver.find_element(By.ID, 'main-2').get_attribute('innerHTML')

    # remove accents from html and transform html into soup
    soup = BeautifulSoup(unidecode(html), 'html.parser')

    return soup


def soup_to_dict(soup):
    '''Get all data from stock soup and return as a dictionary '''
    keys, values = [], []

    # get divs from stock
    soup1 = soup.find('div', class_='pb-3 pb-md-5')
    soup2 = soup.find('div', class_='card rounded text-main-green-dark')
    soup3 = soup.find('div', class_='indicator-today-container')
    soup4 = soup.find(
        'div', class_='top-info info-3 sm d-flex justify-between mb-3')
    soups = [soup1, soup2, soup3, soup4]

    for s in soups:
        # get only titles from a div and append to keys
        titles = s.find_all('h3', re.compile('title m-0[^"]*'))
        titles = [t.get_text() for t in titles]
        keys += titles

        # get only numbers from a div and append to values
        numbers = s.find_all('strong', re.compile('value[^"]*'))
        numbers = [n.get_text()for n in numbers]
        values += numbers

    # remove unused key and insert needed keys
    keys.remove('PART. IBOV')
    keys.insert(6, 'TAG ALONG')
    keys.insert(7, 'LIQUIDEZ MEDIA DIARIA')

    # clean keys list
    keys = [k.replace('\nhelp_outline', '').strip() for k in keys]
    keys = [k for k in keys if k != '']

    # clean values list
    values = [v.replace('\nhelp_outline', '').strip() for v in values]
    values = [v.replace('.', '').replace(',', '.') for v in values]

    # create a dictionary using keys and values from indicators
    d = {k: v for k, v in zip(keys, values)}

    return d


if __name__ == "__main__":
    dict_stocks = {}
    criaPlanilhaIndRentabilidade(wbsaida)
    wsIndiRentabilidade = wbsaida['IndiRentabilidade']

    # start timer
    start = time.time()

    # read file with stocks codes to get stock information
    with open('stocks.txt', 'r') as f:
        stocks = f.read().splitlines()

        # get stock information and create excel sheet
        for stock in stocks:
            #print("stock :"  ,stock)
            try:
                # get data and transform into dictionary
                soup = get_stock_soup(stock)
                dict_stock = soup_to_dict(soup)
                dict_stocks[stock] = dict_stock
                gravaIndiEficiênciaoStaus(wsIndiRentabilidade, dict_stocks, stock)
            except:
                # if we not get the information... just skip it
                print(f'Could not get {stock} information', "    ", metricasts)

    # create dataframe using dictionary of stocks informations
    df = pd.DataFrame(dict_stocks)

    # replace missing values with NaN to facilitate processing
    df = df.replace(['', '-', '--', '-%', '--%'], np.nan)

    # write dataframe into csv file
    df.to_excel('stocks_data.xlsx', index_label='indicadores')

    # exit the driver
    driver.quit()

    # end timer
    end = time.time()
    wbsaida.save("StatusInvest.xlsx")
    print("teste")
    print(f'Brasilian stocks information got in {int(end-start)} s')
# silvio teste