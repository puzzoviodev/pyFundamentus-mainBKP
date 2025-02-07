#!/usr/bin/env python
# encoding: utf-8

# ------------------------------------------------------------------------------
#  Name: run.py
#  Version: 0.0.1
#
#  Summary: Python Fundamentus
#           Python Fundamentus is a Python API that allows you to quickly
#           access the main fundamental indicators of the main stocks
#           in the Brazilian market.
#
#  Author: Alexsander Lopes Camargos
#  Author-email: alcamargos@vivaldi.net
#
#  License: MIT
# ------------------------------------------------------------------------------

"""
Python Fundamentus API: Instant access to key financial indicators of
Brazilian stocks, empowering investors with comprehensive market analysis.
"""

import fundamentus

# statusinvest

import openpyxl
import re
import time
import numpy as np
import pandas as pd
from unidecode import unidecode
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
import requests

TITLES = [
    'Identificação', 'Resumo Financeiro', 'Cotações', 'Informações Básicas',
    'Oscilações', 'Indicadores de Valuation', 'Indicadores de Rentabilidade',
    'Indicadores de Endividamento', 'Balanço Patrimonial', 'Demonstrativo de Resultados'
]

wbsaida = openpyxl.Workbook()
Dicrentabilidade = {}
Dicstock_identification  = {}
Dicfinancial_summary = {}
Dicprice_information = {}
Dicdetailed_information = {}
Dicoscillations = {}
Dicprofitability_indicators = {}
Dicindebtedness_indicators = {}
Dicbalance_sheet = {}
Dicincome_statement = {}
Dicvaluation_indicators = {}



def criaPlanilhaIndiEficiência(IndiEficiência):
    try:
       wbsaida.create_sheet('IndiEficiência')
       IndiEficiência = wbsaida['IndiEficiência']
       IndiEficiência.append(['Fonte','ATIVO','M. Bruta', 'M. EBITDA', 'M. EBIT', 'M. Líquida'])
    except Exception as e:
        print(f"Erro inesperado: {e}")
    finally:
        print("criaPlanilhaIndiEficiência")
    return

def criaPlanilhaIndRentabilidade(IndiRentabilidade):
    wbsaida.create_sheet('IndiRentabilidade')
    IndiRentabilidade = wbsaida['IndiRentabilidade']
    IndiRentabilidade.append(['Fonte','ATIVO','ROE', 'ROA', 'ROIC','Giro ativos',''])
    return

def criaPlanilhaIndEndividamento(IndEndividamento):
    wbsaida.create_sheet('IndEndividamento')
    IndEndividamento = wbsaida['IndEndividamento']
    IndEndividamento.append(['Fonte','ATIVO','Dív. líquida/PL', 'Dív. líquida/EBITDA', 'Dív. líquida/EBIT','PL/Ativos','Passivos/Ativos','Liq. corrente'])
    return
def criaPlanilaIndValuation(wbsaida):
    wbsaida.create_sheet('IndValuation')
    IndValuation = wbsaida['IndValuation']
    IndValuation.append(['Fonte','ATIVO','D.Y', 'P/L', ' PEG Ratio','P/VP','EV/EBITDA','EV/EBIT','P/EBITDA','P/EBIT','VPA','P/Ativo',
                         'LPA','P/SR','P/Ativo Circ. Liq.'])
def criaPlanilhaIndiEmpresa(wbsaida):
    wbsaida.create_sheet('IndEmpresa')
    IndEmpresa = wbsaida['IndEmpresa']
    IndEmpresa.append(['Fonte',
                        "ATIVO",
                        "Valor atual",
                         "Min. 52 semanas",
                          "Max. 52 semanas",
                          "dividend Yield",
                          "Valorizacao (12m)",
                          "Tipo",
                          "TAG ALONG",
                          "LIQUIDEZ MEDIA DIARIA",
                          "PARTICIPACAO NO IBOV",
                          "MERCADO DE OPCOES",
                          "Patrimonio liquido",
                          "Ativos",
                          "Ativo circulante",
                          "Divida bruta",
                          "Disponibilidade",
                          "Divida liquida",
                          "Valor de mercado",
                          "Valor de firma",
                          'Free Float'])
    return

def montadicionario(stock_identification,financial_summary,price_information,detailed_information,oscillations,
                    valuation_indicators,profitability_indicators,indebtedness_indicators,balance_sheet,income_statement):


    try:


        for information in stock_identification:
            Dicstock_identification[stock_identification[information].title] = stock_identification[information].value
        for information in financial_summary:
            Dicfinancial_summary[financial_summary[information].title] = financial_summary[information].value
        for information in price_information:
            Dicprice_information[price_information[information].title] = price_information[information].value
        for information in detailed_information:
            if information != 'variation_52_weeks':
               Dicdetailed_information[detailed_information[information].title] = detailed_information[information].value
            else:
                for sub_information in detailed_information[information]:
                   Dicdetailed_information[detailed_information[information][sub_information].title] = detailed_information[information][sub_information].value
        for information in oscillations:
            Dicoscillations[oscillations[information].title] = oscillations[information].value
        for information in valuation_indicators:
            Dicvaluation_indicators[valuation_indicators[information].title] = valuation_indicators[information].value
        for information in profitability_indicators:
            Dicprofitability_indicators[profitability_indicators[information].title] = profitability_indicators[information].value
        for information in indebtedness_indicators:
            Dicindebtedness_indicators[indebtedness_indicators[information].title] = indebtedness_indicators[information].value
        for information in balance_sheet:
            Dicbalance_sheet[balance_sheet[information].title] = balance_sheet[information].value
        #for information in income_statement['three_months']:
         #   Dicincome_statement[income_statement[information].title] = income_statement[information].value
        #for information in income_statement['twelve_months']:
         #   Dicincome_statement[income_statement[information].title] = income_statement[information].value

    except Exception as e:
        print(f"Erro inesperado: {e}")
    finally:
        print("montadicionario")
    return

def teste():
    fontes = ['StatusInvest', 'Fundamentus']


    try:
         for fonte in fontes:
             print(fonte)
    except Exception as e:
            print(f"Erro inesperado: {e}")
    finally:
            print("teste")
    return


def gravaIndiEficiênciaoStaus(wsIndiEficiência,Fonte, linha,dict_stocks,stock ):
    #fontes ['StatusInvest', 'Fundamentus']
    try:

        MBruta = dict_stocks[stock].get("M. Bruta")
        MEBITDA = dict_stocks[stock].get("M. EBITDA")
        MEBIT = dict_stocks[stock].get("M. EBIT")
        MLiquida = dict_stocks[stock].get("M. Liquida")
        ATIVO =stock

        #MBruta = f"{float(Dicrentabilidade.get("Margem bruta")) * 100}%"
        #MEBITDA = ''
        #MEBIT = f"{float(Dicrentabilidade.get("Margem EBIT")) * 100}%"
        #MLiquida = f"{float(Dicrentabilidade.get("Margem líquida")) * 100}%"

       # Condicional corrigida
        if is_null_zero_or_spaces(MBruta):
            MBruta   = 0
        else:
           MBruta = float(MBruta.strip('%')) /100
        if is_null_zero_or_spaces(MEBITDA):
           MEBITDA =0
        else:
            MEBITDA = float(MEBITDA.strip('%')) / 100


        if is_null_zero_or_spaces(MEBIT):
           MEBIT = 0
        else:

          MEBIT = float(MEBIT.strip('%')) / 100

        if is_null_zero_or_spaces(MLiquida):
           MLiquida =0
        else:
            MLiquida = float(MLiquida.strip('%')) / 100

        wsIndiEficiência.cell(row=linha, column=1, value=Fonte)
        wsIndiEficiência.cell(row=linha, column=2, value=ATIVO)
        wsIndiEficiência.cell(row=linha, column=3, value=MBruta)
        wsIndiEficiência.cell(row=linha, column=4, value=MEBITDA)
        wsIndiEficiência.cell(row=linha, column=5, value=MEBIT)
        wsIndiEficiência.cell(row=linha, column=6, value=MLiquida)
    except Exception as e:
        print(f"Erro inesperado: {e}")
    finally:
        print("gravaIndiEficiênciaoStaus")

def gravaIndiEficiênciaoFund(wsIndiEficiência,Fonte, linha,Dicprofitability_indicators,stock ):
    #fontes ['StatusInvest', 'Fundamentus']
    try:

        ATIVO = stock

        MBruta = f"{float(Dicprofitability_indicators.get("Margem bruta")) * 100}%"
        MEBITDA = ''
        MEBIT = f"{float(Dicprofitability_indicators.get("Margem EBIT")) * 100}%"
        MLiquida = f"{float(Dicprofitability_indicators.get("Margem líquida")) * 100}%"

       # Condicional corrigida
        if is_null_zero_or_spaces(MBruta):
            MBruta   = 0
        else:
           MBruta = float(MBruta.strip('%')) /100
        if is_null_zero_or_spaces(MEBITDA):
           MEBITDA =0
        else:
            MEBITDA = float(MEBITDA.strip('%')) / 100

        if is_null_zero_or_spaces(MEBIT):
           MEBIT = 0
        else:

          MEBIT = float(MEBIT.strip('%')) / 100

        if is_null_zero_or_spaces(MLiquida):
           MLiquida =0
        else:
            MLiquida = float(MLiquida.strip('%')) / 100

        wsIndiEficiência.cell(row=linha, column=1, value=Fonte)
        wsIndiEficiência.cell(row=linha, column=2, value=ATIVO)
        wsIndiEficiência.cell(row=linha, column=3, value=MBruta)
        wsIndiEficiência.cell(row=linha, column=4, value=MEBITDA)
        wsIndiEficiência.cell(row=linha, column=5, value=MEBIT)
        wsIndiEficiência.cell(row=linha, column=6, value=MLiquida)
    except Exception as e:
        print(f"Erro inesperado: {e}")
    finally:
        print("gravaIndiEficiênciaoFund")

def gravaIndiRentabilidadeStaus(wsIndiRentabilidade,Fonte,linha,dict_stocks,stock):
    # Condicional corrigida
    ATIVO = stock
    ROE = dict_stocks[stock].get("ROE")
    ROA = dict_stocks[stock].get("ROA")
    ROIC = dict_stocks[stock].get("ROIC")
    Giroativos = dict_stocks[stock].get("Giro ativos")
    try:
        Fonte = Fonte
        if is_null_zero_or_spaces(ROIC):
            ROIC = 0
        else:
            ROIC = float(ROIC.strip('%')) / 100

        if is_null_zero_or_spaces(ROE):
            ROE = 0
        else:
            ROE = float(ROE.strip('%')) / 100

        if   is_null_zero_or_spaces(ROA) :
            ROA = 0
        else:
            ROA = float(ROA.strip('%')) / 100

        wsIndiRentabilidade.cell(row=linha, column=1, value=Fonte)
        wsIndiRentabilidade.cell(row=linha, column=2, value=ATIVO)
        wsIndiRentabilidade.cell(row=linha, column=3, value=ROE)
        wsIndiRentabilidade.cell(row=linha, column=4, value=ROA)
        wsIndiRentabilidade.cell(row=linha, column=5, value=ROIC)
        wsIndiRentabilidade.cell(row=linha, column=6, value=Giroativos)
    except Exception as e:
        print(f"Erro inesperado: {e}")
    finally:
        print("gravaIndiRentabilidadeStaus")

        # Condicional corrigida
        ATIVO = stock
        ROE = dict_stocks[stock].get("ROE")
        ROA = dict_stocks[stock].get("ROA")
        ROIC = dict_stocks[stock].get("ROIC")
        Giroativos = dict_stocks[stock].get("Giro ativos")
        try:
            Fonte = Fonte
            if is_null_zero_or_spaces(ROIC):
                ROIC = 0
            else:
                ROIC = float(ROIC.strip('%')) / 100

            if is_null_zero_or_spaces(ROE):
                ROE = 0
            else:
                ROE = float(ROE.strip('%')) / 100

            if is_null_zero_or_spaces(ROA):
                ROA = 0
            else:
                ROA = float(ROA.strip('%')) / 100

            wsIndiRentabilidade.cell(row=linha, column=1, value=Fonte)
            wsIndiRentabilidade.cell(row=linha, column=2, value=ATIVO)
            wsIndiRentabilidade.cell(row=linha, column=3, value=ROE)
            wsIndiRentabilidade.cell(row=linha, column=4, value=ROA)
            wsIndiRentabilidade.cell(row=linha, column=5, value=ROIC)
            wsIndiRentabilidade.cell(row=linha, column=6, value=Giroativos)
        except Exception as e:
            print(f"Erro inesperado: {e}")
        finally:
            print("gravaIndiRentabilidadeStaus")


def gravaIndiRentabilidadeFund(wsIndiRentabilidade,Fonte, linha,Dicprofitability_indicators,stock):
    # Condicional corrigida

    ATIVO = stock

    ROE = f"{float(Dicprofitability_indicators.get("ROE")) * 100}%"
    ROA = ''
    ROIC = f"{float(Dicprofitability_indicators.get("ROIC")) * 100}%"
    Giroativos = f"{float(Dicprofitability_indicators.get("Giro ativos")) * 100}%"
    try:
        Fonte = Fonte
        if is_null_zero_or_spaces(ROIC):
            ROIC = 0
        else:
            ROIC = float(ROIC.strip('%')) / 100

        if is_null_zero_or_spaces(ROE):
            ROE = 0
        else:
            ROE = float(ROE.strip('%')) / 100

        if is_null_zero_or_spaces(ROA):
            ROA = 0
        else:
            ROA = float(ROA.strip('%')) / 100

        wsIndiRentabilidade.cell(row=linha, column=1, value=Fonte)
        wsIndiRentabilidade.cell(row=linha, column=2, value=ATIVO)
        wsIndiRentabilidade.cell(row=linha, column=3, value=ROE)
        wsIndiRentabilidade.cell(row=linha, column=4, value=ROA)
        wsIndiRentabilidade.cell(row=linha, column=5, value=ROIC)
        wsIndiRentabilidade.cell(row=linha, column=6, value=Giroativos)
    except Exception as e:
        print(f"Erro inesperado: {e}")
    finally:
        print("gravaIndiRentabilidadeFund")





def gravaIndEndividamentoStaus(wsIndEndividamento, Fonte,linha,dict_stocks,stock):
    ATIVO = stock
    DivliquidaPL = dict_stocks[stock].get("Div. liquida/PL")
    DivliquidaEBITDA = dict_stocks[stock].get("Div. liquida/EBITDA")
    DivliquidaEBIT = dict_stocks[stock].get("Div. liquida/EBIT")
    PLAtivos = dict_stocks[stock].get("PL/Ativos")
    PassivosAtivos = dict_stocks[stock].get("Passivos/Ativos")
    Liqcorrente = dict_stocks[stock].get("Liq. corrente")

    try:
        if DivliquidaPL == "-":
            DivliquidaPL = ""

        if DivliquidaEBITDA == "-":
            DivliquidaEBITDA = ""

        if DivliquidaEBIT == "-":
            DivliquidaEBIT = ""

        if PLAtivos == "-":
            PLAtivos = ""

        if PassivosAtivos == "-":
            PassivosAtivos = ""

        if Liqcorrente == "-":
            Liqcorrente = ""

        if is_null_zero_or_spaces(DivliquidaPL):
           DivliquidaPL =0
        else:
           DivliquidaPL = float(DivliquidaPL.strip('%')) / 100

        if is_null_zero_or_spaces(DivliquidaEBITDA):
           DivliquidaEBITDA =0
        else:
           DivliquidaEBITDA = float(DivliquidaEBITDA.strip('%')) / 100

        if is_null_zero_or_spaces(DivliquidaEBIT):
            DivliquidaEBIT = 0
        else:
            DivliquidaEBIT = float(DivliquidaEBIT.strip('%')) / 100

        if is_null_zero_or_spaces(PLAtivos):
            PLAtivos = 0
        else:
            PLAtivos = float(PLAtivos.strip('%')) / 100

        if is_null_zero_or_spaces(PassivosAtivos):
            PassivosAtivos = 0
        else:
            PassivosAtivos = float(PassivosAtivos.strip('%')) / 100

        if is_null_zero_or_spaces(Liqcorrente):
            Liqcorrente = 0
        else:
            Liqcorrente = float(Liqcorrente.strip('%')) / 100

        wsIndEndividamento.cell(row=linha, column=1, value=Fonte)
        wsIndEndividamento.cell(row=linha, column=2, value=ATIVO)
        wsIndEndividamento.cell(row=linha, column=3, value=DivliquidaPL)
        wsIndEndividamento.cell(row=linha, column=4, value=DivliquidaEBITDA)
        wsIndEndividamento.cell(row=linha, column=5, value=DivliquidaEBIT)
        wsIndEndividamento.cell(row=linha, column=6, value=PLAtivos)
        wsIndEndividamento.cell(row=linha, column=7, value=PassivosAtivos)
        wsIndEndividamento.cell(row=linha, column=8, value=Liqcorrente)
    except Exception as e:
        print(f"Erro inesperado: {e}")
    finally:
        print("gravaIndEndividamentoStaus")


def gravaIndEndividamentosFund(wsIndEndividamento,Fonte, linha,Dicindebtedness_indicators,stock):
    ATIVO = stock

    DivliquidaPL = f"{float(Dicindebtedness_indicators.get("Dívida líquida/Patrim")) * 100}%"
    DivliquidaEBITDA = f"{float(Dicindebtedness_indicators.get("Dívida líquida/EBITDA")) * 100}%"
    DivliquidaEBIT = ""
    PLAtivos = f"{float(Dicindebtedness_indicators.get("PL/Ativos")) * 100}%"
    PassivosAtivos = ""
    Liqcorrente = f"{float(Dicindebtedness_indicators.get("Liquidez corrente")) * 100}%"

    try:
        if DivliquidaPL == "-":
            DivliquidaPL = ""

        if DivliquidaEBITDA == "-":
            DivliquidaEBITDA = ""

        if DivliquidaEBIT == "-":
            DivliquidaEBIT = ""

        if PLAtivos == "-":
            PLAtivos = ""

        if PassivosAtivos == "-":
            PassivosAtivos = ""

        if Liqcorrente == "-":
            Liqcorrente = ""

        if is_null_zero_or_spaces(DivliquidaPL):
           DivliquidaPL =0
        else:
           DivliquidaPL = float(DivliquidaPL.strip('%')) / 100

        if is_null_zero_or_spaces(DivliquidaEBITDA):
           DivliquidaEBITDA =0
        else:
           DivliquidaEBITDA = float(DivliquidaEBITDA.strip('%')) / 100

        if is_null_zero_or_spaces(DivliquidaEBIT):
            DivliquidaEBIT = 0
        else:
            DivliquidaEBIT = float(DivliquidaEBIT.strip('%')) / 100

        if is_null_zero_or_spaces(PLAtivos):
            PLAtivos = 0
        else:
            PLAtivos = float(PLAtivos.strip('%')) / 100

        if is_null_zero_or_spaces(PassivosAtivos):
            PassivosAtivos = 0
        else:
            PassivosAtivos = float(PassivosAtivos.strip('%')) / 100

        if is_null_zero_or_spaces(Liqcorrente):
            Liqcorrente = 0
        else:
            Liqcorrente = float(Liqcorrente.strip('%')) / 100

        wsIndEndividamento.cell(row=linha, column=1, value=Fonte)
        wsIndEndividamento.cell(row=linha, column=2, value=ATIVO)
        wsIndEndividamento.cell(row=linha, column=3, value=DivliquidaPL)
        wsIndEndividamento.cell(row=linha, column=4, value=DivliquidaEBITDA)
        wsIndEndividamento.cell(row=linha, column=5, value=DivliquidaEBIT)
        wsIndEndividamento.cell(row=linha, column=6, value=PLAtivos)
        wsIndEndividamento.cell(row=linha, column=7, value=PassivosAtivos)
        wsIndEndividamento.cell(row=linha, column=8, value=Liqcorrente)
    except Exception as e:
        print(f"Erro inesperado: {e}")
    finally:
        print("gravaIndEndividamentosFund")

def gravaIndValuationStaus(wsIndValuation, Fonte,linha,dict_stocks,stock):
    ATIVO = stock

    DY = dict_stocks[stock].get("D.Y")
    PL = dict_stocks[stock].get("P/L")
    PEGRatio = dict_stocks[stock].get("PEG Ratio")
    PVP = dict_stocks[stock].get("P/VP")
    EVEBITDA = dict_stocks[stock].get("EV/EBITDA")
    EVEBIT = dict_stocks[stock].get("EV/EBIT")
    PEBITDA = dict_stocks[stock].get("P/EBITDA")
    PEBIT = dict_stocks[stock].get("P/EBIT")
    VPA = dict_stocks[stock].get("VPA")
    PAtivo = dict_stocks[stock].get("P/Ativo")
    LPA = dict_stocks[stock].get("LPA")
    PSR = dict_stocks[stock].get("P/SR")
    PCapGiro = dict_stocks[stock].get("P/Cap. Giro")
    PAtivoCircLiq = dict_stocks[stock].get("P/Ativo Circ. Liq.")
    try:

        if DY == "-":
            DY = ""

        if PL == "-":
            PL = ""

        if PEGRatio == "-":
            PEGRatio = ""

        if PVP == "-":
            PVP = ""

        if EVEBITDA == "-":
            EVEBITDA = ""

        if EVEBIT == "-":
            EVEBIT = ""

        if PEBITDA == "-":
            PEBITDA = ""

        if PEBIT == "-":
            PEBIT = ""

        if VPA == "-":
            VPA = ""

        if PAtivo == "-":
            PAtivo = ""

        if LPA == "-":
            LPA = ""

        if PSR == "-":
            PSR = ""

        if PCapGiro == "-":
            PCapGiro = ""

        if PAtivoCircLiq == "-":
            PAtivoCircLiq = ""

        if is_null_zero_or_spaces(DY):
            DY = 0
        else:
            DY = float(DY.strip('%')) / 100

        if is_null_zero_or_spaces(PL):
            PL = 0
        else:
            PL = float(PL.strip('%')) / 100

        if is_null_zero_or_spaces(PEGRatio):
            PEGRatio = 0
        else:
            PEGRatio = float(PEGRatio.strip('%')) / 100

        if is_null_zero_or_spaces(PVP):
            PVP = 0
        else:
            PVP = float(PVP.strip('%')) / 100

        if is_null_zero_or_spaces(EVEBITDA):
            EVEBITDA = 0
        else:
            EVEBITDA = float(EVEBITDA.strip('%')) / 100

        if is_null_zero_or_spaces(EVEBIT):
            EVEBIT = 0
        else:
            EVEBIT = float(EVEBIT.strip('%')) / 100

        if is_null_zero_or_spaces(PEBITDA):
            PEBITDA = 0
        else:
            PEBITDA = float(PEBITDA.strip('%')) / 100

        if is_null_zero_or_spaces(PEBIT):
            PEBIT = 0
        else:
            PEBIT = float(PEBIT.strip('%')) / 100

        if is_null_zero_or_spaces(VPA):
            VPA = 0
        else:
            VPA = float(VPA.strip('%')) / 100

        if is_null_zero_or_spaces(LPA):
            LPA = 0
        else:
            LPA = float(LPA.strip('%')) / 100

        if is_null_zero_or_spaces(PSR):
            PSR = 0
        else:
            PSR = float(PSR.strip('%')) / 100

        if is_null_zero_or_spaces(PCapGiro):
            PCapGiro = 0
        else:
            PCapGiro = float(PCapGiro.strip('%')) / 100

        if is_null_zero_or_spaces(PAtivoCircLiq):
            PAtivoCircLiq = 0
        else:
            PAtivoCircLiq = float(PAtivoCircLiq.strip('%')) / 100


        wsIndValuation.cell(row=linha, column=1, value=Fonte)
        wsIndValuation.cell(row=linha, column=2, value=ATIVO)
        wsIndValuation.cell(row=linha, column=3, value=DY)
        wsIndValuation.cell(row=linha, column=4, value=PL)
        wsIndValuation.cell(row=linha, column=5, value=PEGRatio)
        wsIndValuation.cell(row=linha, column=6, value=PVP)
        wsIndValuation.cell(row=linha, column=7, value=EVEBITDA)
        wsIndValuation.cell(row=linha, column=8, value=EVEBIT)
        wsIndValuation.cell(row=linha, column=9, value=PEBITDA)
        wsIndValuation.cell(row=linha, column=10, value=PEBIT)
        wsIndValuation.cell(row=linha, column=11, value=VPA)
        wsIndValuation.cell(row=linha, column=12, value=PAtivo)
        wsIndValuation.cell(row=linha, column=13, value=LPA)
        wsIndValuation.cell(row=linha, column=14, value=PSR)
        wsIndValuation.cell(row=linha, column=15, value=PCapGiro)
        wsIndValuation.cell(row=linha, column=16, value=PAtivoCircLiq)
    except Exception as e:
        print(f"Erro inesperado: {e}")
    finally:
        print("gravaIndValuationStaus")


def gravaIndValuationFund(wsIndValuation, Fonte,linha,Dicvaluation_indicators,Dicdetailed_information,stock):
    ATIVO = stock

    DY = f"{float(Dicvaluation_indicators.get("Dividend Yield")) * 100}%"
    PL = f"{float(Dicvaluation_indicators.get("P/L")) * 100}%"
    PEGRatio = ""
    PVP = f"{float(Dicvaluation_indicators.get("P/VP")) * 100}%"
    EVEBITDA = f"{float(Dicvaluation_indicators.get("EV/EBITDA")) * 100}%"
    EVEBIT = f"{float(Dicvaluation_indicators.get("EV/EBIT")) * 100}%"
    PEBITDA = ""
    PEBIT = f"{float(Dicvaluation_indicators.get("P/EBIT")) * 100}%"
    VPA = f"{float(Dicdetailed_information.get("VPA")) * 100}%"
    PAtivo = f"{float(Dicvaluation_indicators.get("Preço/Ativos")) * 100}%"
    LPA = f"{float(Dicdetailed_information.get("LPA")) * 100}%"
    PSR = f"{float(Dicvaluation_indicators.get("PSR")) * 100}%"
    PCapGiro = f"{float(Dicvaluation_indicators.get("Preço/Capital de giro")) * 100}%"
    PAtivoCircLiq = f"{float(Dicvaluation_indicators.get("Preço/Ativ circ liq")) * 100}%"
    try:

        if DY == "-":
            DY = ""

        if PL == "-":
            PL = ""

        if PEGRatio == "-":
            PEGRatio = ""

        if PVP == "-":
            PVP = ""

        if EVEBITDA == "-":
            EVEBITDA = ""

        if EVEBIT == "-":
            EVEBIT = ""

        if PEBITDA == "-":
            PEBITDA = ""

        if PEBIT == "-":
            PEBIT = ""

        if VPA == "-":
            VPA = ""

        if PAtivo == "-":
            PAtivo = ""

        if LPA == "-":
            LPA = ""

        if PSR == "-":
            PSR = ""

        if PCapGiro == "-":
            PCapGiro = ""

        if PAtivoCircLiq == "-":
            PAtivoCircLiq = ""

        if is_null_zero_or_spaces(DY):
            DY = 0
        else:
            DY = float(DY.strip('%')) / 100

        if is_null_zero_or_spaces(PL):
            PL = 0
        else:
            PL = float(PL.strip('%')) / 100

        if is_null_zero_or_spaces(PEGRatio):
            PEGRatio = 0
        else:
            PEGRatio = float(PEGRatio.strip('%')) / 100

        if is_null_zero_or_spaces(PVP):
            PVP = 0
        else:
            PVP = float(PVP.strip('%')) / 100

        if is_null_zero_or_spaces(EVEBITDA):
            EVEBITDA = 0
        else:
            EVEBITDA = float(EVEBITDA.strip('%')) / 100

        if is_null_zero_or_spaces(EVEBIT):
            EVEBIT = 0
        else:
            EVEBIT = float(EVEBIT.strip('%')) / 100

        if is_null_zero_or_spaces(PEBITDA):
            PEBITDA = 0
        else:
            PEBITDA = float(PEBITDA.strip('%')) / 100

        if is_null_zero_or_spaces(PEBIT):
            PEBIT = 0
        else:
            PEBIT = float(PEBIT.strip('%')) / 100

        if is_null_zero_or_spaces(VPA):
            VPA = 0
        else:
            VPA = float(VPA.strip('%')) / 100

        if is_null_zero_or_spaces(LPA):
            LPA = 0
        else:
            LPA = float(LPA.strip('%')) / 100

        if is_null_zero_or_spaces(PSR):
            PSR = 0
        else:
            PSR = float(PSR.strip('%')) / 100

        if is_null_zero_or_spaces(PCapGiro):
            PCapGiro = 0
        else:
            PCapGiro = float(PCapGiro.strip('%')) / 100

        if is_null_zero_or_spaces(PAtivoCircLiq):
            PAtivoCircLiq = 0
        else:
            PAtivoCircLiq = float(PAtivoCircLiq.strip('%')) / 100


        wsIndValuation.cell(row=linha, column=1, value=Fonte)
        wsIndValuation.cell(row=linha, column=2, value=ATIVO)
        wsIndValuation.cell(row=linha, column=3, value=DY)
        wsIndValuation.cell(row=linha, column=4, value=PL)
        wsIndValuation.cell(row=linha, column=5, value=PEGRatio)
        wsIndValuation.cell(row=linha, column=6, value=PVP)
        wsIndValuation.cell(row=linha, column=7, value=EVEBITDA)
        wsIndValuation.cell(row=linha, column=8, value=EVEBIT)
        wsIndValuation.cell(row=linha, column=9, value=PEBITDA)
        wsIndValuation.cell(row=linha, column=10, value=PEBIT)
        wsIndValuation.cell(row=linha, column=11, value=VPA)
        wsIndValuation.cell(row=linha, column=12, value=PAtivo)
        wsIndValuation.cell(row=linha, column=13, value=LPA)
        wsIndValuation.cell(row=linha, column=14, value=PSR)
        wsIndValuation.cell(row=linha, column=15, value=PCapGiro)
        wsIndValuation.cell(row=linha, column=16, value=PAtivoCircLiq)
    except Exception as e:
        print(f"Erro inesperado: {e}")
    finally:
        print("gravaIndValuationFund")

def gravaIndVEmpreStaus(wsIndEmpresa,linha,dict_stocks,stock):
    ATIVO = stock

    valoratual  =  dict_stocks[stock].get("Valor atual")
    min52semanas = dict_stocks[stock].get("Min 52 semanas")
    max52semanas = dict_stocks[stock].get("Max 52 semanas")
    dividendyield =  dict_stocks[stock].get("dividend Yield")
    valorizacao12m = dict_stocks[stock].get("Valorizacao (12m)")
    tipo =  dict_stocks[stock].get("Tipo")
    tagalong = dict_stocks[stock].get("TAG ALONG")
    liquidezmediadiaria = dict_stocks[stock].get("LIQUIDEZ MEDIA DIARIA")
    participacaonoibov = dict_stocks[stock].get("PARTICIPACAO NO IBOV")
    mercadodeopcoes = dict_stocks[stock].get("MERCADO DE OPCOES")
    patrimonioliquido = dict_stocks[stock].get("Patrimonio liquido")
    ativos =  dict_stocks[stock].get("Ativos")
    ativocirculante = dict_stocks[stock].get("Ativo circulante")
    dividabruta = dict_stocks[stock].get("Divida bruta")
    disponibilidade =  dict_stocks[stock].get("Disponibilidade")
    dividaliquida = dict_stocks[stock].get("Divida liquida")
    valordemercado = dict_stocks[stock].get("Valor de mercado")
    valordefirma =    dict_stocks[stock].get("Valor de firma")
    freefloat =  dict_stocks[stock].get("Free Float")


    wsIndEmpresa.cell(row=linha, column=2, value=valoratual)
    wsIndEmpresa.cell(row=linha, column=3, value=min52semanas)
    wsIndEmpresa.cell(row=linha, column=4, value=max52semanas)
    wsIndEmpresa.cell(row=linha, column=5, value=dividendyield)
    wsIndEmpresa.cell(row=linha, column=6, value=valorizacao12m)
    wsIndEmpresa.cell(row=linha, column=7, value=tipo)
    wsIndEmpresa.cell(row=linha, column=8, value=tagalong)
    wsIndEmpresa.cell(row=linha, column=9, value=liquidezmediadiaria)
    wsIndEmpresa.cell(row=linha, column=10, value=participacaonoibov)
    wsIndEmpresa.cell(row=linha, column=11, value=mercadodeopcoes)
    wsIndEmpresa.cell(row=linha, column=12, value=patrimonioliquido)
    wsIndEmpresa.cell(row=linha, column=13, value=ativos)
    wsIndEmpresa.cell(row=linha, column=14, value=ativocirculante)
    wsIndEmpresa.cell(row=linha, column=15, value=dividabruta)
    wsIndEmpresa.cell(row=linha, column=16, value=disponibilidade)
    wsIndEmpresa.cell(row=linha, column=17, value=dividaliquida)
    wsIndEmpresa.cell(row=linha, column=18, value=valordemercado)
    wsIndEmpresa.cell(row=linha, column=19, value=valordefirma)
    wsIndEmpresa.cell(row=linha, column=20, value=freefloat)
def gravaIndVEmpreFund(wsIndEmpresa,linha,dict_stocks,stock):
    ATIVO = stock

    valoratual  =  f"{float(Dicvaluation_indicators.get("Dividend Yield")) * 100}%"
    min52semanas = ""
    max52semanas = ""
    dividendyield =  ""
    valorizacao12m = ""
    tipo = f"{float(Dicvaluation_indicators.get("Dividend Yield")) * 100}%"
    tagalong = ""
    liquidezmediadiaria = f"{float(Dicvaluation_indicators.get("Dividend Yield")) * 100}%"
    participacaonoibov = f"{float(Dicvaluation_indicators.get("Dividend Yield")) * 100}%"
    mercadodeopcoes = f"{float(Dicvaluation_indicators.get("Dividend Yield")) * 100}%"
    patrimonioliquido = f"{float(Dicvaluation_indicators.get("Dividend Yield")) * 100}%"
    ativos =  f"{float(Dicvaluation_indicators.get("Dividend Yield")) * 100}%"
    ativocirculante = f"{float(Dicvaluation_indicators.get("Dividend Yield")) * 100}%"
    dividabruta = f"{float(Dicvaluation_indicators.get("Dividend Yield")) * 100}%"
    disponibilidade =  f"{float(Dicvaluation_indicators.get("Dividend Yield")) * 100}%"
    dividaliquida = f"{float(Dicvaluation_indicators.get("Dividend Yield")) * 100}%"
    valordemercado = f"{float(Dicvaluation_indicators.get("Dividend Yield")) * 100}%"
    valordefirma =    ""
    freefloat =  ""


    wsIndEmpresa.cell(row=linha, column=2, value=valoratual)
    wsIndEmpresa.cell(row=linha, column=3, value=min52semanas)
    wsIndEmpresa.cell(row=linha, column=4, value=max52semanas)
    wsIndEmpresa.cell(row=linha, column=5, value=dividendyield)
    wsIndEmpresa.cell(row=linha, column=6, value=valorizacao12m)
    wsIndEmpresa.cell(row=linha, column=7, value=tipo)
    wsIndEmpresa.cell(row=linha, column=8, value=tagalong)
    wsIndEmpresa.cell(row=linha, column=9, value=liquidezmediadiaria)
    wsIndEmpresa.cell(row=linha, column=10, value=participacaonoibov)
    wsIndEmpresa.cell(row=linha, column=11, value=mercadodeopcoes)
    wsIndEmpresa.cell(row=linha, column=12, value=patrimonioliquido)
    wsIndEmpresa.cell(row=linha, column=13, value=ativos)
    wsIndEmpresa.cell(row=linha, column=14, value=ativocirculante)
    wsIndEmpresa.cell(row=linha, column=15, value=dividabruta)
    wsIndEmpresa.cell(row=linha, column=16, value=disponibilidade)
    wsIndEmpresa.cell(row=linha, column=17, value=dividaliquida)
    wsIndEmpresa.cell(row=linha, column=18, value=valordemercado)
    wsIndEmpresa.cell(row=linha, column=19, value=valordefirma)
    wsIndEmpresa.cell(row=linha, column=20, value=freefloat)

#Incio funcionalidades statusinvest

# define selenium webdriver options
options = webdriver.ChromeOptions()

# create selenium webdriver instancee
driver = webdriver.Chrome(options=options)

def get_stock_soup(stock):
    ''' Get raw html from a stock '''

    # access the stock url
    driver.get(f'https://statusinvest.com.br/acoes/{stock}')
    #driver.get(f'https://statusinvest.com.br/acoes/eua/{stock}')

    # get html from stock
    html = driver.find_element(By.ID, 'main-2').get_attribute('innerHTML')

    # remove accents from html and transform html into soup
    soup = BeautifulSoup(unidecode(html), 'html.parser')

    return soup

def is_null_zero_or_spaces(variable):
       # Verifica se a variável é None
       if variable is None:
           return True
       # Verifica se a variável é zero (0)
       elif variable == 0:
           return True
       # Verifica se a variável é uma string e contém apenas espaços
       elif isinstance(variable, str) and variable.strip() == '':
           return True
       elif variable == '-%':
           return True
       else:
           return False

def soup_to_dict(soup):
    '''Get all data from stock soup and return as a dictionary '''
    keys, values = [], []

    # get divs from stock
    soup1 = soup.find('div', class_='pb-3 pb-md-5')
    soup2 = soup.find('div', class_='card rounded text-main-green-dark')
    soup3 = soup.find('div', class_='indicator-today-container')
    soup4 = soup.find(
        'div', class_='top-info info-3 sm d-flex justify-between mb-3')
    soups = [soup1, soup2, soup3, soup4]

    for s in soups:
        # get only titles from a div and append to keys
        titles = s.find_all('h3', re.compile('title m-0[^"]*'))

        titles = [t.get_text() for t in titles]
        keys += titles

        # get only numbers from a div and append to values
        numbers = s.find_all('strong', re.compile('value[^"]*'))
        numbers = [n.get_text()for n in numbers]
        values += numbers

    # remove unused key and insert needed keys
    keys.remove('PART. IBOV')
    keys.insert(6, 'TAG ALONG')
    keys.insert(7, 'LIQUIDEZ MEDIA DIARIA')

    # clean keys list
    keys = [k.replace('\nhelp_outline', '').strip() for k in keys]
    keys = [k for k in keys if k != '']

    # clean values list
    values = [v.replace('\nhelp_outline', '').strip() for v in values]
    values = [v.replace('.', '').replace(',', '.') for v in values]

    # create a dictionary using keys and values from indicators
    d = {k: v for k, v in zip(keys, values)}

    return d

#Fim funcionalidades statusinvest

# pylint: disable=line-too-long
if __name__ == '__main__':
    dict_stocks = {} # dicionario statusinvest
    Dicrentabilidade = {}  # fundamentus
    linhafundamentus = 1  # silvio
    linhastatus = 0


    criaPlanilhaIndiEficiência(wbsaida) # statusinvest + fundamentus
    criaPlanilhaIndRentabilidade(wbsaida)
    criaPlanilhaIndEndividamento(wbsaida)
    criaPlanilaIndValuation(wbsaida)
    criaPlanilhaIndiEmpresa(wbsaida)

    wsIndiEficiência = wbsaida['IndiEficiência']
    wsIndiRentabilidade = wbsaida['IndiRentabilidade']
    wsIndEndividamento = wbsaida['IndEndividamento']
    wsIndValuation = wbsaida['IndValuation']
    wsIndEmpresa = wbsaida['IndEmpresa']
    start = time.time()
    with open('stocks.txt', 'r') as f:
        stocks = f.read().splitlines()
        for stock in stocks:
            linhastatus = linhafundamentus
            linhastatus = linhastatus + 1   # silvio
            linhafundamentus =  linhastatus + 1 # silvio

            main_pipeline = fundamentus.Pipeline(stock.upper())
            response = main_pipeline.get_all_information()

            # Incio statusinvest
            # get data and transform into dictionary
            soup = get_stock_soup(stock)
            dict_stock = soup_to_dict(soup)
            dict_stocks[stock] = dict_stock
            # Fim Statausinvest

            # Extract the information from the response.
            stock_identification = response.transformed_information['stock_identification']
            financial_summary = response.transformed_information['financial_summary']
            price_information = response.transformed_information['price_information']
            detailed_information = response.transformed_information[
                'detailed_information']
            oscillations = response.transformed_information['oscillations']
            valuation_indicators = response.transformed_information[
                'valuation_indicators']
            profitability_indicators = response.transformed_information[
                'profitability_indicators']
            indebtedness_indicators = response.transformed_information[
                'indebtedness_indicators']
            balance_sheet = response.transformed_information['balance_sheet']
            income_statement = response.transformed_information['income_statement']

            montadicionario(stock_identification,financial_summary,price_information,detailed_information,oscillations,
                    valuation_indicators,profitability_indicators,indebtedness_indicators,balance_sheet,income_statement)
           # print(Dicrentabilidade)

            gravaIndiEficiênciaoStaus(wsIndiEficiência,'StatusInvest',linhastatus,dict_stocks,stock)
            gravaIndiEficiênciaoFund(wsIndiEficiência, 'Fundamentus', linhafundamentus, Dicprofitability_indicators, stock)

            gravaIndiRentabilidadeStaus(wsIndiRentabilidade, 'StatusInvest', linhastatus, dict_stocks, stock)
            gravaIndiRentabilidadeFund(wsIndiRentabilidade, 'Fundamentus', linhafundamentus, Dicprofitability_indicators, stock)

            gravaIndEndividamentoStaus(wsIndEndividamento, 'StatusInvest', linhastatus, dict_stocks, stock)
            gravaIndEndividamentosFund(wsIndEndividamento, 'Fundamentus', linhafundamentus, Dicindebtedness_indicators, stock)
            #wbsaida.save("StatusInvest.xlsx")  # silvio
            gravaIndValuationStaus(wsIndValuation,'StatusInvest', linhastatus, dict_stocks, stock)
            gravaIndValuationFund(wsIndValuation, 'Fundamentus', linhafundamentus, Dicvaluation_indicators, Dicdetailed_information, stock)
            print(dict_stocks)
            gravaIndVEmpreStaus(wsIndEmpresa, linhastatus, dict_stocks, stock)
    # exit the driver
    driver.quit()

    # end timer
    end = time.time()
    wbsaida.save("StatusInvest2.xlsx") # silvio