import openpyxl
import re
import time
import numpy as np
import pandas as pd
from unidecode import unidecode
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
import requests
# teste silvio 3e
import warnings
from openpyxl.styles import numbers

#from teste01 import metrica

fillvermelho= PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") # Vermelho

fillverde= PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid") # Verde

fillamarelo =  PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # Amarelo


fillazul = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid") # Azul


#filltitulo =   PatternFill(start_color="#002060", end_color="#002060", fill_type="solid")

filltitulo = PatternFill(start_color="002060", end_color="002060", fill_type="solid")  # Azul escuro

font_branca = Font(color="FFFFFF")  # Branco

TITLES = [
    'Identificação', 'Resumo Financeiro', 'Cotações', 'Informações Básicas',
    'Oscilações', 'Indicadores de Valuation', 'Indicadores de Rentabilidade',
    'Indicadores de Endividamento', 'Balanço Patrimonial', 'Demonstrativo de Resultados'
]

linha2 = 1
metricasts= ""
''''categorias = {
    'otimo': {'min': float('-inf'), 'max': -2},  # Valores muito baixos são ótimos
    'bom': {'min': -2, 'max': 0},               # Valores entre -2 e 0
    'moderado': {'min': 0, 'max': 1.5},         # Valores entre 0 e 1.5
    'ruim': {'min': 1.5, 'max': 3},             # Valores entre 1.5 e 3
    'pessimo': {'min': 3, 'max': 4},            # Valores entre 3 e 4
    'critico': {'min': 4, 'max': float('inf')}   # Valores acima de 4
}'''
MetricasStatus = {
    'Div. liquida/EBIT': {
        'critico': {'min': float('-inf'), 'max': 0},
        'pessimo': {'min': 4, 'max': float('inf')},
        'ruim': {'min': 3, 'max': 4},
        'moderado': {'min': 1.5, 'max': 3},
        'bom': {'min': 0, 'max': 1.5},
        'otimo': {'min': float('-inf'), 'max': 0},
        'descricao': 'Dívida líquida dividida pelo EBIT. Mede tempo para quitar dívida com lucro operacional.',
        'agrupador': 'Eficiência',
        'descrcritico': 'Dívidaw líquida negativa pode indicar EBIT negativo (prejuízo operacional grave) ou excesso de caixa líquido. EBIT negativo, comum em empresas em crise (ex.: OI, OIBR3), reflete ineficiência operacional e risco de insolvência. Excesso de caixa, como em empresas de tecnologia (ex.: Nubank), pode indicar ineficiência na alocação de capital, com caixa ocioso não reinvestido. Em 2025, com juros altos, EBIT negativo agrava a pressão financeira. Investidores devem analisar a causa (EBIT negativo vs. caixa elevado) e evitar empresas sem plano claro de recuperação ou alocação eficiente.',
        'descrpessimo': 'Endividamento crítico, leva mais de 4 anos de EBIT para quitar a dívida líquida, indicando alavancagem extrema. Comum em empresas em crise ou setores cíclicos (ex.: construção). Aumenta vulnerabilidade a choques econômicos ou aumento de juros. Investidores devem evitar, salvo reestruturação robusta com forte geração de caixa futura.',
        'descrruim': 'Endividamento elevado, leva 3-4 anos de EBIT para pagar a dívida, um nível alto para a maioria dos setores. Comum em setores intensivos em capital (ex.: infraestrutura). Investidores devem analisar cobertura de juros e estabilidade do EBIT para avaliar riscos em cenários adversos.',
        'descrmoderado': 'Endividamento moderado, pagável em 1,5-3 anos de EBIT, aceitável em setores estáveis (ex.: utilities). Investidores devem verificar consistência do EBIT e fluxo de caixa livre para garantir que a alavancagem não comprometa a operação.',
        'descrbom': 'Endividamento controlado, dívida quitável em menos de 1,5 anos de EBIT, indicando baixa alavancagem. Comum em empresas maduras com forte geração de caixa (ex.: Ambev). Sinal de estabilidade, mas deve-se avaliar subinvestimento.',
        'descrotimo': 'Caixa líquido supera dívida, refletindo forte saúde financeira. Comum em empresas de tecnologia ou com reservas elevadas (ex.: Nubank). Investidores devem investigar alocação do caixa (ex.: reinvestimento, dividendos) para evitar ineficiências.'
    },
    'Div. liquida/EBITDA': {
        'critico': {'min': float('-inf'), 'max': 0},
        'pessimo': {'min': 3.5, 'max': float('inf')},
        'ruim': {'min': 2.5, 'max': 3.5},
        'moderado': {'min': 1, 'max': 2.5},
        'bom': {'min': 0, 'max': 1},
        'otimo': {'min': float('-inf'), 'max': 0},
        'descricao': 'Dívida líquida dividida pelo EBITDA. Mede tempo para quitar dívida com fluxo de caixa operacional.',
        'agrupador': 'Eficiência',
        'descrcritico': 'Prejuízo operacional, com incapacidade de gerar caixa operacional, indicando problemas estruturais graves. Comum em empresas em crise (ex.: OIBR3) ou setores com alta concorrência (ex.: varejo, CRFB3). Em 2025, com juros de 10-12%, a margem negativa agrava a pressão financeira, dificultando a cobertura de despesas. Riscos incluem insolvência, perda de market share ou aumento de custos fixos. Analise o fluxo de caixa operacional (FCO < 0 é preocupante) e a composição dos custos operacionais. Verifique balanços trimestrais e planos estratégicos para avaliar viabilidade. Setores cíclicos (ex.: construção, MRVE3) podem ter margens negativas temporariamente, mas exigem recuperação clara. Compare com peers (ex.: varejo ~10%) para avaliar gravidade. Evite empresas com EBITDA negativo recorrente ou sem estratégia definida. Oportunidades surgem em reestruturações robustas com potencial de turnaround.',
        'descrpessimo': 'Endividamento crítico, leva mais de 3,5 anos de EBITDA para quitar a dívida líquida, indicando alavancagem extrema. Aumenta vulnerabilidade a recessões ou aumento de juros. Investidores devem evitar, salvo recuperação clara com forte geração de caixa.',
        'descrruim': 'Endividamento elevado, leva 2,5-3,5 anos de EBITDA para pagar a dívida. Comum em setores cíclicos (ex.: commodities). Investidores devem analisar cobertura de juros, volatilidade do EBITDA e riscos macroeconômicos.',
        'descrmoderado': 'Endividamento moderado, pagável em 1-2,5 anos de EBITDA, aceitável em setores com investimentos moderados (ex.: varejo). Investidores devem monitorar tendência do EBITDA e gestão de dívidas para garantir sustentabilidade.',
        'descrbom': 'Endividamento baixo, dívida quitável rapidamente, indicando forte capacidade de pagamento. Comum em empresas com margens altas (ex.: Weg). Sinal positivo, mas deve-se avaliar se limita crescimento.',
        'descrotimo': 'Caixa líquido excede dívida, sugerindo forte geração de caixa ou reservas elevadas. Comum em setores de alto crescimento (ex.: tecnologia). Investidores devem verificar alocação do caixa para evitar ineficiências (ex.: caixa ocioso).'
    },
    'Div. liquida/PL': {
        'critico': {'min': float('-inf'), 'max': 0},
        'pessimo': {'min': 1, 'max': float('inf')},
        'ruim': {'min': 0.7, 'max': 1},
        'moderado': {'min': 0.3, 'max': 0.7},
        'bom': {'min': 0, 'max': 0.3},
        'otimo': {'min': float('-inf'), 'max': 0},
        'descricao': 'Dívida líquida em relação ao patrimônio líquido. Mede alavancagem financeira.',
        'agrupador': 'Eficiência',
        'descrcritico': 'Patrimônio líquido negativo, passivos superam ativos, indicando risco crítico de insolvência. Comum em empresas em crise (ex.: OI). Investidores devem exigir plano robusto de recuperação.',
        'descrpessimo': 'Alavancagem crítica, dívida líquida excede o patrimônio líquido, indicando dependência extrema de financiamento externo. Comum em empresas em crise (ex.: OI). Investidores devem evitar, salvo reestruturação com forte potencial de recuperação.',
        'descrruim': 'Alavancagem elevada, dívida representa 70-100% do patrimônio líquido, um nível alto para a maioria dos setores. Investidores devem analisar capacidade de geração de lucro e prazos da dívida para avaliar sustentabilidade financeira.',
        'descrmoderado': 'Alavancagem moderada, dívida é 30-70% do patrimônio líquido, aceitável em setores como indústria. Investidores devem monitorar cobertura de juros e estabilidade do fluxo de caixa para garantir solvência.',
        'descrbom': 'Baixa alavancagem, estrutura financeira sólida, indicando empresa bem capitalizada. Comum em setores estáveis (ex.: utilities). Sinal de segurança, mas deve-se avaliar se a baixa dívida limita crescimento.',
        'descrotimo': 'Caixa líquido excede dívida, indicando altíssima capitalização. Comum em empresas com reservas elevadas (ex.: Weg). Investidores devem verificar alocação do caixa para evitar ineficiências.'
    },
    'D.Y': {
        'critico': {'min': 0, 'max': 0},
        'pessimo': {'min': 0, 'max': 0},
        'ruim': {'min': 0.001, 'max': 0.02},
        'moderado': {'min': 0.02, 'max': 0.04},
        'bom': {'min': 0.04, 'max': 0.06},
        'otimo': {'min': 0.06, 'max': float('inf')},
        'descricao': 'Dividendos pagos por ação em relação ao preço da ação. Mede retorno via dividendos.',
        'agrupador': 'Outros',
        'descrcritico': 'Sem pagamento de dividendos, indicando reinvestimento total dos lucros (ex.: startups) ou prejuízo financeiro. Investidores focados em renda devem evitar, mas podem considerar se o foco é crescimento de longo prazo.',
        'descrpessimo': 'Sem pagamento de dividendos, indicando reinvestimento total dos lucros (ex.: startups) ou prejuízo financeiro. Investidores focados em renda devem evitar, mas podem considerar se o foco é crescimento de longo prazo.',
        'descrruim': 'Retorno baixo, pouco atrativo para investidores de dividendos. Comum em empresas de crescimento (ex.: Nubank). Investidores devem focar no potencial de valorização das ações em vez de dividendos.',
        'descrmoderado': 'Retorno moderado, oferece equilíbrio entre dividendos e reinvestimento. Comum em empresas maduras (ex.: Ambev). Investidores devem verificar sustentabilidade do payout e consistência dos lucros.',
        'descrbom': 'Retorno atrativo, bom para investidores de renda. Comum em setores como utilities e bancos (ex.: Engie). Investidores devem confirmar sustentabilidade do fluxo de caixa livre e política de dividendos.',
        'descrotimo': 'Alto retorno, muito atrativo, mas pode indicar risco de insustentabilidade, especialmente com lucros voláteis. Investidores devem analisar payout ratio e geração de caixa para garantir continuidade dos dividendos.'
    },
    'EV/EBIT': {
        'critico': {'min': float('-inf'), 'max': 0},
        'pessimo': {'min': 15, 'max': float('inf')},
        'ruim': {'min': 12, 'max': 15},
        'moderado': {'min': 8, 'max': 12},
        'bom': {'min': 4, 'max': 8},
        'otimo': {'min': 0, 'max': 4},
        'descricao': 'Valor da empresa dividido pelo EBIT. Mede valuation em relação ao lucro operacional.',
        'agrupador': 'Valuation',
        'descrcritico': 'Prejuízo operacional, com operação não rentável, indicando ineficiência grave ou investimentos pesados. Comum em empresas em crise (ex.: OIBR3) ou setores com alta concorrência (ex.: varejo, CRFB3). Em 2025, com juros de 10-12%, a margem negativa agrava a pressão financeira, dificultando a cobertura de despesas. Riscos incluem insolvência, perda de market share ou aumento de custos fixos. Analise o fluxo de caixa operacional (FCO < 0 é preocupante) e a composição dos custos operacionais. Verifique balanços trimestrais e planos estratégicos para avaliar viabilidade. Setores cíclicos (ex.: construção, MRVE3) podem ter margens negativas temporariamente, mas exigem recuperação clara. Compare com peers (ex.: varejo ~5%) para avaliar gravidade. Evite empresas com EBIT negativo recorrente ou sem estratégia definida. Oportunidades surgem em reestruturações robustas com potencial de turnaround.',
        'descrpessimo': 'Empresa extremamente cara, EV é mais de 15 vezes o EBIT, sugerindo expectativas irreais de crescimento ou sobrevalorização. Comum em setores de tecnologia em bolha. Investidores devem comparar com peers e avaliar crescimento projetado.',
        'descrruim': 'Empresa cara, múltiplo elevado, comum em setores de crescimento (ex.: saúde). Investidores devem analisar se o crescimento futuro justifica o preço e comparar com média setorial para evitar sobrepreço.',
        'descrmoderado': 'Avaliação razoável, equilibrada para empresas maduras com crescimento moderado (ex.: indústria). Investidores devem verificar consistência do EBIT e perspectivas de crescimento para confirmar atratividade.',
        'descrbom': 'Subvalorizada, EV abaixo de 8x o EBIT indica empresa barata, com potencial de valorização. Comum em setores cíclicos em recuperação (ex.: mineração). Investidores devem confirmar fluxo de caixa livre e tendências setoriais.',
        'descrotimo': 'Extremamente subvalorizada, EV muito baixo em relação ao EBIT, sugerindo oportunidade ou EBIT inflado. Investidores devem verificar fluxo de caixa livre e sustentabilidade do EBIT para confirmar oportunidade.'
    },
    'EV/EBITDA': {
        'critico': {'min': float('-inf'), 'max': 0},
        'pessimo': {'min': 12, 'max': float('inf')},
        'ruim': {'min': 10, 'max': 12},
        'moderado': {'min': 6, 'max': 10},
        'bom': {'min': 3, 'max': 6},
        'otimo': {'min': 0, 'max': 3},
        'descricao': 'Valor da empresa dividido pelo EBITDA. Mede valuation em relação ao fluxo de caixa operacional.',
        'agrupador': 'Valuation',
        'descrcritico': 'Prejuízo operacional, com incapacidade de gerar caixa operacional, indicando problemas estruturais graves. Comum em empresas em crise (ex.: OIBR3) ou setores com alta concorrência (ex.: varejo, CRFB3). Em 2025, com juros de 10-12%, a margem negativa agrava a pressão financeira, dificultando a cobertura de despesas. Riscos incluem insolvência, perda de market share ou aumento de custos fixos. Analise o fluxo de caixa operacional (FCO < 0 é preocupante) e a composição dos custos operacionais. Verifique balanços trimestrais e planos estratégicos para avaliar viabilidade. Setores cíclicos (ex.: construção, MRVE3) podem ter margens negativas temporariamente, mas exigem recuperação clara. Compare com peers (ex.: varejo ~10%) para avaliar gravidade. Evite empresas com EBITDA negativo recorrente ou sem estratégia definida. Oportunidades surgem em reestruturações robustas com potencial de turnaround.',
        'descrpessimo': 'Empresa extremamente cara, EV é mais de 12 vezes o EBITDA, sugerindo sobrevalorização. Comum em setores de tecnologia (ex.: Nubank). Investidores devem comparar com peers e avaliar crescimento projetado.',
        'descrruim': 'Empresa cara, comum em setores de crescimento (ex.: saúde). Investidores devem analisar se o crescimento futuro justifica o preço e comparar com média setorial.',
        'descrmoderado': 'Avaliação equilibrada, comum em setores maduros (ex.: varejo). Investidores devem verificar consistência do EBITDA e perspectivas de crescimento.',
        'descrbom': 'Subvalorizada, oportunidade em setores cíclicos (ex.: mineração). Investidores devem confirmar fluxo de caixa livre e tendências setoriais.',
        'descrotimo': 'Extremamente subvalorizada, sugere oportunidade ou EBITDA inflado. Investidores devem verificar fluxo de caixa livre e sustentabilidade do EBITDA.'
    },
    'Giro do Ativo': {
        'critico': {'min': float('-inf'), 'max': 0},
        'pessimo': {'min': 0, 'max': 0.2},
        'ruim': {'min': 0.2, 'max': 0.4},
        'moderado': {'min': 0.4, 'max': 0.8},
        'bom': {'min': 0.8, 'max': 1.2},
        'otimo': {'min': 1.2, 'max': float('inf')},
        'descricao': 'Receita líquida dividida pelos ativos totais. Mede eficiência na utilização dos ativos.',
        'agrupador': 'Eficiência',
        'descrcritico': 'Receita negativa ou ativos mal contabilizados, indicando problemas estruturais graves ou falha contábil. Comum em empresas em crise severa (ex.: OIBR3). Investidores devem evitar, salvo recuperação robusta com plano claro de reestruturação.',
        'descrpessimo': 'Baixa eficiência, ativos subutilizados, comum em empresas em crise (ex.: OIBR3) ou setores em baixa (ex.: siderurgia). Investidores devem verificar composição dos ativos e fluxo de caixa operacional.',
        'descrruim': 'Eficiência limitada, comum em setores cíclicos (ex.: GGBR4). Investidores devem investigar potencial de melhoria operacional e aumento de vendas.',
        'descrmoderado': 'Eficiência moderada, aceitável em setores industriais (ex.: CSNA3). Investidores devem analisar tendência da receita e composição dos ativos.',
        'descrbom': 'Boa eficiência, ativos bem utilizados, comum em varejo (ex.: LREN3). Investidores devem confirmar sustentabilidade da receita.',
        'descrotimo': 'Alta eficiência, excelente utilização dos ativos, comum em setores dinâmicos (ex.: MGLU3). Investidores devem verificar consistência frente a riscos setoriais.'
    },
    'Liq. corrente': {
        'critico': {'min': float('-inf'), 'max': 0},
        'pessimo': {'min': 0, 'max': 0.8},
        'ruim': {'min': 0.8, 'max': 1.2},
        'moderado': {'min': 1.2, 'max': 1.8},
        'bom': {'min': 1.8, 'max': 2.5},
        'otimo': {'min': 2.5, 'max': float('inf')},
        'descricao': 'Ativo circulante dividido pelo passivo circulante. Mede capacidade de pagar dívidas de curto prazo.',
        'agrupador': 'Liquidez e Solvência',
        'descrcritico': 'Passivo circulante excede ativo circulante de forma crítica, indicando insolvência iminente. Comum em empresas em crise severa (ex.: OIBR3). Investidores devem evitar, salvo plano robusto de recuperação.',
        'descrpessimo': 'Incapacidade de cobrir obrigações de curto prazo, indicando risco de insolvência. Comum em empresas em crise (ex.: OIBR3). Investidores devem evitar, salvo reestruturação robusta.',
        'descrruim': 'Liquidez limitada, risco moderado de dificuldades financeiras. Investidores devem analisar composição do ativo circulante e fluxo de caixa operacional.',
        'descrmoderado': 'Liquidez razoável, aceitável em setores estáveis (ex.: ABEV3). Investidores devem verificar qualidade dos ativos circulantes (ex.: estoques, recebíveis).',
        'descrbom': 'Boa liquidez, capacidade sólida de cobrir obrigações de curto prazo. Comum em empresas maduras (ex.: WEGE3). Investidores devem confirmar sustentabilidade do fluxo de caixa.',
        'descrotimo': 'Alta liquidez, forte capacidade de pagamento, mas pode indicar caixa ocioso. Comum em tecnologia (ex.: TOTS3). Investidores devem avaliar alocação de capital.'
    },
    'LPA': {
        'critico': {'min': float('-inf'), 'max': 0},
        'pessimo': {'min': 0, 'max': 0.1},
        'ruim': {'min': 0.1, 'max': 0.5},
        'moderado': {'min': 0.5, 'max': 1},
        'bom': {'min': 1, 'max': 2},
        'otimo': {'min': 2, 'max': float('inf')},
        'descricao': 'Lucro líquido por ação. Mede rentabilidade por ação emitida.',
        'agrupador': 'Outros',
        'descrcritico': 'Prejuízo por ação, indicando problemas operacionais ou financeiros graves. Comum em empresas em crise ou com investimentos pesados (ex.: OI). Investidores devem analisar plano de recuperação.',
        'descrpessimo': 'Lucro por ação muito baixo, indicando rentabilidade limitada. Comum em setores competitivos (ex.: varejo). Investidores devem investigar potencial de melhoria.',
        'descrruim': 'Lucro baixo, desempenho fraco, comum em empresas em crescimento ou setores cíclicos. Investidores devem analisar tendência do lucro e riscos setoriais.',
        'descrmoderado': 'Lucro moderado, desempenho razoável, aceitável em setores maduros (ex.: indústria). Investidores devem verificar sustentabilidade do lucro.',
        'descrbom': 'Bom lucro por ação, desempenho sólido, comum em setores estáveis (ex.: bancos). Investidores devem confirmar consistência do lucro.',
        'descrotimo': 'Lucro excepcional, refletindo forte rentabilidade. Comum em setores de alta margem (ex.: tecnologia). Investidores devem verificar sustentabilidade frente a riscos setoriais.'
    },
    'M. Bruta': {
        'critico': {'min': float('-inf'), 'max': 0},
        'pessimo': {'min': 0, 'max': 0.15},
        'ruim': {'min': 0.15, 'max': 0.25},
        'moderado': {'min': 0.25, 'max': 0.4},
        'bom': {'min': 0.4, 'max': 0.6},
        'otimo': {'min': 0.6, 'max': float('inf')},
        'descricao': 'Lucro bruto dividido pela receita líquida. Mede eficiência na produção ou vendas.',
        'agrupador': 'Eficiência',
        'descrcritico': 'Custo dos produtos vendidos excede a receita, indicando ineficiência grave ou preços insustentáveis. Comum em empresas em crise (ex.: OIBR3) ou setores com alta concorrência (ex.: varejo, CRFB3). Investidores devem evitar, salvo plano robusto de recuperação.',
        'descrpessimo': 'Baixa margem, indicando ineficiência na produção ou vendas. Comum em setores competitivos (ex.: varejo, MGLU3). Investidores devem analisar composição dos custos e potencial de melhoria.',
        'descrruim': 'Margem limitada, comum em setores com alta concorrência (ex.: CRFB3). Investidores devem investigar estratégias de diferenciação ou redução de custos.',
        'descrmoderado': 'Margem razoável, aceitável em setores industriais (ex.: CSNA3). Investidores devem verificar sustentabilidade da margem e poder de precificação.',
        'descrbom': 'Boa margem, indicando eficiência sólida. Comum em setores com barreiras de entrada (ex.: ABEV3). Investidores devem confirmar consistência da margem.',
        'descrotimo': 'Alta margem, desempenho excepcional, comum em setores de alto valor agregado (ex.: tecnologia, TOTS3). Investidores devem verificar proteção contra concorrência.'
    },
    'M. EBIT': {
        'critico': {'min': float('-inf'), 'max': 0},
        'pessimo': {'min': 0, 'max': 0.05},
        'ruim': {'min': 0.05, 'max': 0.15},
        'moderado': {'min': 0.15, 'max': 0.25},
        'bom': {'min': 0.25, 'max': 0.35},
        'otimo': {'min': 0.35, 'max': float('inf')},
        'descricao': 'Lucro operacional dividido pela receita líquida. Mede rentabilidade antes de juros e impostos.',
        'agrupador': 'Eficiência',
        'descrcritico': 'Prejuízo operacional, com operação não rentável, indicando ineficiência grave ou investimentos pesados. Comum em empresas em crise (ex.: OIBR3) ou setores com alta concorrência (ex.: varejo, CRFB3). Em 2025, com juros de 10-12%, a margem negativa agrava a pressão financeira, dificultando a cobertura de despesas. Riscos incluem insolvência, perda de market share ou aumento de custos fixos. Analise o fluxo de caixa operacional (FCO < 0 é preocupante) e a composição dos custos operacionais. Verifique balanços trimestrais e planos estratégicos para avaliar viabilidade. Setores cíclicos (ex.: construção, MRVE3) podem ter margens negativas temporariamente, mas exigem recuperação clara. Compare com peers (ex.: varejo ~5%) para avaliar gravidade. Evite empresas com EBIT negativo recorrente ou sem estratégia definida. Oportunidades surgem em reestruturações robustas com potencial de turnaround.',
        'descrpessimo': 'Baixa rentabilidade operacional, indicando eficiência limitada. Comum em setores competitivos ou cíclicos (ex.: varejo, MGLU3). Investidores devem analisar potencial de melhoria operacional e redução de custos.',
        'descrruim': 'Rentabilidade operacional limitada, comum em setores com margens moderadas (ex.: indústria, GGBR4). Investidores devem verificar tendência do EBIT e riscos setoriais.',
        'descrmoderado': 'Rentabilidade moderada, desempenho razoável, aceitável em setores com margens moderadas (ex.: varejo, LREN3). Investidores devem analisar sustentabilidade do EBIT e concorrência.',
        'descrbom': 'Boa rentabilidade operacional, desempenho sólido, comum em setores com barreiras de entrada (ex.: utilities, CPFE3). Investidores devem confirmar consistência do EBIT.',
        'descrotimo': 'Alta rentabilidade operacional, desempenho excepcional, comum em setores de alta margem (ex.: tecnologia, TOTS3). Investidores devem verificar proteção contra concorrência e sustentabilidade.'
    },
    'M. EBITDA': {
        'critico': {'min': float('-inf'), 'max': 0},
        'pessimo': {'min': 0, 'max': 0.1},
        'ruim': {'min': 0.1, 'max': 0.2},
        'moderado': {'min': 0.2, 'max': 0.3},
        'bom': {'min': 0.3, 'max': 0.4},
        'otimo': {'min': 0.4, 'max': float('inf')},
        'descricao': 'EBITDA dividido pela receita líquida. Mede geração de caixa operacional.',
        'agrupador': 'Eficiêncial',
        'descrcritico': 'Prejuízo operacional, com incapacidade de gerar caixa operacional, indicando problemas estruturais graves. Comum em empresas em crise (ex.: OIBR3) ou setores com alta concorrência (ex.: varejo, CRFB3). Em 2025, com juros de 10-12%, a margem negativa agrava a pressão financeira, dificultando a cobertura de despesas. Riscos incluem insolvência, perda de market share ou aumento de custos fixos. Analise o fluxo de caixa operacional (FCO < 0 é preocupante) e a composição dos custos operacionais. Verifique balanços trimestrais e planos estratégicos para avaliar viabilidade. Setores cíclicos (ex.: construção, MRVE3) podem ter margens negativas temporariamente, mas exigem recuperação clara. Compare com peers (ex.: varejo ~10%) para avaliar gravidade. Evite empresas com EBITDA negativo recorrente ou sem estratégia definida. Oportunidades surgem em reestruturações robustas com potencial de turnaround.',
        'descrpessimo': 'Baixa geração de caixa, com eficiência limitada, comum em setores intensivos em capital (ex.: infraestrutura, CMIG4). Em 2025, com custos elevados, essa faixa indica desafios para melhorar a eficiência sem aumento de preços ou redução de custos fixos. Investidores devem investigar o potencial de melhoria operacional (ex.: automação, corte de despesas) e o poder de precificação. Riscos incluem concorrência agressiva, aumento de insumos ou queda de demanda. Oportunidades surgem em empresas com estratégias de diferenciação (ex.: marcas fortes para LREN3). Analise o fluxo de caixa livre (FCL > 0) e a tendência do EBITDA nos últimos 3 anos. Compare com peers (ex.: varejo ~15%) para avaliar competitividade. Verifique balanços trimestrais e tendências setoriais (ex.: e-commerce). Evite empresas com EBITDA declinante ou dependência de promoções. Analise o ROIC (> 5%) e o CAPEX para confirmar potencial de melhoria. Setores competitivos exigem cautela extra.',
        'descrruim': 'Geração de caixa moderada, com desempenho razoável, aceitável em setores com margens moderadas (ex.: varejo, LREN3). Em 2025, essa faixa reflete capacidade de cobrir custos operacionais, mas exige análise de pressão competitiva e custos fixos. Investidores devem comparar com peers (ex.: varejo ~15%) e avaliar a sustentabilidade do EBITDA (ex.: variação < 15% em 5 anos). Setores estáveis (ex.: consumo defensivo, ABEV3) são menos arriscados, mas empresas cíclicas (ex.: GGBR4) podem enfrentar volatilidade. Oportunidades surgem em empresas com melhoria operacional, como automação ou expansão de mercado. Riscos incluem aumento de custos ou queda de demanda. Analise o fluxo de caixa livre (FCL > CAPEX) e a composição dos custos operacionais. Verifique balanços trimestrais e tendências setoriais (ex.: consumo estável). Evite empresas com EBITDA instável ou dependência de receitas cíclicas. Analise o ROIC (> 8%) para confirmar eficiência. Setores com barreiras de entrada (ex.: marcas fortes) se destacam.',
        'descrmoderado': 'Boa geração de caixa, desempenho sólido, comum em setores com barreiras de entrada (ex.: utilities, CPFE3). Investidores devem verificar consistência do EBITDA e poder de precificação.',
        'descrbom': 'Alta geração de caixa, desempenho excepcional, comum em setores de alta margem (ex.: tecnologia, TOTS3). Investidores devem confirmar sustentabilidade e proteção contra concorrência.',
        'descrotimo': 'Geração de caixa excepcional, refletindo liderança de mercado e eficiência superior. Comum em setores de alto valor agregado (ex.: Weg). Investidores devem verificar sustentabilidade frente a riscos setoriais.'
    },
    'M. Liquida': {
        'critico': {'min': float('-inf'), 'max': 0},
        'pessimo': {'min': 0, 'max': 0.05},
        'ruim': {'min': 0.05, 'max': 0.1},
        'moderado': {'min': 0.1, 'max': 0.15},
        'bom': {'min': 0.15, 'max': 0.25},
        'otimo': {'min': 0.25, 'max': float('inf')},
        'descricao': 'Lucro líquido dividido pela receita líquida. Mede rentabilidade final após todas as despesas.',
        'agrupador': 'Eficiência',
        'descrcritico': 'Prejuízo líquido, indicando ineficiência operacional ou financeira grave. Comum em empresas em crise (ex.: OIBR3) ou setores com alta concorrência (ex.: varejo, CRFB3). Investidores devem evitar, salvo plano robusto de recuperação.',
        'descrpessimo': 'Baixa rentabilidade, indicando desafios operacionais ou financeiros. Comum em setores competitivos (ex.: varejo, MGLU3). Investidores devem analisar potencial de melhoria e redução de custos.',
        'descrruim': 'Rentabilidade limitada, comum em setores cíclicos (ex.: siderurgia, GGBR4). Investidores devem verificar tendência do lucro líquido e riscos setoriais.',
        'descrmoderado': 'Rentabilidade moderada, aceitável em setores maduros (ex.: indústria, CSNA3). Investidores devem analisar sustentabilidade do lucro e concorrência.',
        'descrbom': 'Boa rentabilidade, desempenho sólido, comum em setores com barreiras de entrada (ex.: ABEV3). Investidores devem confirmar consistência do lucro líquido.',
        'descrotimo': 'Alta rentabilidade, desempenho excepcional, comum em setores de alta margem (ex.: tecnologia, TOTS3). Investidores devem verificar sustentabilidade frente a riscos setoriais.'
    },
    'P/Ativo Circulante Líquido': {
        'critico': {'min': float('-inf'), 'max': 0},
        'pessimo': {'min': 0, 'max': 0},
        'ruim': {'min': 0, 'max': 1},
        'moderado': {'min': 1, 'max': 2},
        'bom': {'min': 2, 'max': 3},
        'otimo': {'min': 3, 'max': float('inf')},
        'descricao': 'Preço da ação dividido pelo ativo circulante líquido por ação. Mede valuation em relação à liquidez.',
        'agrupador': 'Liquidez e Solvência',
        'descrcritico': 'Ativo circulante líquido negativo (passivo circulante excede ativo circulante), indicando risco crítico de insolvência. Comum em empresas em crise (ex.: OIBR3). Investidores devem evitar, salvo plano robusto de recuperação.',
        'descrpessimo': 'Preço da ação não reflete liquidez, indicando passivos elevados. Comum em empresas em crise (ex.: OIBR3). Investidores devem evitar, salvo reestruturação robusta.',
        'descrruim': 'Liquidez muito baixa em relação ao preço da ação, indicando risco financeiro. Investidores devem analisar qualidade do ativo circulante e fluxo de caixa.',
        'descrmoderado': 'Liquidez razoável em relação ao preço, aceitável em setores estáveis (ex.: ABEV3). Investidores devem verificar qualidade dos ativos circulantes.',
        'descrbom': 'Boa liquidez em relação ao preço, indicando equilíbrio financeiro. Comum em empresas maduras (ex.: WEGE3). Investidores devem confirmar sustentabilidade.',
        'descrotimo': 'Alta liquidez em relação ao preço, sugerindo subvalorização ou excesso de caixa. Comum em tecnologia (ex.: TOTS3). Investidores devem avaliar alocação de capital.'
    },
    'P/Capital de Giro': {
        'critico': {'min': float('-inf'), 'max': 0},
        'pessimo': {'min': 0, 'max': 0},
        'ruim': {'min': 0, 'max': 2},
        'moderado': {'min': 2, 'max': 4},
        'bom': {'min': 4, 'max': 6},
        'otimo': {'min': 6, 'max': float('inf')},
        'descricao': 'Preço da ação dividido pelo capital de giro por ação. Mede valuation em relação à capacidade de financiar operações.',
        'agrupador': 'Liquidez e Solvência',
        'descrcritico': 'Capital de giro negativo (ativo circulante menor que passivo circulante), indicando risco crítico de insolvência. Comum em empresas em crise (ex.: OIBR3). Investidores devem evitar, salvo plano robusto de recuperação.',
        'descrpessimo': 'Preço da ação não reflete capital de giro, indicando passivos elevados. Comum em empresas em crise (ex.: OIBR3). Investidores devem evitar, salvo reestruturação robusta.',
        'descrruim': 'Baixa capacidade de financiar operações de curto prazo em relação ao preço. Investidores devem analisar qualidade do capital de giro e fluxo de caixa.',
        'descrmoderado': 'Capacidade moderada de financiar operações, aceitável em setores estáveis (ex.: ABEV3). Investidores devem verificar qualidade do ativo circulante.',
        'descrbom': 'Boa capacidade de financiar operações, indicando equilíbrio financeiro. Comum em empresas maduras (ex.: LREN3). Investidores devem confirmar sustentabilidade.',
        'descrotimo': 'Alta capacidade de financiar operações, sugerindo subvalorização ou excesso de caixa. Comum em tecnologia (ex.: TOTS3). Investidores devem avaliar alocação de capital.'
    },
    # ajustado 28/07
    'P/EBIT': {
        'critico': {'min': float('-inf'), 'max': -0.1 },
        'pessimo': {'min': 20, 'max': float('inf')},
        'ruim': {'min': 15, 'max': 20},
        'moderado': {'min': 10, 'max': 15},
        'bom': {'min': 5, 'max': 10},
        'otimo': {'min': 0, 'max': 5},
        'descricao': 'Preço da ação dividido pelo EBIT por ação. Mede valuation em relação ao lucro operacional.',
        'agrupador': 'Valuation',
        'descrcritico': 'Prejuízo operacional, com EBIT negativo, indicando ineficiência grave ou investimentos pesados. Comum em empresas em crise (ex.: OIBR3). Investidores devem evitar, salvo plano robusto de recuperação.',
        'descrpessimo': 'Empresa extremamente cara, preço é mais de 20 vezes o EBIT, sugerindo sobrevalorização. Comum em setores de tecnologia (ex.: NUBR33). Investidores devem comparar com peers e avaliar crescimento projetado.',
        'descrruim': 'Empresa cara, comum em setores de crescimento (ex.: saúde, RDOR3). Investidores devem analisar se o crescimento justifica o preço e comparar com média setorial.',
        'descrmoderado': 'Avaliação razoável, comum em setores maduros (ex.: varejo, LREN3). Investidores devem verificar consistência do EBIT e tendências setoriais.',
        'descrbom': 'Subvalorizada, oportunidade em setores cíclicos (ex.: mineração, VALE3). Investidores devem confirmar fluxo de caixa livre e sustentabilidade do EBIT.',
        'descrotimo': 'Extremamente subvalorizada, sugere oportunidade ou EBIT inflado. Comum em setores em recuperação (ex.: PETR4). Investidores devem verificar fluxo de caixa livre e ROIC.'
    },
# ajustado 28/07
    'P/EBITDA': {
        'critico': {'min': float('-inf'), 'max': -0.1},
        'pessimo': {'min': 15, 'max': float('inf')},
        'ruim': {'min': 10, 'max': 15},
        'moderado': {'min': 6, 'max': 10},
        'bom': {'min': 3, 'max': 6},
        'otimo': {'min': 0, 'max': 3},
        'descricao': 'Preço da ação dividido pelo EBITDA por ação. Mede valuation em relação ao fluxo de caixa operacional.',
        'agrupador': 'Valuation',
        'descrcritico': 'Prejuízo operacional, com EBITDA negativo, indicando ineficiência grave ou problemas estruturais. Comum em empresas em crise (ex.: OIBR3). Investidores devem evitar, salvo plano robusto de recuperação.',
        'descrpessimo': 'Empresa extremamente cara, preço é mais de 15 vezes o EBITDA, sugerindo sobrevalorização. Comum em setores de tecnologia (ex.: NUBR33). Investidores devem comparar com peers e avaliar crescimento projetado.',
        'descrruim': 'Empresa cara, comum em setores de crescimento (ex.: saúde, RDOR3). Investidores devem analisar se o crescimento justifica o preço e comparar com média setorial.',
        'descrmoderado': 'Avaliação razoável, comum em setores maduros (ex.: varejo, LREN3). Investidores devem verificar consistência do EBITDA e tendências setoriais.',
        'descrbom': 'Subvalorizada, oportunidade em setores cíclicos (ex.: mineração, VALE3). Investidores devem confirmar fluxo de caixa livre e sustentabilidade do EBITDA.',
        'descrotimo': 'Extremamente subvalorizada, sugere oportunidade ou EBITDA inflado. Comum em setores em recuperação (ex.: PETR4). Investidores devem verificar fluxo de caixa livre e ROIC.'
    },
    'P/L': {
        'critico': {'min': float('-inf'), 'max': 0},
        'pessimo': {'min': 20, 'max': float('inf')},
        'ruim': {'min': 15, 'max': 20},
        'moderado': {'min': 10, 'max': 15},
        'bom': {'min': 5, 'max': 10},
        'otimo': {'min': 0, 'max': 5},
        'descricao': 'Preço da ação dividido pelo lucro por ação. Mede valuation em relação ao lucro líquido.',
        'agrupador': 'Valuation',
        'descrcritico': 'Prejuízo por ação, indicando problemas operacionais ou financeiros graves. Comum em empresas em crise (ex.: OIBR3). Investidores devem evitar, salvo plano robusto de recuperação.',
        'descrpessimo': 'Empresa extremamente cara, preço é mais de 20 vezes o lucro por ação, sugerindo sobrevalorização. Comum em setores de tecnologia (ex.: NUBR33). Investidores devem comparar com peers e avaliar crescimento projetado.',
        'descrruim': 'Empresa cara, comum em setores de crescimento (ex.: saúde, RDOR3). Investidores devem analisar se o crescimento justifica o preço e comparar com média setorial.',
        'descrmoderado': 'Avaliação razoável, comum em setores maduros (ex.: varejo, LREN3). Investidores devem verificar consistência do lucro e tendências setoriais.',
        'descrbom': 'Subvalorizada, oportunidade em setores cíclicos (ex.: mineração, VALE3). Investidores devem confirmar fluxo de caixa livre e sustentabilidade do lucro.',
        'descrotimo': 'Extremamente subvalorizada, sugere oportunidade ou lucro inflado. Comum em setores em recuperação (ex.: PETR4). Investidores devem verificar fluxo de caixa livre e ROE.'
    },
    'P/VP': {
        'critico': {'min': float('-inf'), 'max': 0},
        'pessimo': {'min': 2, 'max': float('inf')},
        'ruim': {'min': 1.5, 'max': 2},
        'moderado': {'min': 1, 'max': 1.5},
        'bom': {'min': 0.5, 'max': 1},
        'otimo': {'min': 0, 'max': 0.5},
        'descricao': 'Preço da ação dividido pelo valor patrimonial por ação. Mede valuation em relação ao patrimônio.',
        'agrupador': 'Valuetion',
        'descrcritico': 'Patrimônio líquido por ação negativo, indicando risco crítico de insolvência. Comum em empresas em crise (ex.: OIBR3). Investidores devem evitar, salvo plano robusto de recuperação.',
        'descrpessimo': 'Empresa extremamente cara, preço é mais de 2 vezes o valor patrimonial por ação, sugerindo sobrevalorização. Comum em setores de tecnologia (ex.: NUBR33). Investidores devem comparar com peers e avaliar qualidade do patrimônio.',
        'descrruim': 'Empresa cara, comum em setores de crescimento (ex.: saúde, RDOR3). Investidores devem analisar se o crescimento justifica o preço e comparar com média setorial.',
        'descrmoderado': 'Avaliação razoável, comum em setores maduros (ex.: varejo, LREN3). Investidores devem verificar consistência do patrimônio e tendências setoriais.',
        'descrbom': 'Subvalorizada, oportunidade em setores cíclicos (ex.: mineração, VALE3). Investidores devem confirmar qualidade do patrimônio e fluxo de caixa livre.',
        'descrotimo': 'Extremamente subvalorizada, sugere oportunidade ou patrimônio inflado. Comum em setores em recuperação (ex.: PETR4). Investidores devem verificar fluxo de caixa livre e ROE.'
    },
     # verificar
    'P/Ativo': {
        'critico': {'min': float('-inf'), 'max': 0},
        'pessimo': {'min': 0, 'max': 0.2},
        'ruim': {'min': 0.2, 'max': 0.3},
        'moderado': {'min': 0.3, 'max': 0.5},
        'bom': {'min': 0.5, 'max': 0.8},
        'otimo': {'min': 0.8, 'max': float('inf')},
        'descricao': 'Patrimônio líquido dividido pelos ativos totais. Mede proporção de ativos financiada por capital próprio.',
        'agrupador': 'Alavancagem Financeira',
        'descrcritico': 'Patrimônio líquido negativo, passivos superam ativos, indicando risco crítico de insolvência. Comum em empresas em crise (ex.: OI). Investidores devem exigir plano robusto de recuperação.',
        'descrpessimo': 'Alta alavancagem, patrimônio representa menos de 20% dos ativos, indicando dependência de dívidas. Comum em setores com dívidas elevadas (ex.: construção). Investidores devem analisar capacidade de pagamento de juros.',
        'descrruim': 'Alavancagem elevada, patrimônio financia 20-30% dos ativos, indicando risco moderado. Investidores devem monitorar gestão da dívida e geração de caixa.',
        'descrmoderado': 'Alavancagem moderada, patrimônio financia 30-50% dos ativos, aceitável em setores com alavancagem moderada (ex.: indústria). Investidores devem monitorar gestão da dívida e geração de caixa.',
        'descrbom': 'Baixa alavancagem, estrutura sólida, patrimônio representa 50-80% dos ativos. Comum em setores estáveis (ex.: utilities). Sinal de segurança financeira.',
        'descrotimo': 'Alavancagem mínima, estrutura excepcional, patrimônio financia mais de 80% dos ativos. Ideal para setores voláteis (ex.: commodities). Verificar se excesso de capital próprio limita crescimento.'
    },
    'PL/Ativos': {
        'critico': {'min': float('-inf'), 'max': 0},
        'pessimo': {'min': 0, 'max': 0.2},
        'ruim': {'min': 0.2, 'max': 0.5},
        'moderado': {'min': 0.5, 'max': 0.8},
        'bom': {'min': 0.8, 'max': 1},
        'otimo': {'min': 1, 'max': float('inf')},
        'descricao': 'Patrimônio líquido dividido pelos ativos totais. Mede proporção de ativos financiada por capital próprio.',
        'agrupador': 'Eficiência',
        'descrcritico': 'Patrimônio líquido negativo, passivos superam ativos, indicando risco crítico de insolvência. Comum em empresas em crise (ex.: OI). Investidores devem exigir plano robusto de recuperação.',
        'descrpessimo': 'Alta alavancagem, patrimônio representa menos de 20% dos ativos, indicando dependência de dívidas. Comum em setores com dívidas elevadas (ex.: construção). Investidores devem analisar capacidade de pagamento de juros.',
        'descrruim': 'Alavancagem moderada, patrimônio financia 20-50% dos ativos, aceitável em setores com alavancagem moderada (ex.: indústria). Investidores devem monitorar gestão da dívida e geração de caixa.',
        'descrmoderado': 'Baixa alavancagem, estrutura sólida, patrimônio representa 50-80% dos ativos. Comum em setores estáveis (ex.: utilities). Sinal de segurança financeira.',
        'descrbom': 'Alavancagem mínima, estrutura excepcional, patrimônio financia 80-100% dos ativos. Ideal para setores voláteis (ex.: commodities). Verificar se excesso de capital próprio limita crescimento.',
        'descrotimo': 'Estrutura financeira excepcional, patrimônio excede ativos, indicando altíssima capitalização. Comum em empresas com reservas elevadas (ex.: Weg). Investidores devem verificar alocação de capital.'
    },
    'PSR': {
        'critico': {'min': float('-inf'), 'max': 0},
        'pessimo': {'min': 3, 'max': float('inf')},
        'ruim': {'min': 2, 'max': 3},
        'moderado': {'min': 1, 'max': 2},
        'bom': {'min': 0.5, 'max': 1},
        'otimo': {'min': 0, 'max': 0.5},
        'descricao': 'Valor de mercado dividido pela receita líquida. Mede valuation em relação às vendas.',
        'agrupador': 'Valuation',
        'descrcritico': 'Receita líquida negativa, indicando problemas operacionais graves ou falha contábil. Comum em empresas em crise (ex.: OIBR3). Investidores devem evitar, salvo plano robusto de recuperação.',
        'descrpessimo': 'Empresa extremamente cara, valor de mercado é mais de 3 vezes a receita líquida, sugerindo sobrevalorização. Comum em setores de tecnologia (ex.: NUBR33). Investidores devem comparar com peers e avaliar crescimento projetado.',
        'descrruim': 'Empresa cara, comum em setores de crescimento (ex.: saúde, RDOR3). Investidores devem analisar se o crescimento justifica o preço e comparar com média setorial.',
        'descrmoderado': 'Avaliação razoável, comum em setores maduros (ex.: varejo, LREN3). Investidores devem verificar consistência da receita e tendências setoriais.',
        'descrbom': 'Subvalorizada, oportunidade em setores cíclicos (ex.: mineração, VALE3). Investidores devem confirmar fluxo de caixa livre e sustentabilidade da receita.',
        'descrotimo': 'Extremamente subvalorizada, sugere oportunidade ou receita inflada. Comum em setores em recuperação (ex.: PETR4). Investidores devem verificar fluxo de caixa livre e Margem Bruta.'
    },
    # pendente
    'P/Sales Ratio': {
        'critico': {'min': float('-inf'), 'max': 0},
        'pessimo': {'min': 3, 'max': float('inf')},
        'ruim': {'min': 2, 'max': 3},
        'moderado': {'min': 1, 'max': 2},
        'bom': {'min': 0.5, 'max': 1},
        'otimo': {'min': 0, 'max': 0.5},
        'descricao': 'Valor de mercado dividido pela receita líquida. Mede valuation em relação às vendas.',
        'agrupador': 'Valuation',
        'descrcritico': 'Receita líquida negativa, indicando problemas operacionais graves ou falha contábil. Comum em empresas em crise (ex.: OIBR3). Investidores devem evitar, salvo plano robusto de recuperação.',
        'descrpessimo': 'Empresa extremamente cara, valor de mercado é mais de 3 vezes a receita líquida, sugerindo sobrevalorização. Comum em setores de tecnologia (ex.: NUBR33). Investidores devem comparar com peers e avaliar crescimento projetado.',
        'descrruim': 'Empresa cara, comum em setores de crescimento (ex.: saúde, RDOR3). Investidores devem analisar se o crescimento justifica o preço e comparar com média setorial.',
        'descrmoderado': 'Avaliação razoável, comum em setores maduros (ex.: varejo, LREN3). Investidores devem verificar consistência da receita e tendências setoriais.',
        'descrbom': 'Subvalorizada, oportunidade em setores cíclicos (ex.: mineração, VALE3). Investidores devem confirmar fluxo de caixa livre e sustentabilidade da receita.',
        'descrotimo': 'Extremamente subvalorizada, sugere oportunidade ou receita inflada. Comum em setores em recuperação (ex.: PETR4). Investidores devem verificar fluxo de caixa livre e Margem Bruta.'
    },
    'ROA': {
        'critico': {'min': float('-inf'), 'max': 0},
        'pessimo': {'min': 0, 'max': 0.03},
        'ruim': {'min': 0.03, 'max': 0.08},
        'moderado': {'min': 0.08, 'max': 0.15},
        'bom': {'min': 0.15, 'max': 0.25},
        'otimo': {'min': 0.25, 'max': float('inf')},
        'descricao': 'Lucro líquido dividido pelos ativos totais. Mede retorno sobre os ativos.',
        'agrupador': 'Retorno',
        'descrcritico': 'Prejuízo, ativos gerando retorno negativo, indicando ineficiência grave. Comum em empresas em crise ou com investimentos mal sucedidos (ex.: OI). Investidores devem analisar causas e potencial de recuperação.',
        'descrpessimo': 'Baixa eficiência, retorno limitado dos ativos, comum em setores intensivos em capital (ex.: infraestrutura). Investidores devem investigar potencial de melhoria operacional.',
        'descrruim': 'Eficiência moderada, desempenho razoável, aceitável em setores com margens moderadas (ex.: CSNA3). Investidores devem analisar tendência do ROA e composição dos ativos para avaliar consistência.',
        'descrmoderado': 'Boa eficiência, retorno sólido dos ativos, comum em setores com alta rotatividade (ex.: varejo). Investidores devem confirmar sustentabilidade do ROA e qualidade dos ativos.',
        'descrbom': 'Alta eficiência, retorno robusto dos ativos, comum em setores dinâmicos (ex.: LREN3). Investidores devem verificar consistência frente a riscos setoriais.',
        'descrotimo': 'Eficiência excepcional, retorno elevado dos ativos, comum em setores de alta margem (ex.: TOTS3). Investidores devem verificar sustentabilidade e proteção contra concorrência.'
    },
    'ROE': {
        'critico': {'min': float('-inf'), 'max': 0},
        'pessimo': {'min': 0, 'max': 0.05},
        'ruim': {'min': 0.05, 'max': 0.1},
        'moderado': {'min': 0.1, 'max': 0.15},
        'bom': {'min': 0.15, 'max': 0.25},
        'otimo': {'min': 0.25, 'max': float('inf')},
        'descricao': 'Lucro líquido dividido pelo patrimônio líquido. Mede retorno sobre o capital dos acionistas.',
        'agrupador': 'Retorno',
        'descrcritico': 'Prejuízo, retorno negativo sobre o patrimônio, indicando ineficiência grave. Comum em empresas em crise (ex.: OI). Investidores devem analisar plano de recuperação.',
        'descrpessimo': 'Retorno muito baixo, indicando rentabilidade limitada. Comum em setores competitivos (ex.: varejo). Investidores devem investigar potencial de melhoria.',
        'descrruim': 'Retorno baixo, desempenho fraco, comum em setores cíclicos (ex.: GGBR4). Investidores devem analisar tendência do lucro e riscos setoriais.',
        'descrmoderado': 'Retorno moderado, desempenho razoável, aceitável em setores maduros (ex.: CSNA3). Investidores devem verificar sustentabilidade do lucro.',
        'descrbom': 'Bom retorno, desempenho sólido, comum em setores estáveis (ex.: ABEV3). Investidores devem confirmar consistência do lucro.',
        'descrotimo': 'Retorno excepcional, refletindo forte rentabilidade. Comum em setores de alta margem (ex.: TOTS3). Investidores devem verificar sustentabilidade frente a riscos setoriais.'
    },
    'ROIC': {
        'critico': {'min': float('-inf'), 'max': 0},
        'pessimo': {'min': 0, 'max': 0.05},
        'ruim': {'min': 0.05, 'max': 0.1},
        'moderado': {'min': 0.1, 'max': 0.15},
        'bom': {'min': 0.15, 'max': 0.2},
        'otimo': {'min': 0.2, 'max': float('inf')},
        'descricao': 'Lucro operacional após impostos dividido pelo capital investido. Mede retorno sobre o capital total.',
        'agrupador': 'Retorno',
        'descrcritico': 'Prejuízo operacional, retorno negativo sobre o capital investido, indicando ineficiência grave. Comum em empresas em crise (ex.: OIBR3). Investidores devem evitar, salvo plano robusto de recuperação.',
        'descrpessimo': 'Retorno muito baixo, indicando eficiência limitada. Comum em setores intensivos em capital (ex.: infraestrutura). Investidores devem investigar potencial de melhoria operacional.',
        'descrruim': 'Retorno limitado, comum em setores cíclicos (ex.: GGBR4). Investidores devem analisar tendência do ROIC e riscos setoriais.',
        'descrmoderado': 'Retorno moderado, aceitável em setores maduros (ex.: varejo, LREN3). Investidores devem verificar sustentabilidade do ROIC e composição do capital.',
        'descrbom': 'Bom retorno, desempenho sólido, comum em setores com barreiras de entrada (ex.: ABEV3). Investidores devem confirmar consistência do ROIC.',
        'descrotimo': 'Retorno excepcional, refletindo eficiência superior. Comum em setores de alta margem (ex.: TOTS3). Investidores devem verificar sustentabilidade frente a riscos setoriais.'
    },
    'VPA': {
        'critico': {'min': float('-inf'), 'max': 0},
        'pessimo': {'min': 0, 'max': 0.5},
        'ruim': {'min': 0.5, 'max': 1},
        'moderado': {'min': 1, 'max': 2},
        'bom': {'min': 2, 'max': 5},
        'otimo': {'min': 5, 'max': float('inf')},
        'descricao': 'Patrimônio líquido por ação. Mede valor contábil por ação.',
        'agrupador': 'Outros',
        'descrcritico': 'Patrimônio líquido por ação negativo, indicando risco crítico de insolvência. Comum em empresas em crise (ex.: OIBR3). Investidores devem evitar, salvo plano robusto de recuperação.',
        'descrpessimo': 'Valor patrimonial muito baixo, indicando fragilidade financeira. Comum em empresas com dívidas elevadas (ex.: construção). Investidores devem analisar capacidade de geração de lucro.',
        'descrruim': 'Valor patrimonial limitado, comum em setores cíclicos (ex.: GGBR4). Investidores devem verificar tendência do patrimônio e riscos setoriais.',
        'descrmoderado': 'Valor patrimonial moderado, aceitável em setores maduros (ex.: indústria). Investidores devem confirmar sustentabilidade do patrimônio.',
        'descrbom': 'Bom valor patrimonial, desempenho sólido, comum em setores estáveis (ex.: ABEV3). Investidores devem verificar consistência do patrimônio.',
        'descrotimo': 'Valor patrimonial excepcional, refletindo forte capitalização. Comum em setores de alta margem (ex.: TOTS3). Investidores devem verificar alocação de capital.'
    },
    'CAGR Lucros 5 anos': {
        'critico': {'min': float('-inf'), 'max': 0},
        'pessimo': {'min': 0, 'max': 0.05},
        'ruim': {'min': 0.05, 'max': 0.1},
        'moderado': {'min': 0.1, 'max': 0.15},
        'bom': {'min': 0.15, 'max': 0.2},
        'otimo': {'min': 0.2, 'max': float('inf')},
        'descricao': 'Taxa composta de crescimento anual dos lucros nos últimos 5 anos. Mede crescimento de rentabilidade.',
        'agrupador': 'Outros',
        'descrcritico': 'Crescimento negativo dos lucros, indicando declínio ou prejuízo recorrente. Comum em empresas em crise (ex.: OIBR3). Investidores devem evitar, salvo plano robusto de recuperação.',
        'descrpessimo': 'Crescimento muito baixo, indicando estagnação. Comum em setores maduros com baixa inovação (ex.: varejo tradicional). Investidores devem investigar potencial de melhoria.',
        'descrruim': 'Crescimento limitado, comum em setores cíclicos (ex.: GGBR4). Investidores devem analisar tendências setoriais e estratégias de crescimento.',
        'descrmoderado': 'Crescimento moderado, aceitável em setores estáveis (ex.: ABEV3). Investidores devem verificar sustentabilidade do crescimento e barreiras de entrada.',
        'descrbom': 'Bom crescimento, desempenho sólido, comum em setores com expansão moderada (ex.: LREN3). Investidores devem confirmar consistência do crescimento.',
        'descrotimo': 'Crescimento excepcional, refletindo forte expansão. Comum em setores de alta margem ou inovação (ex.: TOTS3). Investidores devem verificar sustentabilidade frente a riscos setoriais.'
    },
    'Divida bruta': {
        'critico': {'min': float('-inf'), 'max': 0},
        'pessimo': {'min': 0, 'max': 0.05},
        'ruim': {'min': 0.05, 'max': 0.1},
        'moderado': {'min': 0.1, 'max': 0.15},
        'bom': {'min': 0.15, 'max': 0.2},
        'otimo': {'min': 0.2, 'max': float('inf')},
        'descricao': 'Taxa composta de crescimento anual dos lucros nos últimos 5 anos. Mede crescimento de rentabilidade.',
        'agrupador': 'Outros',
        'descrcritico': 'Crescimento negativo dos lucros, indicando declínio ou prejuízo recorrente. Comum em empresas em crise (ex.: OIBR3). Investidores devem evitar, salvo plano robusto de recuperação.',
        'descrpessimo': 'Crescimento muito baixo, indicando estagnação. Comum em setores maduros com baixa inovação (ex.: varejo tradicional). Investidores devem investigar potencial de melhoria.',
        'descrruim': 'Crescimento limitado, comum em setores cíclicos (ex.: GGBR4). Investidores devem analisar tendências setoriais e estratégias de crescimento.',
        'descrmoderado': 'Crescimento moderado, aceitável em setores estáveis (ex.: ABEV3). Investidores devem verificar sustentabilidade do crescimento e barreiras de entrada.',
        'descrbom': 'Bom crescimento, desempenho sólido, comum em setores com expansão moderada (ex.: LREN3). Investidores devem confirmar consistência do crescimento.',
        'descrotimo': 'Crescimento excepcional, refletindo forte expansão. Comum em setores de alta margem ou inovação (ex.: TOTS3). Investidores devem verificar sustentabilidade frente a riscos setoriais.'
        }
}

wbsaida = openpyxl.Workbook()


# define selenium webdriver options
options = webdriver.ChromeOptions()

# create selenium webdriver instance
driver = webdriver.Chrome(options=options)

def classificar_divida_ebit(valor):
        indicadortratado = tratamento2(dict_stocks[stock].get('LPA'))
        print("Pl" , indicadortratado)
        if valor == float('-inf') or valor < 0:
            return {'classificacao': 'pessimo', 'descricao': 'teste'}
        elif 0 <= valor < 1.5:
            return {'classificacao': 'pessimo', 'descricao': 'teste'}
        elif 1.5 <= valor < 3:
            return {'classificacao': 'pessimo', 'descricao': 'teste'}
        elif 3 <= valor < 4:
            return {'classificacao': 'pessimo', 'descricao': 'teste'}
        elif valor >= 4:
            return {'classificacao': 'pessimo', 'descricao': 'teste'}
        else:
            return {'classificacao': 'pessimo', 'descricao': 'teste'}


def categorizar_valor(metrica, valor):
    try:
        if metrica not in MetricasStatus:
            return 'Métrica não reconhecida'
        valor2 = float(valor)
        #print("categoriza_valor", valor2)
        for categoria, limites in MetricasStatus[metrica].items():

            if categoria in ['descricao', 'agrupador','descrcritico','descrpessimo','descrruim','descrmoderado','descrbom','descrotimo']:
                continue
            if limites['min'] <= valor2 < limites['max']:
                return categoria
        return 'Valor fora do alcance definido'
    except Exception as e:
        print("Debug - limites:", limites)
        print(f"Erro inesperado categorizar: {e}")
        print(f"Erro metrica: {e}")
        print("categorizar_valor - Erro" , metrica)
        print("categoriza_valor", valor2)
        print('categoriza_valor' ,stock)
    finally:
       # print("categorizar_valor - OK2")
       pass

def criaPlanilhaIndRentabilidade(wbsaida):
    wbsaida.create_sheet('IndiRentabilidade')
    IndiRentabilidade = wbsaida['IndiRentabilidade']
    IndiRentabilidade.append(
        ['Agrupador', 'Fonte', 'ATIVO', 'Indicador', 'Valor', 'Referencia','Critico', 'Pessimo','Ruim', 'Moderado', 'Bom', 'Otimo', 'Descrição'])

    for cell in IndiRentabilidade[1]:  # Apenas o cabeçalho
        cell.fill = filltitulo
        cell.font = font_branca

def tratamento(indicador):
    indicador2 = indicador

    try:
        if indicador2 in ["-", "--","--%"]:
            indicador2 = 0
        elif indicador2 is None or is_null_zero_or_spaces(indicador2):
            indicador2 = 0
        else:
            if indicador2 == "":  # Verificação adicionada para string vazia
                indicador2 = 0
            else:
                indicador2 = float(indicador2.strip('%')) / 100
        return indicador2

    except Exception as e:
        print(f"Erro inesperado tratamento : {e}", "metrica  ", metricasts, " indicador  ", indicador, " stock  ",stock)
        # print(metrica)  # Certifique-se de que metrica está definida
        # print(indicadortratado)  # Certifique-se de que indicadortratado está definida
        #print('tratamneto - erro', stock, "   ", metrica)


    finally:
       # print('tratamneto OK')
       pass
def tratamento3(indicador):
    indicador2 = indicador

    try:
        if indicador2 in ["-", "--","--%"]:
            indicador2 = 0
        elif indicador2 is None or is_null_zero_or_spaces(indicador2):
            indicador2 = 0
        else:
            if indicador2 == "":  # Verificação adicionada para string vazia
                indicador2 = 0
            else:
                indicador2 = float(indicador2)
        return indicador2

    except Exception as e:
        print(f"Erro inesperado tratamento 3 : {e}", " metrica  ", metricasts, " indicador  ", indicador, " stock  ",stock)
        # print(metrica)  # Certifique-se de que metrica está definida
        # print(indicadortratado)  # Certifique-se de que indicadortratado está definida
        #print('tratamneto3 - erro', stock, "   ", metrica)


    finally:
        #print('tratamneto3 OK')
        pass
# Certifique-se de que as variáveis `metrica` e `indicadortratado` estão definidas corretamente no contexto onde a função é chamada.

def tratamento2(indicador):
    indicador2 = indicador

    try:
        if indicador2 in ["-", "--","--%"]:
            indicador2 = 0
        elif indicador2 is None or is_null_zero_or_spaces(indicador2):
            indicador2 = 0
        elif indicador2 == "":  # Verificação adicionada para string vazia
            indicador2 = 0
        else:
            indicador2 = float(indicador2)
        return indicador2


        return indicador2
    except Exception as e:
      # print(f"Erro inesperado tratamento2: {e}", " metrica  ", metrica, " indicador  ", indicador ," stock ", stock)
        # print(metrica)  # Certifique-se de que metrica está definida
        # print(indicadortratado)  # Certifique-se de que indicadortratado está definida
        print('tratamneto2 - erro', stock)


    finally:
        #print('tratamneto2 OK', indicador)
        pass
def gravaIndiEficiênciaoStaus(wsIndiRentabilidade, dict_stocks, stock):
    # fontes ['StatusInvest', 'Fundamentus']



    global linha2
    global metricasts
    #linha2 = 1
    try:
        #print(dict_stocks)
        for metrica, detalhes in MetricasStatus.items():
    #        print(f'Métrica: {metrica}')
            linha2 += 1
            metricasts = metrica
            if metrica in ['Giro ativos', 'Div. liquida/PL','Div. liquida/EBITDA','Div. liquida/EBIT','PL/Ativos',
                           'Passivos/Ativos','Liq. corrente','P/L','PEG Ratio','P/VP','EV/EBITDA','EV/EBIT',
                            'P/EBITDA','P/EBIT','VPA','P/Ativo','LPA',
                            'P/SR','P/Ativo Circ. Liq.']:
             #   print('IF tratamneto2 ',metrica )

                indicadortratado = tratamento2(dict_stocks[stock].get(metrica))
                valor_pl = indicadortratado
                categoria_pl = categorizar_valor(metrica, (valor_pl))
                teste = classificar_divida_ebit(valor_pl)
                #print("descricao:", teste['descricao'])
                #print("Classificação:", teste['classificacao'])
                #print(teste)
                print(teste['classificacao'])  # saída: 'pessimo'
                print(teste['descricao'])  # saída: 'teste'
              #  print("categoria " + categoria_pl)
               # print("categoria tratamento2 " + categoria_pl)
            elif metrica in ['Valor atual','LIQUIDEZ MEDIA DIARIA','Patrimonio liquido',
                             'Ativos','Ativo circulante','Divida bruta','Disponibilidade',
                             'Divida liquida','Valor de mercado','Valor de firma']:
               # print('IF tratamneto3 ', metrica)
               # print("categoria " + categoria_pl)
                indicadortratado = tratamento3(dict_stocks[stock].get(metrica))
                valor_pl = indicadortratado
                categoria_pl = categorizar_valor(metrica, (valor_pl))
               # print("categoria tratamento3" + categoria_pl)
            else:
               # print('IF tratamneto ', metrica)
                indicadortratado = tratamento(dict_stocks[stock].get(metrica))
                valor_pl = indicadortratado
                categoria_pl = categorizar_valor(metrica,(valor_pl * 100))
                #print("categoria tratamento" + categoria_pl)
                # Certifique-se de que 'ROE' é o valor correto para a métrica
   #         print(f'O índice P/L {valor_pl} é categorizado como: {categoria_pl}')
   #         print(f"  Agrupador: {detalhes['agrupador']}")


            wsIndiRentabilidade.cell(row=linha2, column=1, value=detalhes['agrupador'])
            wsIndiRentabilidade.cell(row=linha2, column=2, value='StausInvest')
            wsIndiRentabilidade.cell(row=linha2, column=3, value=stock)
          #  print('celula - Indicador', metrica)
            wsIndiRentabilidade.cell(row=linha2, column=4, value=metrica)
            if metrica in ['Giro ativos', 'Div. liquida/PL','Div. liquida/EBITDA','Div. liquida/EBITDA',
                           'Div. liquida/EBIT','PL/Ativos','Passivos/Ativos','Liq. corrente',
                           'P/L','PEG Ratio','P/VP','EV/EBITDA','EV/EBIT',
                            'P/EBITDA','P/EBIT','VPA','P/Ativo','LPA',
                            'P/SR','P/Ativo Circ. Liq.']:
               # print('IF da celula - Indicador', metrica )
               # print('IF da celula - valor', valor_pl)
                wsIndiRentabilidade.cell(row=linha2, column=5, value=valor_pl).number_format = numbers.FORMAT_NUMBER_00
            elif  metrica in ['Valor atual','LIQUIDEZ MEDIA DIARIA','Patrimonio liquido',
                             'Ativos','Ativo circulante','Divida bruta','Disponibilidade',
                             'Divida liquida','Valor de mercado','Valor de firma']:
                  #print('IF da celula - valor', valor_pl)
                  #print('IF da celula - Indicador', metrica)
                  wsIndiRentabilidade.cell(row=linha2, column=5, value=valor_pl).number_format = 'R$ #,##0.00'
            else:
                wsIndiRentabilidade.cell(row=linha2, column=5, value=valor_pl).number_format = numbers.FORMAT_PERCENTAGE_00

            wsIndiRentabilidade.cell(row=linha2, column=7,
                             value=f"Mínimo = {detalhes['critico']['min']}, Máximo = {detalhes['critico']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=8,
                             value=f"Mínimo = {detalhes['pessimo']['min']}, Máximo = {detalhes['pessimo']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=9,
                                     value=f"Mínimo = {detalhes['ruim']['min']}, Máximo = {detalhes['ruim']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=10,
                                     value=f"Mínimo = {detalhes['moderado']['min']}, Máximo = {detalhes['moderado']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=11,
                                     value=f"Mínimo = {detalhes['bom']['min']}, Máximo = {detalhes['bom']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=12,
                                     value=f"Mínimo = {detalhes['otimo']['min']}, Máximo = {detalhes['otimo']['max']}")
            #print(detalhes)

            #if metrica == 'Div. liquida/EBITDA':
            if categoria_pl == 'critico':
                wsIndiRentabilidade.cell(row=linha2, column=6, value=categoria_pl).fill = fillvermelho
                wsIndiRentabilidade.cell(row=linha2, column=11, value=f"{detalhes['descrcritico']}")
            if categoria_pl == 'pessimo':
                wsIndiRentabilidade.cell(row=linha2, column=6, value=categoria_pl).fill = fillvermelho
                wsIndiRentabilidade.cell(row=linha2, column=11, value=f"{detalhes['descrpessimo']}")
            if categoria_pl == 'ruim':
                wsIndiRentabilidade.cell(row=linha2, column=6, value=categoria_pl).fill = fillvermelho
                wsIndiRentabilidade.cell(row=linha2, column=11, value=f"{detalhes['descrruim']}")
            if  categoria_pl == 'moderado':
                wsIndiRentabilidade.cell(row=linha2, column=6, value=categoria_pl).fill =fillamarelo
                wsIndiRentabilidade.cell(row=linha2, column=11, value=f"{detalhes['descrmoderado']}")
            if  categoria_pl == 'bom':
                wsIndiRentabilidade.cell(row=linha2, column=6, value=categoria_pl).fill = fillverde
                wsIndiRentabilidade.cell(row=linha2, column=11, value=f"{detalhes['descrbom']}")
            if categoria_pl == 'otimo':
                wsIndiRentabilidade.cell(row=linha2, column=6, value=categoria_pl).fill =fillazul
                wsIndiRentabilidade.cell(row=linha2, column=11, value=f"{detalhes['descrotimo']}")


    except Exception as e:
        print(f"Erro inesperado grava planilha2: {e}")
        print(metrica)
        print(indicadortratado)

        print('gravaIndiEficiênciaoStaus - erro' ,  stock,"    ", metrica)
    finally:
        print('gravaIndiEficiênciaoStaus  OK''', stock)

def is_null_zero_or_spaces(variable):
    # Verifica se a variável é None
    if variable is None:
        return True
    # Verifica se a variável é zero (0)
    elif variable == 0:
        return True
    # Verifica se a variável é uma string e contém apenas espaços
    elif isinstance(variable, str) and variable.strip() == '':
        return True
    elif variable == '-%':
        return True
    else:
        return False


def get_stock_soup(stock):
    ''' Get raw html from a stock '''

    # access the stock urlww
    driver.get(f'https://statusinvest.com.br/acoes/{stock}')

    # get html from stock
    html = driver.find_element(By.ID, 'main-2').get_attribute('innerHTML')

    # remove accents from html and transform html into soup
    soup = BeautifulSoup(unidecode(html), 'html.parser')

    return soup


def soup_to_dict(soup):
    '''Get all data from stock soup and return as a dictionary '''
    keys, values = [], []

    # get divs from stock
    soup1 = soup.find('div', class_='pb-3 pb-md-5')
    soup2 = soup.find('div', class_='card rounded text-main-green-dark')
    soup3 = soup.find('div', class_='indicator-today-container')
    soup4 = soup.find(
        'div', class_='top-info info-3 sm d-flex justify-between mb-3')
    soups = [soup1, soup2, soup3, soup4]

    for s in soups:
        # get only titles from a div and append to keys
        titles = s.find_all('h3', re.compile('title m-0[^"]*'))
        titles = [t.get_text() for t in titles]
        keys += titles

        # get only numbers from a div and append to values
        numbers = s.find_all('strong', re.compile('value[^"]*'))
        numbers = [n.get_text()for n in numbers]
        values += numbers

    # remove unused key and insert needed keys
    keys.remove('PART. IBOV')
    keys.insert(6, 'TAG ALONG')
    keys.insert(7, 'LIQUIDEZ MEDIA DIARIA')

    # clean keys list
    keys = [k.replace('\nhelp_outline', '').strip() for k in keys]
    keys = [k for k in keys if k != '']

    # clean values list
    values = [v.replace('\nhelp_outline', '').strip() for v in values]
    values = [v.replace('.', '').replace(',', '.') for v in values]

    # create a dictionary using keys and values from indicators
    d = {k: v for k, v in zip(keys, values)}

    return d


if __name__ == "__main__":
    dict_stocks = {}
    criaPlanilhaIndRentabilidade(wbsaida)
    wsIndiRentabilidade = wbsaida['IndiRentabilidade']

    # start timer
    start = time.time()

    # read file with stocks codes to get stock information
    with open('stocks.txt', 'r') as f:
        stocks = f.read().splitlines()

        # get stock information and create excel sheet
        for stock in stocks:
            #print("stock :"  ,stock)
            try:
                # get data and transform into dictionary
                soup = get_stock_soup(stock)
                dict_stock = soup_to_dict(soup)
                dict_stocks[stock] = dict_stock
                gravaIndiEficiênciaoStaus(wsIndiRentabilidade, dict_stocks, stock)
            except:
                # if we not get the information... just skip it
                print(f'Could not get {stock} information', "    ", metricasts)

    # create dataframe using dictionary of stocks informations
    df = pd.DataFrame(dict_stocks)

    # replace missing values with NaN to facilitate processing
    df = df.replace(['', '-', '--', '-%', '--%'], np.nan)

    # write dataframe into csv file
    df.to_excel('stocks_data.xlsx', index_label='indicadores')

    # exit the driver
    driver.quit()

    # end timer
    end = time.time()
    wbsaida.save("StatusInvest.xlsx")
    print("teste")
    print(f'Brasilian stocks information got in {int(end-start)} s')
# silvio teste