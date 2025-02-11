import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

# Carregar as planilhas
wb1 = load_workbook('D:/Financeiro/GIT2/pyFundamentus-mainBKP/Fundamentus.xlsx')
wb2 = load_workbook('D:/Financeiro/GIT2/pyFundamentus-mainBKP/StatusInvest.xlsx')
ws1 = wb1['IndiRentabilidade']
ws2 = wb2['IndiRentabilidade']

# Criação de um novo Workbook
wb_resultante = load_workbook('D:/Financeiro/GIT2/pyFundamentus-mainBKP/planilha_juntas_horizontal.xlsx')
ws_resultante = wb_resultante.active

# Copiar os dados e as formatações da primeira planilha
for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row, min_col=1, max_col=ws1.max_column):
    for cell in row:
        new_cell = ws_resultante.cell(row=cell.row, column=cell.column, value=cell.value)
        new_cell.fill = cell.fill

# Ajustar a linha de início para a segunda planilha
start_row = ws1.max_row + 1

# Copiar os dados e as formatações da segunda planilha
for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, min_col=1, max_col=ws2.max_column):
    for cell in row:
        new_cell = ws_resultante.cell(row=cell.row + start_row - 1, column=cell.column, value=cell.value)
        new_cell.fill = cell.fill

# Salvar a planilha resultante com formatações
wb_resultante.save('D:/Financeiro/GIT2/pyFundamentus-mainBKP/planilha_juntas_horizontal.xlsx')
