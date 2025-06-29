import openpyxl
import re
import time
import numpy as np
import pandas as pd
from unidecode import unidecode
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
import requests
# teste silvio 3e
import warnings
from openpyxl.styles import numbers

#from teste01 import metrica

fillvermelho= PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") # Vermelho

fillverde= PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid") # Verde

fillamarelo =  PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # Amarelo


fillazul = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid") # Azul


#filltitulo =   PatternFill(start_color="#002060", end_color="#002060", fill_type="solid")

filltitulo = PatternFill(start_color="002060", end_color="002060", fill_type="solid")  # Azul escuro

font_branca = Font(color="FFFFFF")  # Branco

TITLES = [
    'Identificação', 'Resumo Financeiro', 'Cotações', 'Informações Básicas',
    'Oscilações', 'Indicadores de Valuation', 'Indicadores de Rentabilidade',
    'Indicadores de Endividamento', 'Balanço Patrimonial', 'Demonstrativo de Resultados'
]

linha2 = 1
metricasts= ""

MetricasStatus = {
    'M. Liquida': {
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 3},
        'moderado': {'min': 3, 'max':10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'Prejuízo líquido, indicando problemas financeiros graves após todas as despesas. Comum em empresas em crise ou com alta carga tributária/juros. Investidores devem analisar plano de recuperação.',
        'descrruim': 'Baixa lucratividade líquida, dificuldade em reter lucro após despesas, comum em setores competitivos (ex.: varejo). Investidores devem investigar potencial de redução de custos ou aumento de receita.',
        'descrmoderado': 'Lucratividade moderada, aceitável em setores com margens moderadas (ex.: indústria). Investidores devem analisar exposição a impostos, juros e custos operacionais para avaliar sustentabilidade.',
        'descrbom': 'Boa lucratividade líquida, eficiência sólida em toda a cadeia. Comum em setores com margens moderadas (ex.: bens de consumo). Investidores devem confirmar consistência do lucro.',
        'descotimo': 'Alta lucratividade líquida, desempenho excepcional, comum em setores de alta margem (ex.: tecnologia). Investidores devem verificar sustentabilidade frente a riscos setoriais ou macroeconômicos.'
    },
    'M. EBIT': {
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'pessimo',
        'descrruim': 'ruim',
        'descrmoderado': 'moderado',
        'descrbom': 'bom',
        'descotimo': 'otimo'    },
    'M. EBITDA': {
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'Endividamento crítico, leva mais de 4 anos de EBIT para quitar a dívida líquida, indicando alavancagem extrema e risco financeiro elevado. Comum em empresas em crise (ex.: OIBR3) ou setores cíclicos em baixa (ex.: construção, CSNA3). Em 2025, com juros de 10-12%, o custo da dívida pressiona ainda mais a geração de caixa, aumentando a vulnerabilidade a choques econômicos, como desaceleração ou aumento de taxas. Empresas nessa faixa frequentemente enfrentam dificuldades para cobrir juros, com risco de default ou necessidade de reestruturação. Investidores devem evitar, salvo sinais claros de recuperação, como redução agressiva de dívida, venda de ativos (ex.: Petrobrás em reestruturações passadas) ou aumento robusto do EBIT via eficiência operacional. Analise a cobertura de juros (EBIT/juros) e a tendência do EBIT nos últimos 3-5 anos para avaliar sustentabilidade. Setores intensivos em capital (ex.: infraestrutura) podem justificar alavancagem, mas exigem fluxo de caixa estável. Empresas com histórico de lucros voláteis ou em setores expostos a commodities (ex.: mineração) são particularmente arriscadas.',
        'descrruim': 'ruim',
        'descrmoderado': 'moderado',
        'descrbom': 'bom',
        'descotimo': 'otimo'    },
    'M. Bruta': {
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'pessimo',
        'descrruim': 'ruim',
        'descrmoderado': 'moderado',
        'descrbom': 'bom',
        'descotimo': 'otimo'    },
    'Div. liquida/PL': {
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'pessimo',
        'descrruim': 'ruim',
        'descrmoderado': 'moderado',
        'descrbom': 'bom',
        'descotimo': 'otimo'    },

    'Div. liquida/EBITDA': {
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'pessimo',
        'descrruim': 'ruim',
        'descrmoderado': 'moderado',
        'descrbom': 'bom',
        'descotimo': 'otimo'    },

    'Div. liquida/EBIT': {
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'pessimo',
        'descrruim': 'ruim',
        'descrmoderado': 'moderado',
        'descrbom': 'bom',
        'descotimo': 'otimo'
    },

    'PL/Ativos': {
        'pessimo': {'min': -150, 'max': 0},
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'pessimo',
        'descrruim': 'ruim',
        'descrmoderado': 'moderado',
        'descrbom': 'bom',
        'descotimo': 'otimo'
    },

    'Passivos/Ativos': {
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'pessimo',
        'descrruim': 'ruim',
        'descrmoderado': 'moderado',
        'descrbom': 'bom',
        'descotimo': 'otimo'},

    'Liq. corrente': {
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'pessimo',
        'descrruim': 'ruim',
        'descrmoderado': 'moderado',
        'descrbom': 'bom',
        'descotimo': 'otimo'},
    'D.Y': {
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'pessimo',
        'descrruim': 'ruim',
        'descrmoderado': 'moderado',
        'descrbom': 'bom',
        'descotimo': 'otimo'},

    'P/L': {
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'pessimo',
        'descrruim': 'ruim',
        'descrmoderado': 'moderado',
        'descrbom': 'bom',
        'descotimo': 'otimo'},

    'PEG Ratio': {
        'pessimo': {'min': -150, 'max': -10},
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'pessimo',
        'descrruim': 'ruim',
        'descrmoderado': 'moderado',
        'descrbom': 'bom',
        'descotimo': 'otimo'
    },

    'P/VP': {
        'pessimo': {'min': -150, 'max': -10},
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'pessimo',
        'descrruim': 'ruim',
        'descrmoderado': 'moderado',
        'descrbom': 'bom',
        'descotimo': 'otimo'
    },

    'EV/EBITDA': {
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'pessimo',
        'descrruim': 'ruim',
        'descrmoderado': 'moderado',
        'descrbom': 'bom',
        'descotimo': 'otimo'},

    'EV/EBIT': {
        'pessimo': {'min': -150, 'max': -10},
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'pessimo',
        'descrruim': 'ruim',
        'descrmoderado': 'moderado',
        'descrbom': 'bom',
        'descotimo': 'otimo'
    },

    'P/EBITDA': {
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'pessimo',
        'descrruim': 'ruim',
        'descrmoderado': 'moderado',
        'descrbom': 'bom',
        'descotimo': 'otimo'},

    'P/EBIT': {
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'pessimo',
        'descrruim': 'ruim',
        'descrmoderado': 'moderado',
        'descrbom': 'bom',
        'descotimo': 'otimo'
    },

    'VPA': {
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'pessimo',
        'descrruim': 'ruim',
        'descrmoderado': 'moderado',
        'descrbom': 'bom',
        'descotimo': 'otimo'
    },

    'P/Ativo': {
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'pessimo',
        'descrruim': 'ruim',
        'descrmoderado': 'moderado',
        'descrbom': 'bom',
        'descotimo': 'otimo'
    },

    'LPA': {
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'pessimo',
        'descrruim': 'ruim',
        'descrmoderado': 'moderado',
        'descrbom': 'bom',
        'descotimo': 'otimo'
    },

    'P/SR': {
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'pessimo',
        'descrruim': 'ruim',
        'descrmoderado': 'moderado',
        'descrbom': 'bom',
        'descotimo': 'otimo'
    },

    'P/Ativo Circ. Liq.': {

        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'pessimo',
        'descrruim': 'ruim',
        'descrmoderado': 'moderado',
        'descrbom': 'bom',
        'descotimo': 'otimo'
    },
    'Valor atual': {
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'pessimo',
        'descrruim': 'ruim',
        'descrmoderado': 'moderado',
        'descrbom': 'bom',
        'descotimo': 'otimo'
    },

    'LIQUIDEZ MEDIA DIARIA': {
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'pessimo',
        'descrruim': 'ruim',
        'descrmoderado': 'moderado',
        'descrbom': 'bom',
        'descotimo': 'otimo'
    },

    'Patrimonio liquido': {
        'pessimo': {'min': -150, 'max': -10},
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'pessimo',
        'descrruim': 'ruim',
        'descrmoderado': 'moderado',
        'descrbom': 'bom',
        'descotimo': 'otimo'
    },

    'Ativos': {
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'pessimo',
        'descrruim': 'ruim',
        'descrmoderado': 'moderado',
        'descrbom': 'bom',
        'descotimo': 'otimo'
    },

    'Ativo circulante': {
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'pessimo',
        'descrruim': 'ruim',
        'descrmoderado': 'moderado',
        'descrbom': 'bom',
        'descotimo': 'otimo'
    },

    'Divida bruta': {
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'pessimo',
        'descrruim': 'ruim',
        'descrmoderado': 'moderado',
        'descrbom': 'bom',
        'descotimo': 'otimo'
    },

    'Disponibilidade': {
        'pessimo': {'min': -150, 'max': -10},
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'pessimo',
        'descrruim': 'ruim',
        'descrmoderado': 'moderado',
        'descrbom': 'bom',
        'descotimo': 'otimo'
    },

    'Divida liquida': {
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'pessimo',
        'descrruim': 'ruim',
        'descrmoderado': 'moderado',
        'descrbom': 'bom',
        'descotimo': 'otimo'
    },

    'Valor de mercado': {
        'pessimo': {'min': -150, 'max': -10},
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'pessimo',
        'descrruim': 'ruim',
        'descrmoderado': 'moderado',
        'descrbom': 'bom',
        'descotimo': 'otimo'
    },

    'Valor de firma': {
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'pessimo',
        'descrruim': 'ruim',
        'descrmoderado': 'moderado',
        'descrbom': 'bom',
        'descotimo': 'otimo'
    },

    'Free Float': {
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'pessimo',
        'descrruim': 'ruim',
        'descrmoderado': 'moderado',
        'descrbom': 'bom',
        'descotimo': 'otimo'
    },

    'ROE': {
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'pessimo',
        'descrruim': 'ruim',
        'descrmoderado': 'moderado',
        'descrbom': 'bom',
        'descotimo': 'otimo'
    },
    'ROA': {
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'pessimo',
        'descrruim': 'ruim',
        'descrmoderado': 'moderado',
        'descrbom': 'bom',
        'descotimo': 'otimo'
    },
    'ROIC': {
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'pessimo',
        'descrruim': 'ruim',
        'descrmoderado': 'moderado',
        'descrbom': 'bom',
        'descotimo': 'otimo'
    },

    'Giro ativos': {
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'moderado': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'pessimo',
        'descrruim': 'ruim',
        'descrmoderado': 'moderado',
        'descrbom': 'bom',
        'descotimo': 'otimo'
    }
}

wbsaida = openpyxl.Workbook()


# define selenium webdriver options
options = webdriver.ChromeOptions()

# create selenium webdriver instance
driver = webdriver.Chrome(options=options)

def categorizar_valor(metrica, valor):
    try:
        if metrica not in MetricasStatus:
            return 'Métrica não reconhecida'
        valor2 = float(valor)
        #print("categoriza_valor", valor2)
        for categoria, limites in MetricasStatus[metrica].items():

            if categoria in ['descricao', 'agrupador','descrpessimo','descrruim','descrmoderado','descrbom','descotimo']:
                continue
            if limites['min'] <= valor2 < limites['max']:
                return categoria
        return 'Valor fora do alcance definido'
    except Exception as e:
        print("Debug - limites:", limites)
        print(f"Erro inesperado categorizar: {e}")
        print(f"Erro metrica: {e}")
        print("categorizar_valor - Erro" , metrica)
        print("categoriza_valor", valor2)
        print('categoriza_valor' ,stock)
    finally:
       # print("categorizar_valor - OK2")
       pass

def criaPlanilhaIndRentabilidade(wbsaida):
    wbsaida.create_sheet('IndiRentabilidade')
    IndiRentabilidade = wbsaida['IndiRentabilidade']
    IndiRentabilidade.append(
        ['Agrupador', 'Fonte', 'ATIVO', 'Indicador', 'Valor', 'Referencia', 'Pessimo','Ruim', 'Moderado', 'Bom', 'Otimo', 'Descrição'])

    for cell in IndiRentabilidade[1]:  # Apenas o cabeçalho
        cell.fill = filltitulo
        cell.font = font_branca

def tratamento(indicador):
    indicador2 = indicador

    try:
        if indicador2 in ["-", "--","--%"]:
            indicador2 = 0
        elif indicador2 is None or is_null_zero_or_spaces(indicador2):
            indicador2 = 0
        else:
            if indicador2 == "":  # Verificação adicionada para string vazia
                indicador2 = 0
            else:
                indicador2 = float(indicador2.strip('%')) / 100
        return indicador2

    except Exception as e:
        print(f"Erro inesperado tratamento : {e}", "metrica  ", metricasts, " indicador  ", indicador, " stock  ",stock)
        # print(metrica)  # Certifique-se de que metrica está definida
        # print(indicadortratado)  # Certifique-se de que indicadortratado está definida
        #print('tratamneto - erro', stock, "   ", metrica)


    finally:
       # print('tratamneto OK')
       pass
def tratamento3(indicador):
    indicador2 = indicador

    try:
        if indicador2 in ["-", "--","--%"]:
            indicador2 = 0
        elif indicador2 is None or is_null_zero_or_spaces(indicador2):
            indicador2 = 0
        else:
            if indicador2 == "":  # Verificação adicionada para string vazia
                indicador2 = 0
            else:
                indicador2 = float(indicador2)
        return indicador2

    except Exception as e:
        print(f"Erro inesperado tratamento 3 : {e}", " metrica  ", metricasts, " indicador  ", indicador, " stock  ",stock)
        # print(metrica)  # Certifique-se de que metrica está definida
        # print(indicadortratado)  # Certifique-se de que indicadortratado está definida
        #print('tratamneto3 - erro', stock, "   ", metrica)


    finally:
        #print('tratamneto3 OK')
        pass
# Certifique-se de que as variáveis `metrica` e `indicadortratado` estão definidas corretamente no contexto onde a função é chamada.

def tratamento2(indicador):
    indicador2 = indicador

    try:
        if indicador2 in ["-", "--","--%"]:
            indicador2 = 0
        elif indicador2 is None or is_null_zero_or_spaces(indicador2):
            indicador2 = 0
        elif indicador2 == "":  # Verificação adicionada para string vazia
            indicador2 = 0
        else:
            indicador2 = float(indicador2)
        return indicador2


        return indicador2
    except Exception as e:
      # print(f"Erro inesperado tratamento2: {e}", " metrica  ", metrica, " indicador  ", indicador ," stock ", stock)
        # print(metrica)  # Certifique-se de que metrica está definida
        # print(indicadortratado)  # Certifique-se de que indicadortratado está definida
        print('tratamneto2 - erro', stock)


    finally:
        #print('tratamneto2 OK', indicador)
        pass
def gravaIndiEficiênciaoStaus(wsIndiRentabilidade, dict_stocks, stock):
    # fontes ['StatusInvest', 'Fundamentus']



    global linha2
    global metricasts
    #linha2 = 1
    try:
        #print(dict_stocks)
        for metrica, detalhes in MetricasStatus.items():
    #        print(f'Métrica: {metrica}')
            linha2 += 1
            metricasts = metrica
            if metrica in ['Giro ativos', 'Div. liquida/PL','Div. liquida/EBITDA','Div. liquida/EBIT','PL/Ativos',
                           'Passivos/Ativos','Liq. corrente','P/L','PEG Ratio','P/VP','EV/EBITDA','EV/EBIT',
                            'P/EBITDA','P/EBIT','VPA','P/Ativo','LPA',
                            'P/SR','P/Ativo Circ. Liq.']:
             #   print('IF tratamneto2 ',metrica )

                indicadortratado = tratamento2(dict_stocks[stock].get(metrica))
                valor_pl = indicadortratado
                categoria_pl = categorizar_valor(metrica, (valor_pl))
              #  print("categoria " + categoria_pl)
               # print("categoria tratamento2 " + categoria_pl)
            elif metrica in ['Valor atual','LIQUIDEZ MEDIA DIARIA','Patrimonio liquido',
                             'Ativos','Ativo circulante','Divida bruta','Disponibilidade',
                             'Divida liquida','Valor de mercado','Valor de firma']:
               # print('IF tratamneto3 ', metrica)
               # print("categoria " + categoria_pl)
                indicadortratado = tratamento3(dict_stocks[stock].get(metrica))
                valor_pl = indicadortratado
                categoria_pl = categorizar_valor(metrica, (valor_pl))
               # print("categoria tratamento3" + categoria_pl)
            else:
               # print('IF tratamneto ', metrica)
                indicadortratado = tratamento(dict_stocks[stock].get(metrica))
                valor_pl = indicadortratado
                categoria_pl = categorizar_valor(metrica,(valor_pl * 100))
                #print("categoria tratamento" + categoria_pl)
                # Certifique-se de que 'ROE' é o valor correto para a métrica
   #         print(f'O índice P/L {valor_pl} é categorizado como: {categoria_pl}')
   #         print(f"  Agrupador: {detalhes['agrupador']}")


            wsIndiRentabilidade.cell(row=linha2, column=1, value=detalhes['agrupador'])
            wsIndiRentabilidade.cell(row=linha2, column=2, value='StausInvest')
            wsIndiRentabilidade.cell(row=linha2, column=3, value=stock)
           # print('celula - Indicador', metrica)
            wsIndiRentabilidade.cell(row=linha2, column=4, value=metrica)
            if metrica in ['Giro ativos', 'Div. liquida/PL','Div. liquida/EBITDA','Div. liquida/EBITDA',
                           'Div. liquida/EBIT','PL/Ativos','Passivos/Ativos','Liq. corrente',
                           'P/L','PEG Ratio','P/VP','EV/EBITDA','EV/EBIT',
                            'P/EBITDA','P/EBIT','VPA','P/Ativo','LPA',
                            'P/SR','P/Ativo Circ. Liq.']:
               # print('IF da celula - Indicador', metrica )
               # print('IF da celula - valor', valor_pl)
                wsIndiRentabilidade.cell(row=linha2, column=5, value=valor_pl).number_format = numbers.FORMAT_NUMBER_00
            elif  metrica in ['Valor atual','LIQUIDEZ MEDIA DIARIA','Patrimonio liquido',
                             'Ativos','Ativo circulante','Divida bruta','Disponibilidade',
                             'Divida liquida','Valor de mercado','Valor de firma']:
                  wsIndiRentabilidade.cell(row=linha2, column=5, value=valor_pl).number_format = 'R$ #,##0.00'
            else:
                wsIndiRentabilidade.cell(row=linha2, column=5, value=valor_pl).number_format = numbers.FORMAT_PERCENTAGE_00

            wsIndiRentabilidade.cell(row=linha2, column=7,
                             value=f"Mínimo = {detalhes['pessimo']['min']}, Máximo = {detalhes['pessimo']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=8,
                                     value=f"Mínimo = {detalhes['ruim']['min']}, Máximo = {detalhes['ruim']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=9,
                                     value=f"Mínimo = {detalhes['moderado']['min']}, Máximo = {detalhes['moderado']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=10,
                                     value=f"Mínimo = {detalhes['bom']['min']}, Máximo = {detalhes['bom']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=11,
                                     value=f"Mínimo = {detalhes['otimo']['min']}, Máximo = {detalhes['otimo']['max']}")
            #print(detalhes)

            #if metrica == 'Div. liquida/EBITDA':
            if categoria_pl == 'pessimo':
                wsIndiRentabilidade.cell(row=linha2, column=6, value=categoria_pl).fill = fillvermelho
                wsIndiRentabilidade.cell(row=linha2, column=11, value=f"{detalhes['descrpessimo']}")
            if categoria_pl == 'ruim':
                wsIndiRentabilidade.cell(row=linha2, column=6, value=categoria_pl).fill = fillvermelho
                wsIndiRentabilidade.cell(row=linha2, column=11, value=f"{detalhes['descrruim']}")
            if  categoria_pl == 'moderado':
                wsIndiRentabilidade.cell(row=linha2, column=6, value=categoria_pl).fill =fillamarelo
                wsIndiRentabilidade.cell(row=linha2, column=11, value=f"{detalhes['descrmoderado']}")
            if  categoria_pl == 'bom':
                wsIndiRentabilidade.cell(row=linha2, column=6, value=categoria_pl).fill = fillverde
                wsIndiRentabilidade.cell(row=linha2, column=11, value=f"{detalhes['descrbom']}")
            if categoria_pl == 'otimo':
                wsIndiRentabilidade.cell(row=linha2, column=6, value=categoria_pl).fill =fillazul
                wsIndiRentabilidade.cell(row=linha2, column=11, value=f"{detalhes['descotimo']}")


    except Exception as e:
        print(f"Erro inesperado grava planilha2: {e}")
        print(metrica)
        print(indicadortratado)

        print('gravaIndiEficiênciaoStaus - erro' ,  stock,"    ", metrica)
    finally:
        print('gravaIndiEficiênciaoStaus  OK''', stock)

def is_null_zero_or_spaces(variable):
    # Verifica se a variável é None
    if variable is None:
        return True
    # Verifica se a variável é zero (0)
    elif variable == 0:
        return True
    # Verifica se a variável é uma string e contém apenas espaços
    elif isinstance(variable, str) and variable.strip() == '':
        return True
    elif variable == '-%':
        return True
    else:
        return False


def get_stock_soup(stock):
    ''' Get raw html from a stock '''

    # access the stock urlww
    driver.get(f'https://statusinvest.com.br/acoes/{stock}')

    # get html from stock
    html = driver.find_element(By.ID, 'main-2').get_attribute('innerHTML')

    # remove accents from html and transform html into soup
    soup = BeautifulSoup(unidecode(html), 'html.parser')

    return soup


def soup_to_dict(soup):
    '''Get all data from stock soup and return as a dictionary '''
    keys, values = [], []

    # get divs from stock
    soup1 = soup.find('div', class_='pb-3 pb-md-5')
    soup2 = soup.find('div', class_='card rounded text-main-green-dark')
    soup3 = soup.find('div', class_='indicator-today-container')
    soup4 = soup.find(
        'div', class_='top-info info-3 sm d-flex justify-between mb-3')
    soups = [soup1, soup2, soup3, soup4]

    for s in soups:
        # get only titles from a div and append to keys
        titles = s.find_all('h3', re.compile('title m-0[^"]*'))
        titles = [t.get_text() for t in titles]
        keys += titles

        # get only numbers from a div and append to values
        numbers = s.find_all('strong', re.compile('value[^"]*'))
        numbers = [n.get_text()for n in numbers]
        values += numbers

    # remove unused key and insert needed keys
    keys.remove('PART. IBOV')
    keys.insert(6, 'TAG ALONG')
    keys.insert(7, 'LIQUIDEZ MEDIA DIARIA')

    # clean keys list
    keys = [k.replace('\nhelp_outline', '').strip() for k in keys]
    keys = [k for k in keys if k != '']

    # clean values list
    values = [v.replace('\nhelp_outline', '').strip() for v in values]
    values = [v.replace('.', '').replace(',', '.') for v in values]

    # create a dictionary using keys and values from indicators
    d = {k: v for k, v in zip(keys, values)}

    return d


if __name__ == "__main__":
    dict_stocks = {}
    criaPlanilhaIndRentabilidade(wbsaida)
    wsIndiRentabilidade = wbsaida['IndiRentabilidade']

    # start timer
    start = time.time()

    # read file with stocks codes to get stock information
    with open('stocks.txt', 'r') as f:
        stocks = f.read().splitlines()

        # get stock information and create excel sheet
        for stock in stocks:
            #print("stock :"  ,stock)
            try:
                # get data and transform into dictionary
                soup = get_stock_soup(stock)
                dict_stock = soup_to_dict(soup)
                dict_stocks[stock] = dict_stock
                gravaIndiEficiênciaoStaus(wsIndiRentabilidade, dict_stocks, stock)
            except:
                # if we not get the information... just skip it
                print(f'Could not get {stock} information', "    ", metricasts)

    # create dataframe using dictionary of stocks informations
    df = pd.DataFrame(dict_stocks)

    # replace missing values with NaN to facilitate processing
    df = df.replace(['', '-', '--', '-%', '--%'], np.nan)

    # write dataframe into csv file
    df.to_excel('stocks_data.xlsx', index_label='indicadores')

    # exit the driver
    driver.quit()

    # end timer
    end = time.time()
    wbsaida.save("StatusInvest.xlsx")
    print("teste")
    print(f'Brasilian stocks information got in {int(end-start)} s')
# silvio teste