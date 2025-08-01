import openpyxl
import re
import time
import numpy as np
import pandas as pd
from unidecode import unidecode
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
import requests
# teste silvio 3e
import warnings
from openpyxl.styles import numbers

#from teste01 import metrica

fillvermelho= PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") # Vermelho

fillverde= PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid") # Verde

fillamarelo =  PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # Amarelo


fillazul = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid") # Azul


#filltitulo =   PatternFill(start_color="#002060", end_color="#002060", fill_type="solid")

filltitulo = PatternFill(start_color="002060", end_color="002060", fill_type="solid")  # Azul escuro

font_branca = Font(color="FFFFFF")  # Branco

TITLES = [
    'Identificação', 'Resumo Financeiro', 'Cotações', 'Informações Básicas',
    'Oscilações', 'Indicadores de Valuation', 'Indicadores de Rentabilidade',
    'Indicadores de Endividamento', 'Balanço Patrimonial', 'Demonstrativo de Resultados'
]

linha2 = 1
metricasts= ""
''''categorias = {
    'otimo': {'min': float('-inf'), 'max': -2},  # Valores muito baixos são ótimos
    'bom': {'min': -2, 'max': 0},               # Valores entre -2 e 0
    'moderado': {'min': 0, 'max': 1.5},         # Valores entre 0 e 1.5
    'ruim': {'min': 1.5, 'max': 3},             # Valores entre 1.5 e 3
    'pessimo': {'min': 3, 'max': 4},            # Valores entre 3 e 4
    'critico': {'min': 4, 'max': float('inf')}   # Valores acima de 4
}'''
MetricasStatus = {

    'P/L': {
        'critico': {'min': float('-inf'), 'max': 0},
        'pessimo': {'min': 20, 'max': float('inf')},
        'ruim': {'min': 15, 'max': 20},
        'moderado': {'min': 10, 'max': 15},
        'bom': {'min': 5, 'max': 10},
        'otimo': {'min': 0, 'max': 5},
        'descricao': 'Preço da ação dividido pelo lucro por ação. Mede valuation em relação ao lucro líquido.',
        'agrupador': 'Valuation',
        'descrcritico': 'Prejuízo por ação, indicando problemas operacionais ou financeiros graves. Comum em empresas em crise (ex.: OIBR3). Investidores devem evitar, salvo plano robusto de recuperação.',
        'descrpessimo': 'Empresa extremamente cara, preço é mais de 20 vezes o lucro por ação, sugerindo sobrevalorização. Comum em setores de tecnologia (ex.: NUBR33). Investidores devem comparar com peers e avaliar crescimento projetado.',
        'descrruim': 'Empresa cara, comum em setores de crescimento (ex.: saúde, RDOR3). Investidores devem analisar se o crescimento justifica o preço e comparar com média setorial.',
        'descrmoderado': 'Avaliação razoável, comum em setores maduros (ex.: varejo, LREN3). Investidores devem verificar consistência do lucro e tendências setoriais.',
        'descrbom': 'Subvalorizada, oportunidade em setores cíclicos (ex.: mineração, VALE3). Investidores devem confirmar fluxo de caixa livre e sustentabilidade do lucro.',
        'descrotimo': 'Extremamente subvalorizada, sugere oportunidade ou lucro inflado. Comum em setores em recuperação (ex.: PETR4). Investidores devem verificar fluxo de caixa livre e ROE.'
    }
}

wbsaida = openpyxl.Workbook()


# define selenium webdriver options
options = webdriver.ChromeOptions()

# create selenium webdriver instance
driver = webdriver.Chrome(options=options)

def classificar_divida_ebit(valor):
        indicadortratado = tratamento2(dict_stocks[stock].get('LPA'))
       # print("Pl" , indicadortratado)
        if valor == float('-inf') or valor < 0:
            return {'classificacao': 'pessimo', 'descricao': 'teste'}
        elif 0 <= valor < 1.5:
            return {'classificacao': 'pessimo', 'descricao': 'teste'}
        elif 1.5 <= valor < 3:
            return {'classificacao': 'pessimo', 'descricao': 'teste'}
        elif 3 <= valor < 4:
            return {'classificacao': 'pessimo', 'descricao': 'teste'}
        elif valor >= 4:
            return {'classificacao': 'pessimo', 'descricao': 'teste'}
        else:
            return {'classificacao': 'pessimo', 'descricao': 'teste'}

def evaluate_pl(PL):
    """
    Avalia o Preço/Lucro (P/L). Faixas baseadas no mercado brasileiro:
    - Lucro Líquido ≤ 0: Prejuízo (crítico)
    - 0 < P/L ≤ 8: Subvalorizado (ótimo)
    - 8 < P/L ≤ 12: Atraente (bom)
    - 12 < P/L ≤ 18: Neutro (moderado)
    - 18 < P/L ≤ 25: Caro (ruim)
    - P/L > 25: Sobrevalorizado (péssimo)
    """
    # if lucro_liquido <= 0:  # Lucro Líquido negativo indica prejuízo
    #    return 'critico'   colocar no futuro integração com o fundamentus
    try:
        if PL < 0:
            return {'classificacao': 'Critico', 'faixa': 'PL < 0',
                    'descricao': ' P/L negativo indica que a empresa não está gerando lucro, o que pode sugerir problemas operacionais, má gestão ou dificuldades no mercado'}
        elif 0 <= PL <= 10:
            return {'classificacao': 'Otimo', 'faixa': '0 <= PL <= 10',
                    'descricao': ' P/L negativo indica que a empresa não está gerando lucro, o que pode sugerir problemas operacionais, má gestão ou dificuldades no mercado'}
        elif 10 < PL <= 15:
            return {'classificacao': 'Moderado', 'faixa': '10 < PL <= 15',
                    'descricao': ' P/L negativo indica que a empresa não está gerando lucro, o que pode sugerir problemas operacionais, má gestão ou dificuldades no mercado'}
        elif 15 < PL <= 20:
            return {'classificacao': 'Ruim', 'faixa': '15 < PL <= 20',
                    'descricao': ' P/L negativo indica que a empresa não está gerando lucro, o que pode sugerir problemas operacionais, má gestão ou dificuldades no mercado'}
        elif 20 < PL <= 30:
            return {'classificacao': 'Pessimo', 'faixa': '20 < PL <= 30',
                    'descricao': ' P/L negativo indica que a empresa não está gerando lucro, o que pode sugerir problemas operacionais, má gestão ou dificuldades no mercado'}
        else:  # PL > 30
            return {'classificacao': 'Fora da faixa', 'faixa': 'PL > 30',
                    'descricao': ' P/L negativo indica que a empresa não está gerando lucro, o que pode sugerir problemas operacionais, má gestão ou dificuldades no mercado'}
    except Exception as e:
        print(f"Erro inesperado tratamento : {e}")
        # print(metrica)  # Certifique-se de que metrica está definida
        # print(indicadortratado)  # Certifique-se de que indicadortratado está definida
        #print('tratamneto - erro', stock, "   ", metrica)

def categorizar_valor(metrica, valor):
    try:
        if metrica not in MetricasStatus:
            return 'Métrica não reconhecida'
        valor2 = float(valor)
        #print("categoriza_valor", valor2)
        for categoria, limites in MetricasStatus[metrica].items():

            if categoria in ['descricao', 'agrupador','descrcritico','descrpessimo','descrruim','descrmoderado','descrbom','descrotimo']:
                continue
            if limites['min'] <= valor2 < limites['max']:
                return categoria
        return 'Valor fora do alcance definido'
    except Exception as e:
        print("Debug - limites:", limites)
        print(f"Erro inesperado categorizar: {e}")
        print(f"Erro metrica: {e}")
        print("categorizar_valor - Erro" , metrica)
        print("categoriza_valor", valor2)
        print('categoriza_valor' ,stock)
    finally:
       # print("categorizar_valor - OK2")
       pass

def criaPlanilhaIndRentabilidade(wbsaida):
    wbsaida.create_sheet('IndiRentabilidade')
    IndiRentabilidade = wbsaida['IndiRentabilidade']
    IndiRentabilidade.append(
        ['Agrupador', 'Fonte', 'ATIVO', 'Indicador', 'Valor', 'Referencia','Critico', 'Pessimo','Ruim', 'Moderado', 'Bom', 'Otimo', 'Descrição'])

    for cell in IndiRentabilidade[1]:  # Apenas o cabeçalho
        cell.fill = filltitulo
        cell.font = font_branca

def tratamento(indicador):
    indicador2 = indicador

    try:
        if indicador2 in ["-", "--","--%"]:
            indicador2 = 0
        elif indicador2 is None or is_null_zero_or_spaces(indicador2):
            indicador2 = 0
        else:
            if indicador2 == "":  # Verificação adicionada para string vazia
                indicador2 = 0
            else:
                indicador2 = float(indicador2.strip('%')) / 100
        return indicador2

    except Exception as e:
        print(f"Erro inesperado tratamento : {e}", "metrica  ", metricasts, " indicador  ", indicador, " stock  ",stock)
        # print(metrica)  # Certifique-se de que metrica está definida
        # print(indicadortratado)  # Certifique-se de que indicadortratado está definida
        #print('tratamneto - erro', stock, "   ", metrica)


    finally:
       # print('tratamneto OK')
       pass
def tratamento3(indicador):
    indicador2 = indicador

    try:
        if indicador2 in ["-", "--","--%"]:
            indicador2 = 0
        elif indicador2 is None or is_null_zero_or_spaces(indicador2):
            indicador2 = 0
        else:
            if indicador2 == "":  # Verificação adicionada para string vazia
                indicador2 = 0
            else:
                indicador2 = float(indicador2)
        return indicador2

    except Exception as e:
        print(f"Erro inesperado tratamento 3 : {e}", " metrica  ", metricasts, " indicador  ", indicador, " stock  ",stock)
        # print(metrica)  # Certifique-se de que metrica está definida
        # print(indicadortratado)  # Certifique-se de que indicadortratado está definida
        #print('tratamneto3 - erro', stock, "   ", metrica)


    finally:
        #print('tratamneto3 OK')
        pass
# Certifique-se de que as variáveis `metrica` e `indicadortratado` estão definidas corretamente no contexto onde a função é chamada.

def tratamento2(indicador):
    indicador2 = indicador

    try:
        if indicador2 in ["-", "--","--%"]:
            indicador2 = 0
        elif indicador2 is None or is_null_zero_or_spaces(indicador2):
            indicador2 = 0
        elif indicador2 == "":  # Verificação adicionada para string vazia
            indicador2 = 0
        else:
            indicador2 = float(indicador2)
        return indicador2


        return indicador2
    except Exception as e:
      # print(f"Erro inesperado tratamento2: {e}", " metrica  ", metrica, " indicador  ", indicador ," stock ", stock)
        # print(metrica) g # Certifique-se de que metrica está definida
        # print(indicadortratado)  # Certifique-se de que indicadortratado está definida
        print('tratamneto2 - erro', stock)


    finally:
        #print('tratamneto2 OK', indicador)
        pass
def gravaIndiEficiênciaoStaus(wsIndiRentabilidade, dict_stocks, stock):
    # fontes ['StatusInvest', 'Fundamentus']



    global linha2
    global metricasts
    #linha2 = 1
    try:
        #print(dict_stocks)
        for metrica, detalhes in MetricasStatus.items():
    #        print(f'Métrica: {metrica}')
            linha2 += 1
            metricasts = metrica
            if metrica in ['Giro ativos', 'Div. liquida/PL','Div. liquida/EBITDA','Div. liquida/EBIT','PL/Ativos',
                           'Passivos/Ativos','Liq. corrente','P/L','PEG Ratio','P/VP','EV/EBITDA','EV/EBIT',
                            'P/EBITDA','P/EBIT','VPA','P/Ativo','LPA',
                            'P/SR','P/Ativo Circ. Liq.']:
             #   print('IF tratamneto2 ',metrica )

                indicadortratado = tratamento2(dict_stocks[stock].get(metrica))
                valor_pl = indicadortratado
               # categoria_pl = categorizar_valor(metrica, (valor_pl))

                categoria_pl = evaluate_pl(valor_pl)

                classificacao = categoria_pl['classificacao']
                faixa = categoria_pl['faixa']
                descricao = categoria_pl['descricao']

                print(f"Classificação:" + classificacao)
                print(f"Faixa:" +  faixa)
                #print(retorno)
                #teste = classificar_divida_ebit(valor_pl)
                 #print("descricao:", teste['descricao'])
                 #print("Classificação:", teste['classificacao'])
                #print(teste)
                #print(teste['classificacao'])  # saída: 'pessimo'
                #print(teste['descricao'])  # saída: 'teste'
              #  print("categoria " + categoria_pl)
               # print("categoria tratamento2 " + categoria_pl)
            elif metrica in ['Valor atual','LIQUIDEZ MEDIA DIARIA','Patrimonio liquido',
                             'Ativos','Ativo circulante','Divida bruta','Disponibilidade',
                             'Divida liquida','Valor de mercado','Valor de firma']:
               # print('IF tratamneto3 ', metrica)
               # print("categoria " + categoria_pl)
               # print('Patrimonio liquido ' + dict_stocks[stock].get('Patrimonio liquido'))
                indicadortratado = tratamento3(dict_stocks[stock].get(metrica))
                valor_pl = indicadortratado
                categoria_pl = categorizar_valor(metrica, (valor_pl))
               # print("categoria tratamento3" + categoria_pl)
            else:
               # print('IF tratamneto ', metrica)
                indicadortratado = tratamento(dict_stocks[stock].get(metrica))
                valor_pl = indicadortratado
                categoria_pl = categorizar_valor(metrica,(valor_pl * 100))
                #print("categoria tratamento" + categoria_pl)
                # Certifique-se de que 'ROE' é o valor correto para a métrica
   #         print(f'O índice P/L {valor_pl} é categorizado como: {categoria_pl}')
   #         print(f"  Agrupador: {detalhes['agrupador']}")


            wsIndiRentabilidade.cell(row=linha2, column=1, value=detalhes['agrupador'])
            wsIndiRentabilidade.cell(row=linha2, column=2, value='StausInvest')
            wsIndiRentabilidade.cell(row=linha2, column=3, value=stock)
          #  print('celula - Indicador', metrica)
            wsIndiRentabilidade.cell(row=linha2, column=4, value=metrica)
            if metrica in ['Giro ativos', 'Div. liquida/PL','Div. liquida/EBITDA','Div. liquida/EBITDA',
                           'Div. liquida/EBIT','PL/Ativos','Passivos/Ativos','Liq. corrente',
                           'P/L','PEG Ratio','P/VP','EV/EBITDA','EV/EBIT',
                            'P/EBITDA','P/EBIT','VPA','P/Ativo','LPA',
                            'P/SR','P/Ativo Circ. Liq.']:
               # print('IF da celula - Indicador', metrica )
               # print('IF da celula - valor', valor_pl)
                wsIndiRentabilidade.cell(row=linha2, column=5, value=valor_pl).number_format = numbers.FORMAT_NUMBER_00
            elif  metrica in ['Valor atual','LIQUIDEZ MEDIA DIARIA','Patrimonio liquido',
                             'Ativos','Ativo circulante','Divida bruta','Disponibilidade',
                             'Divida liquida','Valor de mercado','Valor de firma']:
                  #print('IF da celula - valor', valor_pl)
                  #print('IF da celula - Indicador', metrica)
                  wsIndiRentabilidade.cell(row=linha2, column=5, value=valor_pl).number_format = 'R$ #,##0.00'
            else:
                wsIndiRentabilidade.cell(row=linha2, column=5, value=valor_pl).number_format = numbers.FORMAT_PERCENTAGE_00

            wsIndiRentabilidade.cell(row=linha2, column=7,
                             value=f"Mínimo = {detalhes['critico']['min']}, Máximo = {detalhes['critico']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=8,
                             value=f"Mínimo = {detalhes['pessimo']['min']}, Máximo = {detalhes['pessimo']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=9,
                                     value=f"Mínimo = {detalhes['ruim']['min']}, Máximo = {detalhes['ruim']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=10,
                                     value=f"Mínimo = {detalhes['moderado']['min']}, Máximo = {detalhes['moderado']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=11,
                                     value=f"Mínimo = {detalhes['bom']['min']}, Máximo = {detalhes['bom']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=12,
                                     value=f"Mínimo = {detalhes['otimo']['min']}, Máximo = {detalhes['otimo']['max']}")
            #print(detalhes)'''
            desc1 =  ("Lucro Líquido ≤ 0: Prejuízo (crítico)- 0 < P/L ≤ 8: Subvalorizado (ótimo) - 8 < P/L ≤ 12:"
                       " Atraente (bom) - 12 < P/L ≤ 18: Neutro (moderado)    - 18 < P/L ≤ 25: Caro (ruim)  - "
                       "P/L > 25: Sobrevalorizado (péssimo) ")




            #if metrica == 'Div. liquida/EBITDA':
            if classificacao == 'Critico':
                wsIndiRentabilidade.cell(row=linha2, column=6, value=classificacao).fill = fillvermelho
                wsIndiRentabilidade.cell(row=linha2, column=7, value=faixa)
                wsIndiRentabilidade.cell(row=linha2, column=13, value=descricao)
            if classificacao == 'Pessimo':
                wsIndiRentabilidade.cell(row=linha2, column=6, value=classificacao).fill = fillvermelho
                wsIndiRentabilidade.cell(row=linha2, column=8, value=faixa)
                wsIndiRentabilidade.cell(row=linha2, column=13, value=descricao)
            if classificacao == 'Ruim':
                wsIndiRentabilidade.cell(row=linha2, column=6, value=classificacao).fill = fillvermelho
                wsIndiRentabilidade.cell(row=linha2, column=9, value=faixa)
                wsIndiRentabilidade.cell(row=linha2, column=13, value=descricao)
            if  classificacao == 'Moderado':
                wsIndiRentabilidade.cell(row=linha2, column=6, value=classificacao).fill =fillamarelo
                wsIndiRentabilidade.cell(row=linha2, column=10, value=faixa)
                wsIndiRentabilidade.cell(row=linha2, column=13, value=descricao)
            if  classificacao == 'Bom':
                wsIndiRentabilidade.cell(row=linha2, column=6, value=classificacao).fill = fillverde
                wsIndiRentabilidade.cell(row=linha2, column=11, value=faixa)
                wsIndiRentabilidade.cell(row=linha2, column=13, value=descricao)
            if classificacao == 'Otimo':
                wsIndiRentabilidade.cell(row=linha2, column=6, value=classificacao).fill =fillazul
                wsIndiRentabilidade.cell(row=linha2, column=12, value=faixa)
                wsIndiRentabilidade.cell(row=linha2, column=13, value=descricao)
            if classificacao == 'Fora da faixa':
                wsIndiRentabilidade.cell(row=linha2, column=6, value=classificacao).fill = fillazul
                wsIndiRentabilidade.cell(row=linha2, column=12, value=faixa)
                wsIndiRentabilidade.cell(row=linha2, column=13, value=descricao)


    except Exception as e:
        print(f"Erro inesperado grava planilha2: {e}")
        print(metrica)
        print(indicadortratado)

        print('gravaIndiEficiênciaoStaus - erro' ,  stock,"    ", metrica)
    finally:
        print('gravaIndiEficiênciaoStaus  OK''', stock)

def is_null_zero_or_spaces(variable):
    # Verifica se a variável é None
    if variable is None:
        return True
    # Verifica se a variável é zero (0)
    elif variable == 0:
        return True
    # Verifica se a variável é uma string e contém apenas espaços
    elif isinstance(variable, str) and variable.strip() == '':
        return True
    elif variable == '-%':
        return True
    else:
        return False


def get_stock_soup(stock):
    ''' Get raw html from a stock '''

    # access the stock urlww
    driver.get(f'https://statusinvest.com.br/acoes/{stock}')

    # get html from stock
    html = driver.find_element(By.ID, 'main-2').get_attribute('innerHTML')

    # remove accents from html and transform html into soup
    soup = BeautifulSoup(unidecode(html), 'html.parser')

    return soup


def soup_to_dict(soup):
    '''Get all data from stock soup and return as a dictionary '''
    keys, values = [], []

    # get divs from stock
    soup1 = soup.find('div', class_='pb-3 pb-md-5')
    soup2 = soup.find('div', class_='card rounded text-main-green-dark')
    soup3 = soup.find('div', class_='indicator-today-container')
    soup4 = soup.find(
        'div', class_='top-info info-3 sm d-flex justify-between mb-3')
    soups = [soup1, soup2, soup3, soup4]

    for s in soups:
        # get only titles from a div and append to keys
        titles = s.find_all('h3', re.compile('title m-0[^"]*'))
        titles = [t.get_text() for t in titles]
        keys += titles

        # get only numbers from a div and append to values
        numbers = s.find_all('strong', re.compile('value[^"]*'))
        numbers = [n.get_text()for n in numbers]
        values += numbers

    # remove unused key and insert needed keys
    keys.remove('PART. IBOV')
    keys.insert(6, 'TAG ALONG')
    keys.insert(7, 'LIQUIDEZ MEDIA DIARIA')

    # clean keys list
    keys = [k.replace('\nhelp_outline', '').strip() for k in keys]
    keys = [k for k in keys if k != '']

    # clean values list
    values = [v.replace('\nhelp_outline', '').strip() for v in values]
    values = [v.replace('.', '').replace(',', '.') for v in values]

    # create a dictionary using keys and values from indicators
    d = {k: v for k, v in zip(keys, values)}

    return d


if __name__ == "__main__":
    dict_stocks = {}
    criaPlanilhaIndRentabilidade(wbsaida)
    wsIndiRentabilidade = wbsaida['IndiRentabilidade']

    # start timer
    start = time.time()

    # read file with stocks codes to get stock information
    with open('stocks.txt', 'r') as f:
        stocks = f.read().splitlines()

        # get stock information and create excel sheet
        for stock in stocks:
            #print("stock :"  ,stock)
            try:
                # get data and transform into dictionary
                soup = get_stock_soup(stock)
                dict_stock = soup_to_dict(soup)
                dict_stocks[stock] = dict_stock
                gravaIndiEficiênciaoStaus(wsIndiRentabilidade, dict_stocks, stock)
            except:
                # if we not get the information... just skip it
                print(f'Could not get {stock} information', "    ", metricasts)

    # create dataframe using dictionary of stocks informations
    df = pd.DataFrame(dict_stocks)

    # replace missing values with NaN to facilitate processing
    df = df.replace(['', '-', '--', '-%', '--%'], np.nan)

    # write dataframe into csv file
    df.to_excel('stocks_data.xlsx', index_label='indicadores')

    # exit the driver
    driver.quit()

    # end timer
    end = time.time()
    wbsaida.save("StatusInvest.xlsx")
    #print("teste")
    print(f'Brasilian stocks information got in {int(end-start)} s')
# silvio teste