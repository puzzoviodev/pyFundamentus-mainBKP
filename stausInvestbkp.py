import openpyxl
import re
import time
import numpy as np
import pandas as pd
from unidecode import unidecode
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
import requests

import warnings
from openpyxl.styles import numbers

from teste01 import metrica

fillvermelho= PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") # Vermelho

fillverde= PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid") # Verde

fillamarelo =  PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # Amarelo

#filltitulo =   PatternFill(start_color="#002060", end_color="#002060", fill_type="solid")

filltitulo = PatternFill(start_color="002060", end_color="002060", fill_type="solid")  # Azul escuro

font_branca = Font(color="FFFFFF")  # Branco

TITLES = [
    'Identificação', 'Resumo Financeiro', 'Cotações', 'Informações Básicas',
    'Oscilações', 'Indicadores de Valuation', 'Indicadores de Rentabilidade',
    'Indicadores de Endividamento', 'Balanço Patrimonial', 'Demonstrativo de Resultados'
]

linha2 = 1
MetricasStatus = {
    'M. Liquida': {
        'baixo': {'min': -50, 'max': 5},
        'regular': {'min': 5, 'max':10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrbaixo': 'Uma margem líquida abaixo de 5% sugere que a empresa tem uma margem de lucro relativamente baixa. Isso pode ser típico em setores com alta concorrência e margens reduzidas, como o varejo ou o setor de alimentos e bebidas. Pode indicar baixa eficiência na gestão de custos ou baixa rentabilidade',
        'descregular': 'Uma margem líquida nessa faixa é considerada moderada e pode ser típica de empresas que operam em setores com margens mais equilibradas, como algumas indústrias ou serviços. Reflete uma eficiência razoável em converter receita em lucro líquido.',
        'descrbom': 'Uma margem líquida entre 10% e 20% é geralmente considerada boa e indica uma empresa que é eficiente na gestão de suas despesas e custos, resultando em uma rentabilidade sólida. É comum em setores com menos concorrência e mais poder de precificação',
        'descotimo': 'Uma margem líquida acima de 20% sugere alta eficiência e rentabilidade. Empresas com margens líquidas altas geralmente operam em setores com altas barreiras de entrada, produtos ou serviços de alto valor agregado, ou que têm uma forte vantagem competitiva. Isso pode ser visto em setores como tecnologia, software ou bens de luxo'
    },
    'M. EBIT': {
        'baixo': {'min': -50, 'max': 5},
        'regular': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Capacidade de pagar dívidas. Menor que 2.5 é considerado saudável.',
        'agrupador': 'Eficiência',
        'descrbaixo': 'Uma margem EBIT abaixo de 5% sugere que a empresa tem uma baixa eficiência operacional. Isso pode ser comum em setores com alta concorrência e baixos níveis de diferenciação, como o varejo e alguns segmentos industriais. Indica que a empresa pode estar enfrentando desafios significativos em controlar seus custos operacionais',
        'descregular': 'Uma margem EBIT nessa faixa é considerada moderada e pode ser típica de empresas em setores com margens mais equilibradas. Reflete uma eficiência operacional razoável, mas pode haver espaço para melhorar a rentabilidade operacional',
        'descrbom': 'Uma margem EBIT entre 10% e 20% indica uma boa eficiência operacional. A empresa consegue gerar uma quantidade significativa de lucro operacional em relação à receita líquida. Esse nível é comum em setores menos intensivos em capital e com uma boa capacidade de controle de custos',
        'descotimo': 'Uma margem EBIT acima de 20% sugere alta eficiência operacional e uma forte capacidade de gerar lucro operacional a partir das vendas. Isso pode ser típico de setores com altos margens brutas e controle eficaz dos custos operacionais, como tecnologia, software e alguns serviços especializados'
    },
    'M. EBITDA': {
        'baixo': {'min': -50, 'max': 10},
        'regular': {'min':10, 'max': 20},
        'bom': {'min': 20, 'max': 30},
        'otimo': {'min': 30, 'max': float('inf')},
        'descricao': 'Rendimento do fluxo de caixa livre. Acima de 10% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrbaixo': 'Uma margem EBITDA abaixo de 10% pode indicar que a empresa tem uma baixa eficiência operacional ou enfrenta altos custos operacionais. Isso pode ser comum em setores com alta concorrência e margens reduzidas, como o varejo ou indústrias com alta estrutura de custos',
        'descregular': 'Uma margem EBITDA nessa faixa é geralmente considerada moderada e pode ser típica de empresas que operam em setores com uma estrutura de custo mais equilibrada. Reflete uma eficiência operacional razoável e uma capacidade de gerar EBITDA de forma eficaz',
        'descrbom': ' Uma margem EBITDA entre 20% e 30% indica uma boa eficiência operacional, com a empresa gerando uma quantidade significativa de EBITDA em relação à receita líquida. Isso é comum em setores com menos pressão sobre margens e onde as empresas conseguem manter um controle eficiente dos custos.',
        'descotimo': 'OTUma margem EBITDA acima de 30% sugere uma alta eficiência operacional e uma forte capacidade de gerar EBITDA a partir das vendas. Isso pode ser típico de setores com alta margem bruta e baixo custo de produção, como tecnologia, software e serviços financeiros.IMO'
    },
    'M. Bruta': {
        'baixo': {'min': -50, 'max': 20},
        'regular': {'min': 20, 'max': 40},
        'bom': {'min': 40, 'max': 60},
        'otimo': {'min': 60, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Eficiência',
        'descrbaixo': 'Uma margem bruta abaixo de 20% pode indicar que a empresa tem um custo de produção relativamente alto em relação à sua receita. Isso é comum em setores com altos custos de matéria-prima ou produção, como o varejo de alimentos e algumas indústrias manufatureiras.',
        'descregular': 'Uma margem bruta nessa faixa é geralmente considerada moderada e pode ser típica de empresas com uma estrutura de custos mais equilibrada. Indica que a empresa está conseguindo controlar seus custos de produção de maneira razoável.',
        'descrbom': 'Uma margem bruta entre 40% e 60% indica uma boa eficiência na produção e venda de produtos ou serviços. Isso é comum em setores que têm menos custos variáveis e uma estrutura de custo mais favorável, como serviços e algumas indústrias com alta diferenciação de produtos.',
        'descotimo': 'Uma margem bruta acima de 60% sugere uma alta eficiência na produção e venda, com um custo relativamente baixo em relação à receita. Isso é típico em setores com altas barreiras de entrada, baixo custo de produção, ou produtos/serviços de alto valor agregado, como tecnologia e software'

    },
    'Div. liquida/PL': {
        'baixo': {'min': -50, 'max': 2},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 60, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Endividamento',
        'descrbaixo': 'BAIXO',
        'descregular': 'REGULAR',
        'descrbom': 'BOM',
        'descotimo': 'OTIMO'
    },

    'Div. liquida/EBITDA': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Endividamento',
        'descrbaixo': 'BAIXO',
        'descregular': 'REGULAR',
        'descrbom': 'BOM',
        'descotimo': 'OTIMO'

    },

    'Div. liquida/EBIT': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Endividamento',
        'descrbaixo': 'BAIXO',
        'descregular': 'REGULAR',
        'descrbom': 'BOM',
        'descotimo': 'OTIMO'
    },

    'PL/Ativos': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Endividamento',
        'descrbaixo': 'BAIXO',
        'descregular': 'REGULAR',
        'descrbom': 'BOM',
        'descotimo': 'OTIMO'
    },

    'Passivos/Ativos': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Endividamento',
        'descrbaixo': 'BAIXO',
        'descregular': 'REGULAR',
        'descrbom': 'BOM',
        'descotimo': 'OTIMO'
    },
    'Liq. corrente': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Endividamento',
        'descrbaixo': 'BAIXO',
        'descregular': 'REGULAR',
        'descrbom': 'BOM',
        'descotimo': 'OTIMO'
    },
    'D.Y': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation',
        'descrbaixo': 'BAIXO',
        'descregular': 'REGULAR',
        'descrbom': 'BOM',
        'descotimo': 'OTIMO'
    },

    'P/L': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation',
        'descrbaixo': 'BAIXO',
        'descregular': 'REGULAR',
        'descrbom': 'BOM',
        'descotimo': 'OTIMO'
    },

    'PEG Ratio': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation',
        'descrbaixo': 'BAIXO',
        'descregular': 'REGULAR',
        'descrbom': 'BOM',
        'descotimo': 'OTIMO'
    },

    'P/VP': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation',
        'descrbaixo': 'BAIXO',
        'descregular': 'REGULAR',
        'descrbom': 'BOM',
        'descotimo': 'OTIMO'
    },

    'EV/EBITDA': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation',
        'descrbaixo': 'BAIXO',
        'descregular': 'REGULAR',
        'descrbom': 'BOM',
        'descotimo': 'OTIMO'
    },

    'EV/EBIT': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation',
        'descrbaixo': 'BAIXO',
        'descregular': 'REGULAR',
        'descrbom': 'BOM',
        'descotimo': 'OTIMO'
    },

    'P/EBITDA': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation',
        'descrbaixo': 'BAIXO',
        'descregular': 'REGULAR',
        'descrbom': 'BOM',
        'descotimo': 'OTIMO'
    },

    'P/EBIT': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation',
        'descrbaixo': 'BAIXO',
        'descregular': 'REGULAR',
        'descrbom': 'BOM',
        'descotimo': 'OTIMO'
    },

    'VPA': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation',
        'descrbaixo': 'BAIXO',
        'descregular': 'REGULAR',
        'descrbom': 'BOM',
        'descotimo': 'OTIMO'
    },

    'P/Ativo': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation',
        'descrbaixo': 'BAIXO',
        'descregular': 'REGULAR',
        'descrbom': 'BOM',
        'descotimo': 'OTIMO'
    },

    'LPA': {
        'baixo': {'min': -10, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation',
        'descrbaixo': 'BAIXO',
        'descregular': 'REGULAR',
        'descrbom': 'BOM',
        'descotimo': 'OTIMO'
    },

    'P/SR': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation',
        'descrbaixo': 'BAIXO',
        'descregular': 'REGULAR',
        'descrbom': 'BOM',
        'descotimo': 'OTIMO'
    },

    'P/Ativo Circ. Liq.': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation',
        'descrbaixo': 'BAIXO',
        'descregular': 'REGULAR',
        'descrbom': 'BOM',
        'descotimo': 'OTIMO'
    },
    'Valor atual': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa',
        'descrbaixo': 'BAIXO',
        'descregular': 'REGULAR',
        'descrbom': 'BOM',
        'descotimo': 'OTIMO'
    },
'''
    'TAG ALONG': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa',
        'descrbaixo': 'BAIXO',
        'descregular': 'REGULAR',
        'descrbom': 'BOM',
        'descotimo': 'OTIMO'
    },
'''
    'LIQUIDEZ MEDIA DIARIA': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa',
        'descrbaixo': 'BAIXO',
        'descregular': 'REGULAR',
        'descrbom': 'BOM',
        'descotimo': 'OTIMO'
    },

    'Patrimonio liquido': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa',
        'descrbaixo': 'BAIXO',
        'descregular': 'REGULAR',
        'descrbom': 'BOM',
        'descotimo': 'OTIMO'
    },

    'Ativos': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa',
        'descrbaixo': 'BAIXO',
        'descregular': 'REGULAR',
        'descrbom': 'BOM',
        'descotimo': 'OTIMO'
    },

    'Ativo circulante': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa',
        'descrbaixo': 'BAIXO',
        'descregular': 'REGULAR',
        'descrbom': 'BOM',
        'descotimo': 'OTIMO'
    },

    'Divida bruta': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa',
        'descrbaixo': 'BAIXO',
        'descregular': 'REGULAR',
        'descrbom': 'BOM',
        'descotimo': 'OTIMO'
    },

    'Disponibilidade': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa',
        'descrbaixo': 'BAIXO',
        'descregular': 'REGULAR',
        'descrbom': 'BOM',
        'descotimo': 'OTIMO'
    },

    'Divida liquida': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa',
        'descrbaixo': 'BAIXO',
        'descregular': 'REGULAR',
        'descrbom': 'BOM',
        'descotimo': 'OTIMO'
    },

    'Valor de mercado': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa',
        'descrbaixo': 'BAIXO',
        'descregular': 'REGULAR',
        'descrbom': 'BOM',
        'descotimo': 'OTIMO'
    },

    'Valor de firma': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa',
        'descrbaixo': 'BAIXO',
        'descregular': 'REGULAR',
        'descrbom': 'BOM',
        'descotimo': 'OTIMO'
    },

    'Free Float': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa',
        'descrbaixo': 'BAIXO',
        'descregular': 'REGULAR',
        'descrbom': 'BOM',
        'descotimo': 'OTIMO'
    },

    'ROE': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Rentabilidade',
        'descrbaixo': 'BAIXO',
        'descregular': 'REGULAR',
        'descrbom': 'BOM',
        'descotimo': 'OTIMO'
    },
    'ROA': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Rentabilidade',
        'descrbaixo': 'BAIXO',
        'descregular': 'REGULAR',
        'descrbom': 'BOM',
        'descotimo': 'OTIMO'
    },
    'ROIC': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Rentabilidade',
        'descrbaixo': 'BAIXO',
        'descregular': 'REGULAR',
        'descrbom': 'BOM',
        'descotimo': 'OTIMO'
    },

    'Giro ativos': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Rentabilidade',
        'descrbaixo': 'BAIXO',
        'descregular': 'REGULAR',
        'descrbom': 'BOM',
        'descotimo': 'OTIMO'
    }
}

wbsaida = openpyxl.Workbook()


# define selenium webdriver options
options = webdriver.ChromeOptions()

# create selenium webdriver instance
driver = webdriver.Chrome(options=options)

def categorizar_valor(metrica, valor):
    try:
        if metrica not in MetricasStatus:
            return 'Métrica não reconhecida'
        valor2 = float(valor)
        #print("categoriza_valor", valor2)
        for categoria, limites in MetricasStatus[metrica].items():

            if categoria in ['descricao', 'agrupador','descrbaixo','descregular','descrbom','descotimo']:
                continue
            if limites['min'] <= valor2 < limites['max']:
                return categoria
        return 'Valor fora do alcance definido'
    except Exception as e:
        print(f"Erro inesperado categorizar: {e}")
        print(f"Erro metrica: {e}")
        print("categorizar_valor - Erro" , metrica)
        print("categoriza_valor", valor2)
        print('categoriza_valor' ,stock)
    finally:
       # print("categorizar_valor - OK")
       pass

def criaPlanilhaIndRentabilidade(wbsaida):
    wbsaida.create_sheet('IndiRentabilidade')
    IndiRentabilidade = wbsaida['IndiRentabilidade']
    IndiRentabilidade.append(
        ['Agrupador', 'Fonte', 'ATIVO', 'Indicador', 'Valor', 'Referencia', 'Baixo', 'Regular', 'Bom', 'Otimo', 'Descrição'])

    for cell in IndiRentabilidade[1]:  # Apenas o cabeçalho
        cell.fill = filltitulo
        cell.font = font_branca

def tratamento(indicador):
    indicador2 = indicador

    try:
        if indicador2 in ["-", "--"]:
            indicador2 = 0
        elif indicador2 is None or is_null_zero_or_spaces(indicador2):
            indicador2 = 0
        else:
            if indicador2 == "":  # Verificação adicionada para string vazia
                indicador2 = 0
            else:
                indicador2 = float(indicador2.strip('%')) / 100
        return indicador2

    except Exception as e:
        print(f"Erro inesperado tratamento : {e}")
        # print(metrica)  # Certifique-se de que metrica está definida
        # print(indicadortratado)  # Certifique-se de que indicadortratado está definida
        print('tratamneto - erro', stock, "    ", metrica)


    finally:
       # print('tratamneto OK')
       pass
def tratamento3(indicador):
    indicador2 = indicador

    try:
        if indicador2 in ["-", "--"]:
            indicador2 = 0
        elif indicador2 is None or is_null_zero_or_spaces(indicador2):
            indicador2 = 0
        else:
            if indicador2 == "":  # Verificação adicionada para string vazia
                indicador2 = 0
            else:
                indicador2 = float(indicador2)
        return indicador2

    except Exception as e:
        print(f"Erro inesperado tratamento 3 : {e}")
        # print(metrica)  # Certifique-se de que metrica está definida
        # print(indicadortratado)  # Certifique-se de que indicadortratado está definida
        print('tratamneto3 - erro', stock, "   ", metrica)



    finally:
        #print('tratamneto3 OK')
        pass
# Certifique-se de que as variáveis `metrica` e `indicadortratado` estão definidas corretamente no contexto onde a função é chamada.

def tratamento2(indicador):
    indicador2 = indicador

    try:
        if indicador2 in ["-", "--"]:
            indicador2 = 0
        elif indicador2 is None or is_null_zero_or_spaces(indicador2):
            indicador2 = 0
        elif indicador2 == "":  # Verificação adicionada para string vazia
            indicador2 = 0
        else:
            indicador2 = float(indicador2)
        return indicador2


        return indicador2
    except Exception as e:
        print(f"Erro inesperado tratamento2: {e}")
        # print(metrica)  # Certifique-se de que metrica está definida
        # print(indicadortratado)  # Certifique-se de que indicadortratado está definida
        print('tratamneto2 - erro', stock," Metrica  ", metrica)


    finally:
        #print('tratamneto2 OK', indicador)
        pass
def gravaIndiEficiênciaoStaus(wsIndiRentabilidade, dict_stocks, stock):
    # fontes ['StatusInvest', 'Fundamentus']



    global linha2
    #linha2 = 1
    try:
        #print(dict_stocks)
        for metrica, detalhes in MetricasStatus.items():
    #        print(f'Métrica: {metrica}')
            linha2 += 1
            if metrica in ['Giro ativos', 'Div. liquida/PL','Div. liquida/EBITDA','Div. liquida/EBIT','PL/Ativos',
                           'Passivos/Ativos','Liq. corrente','P/L','PEG Ratio','P/VP','EV/EBITDA','EV/EBIT',
                            'P/EBITDA','P/EBIT','VPA','P/Ativo','LPA',
                            'P/SR','P/Ativo Circ. Liq.']:
             #   print('IF tratamneto2 ',metrica )

                indicadortratado = tratamento2(dict_stocks[stock].get(metrica))
                valor_pl = indicadortratado
                categoria_pl = categorizar_valor(metrica, (valor_pl))
              #  print("categoria " + categoria_pl)
               # print("categoria tratamento2 " + categoria_pl)
            elif metrica in ['Valor atual','LIQUIDEZ MEDIA DIARIA','Patrimonio liquido',
                             'Ativos','Ativo circulante','Divida bruta','Disponibilidade',
                             'Divida liquida','Valor de mercado','Valor de firma']:
               # print('IF tratamneto3 ', metrica)
               # print("categoria " + categoria_pl)
                indicadortratado = tratamento3(dict_stocks[stock].get(metrica))
                valor_pl = indicadortratado
                categoria_pl = categorizar_valor(metrica, (valor_pl))
               # print("categoria tratamento3" + categoria_pl)
            else:
               # print('IF tratamneto ', metrica)
                indicadortratado = tratamento(dict_stocks[stock].get(metrica))
                valor_pl = indicadortratado
                categoria_pl = categorizar_valor(metrica,(valor_pl * 100))
                #print("categoria tratamento" + categoria_pl)
                # Certifique-se de que 'ROE' é o valor correto para a métrica
   #         print(f'O índice P/L {valor_pl} é categorizado como: {categoria_pl}')
   #         print(f"  Agrupador: {detalhes['agrupador']}")


            wsIndiRentabilidade.cell(row=linha2, column=1, value=detalhes['agrupador'])
            wsIndiRentabilidade.cell(row=linha2, column=2, value='StausInvest')
            wsIndiRentabilidade.cell(row=linha2, column=3, value=stock)
            wsIndiRentabilidade.cell(row=linha2, column=4, value=metrica)
            if metrica in ['Giro ativos', 'Div. liquida/PL','Div. liquida/EBITDA','Div. liquida/EBITDA',
                           'Div. liquida/EBIT','PL/Ativos','Passivos/Ativos','Liq. corrente',
                           'P/L','PEG Ratio','P/VP','EV/EBITDA','EV/EBIT',
                            'P/EBITDA','P/EBIT','VPA','P/Ativo','LPA',
                            'P/SR','P/Ativo Circ. Liq.']:
               # print('IF da celula - Indicador', metrica )
               # print('IF da celula - valor', valor_pl)
                wsIndiRentabilidade.cell(row=linha2, column=5, value=valor_pl).number_format = numbers.FORMAT_NUMBER_00
            elif  metrica in ['Valor atual','LIQUIDEZ MEDIA DIARIA','Patrimonio liquido',
                             'Ativos','Ativo circulante','Divida bruta','Disponibilidade',
                             'Divida liquida','Valor de mercado','Valor de firma']:
                  wsIndiRentabilidade.cell(row=linha2, column=5, value=valor_pl).number_format = 'R$ #,##0.00'
            else:
                wsIndiRentabilidade.cell(row=linha2, column=5, value=valor_pl).number_format = numbers.FORMAT_PERCENTAGE_00


            wsIndiRentabilidade.cell(row=linha2, column=7,
                                     value=f"Mínimo = {detalhes['baixo']['min']}, Máximo = {detalhes['baixo']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=8,
                                     value=f"Mínimo = {detalhes['regular']['min']}, Máximo = {detalhes['regular']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=9,
                                     value=f"Mínimo = {detalhes['bom']['min']}, Máximo = {detalhes['bom']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=10,
                                     value=f"Mínimo = {detalhes['otimo']['min']}, Máximo = {detalhes['otimo']['max']}")
            #print(detalhes)

            #if metrica == 'Div. liquida/EBITDA':
            if categoria_pl == 'baixo':
                wsIndiRentabilidade.cell(row=linha2, column=6, value=categoria_pl).fill = fillvermelho
                wsIndiRentabilidade.cell(row=linha2, column=11, value=f"{detalhes['descrbaixo']}")
            if  categoria_pl == 'regular':
                wsIndiRentabilidade.cell(row=linha2, column=6, value=categoria_pl)
                wsIndiRentabilidade.cell(row=linha2, column=11, value=f"{detalhes['descregular']}")
            if  categoria_pl == 'bom':
                wsIndiRentabilidade.cell(row=linha2, column=6, value=categoria_pl)
                wsIndiRentabilidade.cell(row=linha2, column=11, value=f"{detalhes['descrbom']}")
            if categoria_pl == 'otimo':
                wsIndiRentabilidade.cell(row=linha2, column=6, value=categoria_pl)
                wsIndiRentabilidade.cell(row=linha2, column=11, value=f"{detalhes['descotimo']}")


    except Exception as e:
        print(f"Erro inesperado grava planilha: {e}")
        print(metrica)
        print(indicadortratado)

        print('gravaIndiEficiênciaoStaus - erro' ,  stock)
    finally:
        print('gravaIndiEficiênciaoStaus  OK''', stock)

def is_null_zero_or_spaces(variable):
    # Verifica se a variável é None
    if variable is None:
        return True
    # Verifica se a variável é zero (0)
    elif variable == 0:
        return True
    # Verifica se a variável é uma string e contém apenas espaços
    elif isinstance(variable, str) and variable.strip() == '':
        return True
    elif variable == '-%':
        return True
    else:
        return False


def get_stock_soup(stock):
    ''' Get raw html from a stock '''

    # access the stock urlww
    driver.get(f'https://statusinvest.com.br/acoes/{stock}')

    # get html from stock
    html = driver.find_element(By.ID, 'main-2').get_attribute('innerHTML')

    # remove accents from html and transform html into soup
    soup = BeautifulSoup(unidecode(html), 'html.parser')

    return soup


def soup_to_dict(soup):
    '''Get all data from stock soup and return as a dictionary '''
    keys, values = [], []

    # get divs from stock
    soup1 = soup.find('div', class_='pb-3 pb-md-5')
    soup2 = soup.find('div', class_='card rounded text-main-green-dark')
    soup3 = soup.find('div', class_='indicator-today-container')
    soup4 = soup.find(
        'div', class_='top-info info-3 sm d-flex justify-between mb-3')
    soups = [soup1, soup2, soup3, soup4]

    for s in soups:
        # get only titles from a div and append to keys
        titles = s.find_all('h3', re.compile('title m-0[^"]*'))
        titles = [t.get_text() for t in titles]
        keys += titles

        # get only numbers from a div and append to values
        numbers = s.find_all('strong', re.compile('value[^"]*'))
        numbers = [n.get_text()for n in numbers]
        values += numbers

    # remove unused key and insert needed keys
    keys.remove('PART. IBOV')
    keys.insert(6, 'TAG ALONG')
    keys.insert(7, 'LIQUIDEZ MEDIA DIARIA')

    # clean keys list
    keys = [k.replace('\nhelp_outline', '').strip() for k in keys]
    keys = [k for k in keys if k != '']

    # clean values list
    values = [v.replace('\nhelp_outline', '').strip() for v in values]
    values = [v.replace('.', '').replace(',', '.') for v in values]

    # create a dictionary using keys and values from indicators
    d = {k: v for k, v in zip(keys, values)}

    return d


if __name__ == "__main__":
    dict_stocks = {}
    criaPlanilhaIndRentabilidade(wbsaida)
    wsIndiRentabilidade = wbsaida['IndiRentabilidade']

    # start timer
    start = time.time()

    # read file with stocks codes to get stock information
    with open('stocks.txt', 'r') as f:
        stocks = f.read().splitlines()

        # get stock information and create excel sheet
        for stock in stocks:
            print(stock)
            try:
                # get data and transform into dictionary
                soup = get_stock_soup(stock)
                dict_stock = soup_to_dict(soup)
                dict_stocks[stock] = dict_stock
                gravaIndiEficiênciaoStaus(wsIndiRentabilidade, dict_stocks, stock)
            except:
                # if we not get the information... just skip it
                print(f'Could not get {stock} information')

    # create dataframe using dictionary of stocks informations
    df = pd.DataFrame(dict_stocks)

    # replace missing values with NaN to facilitate processing
    df = df.replace(['', '-', '--', '-%', '--%'], np.nan)

    # write dataframe into csv file
    df.to_excel('stocks_data.xlsx', index_label='indicadores')

    # exit the driver
    driver.quit()

    # end timer
    end = time.time()
    wbsaida.save("StatusInvest.xlsx")

    print(f'Brasilian stocks information got in {int(end-start)} s')
