import openpyxl
import re
import time
import numpy as np
import pandas as pd
from unidecode import unidecode
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
import requests
# teste silvio 3e
import warnings
from openpyxl.styles import numbers
import analisefundamentalista
import fundamentus2
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


#from teste01 import metrica

fillvermelho= PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") # Vermelho

fillverde= PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid") # Verde

fillamarelo =  PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # Amarelo


fillazul = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid") # Azul


#filltitulo =   PatternFill(start_color="#002060", end_color="#002060", fill_type="solid")

filltitulo = PatternFill(start_color="002060", end_color="002060", fill_type="solid")  # Azul escuro

# Novas faixas
filllaranja = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")   # Laranja
fillroxo = PatternFill(start_color="800080", end_color="800080", fill_type="solid")      # Roxo
fillcinza = PatternFill(start_color="808080", end_color="808080", fill_type="solid")     # Cinza
fillrosa = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")      # Rosa
fillciano = PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid")     # Ciano
fillpreto = PatternFill(start_color="000000", end_color="000000", fill_type="solid")     # Preto
fillbranco = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")    # Branco

# Tons pastéis
fillpastelverde = PatternFill(start_color="B0E57C", end_color="B0E57C", fill_type="solid")   # Verde pastel
fillpastelazul = PatternFill(start_color="A7C7E7", end_color="A7C7E7", fill_type="solid")    # Azul pastel
fillpastelrosa = PatternFill(start_color="FFD1DC", end_color="FFD1DC", fill_type="solid")    # Rosa pastel
fillpastelroxo = PatternFill(start_color="D8BFD8", end_color="D8BFD8", fill_type="solid")    # Roxo pastel
fillpastelamarelo = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid") # Amarelo claro

# Tons metálicos simulados
fillouro = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")          # Ouro
fillprata = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")          # Prata
fillbronze = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")         # Bronze

# Variações de cores básicas
fillvermelhoescuro = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid") # Vermelho escuro
fillverdeescuro = PatternFill(start_color="006400", end_color="006400", fill_type="solid")    # Verde escuro
fillazulescuro = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")      # Azul escuro
fillamareloouro = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")     # Amarelo ouro
filllaranjaescuro = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")   # Laranja escuro

# Tons neutros
fillcinzaclaro = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")      # Cinza claro
fillmarrom = PatternFill(start_color="A52A2A", end_color="A52A2A", fill_type="solid")          # Marrom
fillbege = PatternFill(start_color="F5F5DC", end_color="F5F5DC", fill_type="solid")            # Bege
fillcreme = PatternFill(start_color="FFFDD0", end_color="FFFDD0", fill_type="solid")           # Creme

# Cores vibrantes
fillmagenta = PatternFill(start_color="FF00FF", end_color="FF00FF", fill_type="solid")         # Magenta
filllimão = PatternFill(start_color="CCFF00", end_color="CCFF00", fill_type="solid")           # Verde limão
fillturquesa = PatternFill(start_color="40E0D0", end_color="40E0D0", fill_type="solid")         # Turquesa
fillsalmao = PatternFill(start_color="FA8072", end_color="FA8072", fill_type="solid")           # Salmão
fillcoral = PatternFill(start_color="FF7F50", end_color="FF7F50", fill_type="solid")            # Coral

# Fontes básicas
font_branca = Font(color="FFFFFF")   # Branco
font_preta = Font(color="000000")    # Preto
font_vermelha = Font(color="FF0000") # Vermelho
font_verde = Font(color="00FF00")    # Verde
font_azul = Font(color="0000FF")     # Azul
font_amarela = Font(color="FFFF00")  # Amarelo

# Fontes suaves
font_cinza = Font(color="808080")    # Cinza
font_rosa = Font(color="FFC0CB")     # Rosa
font_roxa = Font(color="800080")     # Roxo
font_laranja = Font(color="FFA500")  # Laranja

# Fontes metálicas simuladas
font_ouro = Font(color="FFD700")     # Ouro
font_prata = Font(color="C0C0C0")    # Prata
font_bronze = Font(color="CD7F32")   # Bronze

# Fontes para destaque
font_titulo = Font(color="002060", bold=True, size=12)  # Azul escuro, negrito
font_alerta = Font(color="FF0000", bold=True)           # Vermelho, negrito
font_sucesso = Font(color="00AA00", italic=True)        # Verde escuro, itálic


font_branca = Font(color="FFFFFF")  # Branco

TITLES = [
    'Identificação', 'Resumo Financeiro', 'Cotações', 'Informações Básicas',
    'Oscilações', 'Indicadores de Valuation', 'Indicadores de Rentabilidade',
    'Indicadores de Endividamento', 'Balanço Patrimonial', 'Demonstrativo de Resultados'
]

linha2 = 1
metricasts= ""
''''categorias = {
    'otimo': {'min': float('-inf'), 'max': -2},  # Valores muito baixos são ótimos
    'bom': {'min': -2, 'max': 0},               # Valores entre -2 e 0
    'moderado': {'min': 0, 'max': 1.5},         # Valores entre 0 e 1.5
    'ruim': {'min': 1.5, 'max': 3},             # Valores entre 1.5 e 3
    'pessimo': {'min': 3, 'max': 4},            # Valores entre 3 e 4
    'critico': {'min': 4, 'max': float('inf')}   # Valores acima de 4
}'''
MetricasStatus = {'Giro ativos', 'Div. liquida/PL','Div. liquida/EBITDA','Div. liquida/EBIT','PL/Ativos',
                           'Passivos/Ativos','Liq. corrente','P/L','PEG Ratio','P/VP','EV/EBITDA','EV/EBIT',
                            'P/EBITDA','P/EBIT','VPA','P/Ativo','LPA',
                            'P/SR','P/Ativo Circ. Liq.','Valor atual','LIQUIDEZ MEDIA DIARIA','Patrimonio liquido',
                             'Ativos','Ativo circulante','Divida bruta','Disponibilidade',
                             'Divida liquida','Valor de mercado','Valor de firma'}

wbsaida = openpyxl.Workbook()


# define selenium webdriver options
options = webdriver.ChromeOptions()

# create selenium webdriver instance
driver = webdriver.Chrome(options=options)


def criaPlanilhaIndRentabilidade(wbsaida):
    wbsaida.create_sheet('IndiRentabilidade')
    IndiRentabilidade = wbsaida['IndiRentabilidade']
    IndiRentabilidade.append(
        ['Agrupador','Fonte','Ativo','Indicador','Formula','Definição','Referencia','Valor','Classificacao','Faixa','Descricao'])

    for cell in IndiRentabilidade[1]:  # Apenas o cabeçalho
        cell.fill = filltitulo
        cell.font = font_branca

def tratamento(indicador):
    indicador2 = indicador

    try:
        if indicador2 in ["-", "--","--%"]:
            indicador2 = 0
        elif indicador2 is None or is_null_zero_or_spaces(indicador2):
            indicador2 = 0
        else:
            if indicador2 == "":  # Verificação adicionada para string vazia
                indicador2 = 0
            else:
                indicador2 = float(indicador2.strip('%')) / 100
        return indicador2

    except Exception as e:
        print(f"Erro inesperado tratamento : {e}", "metrica  ", metricasts, " indicador  ", indicador, " stock  ",stock)


    finally:
       # print('tratamneto OK')
       pass
def tratamento3(indicador):
    indicador2 = indicador

    try:
        if indicador2 in ["-", "--","--%"]:
            indicador2 = 0
        elif indicador2 is None or is_null_zero_or_spaces(indicador2):
            indicador2 = 0
        else:
            if indicador2 == "":  # Verificação adicionada para string vazia
                indicador2 = 0
            else:
                indicador2 = float(indicador2)
        return indicador2

    except Exception as e:
        print(f"Erro inesperado tratamento 3 : {e}", " metrica  ", metricasts, " indicador  ", indicador, " stock  ",stock)



    finally:
        #print('tratamneto3 OK')
        pass


def tratamento2(indicador):
    indicador2 = indicador

    try:
        if indicador2 in ["-", "--","--%"]:
            indicador2 = 0
        elif indicador2 is None or is_null_zero_or_spaces(indicador2):
            indicador2 = 0
        elif indicador2 == "":  # Verificação adicionada para string vazia
            indicador2 = 0
        else:
            indicador2 = float(indicador2)
        return indicador2


        return indicador2
    except Exception as e:

        print('tratamneto2 - erro', stock)


    finally:
        #print('tratamneto2 OK', indicador)
        pass
def gravaIndiEficiênciaoStaus(wsIndiRentabilidade, dict_stocks, stock):




    global linha2
    global metricasts
    #linha2 = 1
    try:
        #print(dict_stocks)
        for metrica in MetricasStatus:
    #        print(f'Métrica: {metrica}')
            linha2 += 1
            metricasts = metrica
            if metrica in ['Giro ativos', 'Div. liquida/PL','Div. liquida/EBITDA','Div. liquida/EBIT','PL/Ativos',
                           'Passivos/Ativos','Liq. corrente','P/L','PEG Ratio','P/VP','EV/EBITDA','EV/EBIT',
                            'P/EBITDA','P/EBIT','VPA','P/Ativo','LPA',
                            'P/SR','P/Ativo Circ. Liq.', 'Disponibilidade','Patrimonio liquido','Divida bruta','Divida liquida','Ativos', 'Ativo circulante','LIQUIDEZ MEDIA DIARIA']:


                if metrica == 'P/L':
                   indicadortratado = tratamento2(dict_stocks[stock].get(metrica))
                   valor_pl = indicadortratado
                   resultado =  analisefundamentalista.evaluate_pl(valor_pl)  # P/L

                elif metrica == 'P/EBITDA':
                   indicadortratado = tratamento2(dict_stocks[stock].get(metrica))
                   valor_pl = indicadortratado
                   resultado = analisefundamentalista.evaluate_pebitda(valor_pl)  # P/EBITDA

                elif metrica == 'P/VP':
                   indicadortratado = tratamento2(dict_stocks[stock].get(metrica))
                   valor_pl = indicadortratado
                   resultado = analisefundamentalista.evaluate_pvp(valor_pl)  # P/VP

                elif metrica == 'P/EBIT':
                    indicadortratado = tratamento2(dict_stocks[stock].get(metrica))
                    valor_pl = indicadortratado
                    resultado = analisefundamentalista.evaluate_pebit(valor_pl)  # P/EBIT

                elif metrica == 'EV/EBITDA':
                    indicadortratado = tratamento2(dict_stocks[stock].get(metrica))
                    valor_pl = indicadortratado
                    resultado = analisefundamentalista.evaluate_evebitda(valor_pl)  # EV/EBITDA

                elif metrica == 'EV/EBIT':
                    indicadortratado = tratamento2(dict_stocks[stock].get(metrica))
                    valor_pl = indicadortratado
                    resultado = analisefundamentalista.evaluate_evebit(valor_pl)  # EV/EBIT

                elif metrica == 'Giro ativos':
                    indicadortratado = tratamento2(dict_stocks[stock].get(metrica))
                    valor_pl = indicadortratado
                    resultado = analisefundamentalista.evaluate_giro_ativos(valor_pl)  # Giro ativos

                elif metrica == 'Div. liquida/PL':
                    indicadortratado = tratamento2(dict_stocks[stock].get(metrica))
                    valor_pl = indicadortratado
                    resultado = analisefundamentalista.evaluate_divida_liquida_pl(valor_pl)  # Div. liquida/PL

                elif metrica == 'Div. liquida/EBITDA':
                    indicadortratado = tratamento2(dict_stocks[stock].get(metrica))
                    valor_pl = indicadortratado
                    resultado = analisefundamentalista.evaluate_divida_liquida_ebitda(valor_pl)  # Div. liquida/EBITDA

                elif metrica == 'Div. liquida/EBIT':
                    indicadortratado = tratamento2(dict_stocks[stock].get(metrica))
                    valor_pl = indicadortratado
                    resultado = analisefundamentalista.evaluate_divida_liquida_ebit(valor_pl)  # Div. liquida/EBIT

                elif metrica == 'PL/Ativos':
                    indicadortratado = tratamento2(dict_stocks[stock].get(metrica))
                    valor_pl = indicadortratado
                    resultado = analisefundamentalista.evaluate_pl_ativos(valor_pl)  # PL/Ativos

                elif metrica == 'Passivos/Ativos':
                    indicadortratado = tratamento2(dict_stocks[stock].get(metrica))
                    valor_pl = indicadortratado
                    resultado = analisefundamentalista.evaluate_passivos_ativos(valor_pl)  # Passivos/Ativos

                elif metrica == 'Liq. corrente':
                    indicadortratado = tratamento2(dict_stocks[stock].get(metrica))
                    valor_pl = indicadortratado
                    resultado = analisefundamentalista.evaluate_liquidez_corrente(valor_pl)  # Liq. corrente

                elif metrica == 'PEG Ratio':
                    indicadortratado = tratamento2(dict_stocks[stock].get(metrica))
                    valor_pl = indicadortratado
                    resultado = analisefundamentalista.evaluate_peg_ratio(valor_pl)  # PEG Ratio

                elif metrica == 'P/Ativo':
                    indicadortratado = tratamento2(dict_stocks[stock].get(metrica))
                    valor_pl = indicadortratado
                    resultado = analisefundamentalista.evaluate_p_ativo(valor_pl)  # P/Ativo

                elif metrica == 'VPA':
                    indicadortratado = tratamento2(dict_stocks[stock].get(metrica))
                    valor_pl = indicadortratado
                    resultado = analisefundamentalista.evaluate_vpa(valor_pl)  # VPA


                elif metrica == 'LPA':
                    indicadortratado = tratamento2(dict_stocks[stock].get(metrica))
                    valor_pl = indicadortratado
                    resultado = analisefundamentalista.evaluate_lpa(valor_pl)  # LPA
                  #  teste = fundamentus2.evaluate_teste(stock)
                   # print("teste retorno" + str(teste))
                   # indicadortratado_fundamentus = tratamento2(teste)
                   # print('indicadortratado_fundamentus ' + str(indicadortratado_fundamentus))
                   # valor_pl_fundamentus = indicadortratado_fundamentus
                   # print('valor_pl_fundamentus ' + str(valor_pl_fundamentus))
                   # resultado_fundamentus = analisefundamentalista.evaluate_vpa(valor_pl_fundamentus)  # VPA
                   # print('resultado_fundamentus' + str(resultado_fundamentus))
                   # wsIndiRentabilidade.cell(row=linha2, column=12,
                    #                         value=valor_pl_fundamentus).number_format = numbers.FORMAT_NUMBER_00

                elif metrica == 'P/SR':
                    indicadortratado = tratamento2(dict_stocks[stock].get(metrica))
                    valor_pl = indicadortratado
                    resultado = analisefundamentalista.evaluate_psr(valor_pl)  # P/SR

                elif metrica == 'P/Ativo Circ. Liq': #P/Ativo Circ. Liq
                    indicadortratado = tratamento2(dict_stocks[stock].get(metrica))
                    valor_pl = indicadortratado
                    resultado =  analisefundamentalista.evaluate_p_ativo_circ_liq(valor_pl)  # P/Ativo Circ. Liq

                elif metrica == 'Disponibilidade': #'Disponibilidade' R$
                    indicadortratado = tratamento3(dict_stocks[stock].get(metrica))
                    valor_pl = indicadortratado
                    resultado = analisefundamentalista.evaluate_disponibilidade(valor_pl)

                elif metrica == 'Patrimonio liquido': #'Patrimonio liquido' R$
                    indicadortratado = tratamento3(dict_stocks[stock].get(metrica))
                    valor_pl = indicadortratado
                    resultado = analisefundamentalista.evaluate_patrimonio_liquido(valor_pl)#Patrimonio liquido

                elif metrica == 'Divida bruta': #'Divida bruta' R$
                    indicadortratado = tratamento3(dict_stocks[stock].get(metrica))
                    valor_pl = indicadortratado
                    resultado = analisefundamentalista.evaluate_divida_bruta(valor_pl)#Divida bruta

                elif metrica == 'Divida liquida':  # 'Divida liquida' R$
                    indicadortratado = tratamento3(dict_stocks[stock].get(metrica))
                    valor_pl = indicadortratado
                    resultado = analisefundamentalista.evaluate_divida_liquida(valor_pl)  # Divida liquida

                elif metrica == 'Ativos':  # 'Ativos' R$
                    indicadortratado = tratamento3(dict_stocks[stock].get(metrica))
                    valor_pl = indicadortratado
                    resultado = analisefundamentalista.evaluate_ativos(valor_pl)  # Ativos

                elif metrica == 'Ativos':  # 'Ativosa' R$
                    indicadortratado = tratamento3(dict_stocks[stock].get(metrica))
                    valor_pl = indicadortratado
                    resultado = analisefundamentalista.evaluate_ativos(valor_pl)  # Ativos

                elif metrica == 'Ativo circulante':  # 'Ativo circulante' R$
                    indicadortratado = tratamento3(dict_stocks[stock].get(metrica))
                    valor_pl = indicadortratado
                    resultado = analisefundamentalista.evaluate_ativo_circulante(valor_pl)  # Ativo circulante

                elif metrica == 'LIQUIDEZ MEDIA DIARIA':  # 'ALIQUIDEZ MEDIA DIARIA' R$
                    indicadortratado = tratamento3(dict_stocks[stock].get(metrica))
                    valor_pl = indicadortratado

                    resultado = analisefundamentalista.evaluate_liquidez_media_diaria(valor_pl)  #LIQUIDEZ MEDIA DIARIA
                elif metrica == 'Valor de firma1':  # 'ALIQUIDEZ MEDIA DIARIA' R$
                    indicadortratado = tratamento3(dict_stocks[stock].get(metrica))
                    valor_pl = indicadortratado
                    resultado = analisefundamentalista.evaluate_liquidez_media_diaria(valor_pl)  # LIQUIDEZ MEDIA DIARIA

                elif metrica == 'Valor atual1':  # 'Valor atual' R$
                    indicadortratado = tratamento3(dict_stocks[stock].get(metrica))
                    valor_pl = indicadortratado
                    resultado = analisefundamentalista.evaluate_liquidez_media_diaria(valor_pl)  # Valor atual

                elif metrica == 'Valor de mercado1':  # 'Valor de mercado' R$
                    indicadortratado = tratamento3(dict_stocks[stock].get(metrica))
                    valor_pl = indicadortratado
                    resultado = analisefundamentalista.evaluate_liquidez_media_diaria(valor_pl)  # Valor de mercado

                faixa = resultado['faixa']
                descricao = resultado['descricao']
                classificacao = resultado['classificacao']
                definicao = resultado['definicao']
                agrupador = resultado['agrupador']
                formula  = resultado['formula']
                #print(faixa)

            wsIndiRentabilidade.cell(row=linha2, column=1, value=agrupador)
            wsIndiRentabilidade.cell(row=linha2, column=2, value='StausInvest')
            wsIndiRentabilidade.cell(row=linha2, column=3, value=stock)
            wsIndiRentabilidade.cell(row=linha2, column=4, value=metrica)
            wsIndiRentabilidade.cell(row=linha2, column=5, value=formula)
            wsIndiRentabilidade.cell(row=linha2, column=6, value=definicao)
            #wsIndiRentabilidade.cell(row=linha2, column=7, value=faixa)
            wsIndiRentabilidade.cell(row=linha2, column=11, value=descricao)

            if metrica in ['Giro ativos', 'Div. liquida/PL','Div. liquida/EBITDA','Div. liquida/EBITDA',
                           'Div. liquida/EBIT','PL/Ativos','Passivos/Ativos','Liq. corrente',
                           'P/L','PEG Ratio','P/VP','EV/EBITDA','EV/EBIT',
                            'P/EBITDA','P/EBIT','VPA','P/Ativo','LPA',
                            'P/SR','P/Ativo Circ. Liq.']:
               wsIndiRentabilidade.cell(row=linha2, column=8, value=valor_pl).number_format = numbers.FORMAT_NUMBER_00

            elif  metrica in ['Valor atual','LIQUIDEZ MEDIA DIARIA','Patrimonio liquido',
                             'Ativos','Ativo circulante','Divida bruta','Disponibilidade',
                             'Divida liquida','Valor de mercado','Valor de firma']:

                  wsIndiRentabilidade.cell(row=linha2, column=8, value=valor_pl).number_format = 'R$ #,##0.00'
            else:
                wsIndiRentabilidade.cell(row=linha2, column=8, value=valor_pl).number_format = numbers.FORMAT_PERCENTAGE_00



            if classificacao == 'Critico':
                wsIndiRentabilidade.cell(row=linha2, column=9, value=classificacao).fill = fillvermelho
                wsIndiRentabilidade.cell(row=linha2, column=10, value=faixa)
            if classificacao == 'Pessimo':
                wsIndiRentabilidade.cell(row=linha2, column=9, value=classificacao).fill = fillvermelho
                wsIndiRentabilidade.cell(row=linha2, column=10, value=faixa)
            if classificacao == 'Ruim':
                wsIndiRentabilidade.cell(row=linha2, column=9, value=classificacao).fill = fillvermelho
                wsIndiRentabilidade.cell(row=linha2, column=10, value=faixa)
            if  classificacao == 'Moderado':
                wsIndiRentabilidade.cell(row=linha2, column=9, value=classificacao).fill =fillamarelo
                wsIndiRentabilidade.cell(row=linha2, column=10, value=faixa)
            if  classificacao == 'Bom':
                wsIndiRentabilidade.cell(row=linha2, column=9, value=classificacao).fill = fillverde
                wsIndiRentabilidade.cell(row=linha2, column=10, value=faixa)
            if classificacao == 'Otimo':
                wsIndiRentabilidade.cell(row=linha2, column=9, value=classificacao).fill =fillverde
                wsIndiRentabilidade.cell(row=linha2, column=10, value=faixa)
            if classificacao == 'Fora da faixa':
                wsIndiRentabilidade.cell(row=linha2, column=9, value=classificacao).fill = fillazul
                wsIndiRentabilidade.cell(row=linha2, column=10, value=faixa)



    except Exception as e:
        print(f"Erro inesperado grava planilha2: {e}")
        print(metrica)
        print(indicadortratado)

        print('gravaIndiEficiênciaoStaus - erro' ,  stock,"    ", metrica)
    finally:
        print('gravaIndiEficiênciaoStaus  OK''', stock)

def is_null_zero_or_spaces(variable):
    # Verifica se a variável é None
    if variable is None:
        return True
    # Verifica se a variável é zero (0)
    elif variable == 0:
        return True
    # Verifica se a variável é uma string e contém apenas espaços
    elif isinstance(variable, str) and variable.strip() == '':
        return True
    elif variable == '-%':
        return True
    else:
        return False


def get_stock_soup(stock):
    ''' Get raw html from a stock '''

    # access the stock urlww
    driver.get(f'https://statusinvest.com.br/acoes/{stock}')

    # get html from stock
    html = driver.find_element(By.ID, 'main-2').get_attribute('innerHTML')

    # remove accents from html and transform html into soup
    soup = BeautifulSoup(unidecode(html), 'html.parser')

    return soup


def soup_to_dict(soup):
    '''Get all data from stock soup and return as a dictionary '''
    keys, values = [], []

    # get divs from stock
    soup1 = soup.find('div', class_='pb-3 pb-md-5')
    soup2 = soup.find('div', class_='card rounded text-main-green-dark')
    soup3 = soup.find('div', class_='indicator-today-container')
    soup4 = soup.find(
        'div', class_='top-info info-3 sm d-flex justify-between mb-3')
    soups = [soup1, soup2, soup3, soup4]

    for s in soups:
        # get only titles from a div and append to keys
        titles = s.find_all('h3', re.compile('title m-0[^"]*'))
        titles = [t.get_text() for t in titles]
        keys += titles

        # get only numbers from a div and append to values
        numbers = s.find_all('strong', re.compile('value[^"]*'))
        numbers = [n.get_text()for n in numbers]
        values += numbers

    # remove unused key and insert needed keys
    keys.remove('PART. IBOV')
    keys.insert(6, 'TAG ALONG')
    keys.insert(7, 'LIQUIDEZ MEDIA DIARIA')

    # clean keys list
    keys = [k.replace('\nhelp_outline', '').strip() for k in keys]
    keys = [k for k in keys if k != '']

    # clean values list
    values = [v.replace('\nhelp_outline', '').strip() for v in values]
    values = [v.replace('.', '').replace(',', '.') for v in values]

    # create a dictionary using keys and values from indicators
    d = {k: v for k, v in zip(keys, values)}

    return d


if __name__ == "__main__":
    dict_stocks = {}
    criaPlanilhaIndRentabilidade(wbsaida)
    wsIndiRentabilidade = wbsaida['IndiRentabilidade']
    #teste = fundamentus2.evaluate_teste(1)
    #print("teste " + str(teste))
    # start t   imer
    start = time.time()

    # read file with stocks codes to get stock information
    with open('stocks.txt', 'r') as f:
        stocks = f.read().splitlines()

        # get stock information and create excel sheet
        for stock in stocks:
            #print("stock :"  ,stock)
            try:
                # get data and transform into dictionary
                soup = get_stock_soup(stock)
                dict_stock = soup_to_dict(soup)
                dict_stocks[stock] = dict_stock
                gravaIndiEficiênciaoStaus(wsIndiRentabilidade, dict_stocks, stock)
            except:
                # if we not get the information... just skip it
                print(f'Could not get {stock} information', "    ", metricasts)

    # create dataframe using dictionary of stocks informations
    df = pd.DataFrame(dict_stocks)

    # replace missing values with NaN to facilitate processing
    df = df.replace(['', '-', '--', '-%', '--%'], np.nan)

    # write dataframe into csv file
    df.to_excel('stocks_data.xlsx', index_label='indicadores')

    # exit the driver
    driver.quit()

    # end timer
    end = time.time()
    altura_padrao = 40
    largura_padrao = 30
    for row in wsIndiRentabilidade.iter_rows():
        wsIndiRentabilidade.row_dimensions[row[0].row].height = altura_padrao  # Altura da linha
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)
            col_letter = get_column_letter(cell.column)
            wsIndiRentabilidade.column_dimensions[col_letter].width = largura_padrao  # Largura da coluna
    wbsaida.save("StatusInvest.xlsx")
    print(f'Brasilian stocks information got in {int(end-start)} s')
# silvio teste