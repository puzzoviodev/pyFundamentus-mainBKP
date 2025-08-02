import openpyxl
import re
import time
import numpy as np
import pandas as pd
from unidecode import unidecode
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
import requests

import warnings
from openpyxl.styles import numbers

TITLES = [
    'Identificação', 'Resumo Financeiro', 'Cotações', 'Informações Básicas',
    'Oscilações', 'Indicadores de Valuation', 'Indicadores de Rentabilidade',
    'Indicadores de Endividamento', 'Balanço Patrimonial', 'Demonstrativo de Resultados'
]

linha2 = 1
MetricasStatus = {
    'Div. liquida/PL': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Endividamento'
    }
}


wbsaida = openpyxl.Workbook()


# define selenium webdriver options
options = webdriver.ChromeOptions()

# create selenium webdriver instance
driver = webdriver.Chrome(options=options)

def categorizar_valor(metrica, valor):
    try:
        if metrica not in MetricasStatus:
            return 'Métrica não reconhecida'
        metrica = 'Dív. líquida/PL'
        if metrica in ['Giro ativos', 'Dív. líquida/PL', 'Dív. líquida/EBITDA', 'Dív. líquida/EBIT', 'PL/Ativos',
                       'Passivos/Ativos'
                       'Liq. corrente']:
            valor2 = valor
        else:
            valor2 = float(valor.strip('%'))

        for categoria, limites in MetricasStatus[metrica].items():
            if categoria in ['descricao', 'agrupador']:
                continue
            if limites['min'] <= valor2 < limites['max']:
                return categoria
        return 'Valor fora do alcance definido'
    except Exception as e:
        print(f"Erro inesperado: {e}")
        print("categorizar_valor - Erro")
    finally:
        print("categorizar_valor - OK")


def criaPlanilhaIndRentabilidade(IndiRentabilidade):
    wbsaida.create_sheet('IndiRentabilidade')
    IndiRentabilidade = wbsaida['IndiRentabilidade']
    IndiRentabilidade.append(
        ['Agrupador', 'Fonte', 'ATIVO', 'Indicador', 'valor', 'referencia', 'Baixo', 'regular', 'bom', 'otimo'])
def tratamento(indicador):
    indicador2 = indicador

    if indicador2 in ["-", "--"]:
        indicador2 = ""
    elif is_null_zero_or_spaces(indicador2):
        indicador2 = 0
    else:
        #indicador2 = float(indicador2.strip('%')) / 100
        indicador2 = indicador
    return indicador2
def tratamento2(indicador):
    indicador2 = indicador
    print("tratamento2 chegada : ",indicador )
    if indicador2 in ["-", "--"]:
        indicador2 = ""
    elif is_null_zero_or_spaces(indicador2):
        indicador2 = 0

    return indicador2

def gravaIndiEficiênciaoStaus(wsIndiRentabilidade, dict_stocks, stock):
    # fontes ['StatusInvest', 'Fundamentus']



    global linha2
    #linha2 = 1
    try:
        print(MetricasStatus.items())
        print(MetricasStatus)
        print(len(MetricasStatus))
        for metrica, detalhes in MetricasStatus.items():
    #        print(f'Métrica: {metrica}')
            linha2 += 1
            print(" metrica chegando1 ", metrica)
            print(" metrica chegando 2" , dict_stocks[stock].get(metrica))
            print("chumbado ", dict_stocks[stock].get('Div. liquida/EBITDA'))
            #metrica = 'Div. liquida/PL'

            if metrica == 'Dív. líquida/PL':
               print("no if: ", metrica)
            if metrica in ['Giro ativos', 'Div. liquida/PL', 'Dív. líquida/EBITDA', 'Dív. líquida/EBIT', 'PL/Ativos',
                           'Passivos/Ativos', 'Liq. corrente']:

                metrica2 = metrica
                metrica3 ='Div. liquida/PL'
                if metrica2.strip().lower() == metrica.strip().lower():
                    print("As strings são iguais1.")
                else:
                    print("As strings são diferentes1.")

                if metrica2.strip().lower() == metrica3.strip().lower():
                    print("As strings são iguais2.")
                else:
                    print("As strings são diferentes2.")

                if metrica3.strip().lower() == metrica.strip().lower():
                    print("As strings são iguais3.")
                else:
                    print("As strings são diferentes3.")
                print("tratamento21: ", metrica3)
                indicadortratado = tratamento2(dict_stocks[stock].get(metrica))
                valor_pl = indicadortratado

                print("tratamento2: ", metrica)
            else:
                indicadortratado = tratamento(dict_stocks[stock].get(metrica))
                valor_pl = indicadortratado
                print("tratamento:   ", metrica)

            categoria_pl = categorizar_valor(metrica,valor_pl)  # Certifique-se de que 'ROE' é o valor correto para a métrica
   #         print(f'O índice P/L {valor_pl} é categorizado como: {categoria_pl}')
   #         print(f"  Agrupador: {detalhes['agrupador']}")


            wsIndiRentabilidade.cell(row=linha2, column=1, value=detalhes['agrupador'])
            wsIndiRentabilidade.cell(row=linha2, column=2, value='StausInvest')
            wsIndiRentabilidade.cell(row=linha2, column=3, value=stock)
            wsIndiRentabilidade.cell(row=linha2, column=4, value=metrica)
            if  metrica in ['Giro ativos','Dív. líquida/PL','Dív. líquida/EBITDA','Dív. líquida/EBIT','PL/Ativos','Passivos/Ativos'
                           'Liq. corrente']:
                wsIndiRentabilidade.cell(row=linha2, column=5, value=valor_pl).number_format = numbers.FORMAT_NUMBER_00
            else:
                wsIndiRentabilidade.cell(row=linha2, column=5, value=valor_pl).number_format = numbers.FORMAT_PERCENTAGE_00
            wsIndiRentabilidade.cell(row=linha2, column=6, value=categoria_pl)
            wsIndiRentabilidade.cell(row=linha2, column=7,
                                     value=f"Mínimo = {detalhes['baixo']['min']}, Máximo = {detalhes['baixo']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=8,
                                     value=f"Mínimo = {detalhes['regular']['min']}, Máximo = {detalhes['regular']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=9,
                                     value=f"Mínimo = {detalhes['bom']['min']}, Máximo = {detalhes['bom']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=10,
                                     value=f"Mínimo = {detalhes['otimo']['min']}, Máximo = {detalhes['otimo']['max']}")
    except Exception as e:
        print(f"Erro inesperado: {e}")
        print(metrica)
        print(indicadortratado)
        print('gravaIndiEficiênciaoStaus - erro')
    finally:
        print('gravaIndiEficiênciaoStaus  OK''')

def is_null_zero_or_spaces(variable):
    # Verifica se a variável é None
    if variable is None:
        return True
    # Verifica se a variável é zero (0)
    elif variable == 0:
        return True
    # Verifica se a variável é uma string e contém apenas espaços
    elif isinstance(variable, str) and variable.strip() == '':
        return True
    elif variable == '-%':
        return True
    else:
        return False


def get_stock_soup(stock):
    ''' Get raw html from a stock '''

    # access the stock urlww
    driver.get(f'https://statusinvest.com.br/acoes/{stock}')

    # get html from stock
    html = driver.find_element(By.ID, 'main-2').get_attribute('innerHTML')

    # remove accents from html and transform html into soup
    soup = BeautifulSoup(unidecode(html), 'html.parser')

    return soup


def soup_to_dict(soup):
    '''Get all data from stock soup and return as a dictionary '''
    keys, values = [], []

    # get divs from stock
    soup1 = soup.find('div', class_='pb-3 pb-md-5')
    soup2 = soup.find('div', class_='card rounded text-main-green-dark')
    soup3 = soup.find('div', class_='indicator-today-container')
    soup4 = soup.find(
        'div', class_='top-info info-3 sm d-flex justify-between mb-3')
    soups = [soup1, soup2, soup3, soup4]

    for s in soups:
        # get only titles from a div and append to keys
        titles = s.find_all('h3', re.compile('title m-0[^"]*'))
        titles = [t.get_text() for t in titles]
        keys += titles

        # get only numbers from a div and append to values
        numbers = s.find_all('strong', re.compile('value[^"]*'))
        numbers = [n.get_text()for n in numbers]
        values += numbers

    # remove unused key and insert needed keys
    keys.remove('PART. IBOV')
    keys.insert(6, 'TAG ALONG')
    keys.insert(7, 'LIQUIDEZ MEDIA DIARIA')

    # clean keys list
    keys = [k.replace('\nhelp_outline', '').strip() for k in keys]
    keys = [k for k in keys if k != '']

    # clean values list
    values = [v.replace('\nhelp_outline', '').strip() for v in values]
    values = [v.replace('.', '').replace(',', '.') for v in values]

    # create a dictionary using keys and values from indicators
    d = {k: v for k, v in zip(keys, values)}

    return d


if __name__ == "__main__":
    dict_stocks = {}
    criaPlanilhaIndRentabilidade(wbsaida)
    wsIndiRentabilidade = wbsaida['IndiRentabilidade']

    # start timer
    start = time.time()

    # read file with stocks codes to get stock information
    with open('../stocks.txt', 'r') as f:
        stocks = f.read().splitlines()

        # get stock information and create excel sheet
        for stock in stocks:
            try:
                # get data and transform into dictionary
                soup = get_stock_soup(stock)
                dict_stock = soup_to_dict(soup)
                dict_stocks[stock] = dict_stock
                #print("antes "  ,dict_stocks[stock].get('Div. liquida/EBITDA'))
                gravaIndiEficiênciaoStaus(wsIndiRentabilidade, dict_stocks, stock)
            except:
                # if we not get the information... just skip it
                print(f'Could not get {stock} information')

    # create dataframe using dictionary of stocks informations
    df = pd.DataFrame(dict_stocks)

    # replace missing values with NaN to facilitate processing
    df = df.replace(['', '-', '--', '-%', '--%'], np.nan)

    # write dataframe into csv file
    df.to_excel('stocks_data.xlsx', index_label='indicadores')

    # exit the driver
    driver.quit()

    # end timer
    end = time.time()
    print(dict_stocks[stock])
    wbsaida.save("StatusInvestbase2.xlsx")

    print(f'Brasilian stocks information got in {int(end-start)} s')
