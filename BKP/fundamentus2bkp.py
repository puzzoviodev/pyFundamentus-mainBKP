import fundamentus

import openpyxl
import time
from openpyxl.styles import numbers

TITLES = [
    'Identificação', 'Resumo Financeiro', 'Cotações', 'Informações Básicas',
    'Oscilações', 'Indicadores de Valuation', 'Indicadores de Rentabilidade',
    'Indicadores de Endividamento', 'Balanço Patrimonial', 'Demonstrativo de Resultados'
]


linha2 = 1
MetricasFund = {
    'ROE': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência'
    },

    'ROIC': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência'
    },
    'Giro ativos': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência'
    },

    'Margem bruta': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência'
    },

    'Margem EBIT': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência'
    },
    'Margem líquida': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência'
    },

'Dívida líquida/Patrim': {
    'baixo': {'min': 0, 'max': 3},
    'regular': {'min': 3, 'max': 6},
    'bom': {'min': 6, 'max': 10},
    'otimo': {'min': 10, 'max': float('inf')},
    'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
    'agrupador': 'Endividamento'
    },
'Dívida líquida/EBITDA': {
    'baixo': {'min': 0, 'max': 3},
    'regular': {'min': 3, 'max': 6},
    'bom': {'min': 6, 'max': 10},
    'otimo': {'min': 10, 'max': float('inf')},
    'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
    'agrupador': 'Endividamento'
    },
'PL/Ativos': {
    'baixo': {'min': 0, 'max': 3},
    'regular': {'min': 3, 'max': 6},
    'bom': {'min': 6, 'max': 10},
    'otimo': {'min': 10, 'max': float('inf')},
    'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
    'agrupador': 'Endividamento'
    },
'Liquidez corrente': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Endividamento'
    },
'Dividend Yield': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Valuation'
    },
'P/L': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Valuation'
    },
'P/VP': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Valuation'
    },
'EV/EBITDA': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Valuation'
    },
'EV/EBIT': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Valuation'
    },

'P/EBIT': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Valuation'
    },

'VPA': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Valuation'
    },
'Preço/Ativos': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Valuation'
    },
'LPA': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Valuation'
    },
'PSR': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Valuation'
    },

'Preço/Capital de giro': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Valuation'
    },

'Preço/Ativ circ liq': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Valuation'
    },
'Cotação': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa'
    },
'Volume negociado por dia': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa'
    },

'Patrimônio líquido': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa'
    },
'Ativo': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa'
    },

'Ativo circulante': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa'
    },

'Dívida bruta': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa'
    },

'Disponibilidades': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa'
    },

'Dívida líquida': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa'
    },

'Valor de mercado': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa'
    }
}

wbsaida = openpyxl.Workbook()
Dicrentabilidade = {}
Dicstock_identification = {}
Dicfinancial_summary = {}
Dicprice_information = {}
Dicdetailed_information = {}
Dicoscillations = {}
Dicprofitability_indicators = {}
Dicindebtedness_indicators = {}
Dicbalance_sheet = {}
Dicincome_statement = {}
Dicvaluation_indicators = {}


def categorizar_valor(metrica, valor):
    try:
        if metrica not in MetricasFund:
            return 'Métrica não reconhecida'

        for categoria, limites in MetricasFund[metrica].items():
            if categoria in ['descricao', 'agrupador']:
                continue
            if limites['min'] <= valor < limites['max']:
                return categoria
        return 'Valor fora do alcance definido'
    except Exception as e:
        print(f"Erro inesperado: {e}")
        print("categorizar_valor - Erro")
    finally:
        print("categorizar_valor - OK")




def criaPlanilhaIndRentabilidade(IndiRentabilidade):
    wbsaida.create_sheet('IndiRentabilidade')
    IndiRentabilidade = wbsaida['IndiRentabilidade']
    IndiRentabilidade.append(
        ['Agrupador', 'Fonte', 'ATIVO', 'Indicador', 'valor', 'referencia', 'Baixo', 'regular', 'bom', 'otimo'])


def montadicionario(stock_identification, financial_summary, price_information, detailed_information, oscillations,
                    valuation_indicators, profitability_indicators, indebtedness_indicators, balance_sheet,
                    income_statement):
    try:

        for information in stock_identification:
            Dicstock_identification[stock_identification[information].title] = stock_identification[information].value
        for information in financial_summary:
            Dicfinancial_summary[financial_summary[information].title] = financial_summary[information].value
        for information in price_information:
            Dicprice_information[price_information[information].title] = price_information[information].value
        for information in detailed_information:
            if information != 'variation_52_weeks':
                Dicdetailed_information[detailed_information[information].title] = detailed_information[
                    information].value
            else:
                for sub_information in detailed_information[information]:
                    Dicdetailed_information[detailed_information[information][sub_information].title] = \
                        detailed_information[information][sub_information].value
        for information in oscillations:
            Dicoscillations[oscillations[information].title] = oscillations[information].value
        for information in valuation_indicators:
            Dicvaluation_indicators[valuation_indicators[information].title] = valuation_indicators[information].value
        for information in profitability_indicators:
            Dicprofitability_indicators[profitability_indicators[information].title] = profitability_indicators[
                information].value
        for information in indebtedness_indicators:
            Dicindebtedness_indicators[indebtedness_indicators[information].title] = indebtedness_indicators[
                information].value
        for information in balance_sheet:
            Dicbalance_sheet[balance_sheet[information].title] = balance_sheet[information].value
        # for information in income_statement['three_months']:
        #   Dicincome_statement[income_statement[information].title] = income_statement[information].value
        # for information in income_statement['twelve_months']:
        #   Dicincome_statement[income_statement[information].title] = income_statement[information].value

    except Exception as e:
        print(f"Erro inesperado: {e}")
        print("montadicionario - Erro")
    finally:
        print("montadicionario - OK")
    return


def teste():
    fontes = ['StatusInvest', 'Fundamentus']

    try:
        for fonte in fontes:
            print(fonte)
    except Exception as e:
        print(f"Erro inesperado: {e}")
    finally:
        print("teste")
    return


def tratamento(indicador):
    indicador2 = indicador

    if indicador2 == "-":
        indicador2 = ""
    if is_null_zero_or_spaces(indicador2):
        indicador2 = 0
    else:
        indicador2 = float(indicador2.strip('%')) / 100
    return indicador2

def gravaIndiRentabilidadeFund(wsIndiRentabilidade, Dicprofitability_indicators,Dicindebtedness_indicators,
                               Dicvaluation_indicators,
                               Dicprice_information,Dicdetailed_information,Dicbalance_sheet,Dicfinancial_summary,stock):
    # Condicional corrigida


    global linha2

    try:


        for metrica, detalhes in MetricasFund.items():
           # print(f'Métrica: {metrica}')
            linha2 += 1
            if detalhes['agrupador'] == 'Eficiência':
                indicadortratado = tratamento(f"{float(Dicprofitability_indicators.get(metrica)) * 100}%")
            elif detalhes['agrupador'] == 'Endividamento':
                indicadortratado = tratamento(f"{float(Dicindebtedness_indicators.get(metrica)) * 100}%")
            elif detalhes['agrupador'] == 'Valuation' and metrica == 'VPA':
                indicadortratado = tratamento(f"{float(Dicdetailed_information.get(metrica)) * 100}%")
            elif detalhes['agrupador'] == 'Valuation' and metrica == 'LPA':
                indicadortratado = tratamento(f"{float(Dicdetailed_information.get(metrica)) * 100}%")
            elif detalhes['agrupador'] == 'Valuation':
                indicadortratado = tratamento(f"{float(Dicvaluation_indicators.get(metrica)) * 100}%")
            elif detalhes['agrupador'] == 'Empresa' and  metrica == 'Cotação' :
                indicadortratado = tratamento(f"{float(Dicprice_information.get(metrica)) * 100}%")
            elif detalhes['agrupador'] == 'Empresa' and metrica == 'Volume negociado por dia':
                indicadortratado = tratamento(f"{float(Dicdetailed_information.get(metrica)) * 100}%")
            elif detalhes['agrupador'] == 'Empresa' and metrica == 'Valor de mercado':
                indicadortratado = tratamento(f"{float(Dicfinancial_summary.get(metrica)) * 100}%")
            elif detalhes['agrupador'] == 'Empresa':
                indicadortratado = tratamento(f"{float(Dicbalance_sheet.get(metrica)) * 100}%")

            valor_pl = indicadortratado
            categoria_pl = categorizar_valor(metrica,
                                             valor_pl)  # Certifique-se de que 'ROE' é o valor correto para a métrica
        #    print(f'O índice P/L {valor_pl} é categorizado como: {categoria_pl}')

            # Certifique-se de que a chave 'Indicador' realmente existe no dicionário



            wsIndiRentabilidade.cell(row=linha2, column=1, value=detalhes['agrupador'])
            wsIndiRentabilidade.cell(row=linha2, column=2, value='Fundamentus')
            wsIndiRentabilidade.cell(row=linha2, column=3, value=stock)
            wsIndiRentabilidade.cell(row=linha2, column=4, value=metrica)
            wsIndiRentabilidade.cell(row=linha2, column=5, value=f"{valor_pl * 100}%").number_format = numbers.FORMAT_PERCENTAGE_00
            wsIndiRentabilidade.cell(row=linha2, column=6, value=categoria_pl)
            wsIndiRentabilidade.cell(row=linha2, column=7,
                                     value=f"Mínimo = {detalhes['baixo']['min']}, Máximo = {detalhes['baixo']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=8,
                                     value=f"Mínimo = {detalhes['regular']['min']}, Máximo = {detalhes['regular']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=9,
                                     value=f"Mínimo = {detalhes['bom']['min']}, Máximo = {detalhes['bom']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=10,
                                     value=f"Mínimo = {detalhes['otimo']['min']}, Máximo = {detalhes['otimo']['max']}")


    except Exception as e:
        print(f"Erro inesperado: {e}")
        print(metrica)
        print(indicadortratado)
        print('gravaIndiRentabilidadeFunds - erro')
    finally:
        print('gravaIndiRentabilidadeFund - OK')

def is_null_zero_or_spaces(variable):
    # Verifica se a variável é None
    if variable is None:
        return True
    # Verifica se a variável é zero (0)
    elif variable == 0:
        return True
    # Verifica se a variável é uma string e contém apenas espaços
    elif isinstance(variable, str) and variable.strip() == '':
        return True
    elif variable == '-%':
        return True
    else:
        return False

def evaluate_teste(pfcl):
    '''Avalia o Preço/Fluxo de Caixa Livre com base em faixas definidas:
    - P/FCL < 0: Crítico (fluxo de caixa livre negativo, risco elevado)
    - 0 ≤ P/FCL ≤ 5: Ótimo (subvalorizado, oportunidade de compra)
    - 5 < P/FCL ≤ 10: Moderado (valuation justo)
    - 10 < P/FCL ≤ 15: Ruim (sobrevalorizado, cautela necessária)
    - 15 < P/FCL ≤ 20: Péssimo (muito caro, alto risco)
    - P/FCL > 20: Fora da faixa (extremamente sobrevalorizado)'''

    #criaPlanilhaIndRentabilidade(wbsaida)
    #wsIndiRentabilidade = wbsaida['IndiRentabilidade']

    # start timer
    start = time.time()

    # get stock information and create excel sheet
    print('valor passado' + str(pfcl))
    stock = pfcl
    main_pipeline = fundamentus.Pipeline(stock)
    response = main_pipeline.get_all_information()

    # Extract the information from the response.
    stock_identification = response.transformed_information['stock_identification']
    financial_summary = response.transformed_information['financial_summary']
    price_information = response.transformed_information['price_information']
    detailed_information = response.transformed_information['detailed_information']
    oscillations = response.transformed_information['oscillations']
    valuation_indicators = response.transformed_information['valuation_indicators']
    profitability_indicators = response.transformed_information['profitability_indicators']
    indebtedness_indicators = response.transformed_information['indebtedness_indicators']
    balance_sheet = response.transformed_information['balance_sheet']
    income_statement = response.transformed_information['income_statement']

    montadicionario(
        stock_identification, financial_summary, price_information,
        detailed_information, oscillations, valuation_indicators,
        profitability_indicators, indebtedness_indicators,
        balance_sheet, income_statement
    )

   # gravaIndiRentabilidadeFund(
   #     wsIndiRentabilidade, Dicprofitability_indicators,
   #     Dicindebtedness_indicators, Dicvaluation_indicators,
   #     Dicprice_information, Dicdetailed_information,
   #     Dicbalance_sheet, Dicfinancial_summary, stock
   # )
    LPA = Dicdetailed_information.get('LPA')
    print('teste saida' + str(LPA) )
    end = time.time()
    #wbsaida.save("Fundamentus_teste1.xlsx")  # silvio
    # print(f'Brasilian stocks information got in {int(end - start)} s')

    #indicadortratado = tratamento(f"{float(Dicdetailed_information.get('LPA')) * 100}%")
    return  LPA

