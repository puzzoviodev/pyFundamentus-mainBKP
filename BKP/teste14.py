#!/usr/bin/env python
# encoding: utf-8

# ------------------------------------------------------------------------------
#  Name: run.py
#  Version: 0.0.1
#
#  Summary: Python Fundamentus
#           Python Fundamentus is a Python API that allows you to quickly
#           access the main fundamental indicators of the main stocks
#           in the Brazilian market.
#
#  Author: Alexsander Lopes Camargos
#  Author-email: alcamargos@vivaldi.net
#
#  License: MIT
# ------------------------------------------------------------------------------

"""
Python Fundamentus API: Instant access to key financial indicators of
Brazilian stocks, empowering investors with comprehensive market analysis.
"""

import fundamentus

# statusinvest

import openpyxl
import re
import time
import numpy as np
import pandas as pd
from unidecode import unidecode
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
import requests
from openpyxl.styles import numbers

TITLES = [
    'Identificação', 'Resumo Financeiro', 'Cotações', 'Informações Básicas',
    'Oscilações', 'Indicadores de Valuation', 'Indicadores de Rentabilidade',
    'Indicadores de Endividamento', 'Balanço Patrimonial', 'Demonstrativo de Resultados'
]

linha2 = 1
MetricasStatus = {
    'M. Liquida': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência'
    },
    'M. EBIT': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Capacidade de pagar dívidas. Menor que 2.5 é considerado saudável.',
        'agrupador': 'Eficiência'
    },
    'M. EBITDA': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento do fluxo de caixa livre. Acima de 10% é considerado bom.',
        'agrupador': 'Eficiência'
    },
    'M. Bruta': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Eficiência'
    },
    'Dív. líquida/PL': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Endividamento'
    },

    'Dív. líquida/EBITDA': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Endividamento'
    },

    'Dív. líquida/EBIT': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Endividamento'
    },

    'PL/Ativos': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Endividamento'
    },

    'Passivos/Ativos': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Endividamento'
    },
    'Liq. corrente': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Endividamento'
    },
    'D.Y': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation'
    },

    'P/L': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation'
    },

    'PEG Ratio': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation'
    },

    'P/VP': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation'
    },

    'EV/EBITDA': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation'
    },

    'EV/EBIT': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation'
    },

    'P/EBITDA': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation'
    },

    'P/EBIT': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation'
    },

    'VPA': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation'
    },

    'P/Ativo': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation'
    },

    'LPA': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation'
    },

    'P/SR': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation'
    },

    'P/Ativo Circ. Liq.': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation'
    },
    'Valor atual': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa'
    },

    'TAG ALONG': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa'
    },

    'LIQUIDEZ MEDIA DIARIA': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa'
    },

    'Patrimonio liquido': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa'
    },

    'Ativos': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa'
    },

    'Ativo circulante': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa'
    },

    'Divida bruta': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa'
    },

    'Disponibilidade': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa'
    },

    'Divida liquida': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa'
    },

    'Valor de mercado': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa'
    },

    'Valor de firma': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa'
    },

    'Free Float': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa'
    },

    'ROE': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Rentabilidade'
    },
    'ROA': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Rentabilidade'
    },
    'ROIC': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Rentabilidade'
    },

    'Giro ativos': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Rentabilidade'
    }
}


MetricasFund = {
    'ROE': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência'
    },

    'ROIC': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência'
    },
    'Giro ativos': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência'
    },

    'Margem bruta': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência'
    },

    'Margem EBIT': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência'
    },
    'Margem líquida': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência'
    },

'Dívida líquida/Patrim': {
    'baixo': {'min': 0, 'max': 3},
    'regular': {'min': 3, 'max': 6},
    'bom': {'min': 6, 'max': 10},
    'otimo': {'min': 10, 'max': float('inf')},
    'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
    'agrupador': 'Endividamento'
    },
'Dívida líquida/EBITDA': {
    'baixo': {'min': 0, 'max': 3},
    'regular': {'min': 3, 'max': 6},
    'bom': {'min': 6, 'max': 10},
    'otimo': {'min': 10, 'max': float('inf')},
    'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
    'agrupador': 'Endividamento'
    },
'PL/Ativos': {
    'baixo': {'min': 0, 'max': 3},
    'regular': {'min': 3, 'max': 6},
    'bom': {'min': 6, 'max': 10},
    'otimo': {'min': 10, 'max': float('inf')},
    'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
    'agrupador': 'Endividamento'
    },
'Liquidez corrente': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Endividamento'
    },
'Dividend Yield': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Valuation'
    },
'P/L': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Valuation'
    },
'P/VP': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Valuation'
    },
'EV/EBITDA': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Valuation'
    },
'EV/EBIT': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Valuation'
    },

'P/EBIT': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Valuation'
    },

'VPA': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Valuation'
    },
'Preço/Ativos': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Valuation'
    },
'LPA': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Valuation'
    },
'PSR': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Valuation'
    },

'Preço/Capital de giro': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Valuation'
    },

'Preço/Ativ circ liq': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Valuation'
    },
'Cotação': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa'
    },
'Volume negociado por dia': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa'
    },

'Patrimônio líquido': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa'
    },
'Ativo': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa'
    },

'Ativo circulante': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa'
    },

'Dívida bruta': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa'
    },

'Disponibilidades': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa'
    },

'Dívida líquida': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa'
    },

'Valor de mercado': {
        'baixo': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa'
    }
}

wbsaida = openpyxl.Workbook()
Dicrentabilidade = {}
Dicstock_identification = {}
Dicfinancial_summary = {}
Dicprice_information = {}
Dicdetailed_information = {}
Dicoscillations = {}
Dicprofitability_indicators = {}
Dicindebtedness_indicators = {}
Dicbalance_sheet = {}
Dicincome_statement = {}
Dicvaluation_indicators = {}


def categorizar_valor(metrica, valor):
    try:
        if metrica not in MetricasStatus:
            return 'Métrica não reconhecida'

        for categoria, limites in MetricasStatus[metrica].items():
            if categoria in ['descricao', 'agrupador']:
                continue
            if limites['min'] <= valor < limites['max']:
                return categoria
        return 'Valor fora do alcance definido'
    except Exception as e:
        print(f"Erro inesperado: {e}")
        print("categorizar_valor - Erro")
    finally:
        print("categorizar_valor - OK")




def criaPlanilhaIndRentabilidade(IndiRentabilidade):
    wbsaida.create_sheet('IndiRentabilidade')
    IndiRentabilidade = wbsaida['IndiRentabilidade']
    IndiRentabilidade.append(
        ['Agrupador', 'Fonte', 'ATIVO', 'Indicador', 'valor', 'referencia', 'Baixo', 'regular', 'bom', 'otimo'])


def montadicionario(stock_identification, financial_summary, price_information, detailed_information, oscillations,
                    valuation_indicators, profitability_indicators, indebtedness_indicators, balance_sheet,
                    income_statement):
    try:

        for information in stock_identification:
            Dicstock_identification[stock_identification[information].title] = stock_identification[information].value
        for information in financial_summary:
            Dicfinancial_summary[financial_summary[information].title] = financial_summary[information].value
        for information in price_information:
            Dicprice_information[price_information[information].title] = price_information[information].value
        for information in detailed_information:
            if information != 'variation_52_weeks':
                Dicdetailed_information[detailed_information[information].title] = detailed_information[
                    information].value
            else:
                for sub_information in detailed_information[information]:
                    Dicdetailed_information[detailed_information[information][sub_information].title] = \
                        detailed_information[information][sub_information].value
        for information in oscillations:
            Dicoscillations[oscillations[information].title] = oscillations[information].value
        for information in valuation_indicators:
            Dicvaluation_indicators[valuation_indicators[information].title] = valuation_indicators[information].value
        for information in profitability_indicators:
            Dicprofitability_indicators[profitability_indicators[information].title] = profitability_indicators[
                information].value
        for information in indebtedness_indicators:
            Dicindebtedness_indicators[indebtedness_indicators[information].title] = indebtedness_indicators[
                information].value
        for information in balance_sheet:
            Dicbalance_sheet[balance_sheet[information].title] = balance_sheet[information].value
        # for information in income_statement['three_months']:
        #   Dicincome_statement[income_statement[information].title] = income_statement[information].value
        # for information in income_statement['twelve_months']:
        #   Dicincome_statement[income_statement[information].title] = income_statement[information].value

    except Exception as e:
        print(f"Erro inesperado: {e}")
        print("montadicionario - Erro")
    finally:
        print("montadicionario - OK")
    return


def teste():
    fontes = ['StatusInvest', 'Fundamentus']

    try:
        for fonte in fontes:
            print(fonte)
    except Exception as e:
        print(f"Erro inesperado: {e}")
    finally:
        print("teste")
    return


def tratamento(indicador):
    indicador2 = indicador

    if indicador2 == "-":
        indicador2 = ""
    if is_null_zero_or_spaces(indicador2):
        indicador2 = 0
    else:
        indicador2 = float(indicador2.strip('%')) / 100
    return indicador2


def gravaIndiEficiênciaoStaus(wsIndiRentabilidade, dict_stocks, stock):
    # fontes ['StatusInvest', 'Fundamentus']



    global linha2
    #linha2 = 1
    try:
        for metrica, detalhes in MetricasStatus.items():
    #        print(f'Métrica: {metrica}')
            linha2 += 1
            indicadortratado = tratamento(dict_stocks[stock].get(metrica))
            valor_pl = indicadortratado
            categoria_pl = categorizar_valor(metrica,
                                             valor_pl)  # Certifique-se de que 'ROE' é o valor correto para a métrica
   #         print(f'O índice P/L {valor_pl} é categorizado como: {categoria_pl}')
   #         print(f"  Agrupador: {detalhes['agrupador']}")


            wsIndiRentabilidade.cell(row=linha2, column=1, value=detalhes['agrupador'])
            wsIndiRentabilidade.cell(row=linha2, column=2, value='StausInvest')
            wsIndiRentabilidade.cell(row=linha2, column=3, value=stock)
            wsIndiRentabilidade.cell(row=linha2, column=4, value=metrica)
            wsIndiRentabilidade.cell(row=linha2, column=5, value=valor_pl).number_format = numbers.FORMAT_PERCENTAGE_00
            wsIndiRentabilidade.cell(row=linha2, column=6, value=categoria_pl)
            wsIndiRentabilidade.cell(row=linha2, column=7,
                                     value=f"Mínimo = {detalhes['baixo']['min']}, Máximo = {detalhes['baixo']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=8,
                                     value=f"Mínimo = {detalhes['regular']['min']}, Máximo = {detalhes['regular']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=9,
                                     value=f"Mínimo = {detalhes['bom']['min']}, Máximo = {detalhes['bom']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=10,
                                     value=f"Mínimo = {detalhes['otimo']['min']}, Máximo = {detalhes['otimo']['max']}")
    except Exception as e:
        print(f"Erro inesperado: {e}")
        print(metrica)
        print(indicadortratado)
        print('gravaIndiEficiênciaoStaus - erro')
    finally:
        print('gravaIndiEficiênciaoStaus  OK''')



def gravaIndiRentabilidadeFund(wsIndiRentabilidade, Dicprofitability_indicators,Dicindebtedness_indicators,
                               Dicvaluation_indicators,
                               Dicprice_information,Dicdetailed_information,Dicbalance_sheet,Dicfinancial_summary,stock):
    # Condicional corrigida


    global linha2

    try:


        for metrica, detalhes in MetricasFund.items():
           # print(f'Métrica: {metrica}')
            linha2 += 1
            if detalhes['agrupador'] == 'Eficiência':
                indicadortratado = tratamento(f"{float(Dicprofitability_indicators.get(metrica)) * 100}%")
            elif detalhes['agrupador'] == 'Endividamento':
                indicadortratado = tratamento(f"{float(Dicindebtedness_indicators.get(metrica)) * 100}%")
            elif detalhes['agrupador'] == 'Valuation' and metrica == 'VPA':
                indicadortratado = tratamento(f"{float(Dicdetailed_information.get(metrica)) * 100}%")
            elif detalhes['agrupador'] == 'Valuation' and metrica == 'LPA':
                indicadortratado = tratamento(f"{float(Dicdetailed_information.get(metrica)) * 100}%")
            elif detalhes['agrupador'] == 'Valuation':
                indicadortratado = tratamento(f"{float(Dicvaluation_indicators.get(metrica)) * 100}%")
            elif detalhes['agrupador'] == 'Empresa' and  metrica == 'Cotação' :
                indicadortratado = tratamento(f"{float(Dicprice_information.get(metrica)) * 100}%")
            elif detalhes['agrupador'] == 'Empresa' and metrica == 'Volume negociado por dia':
                indicadortratado = tratamento(f"{float(Dicdetailed_information.get(metrica)) * 100}%")
            elif detalhes['agrupador'] == 'Empresa' and metrica == 'Valor de mercado':
                indicadortratado = tratamento(f"{float(Dicfinancial_summary.get(metrica)) * 100}%")
            elif detalhes['agrupador'] == 'Empresa':
                indicadortratado = tratamento(f"{float(Dicbalance_sheet.get(metrica)) * 100}%")

            valor_pl = indicadortratado
            categoria_pl = categorizar_valor(metrica,
                                             valor_pl)  # Certifique-se de que 'ROE' é o valor correto para a métrica
        #    print(f'O índice P/L {valor_pl} é categorizado como: {categoria_pl}')

            # Certifique-se de que a chave 'Indicador' realmente existe no dicionário



            wsIndiRentabilidade.cell(row=linha2, column=1, value=detalhes['agrupador'])
            wsIndiRentabilidade.cell(row=linha2, column=2, value='Fundamentus')
            wsIndiRentabilidade.cell(row=linha2, column=3, value=stock)
            wsIndiRentabilidade.cell(row=linha2, column=4, value=metrica)
            wsIndiRentabilidade.cell(row=linha2, column=5, value=f"{valor_pl * 100}%").number_format = numbers.FORMAT_PERCENTAGE_00
            wsIndiRentabilidade.cell(row=linha2, column=6, value=categoria_pl)
            wsIndiRentabilidade.cell(row=linha2, column=7,
                                     value=f"Mínimo = {detalhes['baixo']['min']}, Máximo = {detalhes['baixo']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=8,
                                     value=f"Mínimo = {detalhes['regular']['min']}, Máximo = {detalhes['regular']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=9,
                                     value=f"Mínimo = {detalhes['bom']['min']}, Máximo = {detalhes['bom']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=10,
                                     value=f"Mínimo = {detalhes['otimo']['min']}, Máximo = {detalhes['otimo']['max']}")


    except Exception as e:
        print(f"Erro inesperado: {e}")
        print(metrica)
        print(indicadortratado)
        print('gravaIndiRentabilidadeFund - erro')
    finally:
        print('gravaIndiRentabilidadeFund - OK')


# Incio funcionalidades statusinvest

# define selenium webdriver options
options = webdriver.ChromeOptions()

# create selenium webdriver instancee
driver = webdriver.Chrome(options=options)


def get_stock_soup(stock):
    ''' Get raw html from a stock '''

    # access the stock url
    driver.get(f'https://statusinvest.com.br/acoes/{stock}')
    #driver.get(f'https://statusinvest.com.br/acoes/bbse3')
    # driver.get(f'https://statusinvest.com.br/acoes/eua/{stock}')

    # get html from stock
    html = driver.find_element(By.ID, 'main-2').get_attribute('innerHTML')

    # remove accents from html and transform html into soup
    soup = BeautifulSoup(unidecode(html), 'html.parser')

    return soup


def is_null_zero_or_spaces(variable):
    # Verifica se a variável é None
    if variable is None:
        return True
    # Verifica se a variável é zero (0)
    elif variable == 0:
        return True
    # Verifica se a variável é uma string e contém apenas espaços
    elif isinstance(variable, str) and variable.strip() == '':
        return True
    elif variable == '-%':
        return True
    else:
        return False


def soup_to_dict(soup):
    '''Get all data from stock soup and return as a dictionary '''
    keys, values = [], []

    # get divs from stock
    soup1 = soup.find('div', class_='pb-3 pb-md-5')
    soup2 = soup.find('div', class_='card rounded text-main-green-dark')
    soup3 = soup.find('div', class_='indicator-today-container')
    soup4 = soup.find(
        'div', class_='top-info info-3 sm d-flex justify-between mb-3')
    soups = [soup1, soup2, soup3, soup4]

    for s in soups:
        # get only titles from a div and append to keys
        titles = s.find_all('h3', re.compile('title m-0[^"]*'))

        titles = [t.get_text() for t in titles]
        keys += titles

        # get only numbers from a div and append to values
        numbers = s.find_all('strong', re.compile('value[^"]*'))
        numbers = [n.get_text() for n in numbers]
        values += numbers

    # remove unused key and insert needed keys
    keys.remove('PART. IBOV')
    keys.insert(6, 'TAG ALONG')
    keys.insert(7, 'LIQUIDEZ MEDIA DIARIA')

    # clean keys list
    keys = [k.replace('\nhelp_outline', '').strip() for k in keys]
    keys = [k for k in keys if k != '']

    # clean values list
    values = [v.replace('\nhelp_outline', '').strip() for v in values]
    values = [v.replace('.', '').replace(',', '.') for v in values]

    # create a dictionary using keys and values from indicators
    d = {k: v for k, v in zip(keys, values)}

    return d


# Fim funcionalidades statusinvest

# pylint: disable=line-too-long
if __name__ == '__main__':
    dict_stocks = {}  # dicionario statusinvest
    Dicrentabilidade = {}  # fundamentus
    linhafundamentus = 1  # silvio
    linhastatus = 0


    criaPlanilhaIndRentabilidade(wbsaida)
    wsIndiRentabilidade = wbsaida['IndiRentabilidade']

    start = time.time()
    with open('../stocks.txt', 'r') as f:
        stocks = f.read().splitlines()
        for stock in stocks:
            linhastatus = linhafundamentus
            linhastatus = linhastatus + 1  # silvio
            linhafundamentus = linhastatus + 1  # silvio

            main_pipeline = fundamentus.Pipeline(stock.upper())
            response = main_pipeline.get_all_information()

            # Incio statusinvest
            # get data and transform into dictionary
            soup = get_stock_soup(stock)
            dict_stock = soup_to_dict(soup)
            dict_stocks[stock] = dict_stock
            # Fim Statausinvest

            # Extract the information from the response.
            stock_identification = response.transformed_information['stock_identification']
            financial_summary = response.transformed_information['financial_summary']
            price_information = response.transformed_information['price_information']
            detailed_information = response.transformed_information[
                'detailed_information']
            oscillations = response.transformed_information['oscillations']
            valuation_indicators = response.transformed_information[
                'valuation_indicators']
            profitability_indicators = response.transformed_information[
                'profitability_indicators']
            indebtedness_indicators = response.transformed_information[
                'indebtedness_indicators']
            balance_sheet = response.transformed_information['balance_sheet']
            income_statement = response.transformed_information['income_statement']

            montadicionario(stock_identification, financial_summary, price_information, detailed_information,
                            oscillations,
                            valuation_indicators, profitability_indicators, indebtedness_indicators, balance_sheet,
                            income_statement)
            # print(Dicrentabilidade)

            gravaIndiEficiênciaoStaus(wsIndiRentabilidade, dict_stocks, stock)
            gravaIndiRentabilidadeFund(wsIndiRentabilidade, Dicprofitability_indicators,Dicindebtedness_indicators,
                               Dicvaluation_indicators,
                               Dicprice_information,Dicdetailed_information,Dicbalance_sheet,Dicfinancial_summary,stock)

          #  print(dict_stocks)
          #  print(Dicprice_information)
          #  print(Dicdetailed_information)
          #  print(Dicbalance_sheet)
          #  print(Dicfinancial_summary)

          #  print(MetricasStatus)
    # exit the driver
    driver.quit()

    # end timer
    end = time.time()
    wbsaida.save("StatusInvest2.xlsx")  # silvio