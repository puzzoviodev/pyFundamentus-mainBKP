
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

# Dados da tabela (últimos 9 indicadores, 45 linhas)
data = [

    # Dívida Líquida/EBIT
    {
        "Indicador": "Dívida Líquida/EBIT",
        "Categoria": "Péssimo",
        "Faixa de Referência": "Acima de 4x",
        "Interpretação": "Endividamento crítico, alto risco de insolvência: Um índice Dívida Líquida/EBIT acima de 4x indica que a empresa levaria mais de 4 anos de lucro operacional antes de juros e impostos (EBIT) para quitar sua dívida líquida (dívida total menos caixa e equivalentes). Esse nível de alavancagem é extremamente arriscado, pois reflete dependência excessiva de financiamento externo e vulnerabilidade a choques econômicos, como aumento de taxas de juros ou recessões. Setores comuns: Empresas em crise financeira (ex.: varejo em declínio, companhias aéreas em períodos de baixa demanda) ou setores intensivos em capital com investimentos mal planejados (ex.: construção civil em recessão). Implicações para investidores: Evitar investimento, salvo em casos de reestruturação robusta com sinais claros de recuperação, como redução de custos ou venda de ativos. Fatores a analisar: Cobertura de juros (EBIT/juros), fluxo de caixa livre, volatilidade do EBIT e tendências setoriais. Cenário prático: Uma empresa com Dívida Líquida/EBIT de 5x pode enfrentar dificuldades em refinanciar dívidas em um ambiente de crédito restritivo, aumentando o risco de default.",
        "Agrupador": "Endividamento"
    },
    {
        "Indicador": "Dívida Líquida/EBIT",
        "Categoria": "Ruim",
        "Faixa de Referência": "3 a 4x",
        "Interpretação": "Endividamento elevado, risco financeiro significativo: A empresa levaria de 3 a 4 anos de EBIT para pagar sua dívida líquida, indicando alta alavancagem. Esse nível é tolerável em setores com fluxos de caixa estáveis, mas arriscado em setores cíclicos. Setores comuns: Infraestrutura, siderurgia ou empresas em fase de expansão com dívidas elevadas. Implicações para investidores: Cautela é necessária; o investimento pode ser considerado se houver evidências de EBIT estável ou plano de desalavancagem. Fatores a analisar: Consistência do EBIT, capacidade de geração de caixa, cronograma de vencimento das dívidas e exposição a riscos macroeconômicos (ex.: aumento de juros). Cenário prático: Uma empresa com Dívida Líquida/EBIT de 3,5x em um setor cíclico (ex.: mineração) pode enfrentar dificuldades em períodos de baixa nos preços das commodities, reduzindo o EBIT e comprometendo a capacidade de pagamento. Investidores devem exigir um plano claro de redução de dívidas ou hedge contra volatilidade.",
        "Agrupador": "Endividamento"
    },
    {
        "Indicador": "Dívida Líquida/EBIT",
        "Categoria": "Moderado",
        "Faixa de Referência": "1,5 a 3x",
        "Interpretação": "Endividamento moderado, gestão financeira razoável: A dívida líquida é pagável em 1,5 a 3 anos de EBIT, um nível aceitável para empresas em setores estáveis com fluxos de caixa previsíveis. Setores comuns: Utilities (energia, saneamento), bens de consumo defensivo ou indústrias com margens consistentes. Implicações para investidores: Um investimento relativamente seguro, desde que o EBIT seja estável e o fluxo de caixa livre cubra os pagamentos de juros e principal. Fatores a analisar: Histórico de EBIT, qualidade do fluxo de caixa livre, nível de endividamento em relação a peers setoriais e riscos operacionais. Cenário prático: Uma empresa de energia com Dívida Líquida/EBIT de 2x pode ser vista como estável, especialmente se opera em um setor regulado com receitas garantidas. Contudo, choques regulatórios ou aumento de custos operacionais podem pressionar o EBIT, exigindo monitoramento contínuo.",
        "Agrupador": "Endividamento"
    },
    {
        "Indicador": "Dívida Líquida/EBIT",
        "Categoria": "Bom",
        "Faixa de Referência": "0 a 1,5x",
        "Interpretação": "Endividamento controlado, saúde financeira sólida: A empresa pode quitar sua dívida líquida em menos de 1,5 anos de EBIT, indicando baixa alavancagem e boa capacidade de pagamento. Setores comuns: Empresas maduras com forte geração de caixa, como bens de consumo defensivo (ex.: alimentos, bebidas) ou farmacêuticas. Implicações para investidores: Sinal de estabilidade financeira, atrativo para investidores conservadores. No entanto, deve-se avaliar se a baixa alavancagem reflete subinvestimento ou falta de oportunidades de crescimento. Fatores a analisar: Fluxo de caixa livre, política de investimentos, margem EBIT e potencial de crescimento futuro. Cenário prático: Uma empresa de bens de consumo com Dívida Líquida/EBIT de 1x demonstra capacidade de gerenciar dívidas sem comprometer operações, mas investidores devem verificar se a empresa está reinvestindo lucros de forma eficiente para sustentar crescimento.",
        "Agrupador": "Endividamento"
    },
    {
        "Indicador": "Dívida Líquida/EBIT",
        "Categoria": "Ótimo",
        "Faixa de Referência": "Negativa",
        "Interpretação": "Caixa líquido supera dívida, situação financeira excepcional: A empresa possui mais caixa e equivalentes do que dívidas, resultando em dívida líquida negativa. Isso reflete uma saúde financeira robusta, com flexibilidade para investimentos, dividendos ou aquisições. Setores comuns: Tecnologia (ex.: Apple, Microsoft) ou empresas com reservas elevadas devido a lucros consistentes. Implicações para investidores: Altamente atrativo, pois indica baixo risco financeiro. Porém, investidores devem investigar se o excesso de caixa está sendo alocado eficientemente (ex.: reinvestimento, retorno aos acionistas) ou se reflete ineficiência operacional. Fatores a analisar: Política de alocação de capital, retorno sobre o caixa (ex.: dividendos, recompra de ações) e riscos de caixa ocioso. Cenário prático: Uma empresa de tecnologia com dívida líquida negativa pode usar o caixa para financiar inovação ou aquisições, mas um excesso prolongado de caixa sem uso estratégico pode sinalizar falta de visão de crescimento.",
        "Agrupador": "Endividamento"
    },
    # Dívida Líquida/EBITDA
    {
        "Indicador": "Dívida Líquida/EBITDA",
        "Categoria": "Péssimo",
        "Faixa de Referência": "Acima de 3,5x",
        "Interpretação": "Endividamento crítico, alto risco: Dívida líquida acima de 3,5x o EBITDA indica alavancagem extrema. Setores: empresas em crise (ex.: varejo, aéreas). Evitar investimento, salvo recuperação clara. Analisar: cobertura de juros, fluxo de caixa, volatilidade do EBITDA.",
        "Agrupador": "Endividamento"
    },
    {
        "Indicador": "Dívida Líquida/EBITDA",
        "Categoria": "Ruim",
        "Faixa de Referência": "2,5 a 3,5x",
        "Interpretação": "Endividamento elevado: Pagável em 2,5-3,5 anos de EBITDA, arriscado em setores cíclicos. Setores: mineração, siderurgia. Cautela, verificar plano de desalavancagem. Analisar: estabilidade do EBITDA, fluxo de caixa.",
        "Agrupador": "Endividamento"
    },
    {
        "Indicador": "Dívida Líquida/EBITDA",
        "Categoria": "Moderado",
        "Faixa de Referência": "1 a 2,5x",
        "Interpretação": "Endividamento moderado: Pagável em 1-2,5 anos de EBITDA, aceitável em setores estáveis. Setores: utilities, varejo. Seguro com EBITDA estável. Analisar: fluxo de caixa, riscos setoriais.",
        "Agrupador": "Endividamento"
    },
    {
        "Indicador": "Dívida Líquida/EBITDA",
        "Categoria": "Bom",
        "Faixa de Referência": "0 a 1x",
        "Interpretação": "Endividamento baixo: Dívida quitada em menos de 1 ano de EBITDA, forte capacidade financeira. Setores: bens de consumo, farmacêuticas. Atraente para conservadores. Analisar: qualidade do EBITDA, crescimento.",
        "Agrupador": "Endividamento"
    },
    {
        "Indicador": "Dívida Líquida/EBITDA",
        "Categoria": "Ótimo",
        "Faixa de Referência": "Negativa",
        "Interpretação": "Caixa líquido excede dívida: Saúde financeira robusta, flexibilidade para investimentos. Setores: tecnologia. Atraente, verificar alocação de caixa. Analisar: retorno sobre o capital, riscos de caixa ocioso.",
        "Agrupador": "Endividamento"
    },
    # Dívida Líquida/Patrimônio Líquido
    {
        "Indicador": "Dívida Líquida/Patrimônio Líquido",
        "Categoria": "Péssimo",
        "Faixa de Referência": "Acima de 1",
        "Interpretação": "Alavancagem crítica: Dívida líquida excede o PL, alto risco de insolvência. Setores: varejo em crise, aéreas. Evitar, salvo recuperação clara. Analisar: solvência, fluxo de caixa.",
        "Agrupador": "Endividamento"
    },
    {
        "Indicador": "Dívida Líquida/Patrimônio Líquido",
        "Categoria": "Ruim",
        "Faixa de Referência": "0,7 a 1",
        "Interpretação": "Alavancagem elevada: Dívida líquida é 70-100% do PL, risco significativo. Setores: construção, infraestrutura. Cautela, verificar plano de desalavancagem. Analisar: ROE, fluxo de caixa.",
        "Agrupador": "Endividamento"
    },
    {
        "Indicador": "Dívida Líquida/Patrimônio Líquido",
        "Categoria": "Moderado",
        "Faixa de Referência": "0,3 a 0,7",
        "Interpretação": "Alavancagem moderada: Dívida líquida é 30-70% do PL, aceitável. Setores: indústria, utilities. Seguro com boa gestão. Analisar: cobertura de juros, ROE.",
        "Agrupador": "Endividamento"
    },
    {
        "Indicador": "Dívida Líquida/Patrimônio Líquido",
        "Categoria": "Bom",
        "Faixa de Referência": "0 a 0,3",
        "Interpretação": "Baixa alavancagem: Dívida líquida mínima, estrutura sólida. Setores: bens de consumo, utilities. Atraente para conservadores. Analisar: ROE, crescimento.",
        "Agrupador": "Endividamento"
    },
    {
        "Indicador": "Dívida Líquida/Patrimônio Líquido",
        "Categoria": "Ótimo",
        "Faixa de Referência": "Negativa",
        "Interpretação": "Caixa líquido excede dívida, situação financeira robusta: Mais caixa que dívidas, flexibilidade financeira. Setores: tecnologia. Atraente, verificar uso do caixa. Analisar: alocação de capital, ROE.",
        "Agrupador": "Endividamento"
    },
    # Patrimônio/Ativos
    {
        "Indicador": "Patrimônio/Ativos",
        "Categoria": "Péssimo",
        "Faixa de Referência": "Menor que 0,2",
        "Interpretação": "Estrutura frágil: Patrimônio <20% dos ativos, alta alavancagem. Setores: crise financeira, varejo. Evitar, salvo recuperação. Analisar: solvência, fluxo de caixa.",
        "Agrupador": "Endividamento"
    },
    {
        "Indicador": "Patrimônio/Ativos",
        "Categoria": "Ruim",
        "Faixa de Referência": "0,2 a 0,3",
        "Interpretação": "Estrutura limitada: Patrimônio 20-30% dos ativos, risco moderado. Setores: construção. Cautela, verificar solvência. Analisar: fluxo de caixa, ROE.",
        "Agrupador": "Endividamento"
    },
    {
        "Indicador": "Patrimônio/Ativos",
        "Categoria": "Moderado",
        "Faixa de Referência": "0,3 a 0,5",
        "Interpretação": "Estrutura equilibrada: Patrimônio 30-50% dos ativos, aceitável. Setores: indústria, utilities. Seguro com boa gestão. Analisar: ROE, riscos setoriais.",
        "Agrupador": "Endividamento"
    },
    {
        "Indicador": "Patrimônio/Ativos",
        "Categoria": "Bom",
        "Faixa de Referência": "0,5 a 0,7",
        "Interpretação": "Estrutura forte: Patrimônio 50-70% dos ativos, baixa alavancagem. Setores: bens de consumo. Atraente para conservadores. Analisar: ROE, crescimento.",
        "Agrupador": "Endividamento"
    },
    {
        "Indicador": "Patrimônio/Ativos",
        "Categoria": "Ótimo",
        "Faixa de Referência": "Acima de 0,7",
        "Interpretação": "Estrutura robusta: Patrimônio >70% dos ativos, mínima alavancagem. Setores: tecnologia. Atraente, verificar alocação de capital. Analisar: ROE, crescimento.",
        "Agrupador": "Endividamento"
    },
    # PL/Ativos
    {
        "Indicador": "PL/Ativos",
        "Categoria": "Péssimo",
        "Faixa de Referência": "Menor que 0,2",
        "Interpretação": "Patrimônio muito baixo: PL <20% dos ativos, alto risco. Setores: crise financeira. Evitar, salvo recuperação. Analisar: solvência, fluxo de caixa.",
        "Agrupador": "Endividamento"
    },
    {
        "Indicador": "PL/Ativos",
        "Categoria": "Ruim",
        "Faixa de Referência": "0,2 a 0,3",
        "Interpretação": "Patrimônio limitado: PL 20-30% dos ativos, risco moderado. Setores: construção. Cautela, verificar solvência. Analisar: ROE, fluxo de caixa.",
        "Agrupador": "Endividamento"
    },
    {
        "Indicador": "PL/Ativos",
        "Categoria": "Moderado",
        "Faixa de Referência": "0,3 a 0,5",
        "Interpretação": "Patrimônio equilibrado: PL 30-50% dos ativos, aceitável. Setores: indústria. Seguro com boa gestão. Analisar: ROE, riscos setoriais.",
        "Agrupador": "Endividamento"
    },
    {
        "Indicador": "PL/Ativos",
        "Categoria": "Bom",
        "Faixa de Referência": "0,5 a 0,7",
        "Interpretação": "Patrimônio forte: PL 50-70% dos ativos, baixa alavancagem. Setores: bens de consumo. Atraente para conservadores. Analisar: ROE, crescimento.",
        "Agrupador": "Endividamento"
    },
    {
        "Indicador": "PL/Ativos",
        "Categoria": "Ótimo",
        "Faixa de Referência": "Acima de 0,7",
        "Interpretação": "Patrimônio robusto: PL >70% dos ativos, mínima alavancagem. Setores: tecnologia. Atraente, verificar alocação de capital. Analisar: ROE, crescimento.",
        "Agrupador": "Endividamento"
    },
    # Dividend Yield (DY)
    {
        "Indicador": "Dividend Yield (DY)",
        "Categoria": "Péssimo",
        "Faixa de Referência": "0%",
        "Interpretação": "Sem dividendos: Foco em reinvestimento ou prejuízo. Setores: startups, tecnologia. Evitar para renda, considerar para crescimento. Analisar: fluxo de caixa, lucros.",
        "Agrupador": "Renda"
    },
    {
        "Indicador": "Dividend Yield (DY)",
        "Categoria": "Ruim",
        "Faixa de Referência": "0,1 a 2%",
        "Interpretação": "Retorno baixo: Dividendos mínimos, foco em reinvestimento. Setores: tecnologia, saúde. Pouco atrativo para renda. Analisar: payout ratio, sustentabilidade.",
        "Agrupador": "Renda"
    },
    {
        "Indicador": "Dividend Yield (DY)",
        "Categoria": "Moderado",
        "Faixa de Referência": "2 a 4%",
        "Interpretação": "Retorno moderado: Equilíbrio entre dividendos e reinvestimento. Setores: utilities, bens de consumo. Atraente para renda moderada. Analisar: payout ratio, fluxo de caixa.",
        "Agrupador": "Renda"
    },
    {
        "Indicador": "Dividend Yield (DY)",
        "Categoria": "Bom",
        "Faixa de Referência": "4 a 6%",
        "Interpretação": "Retorno atrativo: Boa política de dividendos, lucros sólidos. Setores: bancos, utilities. Ideal para renda. Analisar: sustentabilidade, fluxo de caixa.",
        "Agrupador": "Renda"
    },
    {
        "Indicador": "Dividend Yield (DY)",
        "Categoria": "Ótimo",
        "Faixa de Referência": "Acima de 6%",
        "Interpretação": "Alto retorno: Muito atrativo, mas verificar sustentabilidade. Setores: utilities. Exige análise rigorosa para evitar armadilhas. Analisar: payout ratio, lucros.",
        "Agrupador": "Renda"
    },
    # EV/EBIT
    {
        "Indicador": "EV/EBIT",
        "Categoria": "Péssimo",
        "Faixa de Referência": "Acima de 20x",
        "Interpretação": "Valuation elevado: Risco de supervalorização. Setores: tecnologia em bolha. Evitar, salvo crescimento excepcional. Analisar: EBIT, fluxo de caixa.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "EV/EBIT",
        "Categoria": "Ruim",
        "Faixa de Referência": "15 a 20x",
        "Interpretação": "Valuation alto: Precificação acima da média. Setores: tecnologia, saúde. Cautela, verificar crescimento. Analisar: EBIT, concorrência.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "EV/EBIT",
        "Categoria": "Moderado",
        "Faixa de Referência": "10 a 15x",
        "Interpretação": "Valuation razoável: Precificação justa. Setores: indústria, bens de consumo. Atraente com EBIT estável. Analisar: fluxo de caixa, riscos setoriais.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "EV/EBIT",
        "Categoria": "Bom",
        "Faixa de Referência": "5 a 10x",
        "Interpretação": "Valuation atrativo: Subvalorização ou EBIT sólido. Setores: utilities. Ideal para valor. Analisar: consistência do EBIT, crescimento.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "EV/EBIT",
        "Categoria": "Ótimo",
        "Faixa de Referência": "Menor que 5x",
        "Interpretação": "Valuation extremamente atrativo: Subvalorização significativa. Setores: cíclicos, recuperação. Oportunidade, verificar riscos. Analisar: recuperação do EBIT.",
        "Agrupador": "Valuation"
    },
    # EV/EBITDA
    {
        "Indicador": "EV/EBITDA",
        "Categoria": "Péssimo",
        "Faixa de Referência": "Acima de 15x",
        "Interpretação": "Valuation elevado: Risco de supervalorização. Setores: tecnologia. Evitar, salvo crescimento excepcional. Analisar: EBITDA, fluxo de caixa.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "EV/EBITDA",
        "Categoria": "Ruim",
        "Faixa de Referência": "10 a 15x",
        "Interpretação": "Valuation alto: Precificação acima da média. Setores: tecnologia, saúde. Cautela, verificar crescimento. Analisar: EBITDA, concorrência.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "EV/EBITDA",
        "Categoria": "Moderado",
        "Faixa de Referência": "7 a 10x",
        "Interpretação": "Valuation razoável: Precificação justa. Setores: indústria. Atraente com EBITDA estável. Analisar: fluxo de caixa, riscos setoriais.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "EV/EBITDA",
        "Categoria": "Bom",
        "Faixa de Referência": "4 a 7x",
        "Interpretação": "Valuation atrativo: Subvalorização ou EBITDA sólido. Setores: utilities. Ideal para valor. Analisar: consistência do EBITDA, crescimento.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "EV/EBITDA",
        "Categoria": "Ótimo",
        "Faixa de Referência": "Menor que 4x",
        "Interpretação": "Valuation extremamente atrativo: Subvalorização significativa. Setores: cíclicos, recuperação. Oportunidade, verificar riscos. Analisar: recuperação do EBITDA.",
        "Agrupador": "Valuation"
    },
    # P/ATIVO
    {
        "Indicador": "P/ATIVO",
        "Categoria": "Péssimo",
        "Faixa de Referência": "Acima de 2x",
        "Interpretação": "Valuation elevado: Preço alto em relação aos ativos. Setores: tecnologia. Evitar, salvo crescimento excepcional. Analisar: ROA, qualidade dos ativos.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "P/ATIVO",
        "Categoria": "Ruim",
        "Faixa de Referência": "1,5 a 2x",
        "Interpretação": "Valuation alto: Precificação acima da média. Setores: tecnologia, saúde. Cautela, verificar crescimento. Analisar: ROA, composição dos ativos.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "P/ATIVO",
        "Categoria": "Moderado",
        "Faixa de Referência": "1 a 1,5x",
        "Interpretação": "Valuation razoável: Precificação justa. Setores: indústria. Atraente com boa gestão. Analisar: ROA, tendências setoriais.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "P/ATIVO",
        "Categoria": "Bom",
        "Faixa de Referência": "0,5 a 1x",
        "Interpretação": "Valuation atrativo: Subvalorização ou ativos sólidos. Setores: utilities. Ideal para valor. Analisar: qualidade dos ativos, ROA.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "P/ATIVO",
        "Categoria": "Ótimo",
        "Faixa de Referência": "Menor que 0,5x",
        "Interpretação": "Valuation extremamente atrativo: Subvalorização significativa. Setores: cíclicos, recuperação. Oportunidade, verificar riscos. Analisar: recuperação dos ativos.",
        "Agrupador": "Valuation"
    },
    # P/Ativo Circulante Líquido
    {
        "Indicador": "P/Ativo Circulante Líquido",
        "Categoria": "Péssimo",
        "Faixa de Referência": "Acima de 3x",
        "Interpretação": "Valuation elevado: Preço alto em relação aos ativos circulantes líquidos. Setores: tecnologia. Evitar, salvo liquidez excepcional. Analisar: liquidez corrente.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "P/Ativo Circulante Líquido",
        "Categoria": "Ruim",
        "Faixa de Referência": "2 a 3x",
        "Interpretação": "Valuation alto: Precificação acima da média. Setores: tecnologia, saúde. Cautela, verificar liquidez. Analisar: composição dos ativos.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "P/Ativo Circulante Líquido",
        "Categoria": "Moderado",
        "Faixa de Referência": "1 a 2x",
        "Interpretação": "Valuation razoável: Precificação justa. Setores: indústria. Atraente com boa liquidez. Analisar: liquidez corrente, tendências setoriais.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "P/Ativo Circulante Líquido",
        "Categoria": "Bom",
        "Faixa de Referência": "0,5 a 1x",
        "Interpretação": "Valuation atrativo: Subvalorização ou liquidez sólida. Setores: utilities. Ideal para valor. Analisar: qualidade dos ativos, liquidez.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "P/Ativo Circulante Líquido",
        "Categoria": "Ótimo",
        "Faixa de Referência": "Menor que 0,5x",
        "Interpretação": "Valuation extremamente atrativo: Subvalorização significativa. Setores: recuperação. Oportunidade, verificar riscos. Analisar: recuperação da liquidez.",
        "Agrupador": "Valuation"
    },
# P/Capital de Giro
    {
        "Indicador": "P/Capital de Giro",
        "Categoria": "Péssimo",
        "Faixa de Referência": "Acima de 5x",
        "Interpretação": "Valuation elevado: Preço alto em relação ao capital de giro. Setores: tecnologia. Evitar, salvo forte geração de caixa. Analisar: liquidez, gestão de capital.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "P/Capital de Giro",
        "Categoria": "Ruim",
        "Faixa de Referência": "3 a 5x",
        "Interpretação": "Valuation alto: Precificação acima da média. Setores: tecnologia, saúde. Cautela, verificar gestão de capital. Analisar: liquidez, eficiência.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "P/Capital de Giro",
        "Categoria": "Moderado",
        "Faixa de Referência": "1,5 a 3x",
        "Interpretação": "Valuation razoável: Precificação justa. Setores: indústria. Atraente com boa gestão. Analisar: liquidez, tendências setoriais.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "P/Capital de Giro",
        "Categoria": "Bom",
        "Faixa de Referência": "0,5 a 1,5x",
        "Interpretação": "Valuation atrativo: Subvalorização ou capital sólido. Setores: utilities. Ideal para valor. Analisar: gestão de capital, liquidez.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "P/Capital de Giro",
        "Categoria": "Ótimo",
        "Faixa de Referência": "Menor que 0,5x",
        "Interpretação": "Valuation extremamente atrativo: Subvalorização significativa. Setores: recuperação. Oportunidade, verificar riscos. Analisar: recuperação do capital.",
        "Agrupador": "Valuation"
    },
    # P/EBIT
    {
        "Indicador": "P/EBIT",
        "Categoria": "Péssimo",
        "Faixa de Referência": "Acima de 25x",
        "Interpretação": "Valuation elevado: Preço alto em relação ao EBIT. Setores: tecnologia. Evitar, salvo crescimento excepcional. Analisar: EBIT, fluxo de caixa.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "P/EBIT",
        "Categoria": "Ruim",
        "Faixa de Referência": "20 a 25x",
        "Interpretação": "Valuation alto: Precificação acima da média. Setores: tecnologia, saúde. Cautela, verificar crescimento. Analisar: EBIT, concorrência.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "P/EBIT",
        "Categoria": "Moderado",
        "Faixa de Referência": "15 a 20x",
        "Interpretação": "Valuation razoável: Precificação justa. Setores: indústria. Atraente com EBIT estável. Analisar: fluxo de caixa, riscos setoriais.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "P/EBIT",
        "Categoria": "Bom",
        "Faixa de Referência": "10 a 15x",
        "Interpretação": "Valuation atrativo: Subvalorização ou EBIT sólido. Setores: utilities. Ideal para valor. Analisar: consistência do EBIT, crescimento.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "P/EBIT",
        "Categoria": "Ótimo",
        "Faixa de Referência": "Menor que 10x",
        "Interpretação": "Valuation extremamente atrativo: Subvalorização significativa. Setores: cíclicos, recuperação. Oportunidade, verificar riscos. Analisar: recuperação do EBIT.",
        "Agrupador": "Valuation"
    },
    # P/EBITDA
    {
        "Indicador": "P/EBITDA",
        "Categoria": "Péssimo",
        "Faixa de Referência": "Acima de 15x",
        "Interpretação": "Valuation elevado: Preço alto em relação ao EBITDA. Setores: tecnologia. Evitar, salvo crescimento excepcional. Analisar: EBITDA, fluxo de caixa.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "P/EBITDA",
        "Categoria": "Ruim",
        "Faixa de Referência": "10 a 15x",
        "Interpretação": "Valuation alto: Precificação acima da média. Setores: tecnologia, saúde. Cautela, verificar crescimento. Analisar: EBITDA, concorrência.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "P/EBITDA",
        "Categoria": "Moderado",
        "Faixa de Referência": "7 a 10x",
        "Interpretação": "Valuation razoável: Precificação justa. Setores: indústria. Atraente com EBITDA estável. Analisar: fluxo de caixa, riscos setoriais.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "P/EBITDA",
        "Categoria": "Bom",
        "Faixa de Referência": "4 a 7x",
        "Interpretação": "Valuation atrativo: Subvalorização ou EBITDA sólido. Setores: utilities. Ideal para valor. Analisar: consistência do EBITDA, crescimento.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "P/EBITDA",
        "Categoria": "Ótimo",
        "Faixa de Referência": "Menor que 4x",
        "Interpretação": "Valuation extremamente atrativo: Subvalorização significativa. Setores: cíclicos, recuperação. Oportunidade, verificar riscos. Analisar: recuperação do EBITDA.",
        "Agrupador": "Valuation"
    },
    # P/L
    {
        "Indicador": "P/L",
        "Categoria": "Péssimo",
        "Faixa de Referência": "Acima de 30x",
        "Interpretação": "Valuation elevado: Preço alto em relação ao lucro líquido. Setores: tecnologia. Evitar, salvo crescimento excepcional. Analisar: lucro, fluxo de caixa.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "P/L",
        "Categoria": "Ruim",
        "Faixa de Referência": "20 a 30x",
        "Interpretação": "Valuation alto: Precificação acima da média. Setores: tecnologia, saúde. Cautela, verificar crescimento. Analisar: lucro, concorrência.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "P/L",
        "Categoria": "Moderado",
        "Faixa de Referência": "15 a 20x",
        "Interpretação": "Valuation razoável: Precificação justa. Setores: indústria. Atraente com lucro estável. Analisar: fluxo de caixa, riscos setoriais.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "P/L",
        "Categoria": "Bom",
        "Faixa de Referência": "10 a 15x",
        "Interpretação": "Valuation atrativo: Subvalorização ou lucro sólido. Setores: utilities. Ideal para valor. Analisar: consistência do lucro, crescimento.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "P/L",
        "Categoria": "Ótimo",
        "Faixa de Referência": "Menor que 10x",
        "Interpretação": "Valuation extremamente atrativo: Subvalorização significativa. Setores: cíclicos, recuperação. Oportunidade, verificar riscos. Analisar: recuperação do lucro.",
        "Agrupador": "Valuation"
    },
    # P/VPA
    {
        "Indicador": "P/VPA",
        "Categoria": "Péssimo",
        "Faixa de Referência": "Acima de 3x",
        "Interpretação": "Valuation elevado: Preço alto em relação ao valor patrimonial. Setores: tecnologia. Evitar, salvo crescimento excepcional. Analisar: ROE, patrimônio.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "P/VPA",
        "Categoria": "Ruim",
        "Faixa de Referência": "2 a 3x",
        "Interpretação": "Valuation alto: Precificação acima da média. Setores: tecnologia, saúde. Cautela, verificar crescimento. Analisar: ROE, composição do patrimônio.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "P/VPA",
        "Categoria": "Moderado",
        "Faixa de Referência": "1,5 a 2x",
        "Interpretação": "Valuation razoável: Precificação justa. Setores: indústria. Atraente com patrimônio estável. Analisar: ROE, tendências setoriais.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "P/VPA",
        "Categoria": "Bom",
        "Faixa de Referência": "1 a 1,5x",
        "Interpretação": "Valuation atrativo: Subvalorização ou patrimônio sólido. Setores: utilities. Ideal para valor. Analisar: consistência do patrimônio, ROE.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "P/VPA",
        "Categoria": "Ótimo",
        "Faixa de Referência": "Menor que 1x",
        "Interpretação": "Valuation extremamente atrativo: Subvalorização significativa. Setores: cíclicos, recuperação. Oportunidade, verificar riscos. Analisar: recuperação do patrimônio.",
        "Agrupador": "Valuation"
    },
    # PSR
    {
        "Indicador": "PSR",
        "Categoria": "Péssimo",
        "Faixa de Referência": "Acima de 3x",
        "Interpretação": "Valuation elevado: Preço alto em relação à receita. Setores: tecnologia. Evitar, salvo crescimento excepcional. Analisar: receita, margens.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "PSR",
        "Categoria": "Ruim",
        "Faixa de Referência": "2 a 3x",
        "Interpretação": "Valuation alto: Precificação acima da média. Setores: tecnologia, saúde. Cautela, verificar crescimento. Analisar: receita, margens.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "PSR",
        "Categoria": "Moderado",
        "Faixa de Referência": "1 a 2x",
        "Interpretação": "Valuation razoável: Precificação justa. Setores: indústria. Atraente com receita estável. Analisar: receita, margens.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "PSR",
        "Categoria": "Bom",
        "Faixa de Referência": "0,5 a 1x",
        "Interpretação": "Valuation atrativo: Subvalorização ou receita sólida. Setores: utilities. Ideal para valor. Analisar: consistência da receita, margens.",
        "Agrupador": "Valuation"
    },
    {
        "Indicador": "PSR",
        "Categoria": "Ótimo",
        "Faixa de Referência": "Menor que 0,5x",
        "Interpretação": "Valuation extremamente atrativo: Subvalorização significativa. Setores: cíclicos, recuperação. Oportunidade, verificar riscos. Analisar: recuperação da receita.",
        "Agrupador": "Valuation"
    },
    # Giro do Ativo
    {
        "Indicador": "Giro do Ativo",
        "Categoria": "Péssimo",
        "Faixa de Referência": "Menor que 0,5x",
        "Interpretação": "Baixa eficiência: Pouca receita gerada por ativo. Setores: crise, intensivos em capital. Evitar, salvo melhoria operacional. Analisar: eficiência, ativos.",
        "Agrupador": "Eficiência"
    },
    {
        "Indicador": "Giro do Ativo",
        "Categoria": "Ruim",
        "Faixa de Referência": "0,5 a 1x",
        "Interpretação": "Eficiência limitada: Utilização moderada dos ativos. Setores: indústria pesada. Cautela, verificar melhoria. Analisar: gestão de ativos, receita.",
        "Agrupador": "Eficiência"
    },
    {
        "Indicador": "Giro do Ativo",
        "Categoria": "Moderado",
        "Faixa de Referência": "1 a 1,5x",
        "Interpretação": "Eficiência razoável: Utilização adequada dos ativos. Setores: indústria, varejo. Atraente com boa gestão. Analisar: receita, eficiência.",
        "Agrupador": "Eficiência"
    },
    {
        "Indicador": "Giro do Ativo",
        "Categoria": "Bom",
        "Faixa de Referência": "1,5 a 2x",
        "Interpretação": "Alta eficiência: Boa utilização dos ativos. Setores: bens de consumo. Ideal para eficiência. Analisar: receita, gestão de ativos.",
        "Agrupador": "Eficiência"
    },
    {
        "Indicador": "Giro do Ativo",
        "Categoria": "Ótimo",
        "Faixa de Referência": "Acima de 2x",
        "Interpretação": "Eficiência excepcional: Alta utilização dos ativos. Setores: tecnologia, varejo. Atraente, sinal de gestão superior. Analisar: sustentabilidade, receita.",
        "Agrupador": "Eficiência"
    },
    # Liquidez Corrente
    {
        "Indicador": "Liquidez Corrente",
        "Categoria": "Péssimo",
        "Faixa de Referência": "Menor que 1",
        "Interpretação": "Baixa liquidez: Ativos circulantes não cobrem passivos de curto prazo. Setores: crise, varejo. Evitar, salvo recuperação. Analisar: fluxo de caixa, passivos.",
        "Agrupador": "Liquidez"
    },
    {
        "Indicador": "Liquidez Corrente",
        "Categoria": "Ruim",
        "Faixa de Referência": "1 a 1,2",
        "Interpretação": "Liquidez limitada: Ativos cobrem passivos com margem mínima. Setores: indústria, varejo. Cautela, verificar gestão de passivos. Analisar: fluxo de caixa, pagamentos.",
        "Agrupador": "Liquidez"
    },
    {
        "Indicador": "Liquidez Corrente",
        "Categoria": "Moderado",
        "Faixa de Referência": "1,2 a 1,5",
        "Interpretação": "Liquidez razoável: Ativos cobrem passivos com folga moderada. Setores: indústria, utilities. Seguro com boa gestão. Analisar: qualidade dos ativos, riscos setoriais.",
        "Agrupador": "Liquidez"
    },
    {
        "Indicador": "Liquidez Corrente",
        "Categoria": "Bom",
        "Faixa de Referência": "1,5 a 2",
        "Interpretação": "Boa liquidez: Ativos cobrem passivos com folga confortável. Setores: bens de consumo, utilities. Atraente para conservadores. Analisar: fluxo de caixa, gestão de passivos.",
        "Agrupador": "Liquidez"
    },
    {
        "Indicador": "Liquidez Corrente",
        "Categoria": "Ótimo",
        "Faixa de Referência": "Acima de 2",
        "Interpretação": "Liquidez excepcional: Forte capacidade de cobrir passivos de curto prazo. Setores: tecnologia, bens de consumo. Atraente, verificar alocação de ativos. Analisar: fluxo de caixa, eficiência.",
        "Agrupador": "Liquidez"
    },
    # LPA (Lucro por Ação)
    {
        "Indicador": "LPA",
        "Categoria": "Péssimo",
        "Faixa de Referência": "Negativo",
        "Interpretação": "Prejuízo por ação: Empresa não gera lucro, alto risco. Setores: crise, startups. Evitar, salvo recuperação clara. Analisar: fluxo de caixa, recuperação de lucros.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "LPA",
        "Categoria": "Ruim",
        "Faixa de Referência": "0 a 0,5",
        "Interpretação": "Lucro baixo: LPA positivo, mas limitado, risco moderado. Setores: cíclicos, varejo. Cautela, verificar crescimento. Analisar: lucros, tendências setoriais.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "LPA",
        "Categoria": "Moderado",
        "Faixa de Referência": "0,5 a 1",
        "Interpretação": "Lucro razoável: LPA aceitável, estabilidade moderada. Setores: indústria, bens de consumo. Atraente com lucros estáveis. Analisar: fluxo de caixa, crescimento.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "LPA",
        "Categoria": "Bom",
        "Faixa de Referência": "1 a 2",
        "Interpretação": "Lucro sólido: LPA indica boa rentabilidade por ação. Setores: utilities, bens de consumo. Ideal para investidores de valor. Analisar: consistência, crescimento.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "LPA",
        "Categoria": "Ótimo",
        "Faixa de Referência": "Acima de 2",
        "Interpretação": "Lucro excepcional: LPA alto, forte rentabilidade. Setores: tecnologia, bens de consumo. Atraente, verificar sustentabilidade. Analisar: lucros, escalabilidade.",
        "Agrupador": "Rentabilidade"
    },
    # Margem Bruta
    {
        "Indicador": "Margem Bruta",
        "Categoria": "Péssimo",
        "Faixa de Referência": "Menor que 10%",
        "Interpretação": "Margem muito baixa: Lucro bruto insuficiente após custos diretos. Setores: varejo competitivo, commodities. Evitar, salvo melhoria operacional. Analisar: custos, eficiência.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "Margem Bruta",
        "Categoria": "Ruim",
        "Faixa de Referência": "10 a 20%",
        "Interpretação": "Margem limitada: Lucro bruto moderado, risco em setores competitivos. Setores: varejo, indústria. Cautela, verificar redução de custos. Analisar: eficiência, concorrência.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "Margem Bruta",
        "Categoria": "Moderado",
        "Faixa de Referência": "20 a 30%",
        "Interpretação": "Margem razoável: Lucro bruto aceitável, estabilidade moderada. Setores: indústria, bens de consumo. Atraente com boa gestão. Analisar: custos, receita.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "Margem Bruta",
        "Categoria": "Bom",
        "Faixa de Referência": "30 a 50%",
        "Interpretação": "Margem sólida: Lucro bruto robusto, boa eficiência. Setores: bens de consumo, utilities. Ideal para investidores de valor. Analisar: consistência, crescimento.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "Margem Bruta",
        "Categoria": "Ótimo",
        "Faixa de Referência": "Acima de 50%",
        "Interpretação": "Margem excepcional: Lucro bruto muito alto, forte eficiência. Setores: tecnologia, saúde. Atraente, verificar sustentabilidade. Analisar: escalabilidade, concorrência.",
        "Agrupador": "Rentabilidade"
    },
    # Margem EBIT
    {
        "Indicador": "Margem EBIT",
        "Categoria": "Péssimo",
        "Faixa de Referência": "Menor que 5%",
        "Interpretação": "Margem operacional muito baixa: Lucro operacional insuficiente. Setores: crise, commodities. Evitar, salvo recuperação. Analisar: custos, eficiência.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "Margem EBIT",
        "Categoria": "Ruim",
        "Faixa de Referência": "5 a 10%",
        "Interpretação": "Margem operacional limitada: Lucro operacional moderado, risco em setores competitivos. Setores: varejo, indústria. Cautela, verificar eficiência. Analisar: custos, receita.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "Margem EBIT",
        "Categoria": "Moderado",
        "Faixa de Referência": "10 a 15%",
        "Interpretação": "Margem operacional razoável: Lucro operacional aceitável. Setores: indústria, bens de consumo. Atraente com boa gestão. Analisar: receita, eficiência.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "Margem EBIT",
        "Categoria": "Bom",
        "Faixa de Referência": "15 a 25%",
        "Interpretação": "Margem operacional sólida: Lucro operacional robusto. Setores: bens de consumo, utilities. Ideal para valor. Analisar: consistência, crescimento.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "Margem EBIT",
        "Categoria": "Ótimo",
        "Faixa de Referência": "Acima de 25%",
        "Interpretação": "Margem operacional excepcional: Lucro operacional muito alto. Setores: tecnologia, saúde. Atraente, verificar sustentabilidade. Analisar: escalabilidade, concorrência.",
        "Agrupador": "Rentabilidade"
    },
    # Margem EBITDA
    {
        "Indicador": "Margem EBITDA",
        "Categoria": "Péssimo",
        "Faixa de Referência": "Menor que 10%",
        "Interpretação": "Margem operacional muito baixa: Fluxo de caixa operacional insuficiente. Setores: crise, commodities. Evitar, salvo recuperação. Analisar: custos, eficiência.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "Margem EBITDA",
        "Categoria": "Ruim",
        "Faixa de Referência": "10 a 15%",
        "Interpretação": "Margem operacional limitada: Fluxo de caixa moderado, risco em setores competitivos. Setores: varejo, indústria. Cautela, verificar eficiência. Analisar: custos, receita.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "Margem EBITDA",
        "Categoria": "Moderado",
        "Faixa de Referência": "15 a 20%",
        "Interpretação": "Margem operacional razoável: Fluxo de caixa aceitável. Setores: indústria, bens de consumo. Atraente com boa gestão. Analisar: receita, eficiência.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "Margem EBITDA",
        "Categoria": "Bom",
        "Faixa de Referência": "20 a 30%",
        "Interpretação": "Margem operacional sólida: Fluxo de caixa robusto. Setores: bens de consumo, utilities. Ideal para valor. Analisar: consistência, crescimento.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "Margem EBITDA",
        "Categoria": "Ótimo",
        "Faixa de Referência": "Acima de 30%",
        "Interpretação": "Margem operacional excepcional: Fluxo de caixa muito alto. Setores: tecnologia, saúde. Atraente, verificar sustentabilidade. Analisar: escalabilidade, concorrência.",
        "Agrupador": "Rentabilidade"
    },
    # Margem Líquida
    {
        "Indicador": "Margem Líquida",
        "Categoria": "Péssimo",
        "Faixa de Referência": "Menor que 0%",
        "Interpretação": "Prejuízo líquido: Empresa não gera lucro após todas as despesas. Setores: crise, startups. Evitar, salvo recuperação clara. Analisar: fluxo de caixa, recuperação.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "Margem Líquida",
        "Categoria": "Ruim",
        "Faixa de Referência": "0 a 5%",
        "Interpretação": "Margem líquida baixa: Lucro líquido limitado, risco moderado. Setores: varejo, indústria. Cautela, verificar eficiência. Analisar: custos, receita.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "Margem Líquida",
        "Categoria": "Moderado",
        "Faixa de Referência": "5 a 10%",
        "Interpretação": "Margem líquida razoável: Lucro líquido aceitável. Setores: indústria, bens de consumo. Atraente com boa gestão. Analisar: receita, eficiência.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "Margem Líquida",
        "Categoria": "Bom",
        "Faixa de Referência": "10 a 20%",
        "Interpretação": "Margem líquida sólida: Lucro líquido robusto. Setores: bens de consumo, utilities. Ideal para valor. Analisar: consistência, crescimento.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "Margem Líquida",
        "Categoria": "Ótimo",
        "Faixa de Referência": "Acima de 20%",
        "Interpretação": "Margem líquida excepcional: Lucro líquido muito alto. Setores: tecnologia, saúde. Atraente, verificar sustentabilidade. Analisar: escalabilidade, concorrência.",
        "Agrupador": "Rentabilidade"
    },
    # ROA
    {
        "Indicador": "ROA",
        "Categoria": "Péssimo",
        "Faixa de Referência": "Menor que 0%",
        "Interpretação": "Retorno sobre ativos negativo: Empresa não gera lucro com ativos. Setores: crise, startups. Evitar, salvo recuperação. Analisar: eficiência, recuperação.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "ROA",
        "Categoria": "Ruim",
        "Faixa de Referência": "0 a 5%",
        "Interpretação": "Retorno baixo: ROA limitado, eficiência moderada. Setores: varejo, indústria. Cautela, verificar melhoria. Analisar: eficiência, ativos.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "ROA",
        "Categoria": "Moderado",
        "Faixa de Referência": "5 a 10%",
        "Interpretação": "Retorno razoável: ROA aceitável, eficiência moderada. Setores: indústria, bens de consumo. Atraente com boa gestão. Analisar: receita, eficiência.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "ROA",
        "Categoria": "Bom",
        "Faixa de Referência": "10 a 15%",
        "Interpretação": "Retorno sólido: ROA robusto, boa eficiência. Setores: bens de consumo, utilities. Ideal para valor. Analisar: consistência, crescimento.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "ROA",
        "Categoria": "Ótimo",
        "Faixa de Referência": "Acima de 15%",
        "Interpretação": "Retorno excepcional: ROA muito alto, eficiência superior. Setores: tecnologia, saúde. Atraente, verificar sustentabilidade. Analisar: escalabilidade, concorrência.",
        "Agrupador": "Rentabilidade"
    },
    # ROE
    {
        "Indicador": "ROE",
        "Categoria": "Péssimo",
        "Faixa de Referência": "Menor que 0%",
        "Interpretação": "Retorno sobre patrimônio negativo: Empresa não gera lucro para acionistas. Setores: crise, startups. Evitar, salvo recuperação. Analisar: fluxo de caixa, recuperação.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "ROE",
        "Categoria": "Ruim",
        "Faixa de Referência": "0 a 10%",
        "Interpretação": "Retorno baixo: ROE limitado, rentabilidade moderada. Setores: varejo, indústria. Cautela, verificar melhoria. Analisar: eficiência, patrimônio.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "ROE",
        "Categoria": "Moderado",
        "Faixa de Referência": "10 a 15%",
        "Interpretação": "Retorno razoável: ROE aceitável, rentabilidade moderada. Setores: indústria, bens de consumo. Atraente com boa gestão. Analisar: receita, eficiência.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "ROE",
        "Categoria": "Bom",
        "Faixa de Referência": "15 a 20%",
        "Interpretação": "Retorno sólido: ROE robusto, boa rentabilidade. Setores: bens de consumo, utilities. Ideal para valor. Analisar: consistência, crescimento.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "ROE",
        "Categoria": "Ótimo",
        "Faixa de Referência": "Acima de 20%",
        "Interpretação": "Retorno excepcional: ROE muito alto, rentabilidade superior. Setores: tecnologia, saúde. Atraente, verificar sustentabilidade. Analisar: escalabilidade, concorrência.",
        "Agrupador": "Rentabilidade"
    },
    # CAGR Lucros 5 anos
    {
        "Indicador": "CAGR Lucros 5 anos",
        "Categoria": "Péssimo",
        "Faixa de Referência": "Menor que 0%",
        "Interpretação": "Crescimento negativo: Lucros diminuíram nos últimos 5 anos. Setores: crise, cíclicos. Evitar, salvo recuperação. Analisar: tendências, recuperação.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "CAGR Lucros 5 anos",
        "Categoria": "Ruim",
        "Faixa de Referência": "0 a 5%",
        "Interpretação": "Crescimento baixo: Lucros cresceram pouco, risco moderado. Setores: varejo, indústria. Cautela, verificar tendências. Analisar: receita, eficiência.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "CAGR Lucros 5 anos",
        "Categoria": "Moderado",
        "Faixa de Referência": "5 a 10%",
        "Interpretação": "Crescimento razoável: Lucros com crescimento moderado. Setores: indústria, bens de consumo. Atraente com boa gestão. Analisar: receita, crescimento.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "CAGR Lucros 5 anos",
        "Categoria": "Bom",
        "Faixa de Referência": "10 a 20%",
        "Interpretação": "Crescimento sólido: Lucros com crescimento robusto. Setores: bens de consumo, utilities. Ideal para valor. Analisar: consistência, escalabilidade.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "CAGR Lucros 5 anos",
        "Categoria": "Ótimo",
        "Faixa de Referência": "Acima de 20%",
        "Interpretação": "Crescimento excepcional: Lucros com crescimento muito alto. Setores: tecnologia, saúde. Atraente, verificar sustentabilidade. Analisar: escalabilidade, concorrência.",
        "Agrupador": "Rentabilidade"
    },
    # ROIC
    {
        "Indicador": "ROIC",
        "Categoria": "Péssimo",
        "Faixa de Referência": "Menor que 0%",
        "Interpretação": "Retorno sobre capital investido negativo: Empresa não gera retorno. Setores: crise, startups. Evitar, salvo recuperação. Analisar: eficiência, recuperação.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "ROIC",
        "Categoria": "Ruim",
        "Faixa de Referência": "0 a 5%",
        "Interpretação": "Retorno baixo: ROIC limitado, eficiência moderada. Setores: varejo, indústria. Cautela, verificar melhoria. Analisar: eficiência, capital.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "ROIC",
        "Categoria": "Moderado",
        "Faixa de Referência": "5 a 10%",
        "Interpretação": "Retorno razoável: ROIC aceitável, eficiência moderada. Setores: indústria, bens de consumo. Atraente com boa gestão. Analisar: receita, eficiência.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "ROIC",
        "Categoria": "Bom",
        "Faixa de Referência": "10 a 15%",
        "Interpretação": "Retorno sólido: ROIC robusto, boa eficiência. Setores: bens de consumo, utilities. Ideal para valor. Analisar: consistência, crescimento.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "ROIC",
        "Categoria": "Ótimo",
        "Faixa de Referência": "Acima de 15%",
        "Interpretação": "Retorno excepcional: ROIC muito alto, eficiência superior. Setores: tecnologia, saúde. Atraente, verificar sustentabilidade. Analisar: escalabilidade, concorrência.",
        "Agrupador": "Rentabilidade"
    },
    # VPA
    {
        "Indicador": "VPA",
        "Categoria": "Péssimo",
        "Faixa de Referência": "Negativo",
        "Interpretação": "Valor patrimonial por ação negativo: Patrimônio líquido insuficiente. Setores: crise, startups. Evitar, salvo recuperação. Analisar: fluxo de caixa, recuperação.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "VPA",
        "Categoria": "Ruim",
        "Faixa de Referência": "0 a 1",
        "Interpretação": "Valor patrimonial baixo: VPA limitado, risco moderado. Setores: varejo, indústria. Cautela, verificar crescimento. Analisar: patrimônio, tendências.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "VPA",
        "Categoria": "Moderado",
        "Faixa de Referência": "1 a 5",
        "Interpretação": "Valor patrimonial razoável: VPA aceitável, estabilidade moderada. Setores: indústria, bens de consumo. Atraente com boa gestão. Analisar: patrimônio, crescimento.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "VPA",
        "Categoria": "Bom",
        "Faixa de Referência": "5 a 10",
        "Interpretação": "Valor patrimonial sólido: VPA robusto, boa estrutura. Setores: bens de consumo, utilities. Ideal para valor. Analisar: consistência, crescimento.",
        "Agrupador": "Rentabilidade"
    },
    {
        "Indicador": "VPA",
        "Categoria": "Ótimo",
        "Faixa de Referência": "Acima de 10",
        "Interpretação": "Valor patrimonial excepcional: VPA muito alto, forte estrutura. Setores: tecnologia, saúde. Atraente, verificar sustentabilidade. Analisar: escalabilidade, patrimônio.",
        "Agrupador": "Rentabilidade"
    }
]

# Criar DataFrame
df = pd.DataFrame(data)

# Criar arquivo Excel
wb = Workbook()
ws = wb.active
ws.title = "Indicadores Financeiros"

# Adicionar cabeçalhos
headers = ["Indicador", "Categoria", "Faixa de Referência", "Interpretação", "Agrupador"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.value = header
    cell.font = Font(bold=True)

# Adicionar dados
for r, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
    for c, value in enumerate(row, 1):
        ws.cell(row=r, column=c).value = value

# Ajustar largura das colunas
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)  # Limite de 50 para Interpretação
    ws.column_dimensions[column].width = adjusted_width

# Salvar o arquivo
wb.save("Indicadores_Financeiros_Detalhados2.xlsx")
print("Arquivo 'Indicadores_Financeiros_Detalhados.xlsx' gerado com sucesso!")

# Nota: Esta é a terceira parte dos dados. Combine com as partes 1 e 2 para a tabela completa.
