import openpyxl
import re
import time
import numpy as np
import pandas as pd
from unidecode import unidecode
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
import requests
# teste silvio 3e
import warnings
from openpyxl.styles import numbers

#from teste01 import metrica

fillvermelho= PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") # Vermelho

fillverde= PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid") # Verde

fillamarelo =  PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # Amarelo


fillazul = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid") # Azul


#filltitulo =   PatternFill(start_color="#002060", end_color="#002060", fill_type="solid")

filltitulo = PatternFill(start_color="002060", end_color="002060", fill_type="solid")  # Azul escuro

font_branca = Font(color="FFFFFF")  # Branco

TITLES = [
    'Identificação', 'Resumo Financeiro', 'Cotações', 'Informações Básicas',
    'Oscilações', 'Indicadores de Valuation', 'Indicadores de Rentabilidade',
    'Indicadores de Endividamento', 'Balanço Patrimonial', 'Demonstrativo de Resultados'
]

linha2 = 1
metricasts= ""

MetricasStatus = {
    'M. Liquida': {
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'regular': {'min': 5, 'max':10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Rendimento de dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'Uma margem líquida negativa revela que a empresa está gastando mais do que arrecada, seja por custos operacionais elevados, despesas financeiras, impostos altos, ou queda nas receitas',
        'descrruim': 'Uma margem líquida abaixo de 5% sugere que a empresa tem uma margem de lucro relativamente baixa. Isso pode ser típico em setores com alta concorrência e margens reduzidas, como o varejo ou o setor de alimentos e bebidas. Pode indicar baixa eficiência na gestão de custos ou baixa rentabilidade',
        'descregular': 'Uma margem líquida nessa faixa é considerada moderada e pode ser típica de empresas que operam em setores com margens mais equilibradas, como algumas indústrias ou serviços. Reflete uma eficiência razoável em converter receita em lucro líquido.',
        'descrbom': ' Uma margem líquida entre 10% e 20% é geralmente considerada boa e indica uma empresa que é eficiente na gestão de suas despesas e custos, resultando em uma rentabilidade sólida. É comum em setores com menos concorrência e mais poder de precificação',
        'descotimo': 'Uma margem líquida acima de 20% sugere alta eficiência e rentabilidade. Empresas com margens líquidas altas geralmente operam em setores com altas barreiras de entrada, produtos ou serviços de alto valor agregado, ou que têm uma forte vantagem competitiva. Isso pode ser visto em setores como tecnologia, software ou bens de luxo'
    },
    'M. EBIT': {
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 5},
        'regular': {'min': 5, 'max': 10},
        'bom': {'min': 10, 'max': 20},
        'otimo': {'min': 20, 'max': float('inf')},
        'descricao': 'Capacidade de pagar dívidas. Menor que 2.5 é considerado saudável.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'Uma margem EBIT negativa indica que a empresa está enfrentando problemas em sua operação principal, não conseguindo gerar lucro operacional. Isso pode ser resultado de custos elevados, queda nas vendas ou ineficiência na gestão',
        'descrruim': 'Uma margem EBIT abaixo de 5% sugere que a empresa tem uma baixa eficiência operacional. Isso pode ser comum em setores com alta concorrência e baixos níveis de diferenciação, como o varejo e alguns segmentos industriais. Indica que a empresa pode estar enfrentando desafios significativos em controlar seus custos operacionais.',
        'descregular': 'Uma margem EBIT nessa faixa é considerada moderada e pode ser típica de empresas em setores com margens mais equilibradas. Reflete uma eficiência operacional razoável, mas pode haver espaço para melhorar a rentabilidade operacional.',
        'descrbom': 'Uma margem EBIT entre 10% e 20% indica uma boa eficiência operacional. A empresa consegue gerar uma quantidade significativa de lucro operacional em relação à receita líquida. Esse nível é comum em setores menos intensivos em capital e com uma boa capacidade de controle de custos',
        'descotimo': 'Uma margem EBIT acima de 20% sugere alta eficiência operacional e uma forte capacidade de gerar lucro operacional a partir das vendas. Isso pode ser típico de setores com altos margens brutas e controle eficaz dos custos operacionais, como tecnologia, software e alguns serviços especializados'
    },
    'M. EBITDA': {
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 10},
        'regular': {'min':10, 'max': 20},
        'bom': {'min': 20, 'max': 30},
        'otimo': {'min': 30, 'max': float('inf')},
        'descricao': 'Rendimento do fluxo de caixa livre. Acima de 10% é considerado bom.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'A empresa está operando no prejuízo, pois suas despesas superam as receitas geradas pelas atividades principais',
        'descrruim': 'Uma margem EBITDA abaixo de 10% pode indicar que a empresa tem uma baixa eficiência operacional ou enfrenta altos custos operacionais. Isso pode ser comum em setores com alta concorrência e margens reduzidas, como o varejo ou indústrias com alta estrutura de custos',
        'descregular': ' Uma margem EBITDA nessa faixa é geralmente considerada moderada e pode ser típica de empresas que operam em setores com uma estrutura de custo mais equilibrada. Reflete uma eficiência operacional razoável e uma capacidade de gerar EBITDA de forma eficaz.',
        'descrbom': ' Uma margem EBITDA entre 20% e 30% indica uma boa eficiência operacional, com a empresa gerando uma quantidade significativa de EBITDA em relação à receita líquida. Isso é comum em setores com menos pressão sobre margens e onde as empresas conseguem manter um controle eficiente dos custos',
        'descotimo': 'Uma margem EBITDA acima de 30% sugere uma alta eficiência operacional e uma forte capacidade de gerar EBITDA a partir das vendas. Isso pode ser típico de setores com alta margem bruta e baixo custo de produção, como tecnologia, software e serviços financeiros.'
    },
    'M. Bruta': {
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 20},
        'regular': {'min': 20, 'max': 40},
        'bom': {'min': 40, 'max': 60},
        'otimo': {'min': 60, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Eficiência',
        'descrpessimo': 'A empresa não consegue cobrir nem os custos básicos de produção ou aquisição dos produtos vendidos ou dificuldades de mercado',
        'descrruim': 'Uma margem bruta abaixo de 20% pode indicar que a empresa tem um custo de produção relativamente alto em relação à sua receita. Isso é comum em setores com altos custos de matéria-prima ou produção, como o varejo de alimentos e algumas indústrias manufatureiras',
        'descregular': ' Uma margem bruta nessa faixa é geralmente considerada moderada e pode ser típica de empresas com uma estrutura de custos mais equilibrada. Indica que a empresa está conseguindo controlar seus custos de produção de maneira razoável.',
        'descrbom': 'Uma margem bruta entre 40% e 60% indica uma boa eficiência na produção e venda de produtos ou serviços. Isso é comum em setores que têm menos custos variáveis e uma estrutura de custo mais favorável, como serviços e algumas indústrias com alta diferenciação de produtos',
        'descotimo': 'Uma margem bruta acima de 60% sugere uma alta eficiência na produção e venda, com um custo relativamente baixo em relação à receita. Isso é típico em setores com altas barreiras de entrada, baixo custo de produção, ou produtos/serviços de alto valor agregado, como tecnologia e software'
    },
    'Div. liquida/PL': {
        'pessimo': {'min': 2,'max': float('inf')},
        'ruim': {'min': 1,'max': 2 },
        'regular': {'min': 0.5, 'max': 1},
        'bom': {'min': 0.4, 'max': 0.5},
        'otimo': {'min': -150, 'max': 0.4},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Endividamento',
        'descrpessimo': 'Este nível indica que a empresa está altamente alavancada, com um montante significativo de dívida em relação ao patrimônio líquido. Isso pode sinalizar um alto risco financeiro e a possibilidade de dificuldades para cumprir obrigações de dívida, especialmente se a empresa enfrentar quedas nos lucros ou aumento nos custos de financiamento',
        'descrruim': 'Empresas com múltiplos nessa faixa estão mais alavancadas, o que pode aumentar o risco financeiro, especialmente em períodos de incerteza econômica. Essas empresas dependem mais de financiamento por dívida, o que pode afetar sua capacidade de enfrentar crises ou de expandir',
        'descregular': ' Um múltiplo nessa faixa sugere que a empresa tem um equilíbrio saudável entre dívida e patrimônio líquido. Isso é comum em empresas que utilizam algum nível de alavancagem para financiar seu crescimento, mas que ainda mantêm um risco financeiro relativamente controlado.',
        'descrbom': 'Indica que a empresa tem um nível relativamente baixo de dívida em comparação ao seu patrimônio líquido. Empresas nessa faixa são geralmente consideradas financeiramente conservadoras e menos arriscadas, com menor dependência de financiamento por dívida',
        'descotimo': 'Uma relação negativa sugere que a empresa está em uma posição financeira confortável, com recursos suficientes para quitar todas as suas dívidas e ainda manter excedentes em caixa'
    },

    'Div. liquida/EBITDA': {
        'pessimo': {'min': 4,'max': float('inf')},
        'ruim': {'min': 3,'max': 4 },
        'regular': {'min': 2, 'max': 3},
        'bom': {'min':1.9, 'max': 2},
        'otimo': {'min': -150, 'max': 1.9},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Endividamento',
        'descrpessimo': 'Este múltiplo sugere que a empresa está altamente alavancada e pode enfrentar dificuldades em cumprir suas obrigações financeiras. Empresas com múltiplos acima de 4 estão em uma posição financeira mais vulnerável e podem estar em risco de insolvência ou precisar de reestruturação financeira se o EBITDA não for sustentável.',
        'descrruim': 'Indica uma alavancagem significativa. Empresas com múltiplos nesta faixa devem ser monitoradas de perto, pois têm menos margem de segurança para lidar com quedas no EBITDA ou aumentos de dívida. Em setores mais voláteis, esse nível de alavancagem pode ser preocupante.',
        'descregular': 'Um múltiplo nesta faixa é considerado aceitável, especialmente em setores onde é comum operar com algum nível de alavancagem. Indica que a empresa está gerenciando bem sua dívida, mas existe uma exposição maior ao risco se houver uma queda no EBITDA ou aumento na dívida.',
        'descrbom': 'Indica uma posição financeira forte, com baixa alavancagem. Empresas nesta faixa são geralmente vistas como financeiramente saudáveis, com uma sólida capacidade de pagar suas dívidas rapidamente utilizando seus lucros operacionais antes de despesas não operacionais. Essas empresas têm menor risco de problemas financeiros, mesmo em períodos de crise',
        'descotimo': 'A empresa possui mais caixa e ativos líquidos do que o total de suas dívidas. Nesse caso, a dívida líquida é menor que zero, indicando uma posição financeira confortável e baixo risco de insolvência'
    },

    'Div. liquida/EBIT': {
        'pessimo': {'min': 4,'max': float('inf')},
        'ruim': {'min': 3,'max': 4 },
        'regular': {'min': 2, 'max': 3},
        'bom': {'min':1.9, 'max': 2},
        'otimo': {'min': -150, 'max': 1.9},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Endividamento',
        'descrpessimo': 'Este valor sugere que a empresa está altamente alavancada e pode ter dificuldades para cumprir suas obrigações de dívida se o EBIT não for sustentável. Empresas com múltiplos superiores a 4 estão em uma posição financeira mais vulnerável e podem enfrentar problemas significativos se não conseguirem aumentar seus lucros ou reduzir suas dívidas.',
        'descrruim': 'Um múltiplo nesta faixa começa a indicar um nível elevado de alavancagem. Isso pode ser preocupante, especialmente se a empresa tiver fluxos de caixa voláteis ou se houver uma queda inesperada nos lucros. Embora possa ser administrável em setores com fluxos de caixa estáveis, é necessário cautela',
        'descregular': 'Esta faixa é considerada aceitável, especialmente em setores onde o uso de alavancagem é comum, como utilities ou infraestrutura. Indica que a empresa está gerenciando bem sua alavancagem, mas qualquer aumento significativo na dívida ou uma queda no EBIT pode levar a preocupações financeiras.',
        'descrbom': ' Este valor é considerado saudável. Indica que a empresa está em uma posição financeira sólida, com baixa alavancagem e boa capacidade de pagar suas dívidas rapidamente com os lucros operacionais. Empresas com este múltiplo são vistas como menos arriscadas e geralmente têm mais facilidade para enfrentar períodos de incerteza econômica',
        'descotimo': 'A empresa possui mais caixa e equivalentes de caixa do que dívidas financeiras. Nesse caso, o numerador (dívida líquida) é negativo, o que geralmente é considerado um sinal positivo, indicando solidez financeira e baixo risco de insolvência. A empresa teria recursos suficientes para quitar todas as dívidas e ainda manter excedente em caixa'    },

    'PL/Ativos': {
        'pessimo': {'min': -150, 'max': 0},
        'ruim': {'min': 0, 'max': 0.30},
        'regular': {'min': 0.30, 'max': 0.50},
        'bom': {'min': 0.50, 'max': 0.70},
        'otimo': {'min': 0.70, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Endividamento',
        'descrpessimo': 'A empresa possui mais dívidas do que bens e direitos, ou seja, se liquidasse todos os seus ativos, ainda assim não conseguiria pagar todas as suas obrigações',
        'descrruim': 'Um índice abaixo de 0,30 indica que a empresa tem uma alta proporção de ativos financiados por dívidas, o que pode indicar uma estrutura de capital altamente alavancada. Isso pode ser um sinal de risco financeiro elevado, pois a empresa depende fortemente de financiamento externo para operar. Em setores onde o capital é intensivo e o endividamento é comum, como o setor imobiliário ou de infraestrutura, valores baixos podem ser mais aceitáveis',
        'descregular': ' Um índice nessa faixa sugere uma estrutura de capital com uma proporção razoável de patrimônio líquido em relação aos ativos. A empresa tem um equilíbrio entre o financiamento próprio e o financiamento externo. Esse nível de alavancagem é geralmente considerado moderado e pode ser típico em muitos setores.',
        'descrbom': 'Um índice entre 0,50 e 0,70 indica que a empresa possui uma boa proporção de patrimônio líquido em relação aos ativos, o que sugere uma sólida base financeira e uma menor dependência de dívidas. Empresas com valores nessa faixa geralmente têm uma estrutura de capital mais conservadora e são menos suscetíveis a riscos financeiros',
        'descotimo': 'Um índice acima de 0,70 mostra que a empresa possui uma alta proporção de patrimônio líquido em relação aos seus ativos, indicando baixa alavancagem. Isso sugere que a empresa é financeiramente sólida e menos dependente de financiamento externo. No entanto, uma estrutura de capital com alta proporção de patrimônio líquido também pode indicar que a empresa não está aproveitando oportunidades de alavancagem financeira para expandir seus negócios'    },

    'Passivos/Ativos': {
        'pessimo': {'min': 1.5, 'max': float('inf')},
        'ruim': {'min': 1, 'max': 1.5},
        'regular': {'min': 1, 'max': 1},
        'bom': {'min': 0.9 , 'max': 1},
        'otimo': {'min': 0, 'max': 0.9},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Endividamento',
        'descrpessimo': 'Alto risco-Atenção',
        'descrruim': ' A empresa tem mais passivos do que ativos, o que significa que, mesmo vendendo todos os seus ativos, não conseguiria quitar todas as dívidas. Esse cenário representa risco elevado de insolvência e pode dificultar o acesso a crédito ou investimentos',
        'descregular': 'Indica que todos os ativos da empresa estão comprometidos com dívidas. A empresa esgotou os recursos próprios para pagamento das obrigações e não teria margem de segurança para enfrentar oscilações financeiras.',
        'descrbom': ' empresa possui mais ativos do que passivos e, portanto, tem capacidade de saldar todas as dívidas com seus próprios recursos',
        'descotimo': 'Considerado saudável. Significa que a empresa possui mais ativos do que passivos e, portanto, tem capacidade de saldar todas as dívidas com seus próprios recursos'
    },

    'Liq. corrente': {
        'pessimo': {'min': 0, 'max': 0.5},
        'ruim': {'min': 0.5, 'max': 1},
        'regular': {'min': 1 , 'max': 1.5},
        'bom': {'min': 1.5, 'max': 2},
        'otimo': {'min': 2, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Endividamento',
        'descrpessimo': 'Atenção',
        'descrruim': 'Um valor abaixo de 1 indica que a empresa possui mais passivos circulantes do que ativos circulantes, sugerindo que ela pode ter dificuldades em cumprir suas obrigações de curto prazo. Esse é um sinal de alerta, pois a empresa pode enfrentar problemas de liquidez, especialmente em momentos de pressão financeira',
        'descregular': 'Um múltiplo nesta faixa é geralmente considerado aceitável, significando que a empresa possui ativos circulantes suficientes para cobrir suas obrigações de curto prazo. No entanto, esse nível de liquidez pode ser considerado baixo em setores com alta volatilidade nos fluxos de caixa ou onde os ciclos de pagamento e recebimento são longos.',
        'descrbom': ' Essa faixa é vista como saudável, indicando que a empresa tem uma margem de segurança confortável para cobrir suas obrigações de curto prazo. Empresas com liquidez corrente nesta faixa geralmente são consideradas financeiramente estáveis',
        'descotimo': 'Um valor acima de 2 sugere que a empresa tem uma posição de liquidez muito forte, com ativos circulantes significativamente superiores aos passivos circulantes. Enquanto isso pode indicar uma gestão financeira conservadora, em alguns casos pode também sinalizar que a empresa não está utilizando seus ativos de forma eficiente para gerar retornos'
    },
    'D.Y': {
        'pessimo': {'min': -150, 'max': -10},
        'ruim': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation',
        'descrpessimo': 'descrpessimo',
        'descrruim': 'descrruim',
        'descregular': 'descregular.',
        'descrbom': 'descrbom',
        'descotimo': 'descotimo'    },

    'P/L': {
        'pessimo': {'min': -150, 'max': -10},
        'ruim': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation',
        'descrpessimo': 'descrpessimo',
        'descrruim': 'descrruim',
        'descregular': 'descregular.',
        'descrbom': 'descrbom',
        'descotimo': 'descotimo'    },

    'PEG Ratio': {
        'pessimo': {'min': -150, 'max': -10},
        'ruim': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation',
        'descrpessimo': 'descrpessimo',
        'descrruim': 'descrruim',
        'descregular': 'descregular.',
        'descrbom': 'descrbom',
        'descotimo': 'descotimo'    },

    'P/VP': {
        'pessimo': {'min': -150, 'max': -10},
        'ruim': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation',
        'descrpessimo': 'descrpessimo',
        'descrruim': 'descrruim',
        'descregular': 'descregular.',
        'descrbom': 'descrbom',
        'descotimo': 'descotimo'
    },

    'EV/EBITDA': {
        'pessimo': {'min': -150, 'max': -10},
        'ruim': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation',
        'descrpessimo': 'descrpessimo',
        'descrruim': 'descrruim',
        'descregular': 'descregular.',
        'descrbom': 'descrbom',
        'descotimo': 'descotimo'    },

    'EV/EBIT': {
        'pessimo': {'min': -150, 'max': -10},
        'ruim': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation',
        'descrpessimo': 'descrpessimo',
        'descrruim': 'descrruim',
        'descregular': 'descregular.',
        'descrbom': 'descrbom',
        'descotimo': 'descotimo'    },

    'P/EBITDA': {
        'pessimo': {'min': -150, 'max': -10},
        'ruim': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation',
        'descrpessimo': 'descrpessimo',
        'descrruim': 'descrruim',
        'descregular': 'descregular.',
        'descrbom': 'descrbom',
        'descotimo': 'descotimo'    },

    'P/EBIT': {
        'pessimo': {'min': -150, 'max': -10},
        'ruim': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation',
        'descrpessimo': 'descrpessimo',
        'descrruim': 'descrruim',
        'descregular': 'descregular.',
        'descrbom': 'descrbom',
        'descotimo': 'descotimo'    },

    'VPA': {
        'pessimo': {'min': -150, 'max': -10},
        'ruim': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation',
        'descrpessimo': 'descrpessimo',
        'descrruim': 'descrruim',
        'descregular': 'descregular.',
        'descrbom': 'descrbom',
        'descotimo': 'descotimo'    },

    'P/Ativo': {
        'pessimo': {'min': -150, 'max': -10},
        'ruim': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation',
        'descrpessimo': 'descrpessimo',
        'descrruim': 'descrruim',
        'descregular': 'descregular.',
        'descrbom': 'descrbom',
        'descotimo': 'descotimo'    },

    'LPA': {
        'pessimo': {'min': -150, 'max': -10},
        'ruim': {'min': -10, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation',
        'descrpessimo': 'descrpessimo',
        'descrruim': 'descrruim',
        'descregular': 'descregular.',
        'descrbom': 'descrbom',
        'descotimo': 'descotimo'    },

    'P/SR': {
        'pessimo': {'min': -150, 'max': -10},
        'ruim': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation',
        'descrpessimo': 'descrpessimo',
        'descrruim': 'descrruim',
        'descregular': 'descregular.',
        'descrbom': 'descrbom',
        'descotimo': 'descotimo'    },

    'P/Ativo Circ. Liq.': {

        'pessimo': {'min': -150, 'max': -10},
        'ruim': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation',
        'descrpessimo': 'descrpessimo',
        'descrruim': 'descrruim',
        'descregular': 'descregular.',
        'descrbom': 'descrbom',
        'descotimo': 'descotimo'    },
    'Valor atual': {
        'pessimo': {'min': -150, 'max': -10},
        'ruim': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa',
        'descrpessimo': 'descrpessimo',
        'descrruim': 'descrruim',
        'descregular': 'descregular.',
        'descrbom': 'descrbom',
        'descotimo': 'descotimo'    },

    'LIQUIDEZ MEDIA DIARIA': {
        'pessimo': {'min': -150, 'max': -10},
        'ruim': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa',
        'descrpessimo': 'descrpessimo',
        'descrruim': 'descrruim',
        'descregular': 'descregular.',
        'descrbom': 'descrbom',
        'descotimo': 'descotimo'    },

    'Patrimonio liquido': {
        'pessimo': {'min': -150, 'max': -10},
        'ruim': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa',
        'descrpessimo': 'descrpessimo',
        'descrruim': 'descrruim',
        'descregular': 'descregular.',
        'descrbom': 'descrbom',
        'descotimo': 'descotimo'
    },

    'Ativos': {
        'pessimo': {'min': -150, 'max': -10},
        'ruim': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa',
        'descrpessimo': 'descrpessimo',
        'descrruim': 'descrruim',
        'descregular': 'descregular.',
        'descrbom': 'descrbom',
        'descotimo': 'descotimo'    },

    'Ativo circulante': {
        'pessimo': {'min': -150, 'max': -10},
        'ruim': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa',
        'descrpessimo': 'descrpessimo',
        'descrruim': 'descrruim',
        'descregular': 'descregular.',
        'descrbom': 'descrbom',
        'descotimo': 'descotimo'    },

    'Divida bruta': {
        'pessimo': {'min': -150, 'max': -10},
        'ruim': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa',
        'descrpessimo': 'descrpessimo',
        'descrruim': 'descrruim',
        'descregular': 'descregular.',
        'descrbom': 'descrbom',
        'descotimo': 'descotimo'    },

    'Disponibilidade': {
        'pessimo': {'min': -150, 'max': -10},
        'ruim': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa',
        'descrpessimo': 'descrpessimo',
        'descrruim': 'descrruim',
        'descregular': 'descregular.',
        'descrbom': 'descrbom',
        'descotimo': 'descotimo'    },

    'Divida liquida': {
        'pessimo': {'min': -150, 'max': -10},
        'ruim': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa',
        'descrpessimo': 'descrpessimo',
        'descrruim': 'descrruim',
        'descregular': 'descregular.',
        'descrbom': 'descrbom',
        'descotimo': 'descotimo'    },

    'Valor de mercado': {
        'pessimo': {'min': -150, 'max': -10},
        'ruim': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa',
        'descrpessimo': 'descrpessimo',
        'descrruim': 'descrruim',
        'descregular': 'descregular.',
        'descrbom': 'descrbom',
        'descotimo': 'descotimo'
    },

    'Valor de firma': {
        'pessimo': {'min': -150, 'max': -10},
        'ruim': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa',
        'descrpessimo': 'descrpessimo',
        'descrruim': 'descrruim',
        'descregular': 'descregular.',
        'descrbom': 'descrbom',
        'descotimo': 'descotimo'    },

    'Free Float': {
        'pessimo': {'min': -150, 'max': -10},
        'ruim': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Empresa',
        'descrpessimo': 'descrpessimo',
        'descrruim': 'descrruim',
        'descregular': 'descregular.',
        'descrbom': 'descrbom',
        'descotimo': 'descotimo'    },

    'ROE': {
        'pessimo': {'min': -150, 'max': -10},
        'ruim': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Rentabilidade',
        'descrpessimo': 'descrpessimo',
        'descrruim': 'descrruim',
        'descregular': 'descregular.',
        'descrbom': 'descrbom',
        'descotimo': 'descotimo'    },
    'ROA': {
        'pessimo': {'min': -150, 'max': -10},
        'ruim': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Rentabilidade',
        'descrpessimo': 'descrpessimo',
        'descrruim': 'descrruim',
        'descregular': 'descregular.',
        'descrbom': 'descrbom',
        'descotimo': 'descotimo'    },
    'ROIC': {
        'pessimo': {'min': -150, 'max': -10},
        'ruim': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Rentabilidade',
        'descrpessimo': 'descrpessimo',
        'descrruim': 'descrruim',
        'descregular': 'descregular.',
        'descrbom': 'descrbom',
        'descotimo': 'descotimo'    },

    'Giro ativos': {
        'pessimo': {'min': -150, 'max': -10},
        'ruim': {'min': 0, 'max': 3},
        'regular': {'min': 3, 'max': 6},
        'bom': {'min': 6, 'max': 10},
        'otimo': {'min': 10, 'max': float('inf')},
        'descricao': 'Preço em relação ao lucro. Quanto menor, mais barata a ação.',
        'agrupador': 'Rentabilidade',
        'descrpessimo': 'descrpessimo',
        'descrruim': 'descrruim',
        'descregular': 'descregular.',
        'descrbom': 'descrbom',
        'descotimo': 'descotimo'    }
}

wbsaida = openpyxl.Workbook()


# define selenium webdriver options
options = webdriver.ChromeOptions()

# create selenium webdriver instance
driver = webdriver.Chrome(options=options)

def categorizar_valor(metrica, valor):
    try:
        if metrica not in MetricasStatus:
            return 'Métrica não reconhecida'
        valor2 = float(valor)
        #print("categoriza_valor", valor2)
        for categoria, limites in MetricasStatus[metrica].items():

            if categoria in ['descricao', 'agrupador','descrpessimo','descrruim','descregular','descrbom','descotimo']:
                continue
            if limites['min'] <= valor2 < limites['max']:
                return categoria
        return 'Valor fora do alcance definido'
    except Exception as e:
        print("Debug - limites:", limites)
        print(f"Erro inesperado categorizar: {e}")
        print(f"Erro metrica: {e}")
        print("categorizar_valor - Erro" , metrica)
        print("categoriza_valor", valor2)
        print('categoriza_valor' ,stock)
    finally:
       # print("categorizar_valor - OK2")
       pass

def criaPlanilhaIndRentabilidade(wbsaida):
    wbsaida.create_sheet('IndiRentabilidade')
    IndiRentabilidade = wbsaida['IndiRentabilidade']
    IndiRentabilidade.append(
        ['Agrupador', 'Fonte', 'ATIVO', 'Indicador', 'Valor', 'Referencia', 'Pessimo','Ruim', 'Regular', 'Bom', 'Otimo', 'Descrição'])

    for cell in IndiRentabilidade[1]:  # Apenas o cabeçalho
        cell.fill = filltitulo
        cell.font = font_branca

def tratamento(indicador):
    indicador2 = indicador

    try:
        if indicador2 in ["-", "--","--%"]:
            indicador2 = 0
        elif indicador2 is None or is_null_zero_or_spaces(indicador2):
            indicador2 = 0
        else:
            if indicador2 == "":  # Verificação adicionada para string vazia
                indicador2 = 0
            else:
                indicador2 = float(indicador2.strip('%')) / 100
        return indicador2

    except Exception as e:
        print(f"Erro inesperado tratamento : {e}", "metrica  ", metricasts, " indicador  ", indicador, " stock  ",stock)
        # print(metrica)  # Certifique-se de que metrica está definida
        # print(indicadortratado)  # Certifique-se de que indicadortratado está definida
        #print('tratamneto - erro', stock, "   ", metrica)


    finally:
       # print('tratamneto OK')
       pass
def tratamento3(indicador):
    indicador2 = indicador

    try:
        if indicador2 in ["-", "--","--%"]:
            indicador2 = 0
        elif indicador2 is None or is_null_zero_or_spaces(indicador2):
            indicador2 = 0
        else:
            if indicador2 == "":  # Verificação adicionada para string vazia
                indicador2 = 0
            else:
                indicador2 = float(indicador2)
        return indicador2

    except Exception as e:
        print(f"Erro inesperado tratamento 3 : {e}", " metrica  ", metricasts, " indicador  ", indicador, " stock  ",stock)
        # print(metrica)  # Certifique-se de que metrica está definida
        # print(indicadortratado)  # Certifique-se de que indicadortratado está definida
        #print('tratamneto3 - erro', stock, "   ", metrica)


    finally:
        #print('tratamneto3 OK')
        pass
# Certifique-se de que as variáveis `metrica` e `indicadortratado` estão definidas corretamente no contexto onde a função é chamada.

def tratamento2(indicador):
    indicador2 = indicador

    try:
        if indicador2 in ["-", "--","--%"]:
            indicador2 = 0
        elif indicador2 is None or is_null_zero_or_spaces(indicador2):
            indicador2 = 0
        elif indicador2 == "":  # Verificação adicionada para string vazia
            indicador2 = 0
        else:
            indicador2 = float(indicador2)
        return indicador2


        return indicador2
    except Exception as e:
      # print(f"Erro inesperado tratamento2: {e}", " metrica  ", metrica, " indicador  ", indicador ," stock ", stock)
        # print(metrica)  # Certifique-se de que metrica está definida
        # print(indicadortratado)  # Certifique-se de que indicadortratado está definida
        print('tratamneto2 - erro', stock)


    finally:
        #print('tratamneto2 OK', indicador)
        pass
def gravaIndiEficiênciaoStaus(wsIndiRentabilidade, dict_stocks, stock):
    # fontes ['StatusInvest', 'Fundamentus']



    global linha2
    global metricasts
    #linha2 = 1
    try:
        #print(dict_stocks)
        for metrica, detalhes in MetricasStatus.items():
    #        print(f'Métrica: {metrica}')
            linha2 += 1
            metricasts = metrica
            if metrica in ['Giro ativos', 'Div. liquida/PL','Div. liquida/EBITDA','Div. liquida/EBIT','PL/Ativos',
                           'Passivos/Ativos','Liq. corrente','P/L','PEG Ratio','P/VP','EV/EBITDA','EV/EBIT',
                            'P/EBITDA','P/EBIT','VPA','P/Ativo','LPA',
                            'P/SR','P/Ativo Circ. Liq.']:
             #   print('IF tratamneto2 ',metrica )

                indicadortratado = tratamento2(dict_stocks[stock].get(metrica))
                valor_pl = indicadortratado
                categoria_pl = categorizar_valor(metrica, (valor_pl))
              #  print("categoria " + categoria_pl)
               # print("categoria tratamento2 " + categoria_pl)
            elif metrica in ['Valor atual','LIQUIDEZ MEDIA DIARIA','Patrimonio liquido',
                             'Ativos','Ativo circulante','Divida bruta','Disponibilidade',
                             'Divida liquida','Valor de mercado','Valor de firma']:
               # print('IF tratamneto3 ', metrica)
               # print("categoria " + categoria_pl)
                indicadortratado = tratamento3(dict_stocks[stock].get(metrica))
                valor_pl = indicadortratado
                categoria_pl = categorizar_valor(metrica, (valor_pl))
               # print("categoria tratamento3" + categoria_pl)
            else:
               # print('IF tratamneto ', metrica)
                indicadortratado = tratamento(dict_stocks[stock].get(metrica))
                valor_pl = indicadortratado
                categoria_pl = categorizar_valor(metrica,(valor_pl * 100))
                #print("categoria tratamento" + categoria_pl)
                # Certifique-se de que 'ROE' é o valor correto para a métrica
   #         print(f'O índice P/L {valor_pl} é categorizado como: {categoria_pl}')
   #         print(f"  Agrupador: {detalhes['agrupador']}")


            wsIndiRentabilidade.cell(row=linha2, column=1, value=detalhes['agrupador'])
            wsIndiRentabilidade.cell(row=linha2, column=2, value='StausInvest')
            wsIndiRentabilidade.cell(row=linha2, column=3, value=stock)
           # print('celula - Indicador', metrica)
            wsIndiRentabilidade.cell(row=linha2, column=4, value=metrica)
            if metrica in ['Giro ativos', 'Div. liquida/PL','Div. liquida/EBITDA','Div. liquida/EBITDA',
                           'Div. liquida/EBIT','PL/Ativos','Passivos/Ativos','Liq. corrente',
                           'P/L','PEG Ratio','P/VP','EV/EBITDA','EV/EBIT',
                            'P/EBITDA','P/EBIT','VPA','P/Ativo','LPA',
                            'P/SR','P/Ativo Circ. Liq.']:
               # print('IF da celula - Indicador', metrica )
               # print('IF da celula - valor', valor_pl)
                wsIndiRentabilidade.cell(row=linha2, column=5, value=valor_pl).number_format = numbers.FORMAT_NUMBER_00
            elif  metrica in ['Valor atual','LIQUIDEZ MEDIA DIARIA','Patrimonio liquido',
                             'Ativos','Ativo circulante','Divida bruta','Disponibilidade',
                             'Divida liquida','Valor de mercado','Valor de firma']:
                  wsIndiRentabilidade.cell(row=linha2, column=5, value=valor_pl).number_format = 'R$ #,##0.00'
            else:
                wsIndiRentabilidade.cell(row=linha2, column=5, value=valor_pl).number_format = numbers.FORMAT_PERCENTAGE_00

            wsIndiRentabilidade.cell(row=linha2, column=7,
                             value=f"Mínimo = {detalhes['pessimo']['min']}, Máximo = {detalhes['pessimo']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=8,
                                     value=f"Mínimo = {detalhes['ruim']['min']}, Máximo = {detalhes['ruim']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=9,
                                     value=f"Mínimo = {detalhes['regular']['min']}, Máximo = {detalhes['regular']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=10,
                                     value=f"Mínimo = {detalhes['bom']['min']}, Máximo = {detalhes['bom']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=11,
                                     value=f"Mínimo = {detalhes['otimo']['min']}, Máximo = {detalhes['otimo']['max']}")
            #print(detalhes)

            #if metrica == 'Div. liquida/EBITDA':
            if categoria_pl == 'pessimo':
                wsIndiRentabilidade.cell(row=linha2, column=6, value=categoria_pl).fill = fillvermelho
                wsIndiRentabilidade.cell(row=linha2, column=11, value=f"{detalhes['descrpessimo']}")
            if categoria_pl == 'ruim':
                wsIndiRentabilidade.cell(row=linha2, column=6, value=categoria_pl).fill = fillvermelho
                wsIndiRentabilidade.cell(row=linha2, column=11, value=f"{detalhes['descrruim']}")
            if  categoria_pl == 'regular':
                wsIndiRentabilidade.cell(row=linha2, column=6, value=categoria_pl).fill =fillamarelo
                wsIndiRentabilidade.cell(row=linha2, column=11, value=f"{detalhes['descregular']}")
            if  categoria_pl == 'bom':
                wsIndiRentabilidade.cell(row=linha2, column=6, value=categoria_pl).fill = fillverde
                wsIndiRentabilidade.cell(row=linha2, column=11, value=f"{detalhes['descrbom']}")
            if categoria_pl == 'otimo':
                wsIndiRentabilidade.cell(row=linha2, column=6, value=categoria_pl).fill =fillazul
                wsIndiRentabilidade.cell(row=linha2, column=11, value=f"{detalhes['descotimo']}")


    except Exception as e:
        print(f"Erro inesperado grava planilha2: {e}")
        print(metrica)
        print(indicadortratado)

        print('gravaIndiEficiênciaoStaus - erro' ,  stock,"    ", metrica)
    finally:
        print('gravaIndiEficiênciaoStaus  OK''', stock)

def is_null_zero_or_spaces(variable):
    # Verifica se a variável é None
    if variable is None:
        return True
    # Verifica se a variável é zero (0)
    elif variable == 0:
        return True
    # Verifica se a variável é uma string e contém apenas espaços
    elif isinstance(variable, str) and variable.strip() == '':
        return True
    elif variable == '-%':
        return True
    else:
        return False


def get_stock_soup(stock):
    ''' Get raw html from a stock '''

    # access the stock urlww
    driver.get(f'https://statusinvest.com.br/acoes/{stock}')

    # get html from stock
    html = driver.find_element(By.ID, 'main-2').get_attribute('innerHTML')

    # remove accents from html and transform html into soup
    soup = BeautifulSoup(unidecode(html), 'html.parser')

    return soup


def soup_to_dict(soup):
    '''Get all data from stock soup and return as a dictionary '''
    keys, values = [], []

    # get divs from stock
    soup1 = soup.find('div', class_='pb-3 pb-md-5')
    soup2 = soup.find('div', class_='card rounded text-main-green-dark')
    soup3 = soup.find('div', class_='indicator-today-container')
    soup4 = soup.find(
        'div', class_='top-info info-3 sm d-flex justify-between mb-3')
    soups = [soup1, soup2, soup3, soup4]

    for s in soups:
        # get only titles from a div and append to keys
        titles = s.find_all('h3', re.compile('title m-0[^"]*'))
        titles = [t.get_text() for t in titles]
        keys += titles

        # get only numbers from a div and append to values
        numbers = s.find_all('strong', re.compile('value[^"]*'))
        numbers = [n.get_text()for n in numbers]
        values += numbers

    # remove unused key and insert needed keys
    keys.remove('PART. IBOV')
    keys.insert(6, 'TAG ALONG')
    keys.insert(7, 'LIQUIDEZ MEDIA DIARIA')

    # clean keys list
    keys = [k.replace('\nhelp_outline', '').strip() for k in keys]
    keys = [k for k in keys if k != '']

    # clean values list
    values = [v.replace('\nhelp_outline', '').strip() for v in values]
    values = [v.replace('.', '').replace(',', '.') for v in values]

    # create a dictionary using keys and values from indicators
    d = {k: v for k, v in zip(keys, values)}

    return d


if __name__ == "__main__":
    dict_stocks = {}
    criaPlanilhaIndRentabilidade(wbsaida)
    wsIndiRentabilidade = wbsaida['IndiRentabilidade']

    # start timer
    start = time.time()

    # read file with stocks codes to get stock information
    with open('stocks.txt', 'r') as f:
        stocks = f.read().splitlines()

        # get stock information and create excel sheet
        for stock in stocks:
            #print("stock :"  ,stock)
            try:
                # get data and transform into dictionary
                soup = get_stock_soup(stock)
                dict_stock = soup_to_dict(soup)
                dict_stocks[stock] = dict_stock
                gravaIndiEficiênciaoStaus(wsIndiRentabilidade, dict_stocks, stock)
            except:
                # if we not get the information... just skip it
                print(f'Could not get {stock} information', "    ", metricasts)

    # create dataframe using dictionary of stocks informations
    df = pd.DataFrame(dict_stocks)

    # replace missing values with NaN to facilitate processing
    df = df.replace(['', '-', '--', '-%', '--%'], np.nan)

    # write dataframe into csv file
    df.to_excel('stocks_data.xlsx', index_label='indicadores')

    # exit the driver
    driver.quit()

    # end timer
    end = time.time()
    wbsaida.save("StatusInvest.xlsx")
    print("teste")
    print(f'Brasilian stocks information got in {int(end-start)} s')
# silvio teste