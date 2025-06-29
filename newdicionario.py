import openpyxl
import re
import time
import numpy as np
import pandas as pd
from unidecode import unidecode
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
import requests
# teste silvio 3e
import warnings
from openpyxl.styles import numbers

#from teste01 import metrica

fillvermelho= PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") # Vermelho

fillverde= PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid") # Verde

fillamarelo =  PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # Amarelo


fillazul = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid") # Azul


#filltitulo =   PatternFill(start_color="#002060", end_color="#002060", fill_type="solid")

filltitulo = PatternFill(start_color="002060", end_color="002060", fill_type="solid")  # Azul escuro

font_branca = Font(color="FFFFFF")  # Branco

TITLES = [
    'Identificação', 'Resumo Financeiro', 'Cotações', 'Informações Básicas',
    'Oscilações', 'Indicadores de Valuation', 'Indicadores de Rentabilidade',
    'Indicadores de Endividamento', 'Balanço Patrimonial', 'Demonstrativo de Resultados'
]

linha2 = 1
metricasts= ""

MetricasStatus = {

    'Dívida Líquida/EBIT': {
        'pessimo': {'min': 4, 'max': float('inf'), 'descrpessimo': 'Endividamento crítico, alto risco de insolvência: A empresa levaria mais de 4 anos de EBIT para quitar a dívida líquida, indicando alavancagem extrema. Comum em empresas em crise ou setores cíclicos (ex.: construção em recessão). Investidores devem evitar, salvo reestruturação robusta com forte geração de caixa. Análise de fluxo de caixa e cobertura de juros é essencial.'},
        'ruim': {'min': 3, 'max': 4, 'descrruim': 'Endividamento elevado, risco financeiro significativo: Leva 3-4 anos de EBIT para pagar a dívida, um nível alto para a maioria dos setores. Comum em setores intensivos em capital (ex.: infraestrutura). Investidores devem analisar a estabilidade do EBIT e a exposição a riscos macroeconômicos (ex.: aumento de juros).'},
        'moderado': {'min': 1.5, 'max': 3, 'descrmoderado': 'Endividamento moderado, gestão financeira razoável: A dívida é pagável em 1,5-3 anos de EBIT, aceitável em setores estáveis (ex.: utilities). Investidores devem verificar a consistência do EBIT e o fluxo de caixa livre para garantir que a alavancagem não comprometa a operação.'},
        'bom': {'min': 0, 'max': 1.5, 'descrbom': 'Endividamento controlado, saúde financeira sólida: A empresa pode quitar a dívida em menos de 1,5 anos de EBIT, indicando baixa alavancagem. Comum em empresas maduras com forte geração de caixa (ex.: consumo defensivo). Investidores devem avaliar se a baixa dívida reflete subinvestimento.'},
        'otimo': {'min': float('-inf'), 'max': 0, 'descotimo': 'Caixa líquido supera dívida, situação financeira excepcional: A empresa tem mais caixa que dívida, refletindo forte saúde financeira. Comum em empresas de tecnologia ou com reservas elevadas (ex.: Apple). Investidores devem investigar a alocação do caixa (ex.: reinvestimento, dividendos).'},
        'descricao': 'Relação entre dívida líquida e EBIT. Quanto menor, melhor a saúde financeira.',
        'agrupador': 'Endividamento'
    },
    'Dívida Líquida/EBITDA': {
        'pessimo': {'min': 3.5, 'max': float('inf'), 'descrpessimo': 'Endividamento crítico, risco elevado de insolvência: A empresa precisa de mais de 3,5 anos de EBITDA para quitar a dívida líquida, indicando alavancagem extrema. Comum em empresas em crise (ex.: varejo em declínio). Investidores devem evitar, salvo recuperação clara com forte geração de caixa.'},
        'ruim': {'min': 2.5, 'max': 3.5, 'descrruim': 'Endividamento elevado, risco financeiro significativo: Leva 2,5-3,5 anos de EBITDA para pagar a dívida, um nível alto para a maioria dos setores. Comum em setores cíclicos (ex.: commodities). Investidores devem analisar a cobertura de juros e a volatilidade do EBITDA.'},
        'moderado': {'min': 1, 'max': 2.5, 'descrmoderado': 'Endividamento moderado, dependência razoável: A dívida é pagável em 1-2,5 anos de EBITDA, aceitável em setores com investimentos moderados (ex.: varejo). Investidores devem monitorar a tendência do EBITDA e a gestão de dívidas.'},
        'bom': {'min': 0, 'max': 1, 'descrbom': 'Endividamento baixo, forte capacidade de pagamento: A dívida líquida é mínima em relação ao EBITDA, indicando quitação rápida. Comum em empresas com margens altas (ex.: bens de consumo). Investidores devem avaliar se a baixa alavancagem limita crescimento.'},
        'otimo': {'min': float('-inf'), 'max': 0, 'descotimo': 'Caixa líquido excede dívida, saúde financeira robusta: A empresa tem mais caixa que dívida, sugerindo forte geração de caixa. Comum em setores de alto crescimento (ex.: tecnologia). Investidores devem verificar a alocação do caixa para evitar ineficiências.'},
        'descricao': 'Relação entre dívida líquida e EBITDA. Quanto menor, melhor a capacidade de pagamento.',
        'agrupador': 'Endividamento'
    },
    'Dívida Líquida/Patrimônio Líquido': {
        'pessimo': {'min': 1, 'max': float('inf'), 'descrpessimo': 'Alavancagem crítica, alto risco de insolvência: A dívida líquida excede o patrimônio líquido, indicando dependência extrema de financiamento externo. Comum em empresas em crise (ex.: varejo em declínio). Investidores devem evitar, salvo reestruturação com forte potencial de recuperação.'},
        'ruim': {'min': 0.7, 'max': 1, 'descrruim': 'Alavancagem elevada, risco financeiro significativo: A dívida representa 70-100% do patrimônio líquido, um nível alto para a maioria dos setores. Investidores devem analisar a capacidade de geração de lucro e a estrutura de prazos da dívida.'},
        'moderado': {'min': 0.3, 'max': 0.7, 'descrmoderado': 'Alavancagem moderada, dependência razoável: A dívida é 30-70% do patrimônio líquido, aceitável em setores com alavancagem moderada (ex.: indústria). Investidores devem monitorar a cobertura de juros e o fluxo de caixa.'},
        'bom': {'min': 0, 'max': 0.3, 'descrbom': 'Baixa alavancagem, estrutura financeira sólida: A dívida líquida é mínima em relação ao patrimônio, indicando uma empresa bem capitalizada. Comum em setores estáveis (ex.: utilities). Investidores devem avaliar se a baixa dívida limita crescimento.'},
        'otimo': {'min': float('-inf'), 'max': 0, 'descotimo': 'Patrimônio líquido supera dívida, situação financeira excepcional: A empresa tem mais caixa que dívida, refletindo forte capitalização. Comum em empresas de tecnologia ou maduras. Investidores devem verificar a alocação do caixa.'},
        'descricao': 'Relação entre dívida líquida e patrimônio líquido. Quanto menor, menor a alavancagem.',
        'agrupador': 'Endividamento'
    },
    'Dividend Yield (DY)': {
        'pessimo': {'min': 0, 'max': 0, 'descrpessimo': 'Sem pagamento de dividendos, foco em reinvestimento ou prejuízo: A empresa não paga dividendos, indicando reinvestimento total dos lucros (ex.: startups) ou prejuízo financeiro. Investidores focados em renda devem evitar, mas podem considerar para crescimento de longo prazo.'},
        'ruim': {'min': 0.1, 'max': 2, 'descrruim': 'Retorno baixo, pouco atrativo para investidores de dividendos: Um yield de 0,1-2% sugere que a empresa prioriza reinvestimento ou tem lucros limitados. Comum em empresas de crescimento (ex.: tecnologia). Investidores devem focar no potencial de valorização das ações.'},
        'moderado': {'min': 2, 'max': 4, 'descrmoderado': 'Retorno moderado, equilíbrio razoável: Um yield de 2-4% oferece retorno aceitável, comum em empresas maduras (ex.: consumo defensivo). Investidores devem verificar a sustentabilidade do payout e a consistência dos lucros.'},
        'bom': {'min': 4, 'max': 6, 'descrbom': 'Retorno atrativo, bom para investidores de renda: Um yield de 4-6% é atraente, indicando equilíbrio entre dividendos e reinvestimento. Comum em setores como utilities e bancos. Investidores devem confirmar a sustentabilidade do fluxo de caixa livre.'},
        'otimo': {'min': 6, 'max': float('inf'), 'descotimo': 'Alto retorno, mas verificar sustentabilidade: Um yield acima de 6% é muito atrativo, mas pode indicar risco de insustentabilidade, especialmente em empresas com lucros voláteis. Investidores devem analisar o payout ratio e a geração de caixa.'},
        'descricao': 'Percentual do preço da ação pago como dividendos. Acima de 6% é considerado bom.',
        'agrupador': 'Renda'
    },
    'EV/EBIT': {
        'pessimo': {'min': 15, 'max': float('inf'), 'descrpessimo': 'Empresa extremamente cara, alto risco de sobrevalorização: O valor da empresa (EV) é mais de 15 vezes o EBIT, sugerindo expectativas irreais de crescimento. Comum em setores de tecnologia em bolha. Investidores devem comparar com peers e avaliar o crescimento projetado.'},
        'ruim': {'min': 12, 'max': 15, 'descrruim': 'Empresa cara, múltiplo elevado: Um EV/EBIT de 12-15 indica valuation alta, comum em setores de crescimento (ex.: saúde). Investidores devem analisar se o crescimento futuro justifica o preço e comparar com a média setorial.'},
        'moderado': {'min': 8, 'max': 12, 'descrmoderado': 'Avaliação razoável, depende do setor: Um EV/EBIT de 8-12 é equilibrado, comum em empresas maduras com crescimento moderado (ex.: indústria). Investidores devem verificar a consistência do EBIT e as perspectivas de crescimento.'},
        'bom': {'min': 4, 'max': 8, 'descrbom': 'Subvalorizada, boa oportunidade: Um EV/EBIT de 4-8 sugere que a empresa está barata em relação ao lucro operacional. Comum em setores cíclicos ou em recuperação. Investidores devem investigar riscos operacionais.'},
        'otimo': {'min': float('-inf'), 'max': 4, 'descotimo': 'Extremamente subvalorizada ou EBIT negativo: Um EV/EBIT abaixo de 4 indica subvalorização significativa, mas um EBIT negativo torna o múltiplo inválido, sugerindo prejuízo operacional. Investidores devem analisar o potencial de recuperação.'},
        'descricao': 'Relação entre valor da empresa e EBIT. Quanto menor, mais barata a empresa.',
        'agrupador': 'Valuation'
    },
    'EV/EBITDA': {
        'pessimo': {'min': 12, 'max': float('inf'), 'descrpessimo': 'Empresa extremamente cara, risco de sobrevalorização: O EV/EBITDA acima de 12 sugere expectativas irreais de crescimento. Comum em setores de tecnologia em alta. Investidores devem comparar com peers e avaliar o crescimento projetado.'},
        'ruim': {'min': 10, 'max': 12, 'descrruim': 'Empresa cara, múltiplo elevado: Um EV/EBITDA de 10-12 indica valuation alta, comum em setores de crescimento moderado (ex.: saúde). Investidores devem analisar a estabilidade do EBITDA e o potencial de crescimento.'},
        'moderado': {'min': 6, 'max': 10, 'descrmoderado': 'Avaliação equilibrada, depende do crescimento: Um EV/EBITDA de 6-10 é aceitável em setores com crescimento moderado (ex.: varejo, indústria). Investidores devem verificar a tendência do EBITDA e o potencial de expansão.'},
        'bom': {'min': 3, 'max': 6, 'descrbom': 'Subvalorizada, oportunidade atrativa: Um EV/EBITDA de 3-6 indica que a empresa está barata em relação à geração de caixa. Comum em setores cíclicos ou em recuperação. Investidores devem verificar riscos operacionais.'},
        'otimo': {'min': float('-inf'), 'max': 3, 'descotimo': 'Extremamente subvalorizada ou EBITDA negativo: Um EV/EBITDA abaixo de 3 sugere subvalorização significativa, mas um EBITDA negativo indica prejuízo operacional. Investidores devem analisar as causas e o potencial de recuperação.'},
        'descricao': 'Relação entre valor da empresa e EBITDA. Quanto menor, mais barata a empresa.',
        'agrupador': 'Valuation'
    },
    'Giro do Ativo': {
        'pessimo': {'min': float('-inf'), 'max': 0.2, 'descrpessimo': 'Eficiência extremamente baixa, ativos subutilizados: A empresa gera pouca receita em relação aos ativos, indicando ineficiência grave. Comum em empresas com ativos ociosos ou em crise (ex.: infraestrutura em recessão). Investidores devem evitar, salvo reestruturação.'},
        'ruim': {'min': 0.2, 'max': 0.4, 'descrruim': 'Baixa eficiência, utilização limitada de ativos: O giro baixo sugere que os ativos não estão sendo bem utilizados, comum em setores com ativos pesados (ex.: indústria). Investidores devem investigar o potencial de melhoria operacional.'},
        'moderado': {'min': 0.4, 'max': 0.8, 'descrmoderado': 'Eficiência moderada, utilização razoável: A empresa utiliza seus ativos de forma aceitável, mas há espaço para melhorias. Comum em setores industriais. Investidores devem avaliar estratégias para aumentar o giro.'},
        'bom': {'min': 0.8, 'max': 1.2, 'descrbom': 'Boa eficiência, utilização sólida de ativos: A empresa gera receita significativa em relação aos ativos. Comum em setores intensivos em vendas (ex.: varejo). Investidores devem confirmar a sustentabilidade do giro.'},
        'otimo': {'min': 1.2, 'max': float('inf'), 'descotimo': 'Alta eficiência, utilização excepcional de ativos: A empresa maximiza a receita com seus ativos, indicando excelente eficiência operacional. Comum em setores com alta rotatividade (ex.: tecnologia, varejo). Investidores devem verificar margens.'},
        'descricao': 'Relação entre receita e ativos totais. Quanto maior, mais eficiente o uso dos ativos.',
        'agrupador': 'Eficiência'
    },
    'Liquidez Corrente': {
        'pessimo': {'min': float('-inf'), 'max': 0.8, 'descrpessimo': 'Risco crítico de insolvência, liquidez insuficiente: Os ativos circulantes não cobrem os passivos de curto prazo, indicando dificuldades graves. Comum em empresas em crise financeira. Investidores devem evitar, salvo recuperação clara.'},
        'ruim': {'min': 0.8, 'max': 1.2, 'descrruim': 'Liquidez frágil, risco moderado de insolvência: Os ativos circulantes estão próximos ou ligeiramente abaixo dos passivos, sugerindo risco moderado. Investidores devem analisar a qualidade dos ativos circulantes e o fluxo de caixa.'},
        'moderado': {'min': 1.2, 'max': 1.8, 'descrmoderado': 'Liquidez razoável, capacidade moderada de pagamento: A empresa pode cobrir obrigações de curto prazo com folga moderada. Investidores devem monitorar a composição dos ativos circulantes (ex.: caixa vs. estoques).'},
        'bom': {'min': 1.8, 'max': 2.5, 'descrbom': 'Boa liquidez, forte capacidade de pagamento: A empresa tem ativos circulantes suficientes para cobrir obrigações de curto prazo com folga. Comum em setores estáveis (ex.: consumo defensivo). Investidores veem isso como sinal de segurança.'},
        'otimo': {'min': 2.5, 'max': float('inf'), 'descotimo': 'Excelente liquidez, mínima chance de insolvência: A empresa tem mais que o dobro de ativos circulantes em relação aos passivos. Ideal para setores voláteis (ex.: varejo). Investidores devem verificar se o excesso de liquidez reflete ineficiência.'},
        'descricao': 'Relação entre ativos e passivos circulantes. Quanto maior, maior a capacidade de cumprir obrigações de curto prazo.',
        'agrupador': 'Liquidez'
    },
    'LPA (Lucro por Ação)': {
        'pessimo': {'min': float('-inf'), 'max': 0, 'descrpessimo': 'Prejuízo por ação, situação financeira crítica: A empresa gera prejuízo, indicando problemas operacionais ou financeiros graves. Comum em empresas em crise ou startups em fase de investimento. Investidores devem analisar o plano de recuperação.'},
        'ruim': {'min': 0, 'max': 0.1, 'descrruim': 'Lucro muito baixo, rentabilidade limitada: O LPA próximo de zero sugere lucros insignificantes, comum em setores competitivos ou empresas em recuperação. Investidores devem investigar se a baixa lucratividade é temporária.'},
        'moderado': {'min': 0.1, 'max': 0.5, 'descrmoderado': 'Lucro moderado, desempenho razoável: O LPA indica lucratividade limitada, mas positiva, comum em empresas em crescimento ou setores com margens menores (ex.: varejo). Investidores devem analisar o potencial de aumento do LPA.'},
        'bom': {'min': 0.5, 'max': 1, 'descrbom': 'Boa lucratividade, desempenho sólido: Um LPA de 0,5-1 reflete uma empresa lucrativa, comum em setores maduros (ex.: bancos). Investidores devem comparar com o P/L para avaliar a atratividade.'},
        'otimo': {'min': 1, 'max': float('inf'), 'descotimo': 'Alta lucratividade, desempenho excepcional: Um LPA acima de 1 indica forte lucro por ação, atraente para investidores. Comum em empresas com alta rentabilidade (ex.: tecnologia). Investidores devem confirmar a consistência dos lucros.'},
        'descricao': 'Lucro líquido por ação. Quanto maior, mais lucrativa a empresa por ação.',
        'agrupador': 'Rentabilidade'
    },
    'Margem Bruta': {
        'pessimo': {'min': float('-inf'), 'max': 0, 'descrpessimo': 'Prejuízo bruto, operação não rentável: Os custos diretos superam a receita, indicando ineficiência operacional grave. Comum em empresas em crise ou setores altamente competitivos (ex.: commodities em baixa). Investidores devem evitar, salvo recuperação.'},
        'ruim': {'min': 0, 'max': 15, 'descrruim': 'Baixa rentabilidade bruta, eficiência limitada: A margem bruta baixa sugere dificuldade em cobrir custos diretos, comum em setores com pressão de preços (ex.: varejo de baixa margem). Investidores devem investigar melhorias na precificação.'},
        'moderado': {'min': 15, 'max': 30, 'descrmoderado': 'Rentabilidade moderada, depende do setor: Uma margem bruta de 15-30% é aceitável em setores com concorrência moderada (ex.: indústria). Investidores devem comparar com peers e avaliar a pressão de custos.'},
        'bom': {'min': 30, 'max': 50, 'descrbom': 'Boa rentabilidade bruta, eficiência sólida: A empresa retém 30-50% da receita após custos diretos, indicando bom poder de precificação. Comum em setores com margens moderadas (ex.: bens de consumo). Investidores devem verificar a sustentabilidade.'},
        'otimo': {'min': 50, 'max': float('inf'), 'descotimo': 'Alta rentabilidade bruta, eficiência excepcional: Uma margem bruta acima de 50% reflete forte eficiência operacional. Comum em setores como tecnologia ou bens premium. Investidores devem confirmar a consistência frente a concorrência.'},
        'descricao': 'Percentual da receita que sobra após custos diretos. Quanto maior, mais eficiente a operação.',
        'agrupador': 'Rentabilidade'
    },
    'M. EBIT': {
        'pessimo': {'min': -150, 'max': 0, 'descrpessimo': 'Uma margem EBIT negativa indica que a empresa está enfrentando problemas em sua operação principal, não conseguindo gerar lucro operacional. Isso pode ser resultado de custos elevados, queda nas vendas ou ineficiência na gestão.'},
        'ruim': {'min': 0, 'max': 5, 'descrruim': 'Uma margem EBIT abaixo de 5% sugere que a empresa tem uma baixa eficiência operacional. Isso pode ser comum em setores com alta concorrência e baixos níveis de diferenciação, como o varejo e alguns segmentos industriais. Indica que a empresa pode estar enfrentando desafios significativos em controlar seus custos operacionais.'},
        'moderado': {'min': 5, 'max': 10, 'descrmoderado': 'Uma margem EBIT nessa faixa é considerada moderada e pode ser típica de empresas em setores com margens mais equilibradas. Reflete uma eficiência operacional razoável, mas pode haver espaço para melhorar a rentabilidade operacional.'},
        'bom': {'min': 10, 'max': 20, 'descrbom': 'Uma margem EBIT entre 10% e 20% indica uma boa eficiência operacional. A empresa consegue gerar uma quantidade significativa de lucro operacional em relação à receita líquida. Esse nível é comum em setores menos intensivos em capital e com uma boa capacidade de controle de custos.'},
        'otimo': {'min': 20, 'max': float('inf'), 'descotimo': 'Uma margem EBIT acima de 20% sugere alta eficiência operacional e uma forte capacidade de gerar lucro operacional a partir das vendas. Isso pode ser típico de setores com altos margens brutas e controle eficaz dos custos operacionais, como tecnologia, software e alguns serviços especializados.'},
        'descricao': 'Percentual da receita que sobra como lucro operacional. Quanto maior, mais eficiente a operação.',
        'agrupador': 'Rentabilidade'
    },
    'M. EBITDA': {
        'pessimo': {'min': -150, 'max': 0, 'descrpessimo': 'Uma margem EBITDA negativa indica incapacidade de gerar caixa operacional, um sinal crítico de problemas estruturais. Comum em empresas em crise ou setores com margens apertadas (ex.: varejo em declínio). Investidores devem evitar, salvo recuperação clara.'},
        'ruim': {'min': 0, 'max': 5, 'descrruim': 'Uma margem EBITDA abaixo de 5% sugere baixa eficiência na geração de caixa operacional, comum em setores intensivos em capital ou com alta concorrência (ex.: infraestrutura, varejo). Investidores devem investigar o potencial de melhoria operacional.'},
        'moderado': {'min': 5, 'max': 10, 'descrmoderado': 'Uma margem EBITDA nessa faixa é considerada moderada e pode ser típica de empresas em setores com margens equilibradas (ex.: indústria, varejo). Reflete uma capacidade razoável de gerar caixa, mas há espaço para melhorias na eficiência.'},
        'bom': {'min': 10, 'max': 20, 'descrbom': 'Uma margem EBITDA entre 10% e 20% indica boa eficiência na geração de caixa operacional. Comum em setores com margens moderadas e boa gestão de custos (ex.: bens de consumo). Investidores devem verificar a qualidade do EBITDA e a sustentabilidade.'},
        'otimo': {'min': 20, 'max': float('inf'), 'descotimo': 'Uma margem EBITDA acima de 20% sugere alta eficiência na geração de caixa, típica de setores com altas barreiras de entrada ou forte vantagem competitiva (ex.: tecnologia, software). Investidores devem confirmar a consistência frente a riscos setoriais.'},
        'descricao': 'Percentual da receita que sobra como EBITDA. Quanto maior, maior a geração de caixa operacional.',
        'agrupador': 'Rentabilidade'
    },
    'M. Líquida': {
        'pessimo': {'min': -150, 'max': 0, 'descrpessimo': 'Uma margem líquida negativa revela que a empresa está gastando mais do que arrecada, seja por custos operacionais elevados, despesas financeiras, impostos altos, ou queda nas receitas.'},
        'ruim': {'min': 0, 'max': 5, 'descrruim': 'Uma margem líquida abaixo de 5% sugere que a empresa tem uma margem de lucro relativamente baixa. Isso pode ser típico em setores com alta concorrência e margens reduzidas, como o varejo ou o setor de alimentos e bebidas. Pode indicar baixa eficiência na gestão de custos ou baixa rentabilidade.'},
        'moderado': {'min': 5, 'max': 10, 'descrmoderado': 'Uma margem líquida nessa faixa é considerada moderada e pode ser típica de empresas que operam em setores com margens mais equilibradas, como algumas indústrias ou serviços. Reflete uma eficiência razoável em converter receita em lucro líquido.'},
        'bom': {'min': 10, 'max': 20, 'descrbom': 'Uma margem líquida entre 10% e 20% é geralmente considerada boa e indica uma empresa que é eficiente na gestão de suas despesas e custos, resultando em uma rentabilidade sólida. É comum em setores com menos concorrência e mais poder de precificação.'},
        'otimo': {'min': 20, 'max': float('inf'), 'descotimo': 'Uma margem líquida acima de 20% sugere alta eficiência e rentabilidade. Empresas com margens líquidas altas geralmente operam em setores com altas barreiras de entrada, produtos ou serviços de alto valor agregado, ou que têm uma forte vantagem competitiva. Isso pode ser visto em setores como tecnologia, software ou bens de luxo.'},
        'descricao': 'Percentual da receita que sobra como lucro líquido. Quanto maior, mais rentável a empresa.',
        'agrupador': 'Rentabilidade'
    },
    'P/ATIVO': {
        'pessimo': {'min': 2, 'max': float('inf'), 'descrpessimo': 'Empresa extremamente cara, alto risco de sobrevalorização: O preço de mercado excede o dobro do valor dos ativos, sugerindo expectativas irreais de crescimento. Comum em setores intangíveis (ex.: tecnologia em bolha). Investidores devem comparar com peers e avaliar o crescimento projetado.'},
        'ruim': {'min': 1.5, 'max': 2, 'descrruim': 'Empresa cara, possível sobrevalorização: O preço é 1,5-2 vezes o valor dos ativos, indicando valuation alta. Investidores devem analisar a qualidade dos ativos (ex.: intangíveis vs. tangíveis) e o potencial de geração de receita.'},
        'moderado': {'min': 0.5, 'max': 1.5, 'descrmoderado': 'Avaliação equilibrada, depende do setor: O preço está alinhado com o valor dos ativos, aceitável na maioria dos setores. Investidores devem verificar a composição dos ativos e o potencial de geração de receita.'},
        'bom': {'min': 0.2, 'max': 0.5, 'descrbom': 'Subvalorizada, boa oportunidade: O preço é inferior ao valor dos ativos, indicando possível subvalorização. Comum em setores cíclicos ou em recuperação. Investidores devem investigar riscos operacionais.'},
        'otimo': {'min': float('-inf'), 'max': 0.2, 'descotimo': 'Extremamente subvalorizada: Um P/ATIVO abaixo de 0,2 sugere subvalorização significativa, indicando que o mercado precifica a empresa muito abaixo do valor de seus ativos. Investidores devem analisar riscos patrimoniais ou operacionais.'},
        'descricao': 'Relação entre preço de mercado e valor dos ativos. Quanto menor, mais barata a empresa.',
        'agrupador': 'Valuation'
    },
    'P/Ativo Circulante Líquido': {
        'pessimo': {'min': 3, 'max': float('inf'), 'descrpessimo': 'Avaliação extremamente elevada, risco de sobrevalorização: O preço excede o triplo do ativo circulante líquido, sugerindo expectativas irreais de crescimento. Investidores devem comparar com o fluxo de caixa e avaliar se o preço é justificado.'},
        'ruim': {'min': 2, 'max': 3, 'descrruim': 'Avaliação elevada, possível sobrevalorização: O preço é 2-3 vezes o ativo circulante líquido, indicando valuation alta. Investidores devem analisar a qualidade dos ativos circulantes (ex.: caixa vs. estoques) e o fluxo de caixa.'},
        'moderado': {'min': 1, 'max': 2, 'descrmoderado': 'Avaliação razoável, depende do contexto: O preço está alinhado com o ativo circulante líquido, aceitável em setores com alta liquidez (ex.: varejo). Investidores devem verificar a composição dos ativos circulantes.'},
        'bom': {'min': 0.5, 'max': 1, 'descrbom': 'Subvalorizada, boa oportunidade: O preço é inferior ao ativo circulante líquido, indicando possível subvalorização. Comum em empresas com forte liquidez. Investidores devem verificar riscos operacionais.'},
        'otimo': {'min': float('-inf'), 'max': 0.5, 'descotimo': 'Extremamente subvalorizada: Um P/ACL abaixo de 0,5 sugere subvalorização significativa, indicando que a empresa está precificada muito abaixo de seus ativos líquidos. Investidores devem analisar riscos de liquidez.'},
        'descricao': 'Relação entre preço de mercado e ativo circulante líquido. Quanto menor, mais barata a empresa.',
        'agrupador': 'Valuation'
    },
    'P/Capital de Giro': {
        'pessimo': {'min': 4, 'max': float('inf'), 'descrpessimo': 'Avaliação extremamente elevada, alto risco de liquidez: O preço é mais de 4 vezes o capital de giro, sugerindo sobrevalorização ou expectativas irreais. Investidores devem analisar a sustentabilidade do capital de giro.'},
        'ruim': {'min': 3, 'max': 4, 'descrruim': 'Avaliação elevada, risco de liquidez: O preço é 3-4 vezes o capital de giro, indicando valuation alta. Investidores devem verificar a qualidade do capital de giro e a necessidade do setor.'},
        'moderado': {'min': 1, 'max': 3, 'descrmoderado': 'Avaliação equilibrada, razoável para o setor: O preço está alinhado com o capital de giro, aceitável em setores com necessidade moderada de capital (ex.: varejo). Investidores devem monitorar a gestão do capital de giro.'},
        'bom': {'min': 0.5, 'max': 1, 'descrbom': 'Subvalorizada, oportunidade atrativa: O preço é inferior ao capital de giro, indicando possível subvalorização. Comum em empresas com forte liquidez. Investidores devem verificar riscos operacionais.'},
        'otimo': {'min': float('-inf'), 'max': 0.5, 'descotimo': 'Extremamente subvalorizada: Um P/Capital de Giro abaixo de 0,5 sugere subvalorização significativa, indicando que a empresa está precificada abaixo do capital de giro. Investidores devem analisar riscos de liquidez.'},
        'descricao': 'Relação entre preço de mercado e capital de giro. Quanto menor, mais barata a empresa.',
        'agrupador': 'Valuation'
    },
    'P/EBIT': {
        'pessimo': {'min': 20, 'max': float('inf'), 'descrpessimo': 'Empresa extremamente cara, alto risco de sobrevalorização: O preço é mais de 20 vezes o EBIT, sugestivo de expectativas irreais de crescimento. Comum em setores de tecnologia em bolha. Investidores devem comparar com peers e avaliar o crescimento projetado.'},
        'ruim': {'min': 15, 'max': 20, 'descrruim': 'Empresa cara, múltiplo elevado: Um P/EBIT de 15-20 indica valuation alta, comum em setores de crescimento (ex.: saúde). Investidores devem analisar a consistência do EBIT e o potencial de crescimento.'},
        'moderado': {'min': 8, 'max': 15, 'descrmoderado': 'Avaliação razoável, depende do crescimento: Um P/EBIT de 8-15 é aceitável em setores com crescimento moderado (ex.: indústria). Investidores devem verificar a tendência do EBIT e as perspectivas de expansão.'},
        'bom': {'min': 4, 'max': 8, 'descrbom': 'Subvalorizada, boa oportunidade: Um P/EBIT de 4-8 indica que a empresa está barata em relação ao lucro operacional. Comum em setores cíclicos ou em recuperação. Investidores devem investigar riscos operacionais.'},
        'otimo': {'min': float('-inf'), 'max': 4, 'descotimo': 'Extremamente subvalorizada ou EBIT negativo: Um P/EBIT abaixo de 4 sugere subvalorização significativa, mas um EBIT negativo torna o múltiplo inválido, indicando prejuízo operacional. Investidores devem analisar o potencial de recuperação.'},
        'descricao': 'Relação entre preço de mercado e EBIT. Quanto menor, mais barata a empresa.',
        'agrupador': 'Valuation'
    },
    'P/EBITDA': {
        'pessimo': {'min': 15, 'max': float('inf'), 'descrpessimo': 'Empresa extremamente cara, risco de sobrevalorização: O preço é mais de 15 vezes o EBITDA, sugerindo expectativas irreais de crescimento. Comum em setores de tecnologia em alta. Investidores devem comparar com peers e avaliar o crescimento projetado.'},
        'ruim': {'min': 10, 'max': 15, 'descrruim': 'Empresa cara, múltiplo elevado: Um P/EBITDA de 10-15 indica valuation alta, comum em setores de crescimento moderado (ex.: saúde). Investidores devem analisar a estabilidade do EBITDA e o potencial de crescimento.'},
        'moderado': {'min': 6, 'max': 10, 'descrmoderado': 'Avaliação equilibrada, depende do crescimento: Um P/EBITDA de 6-10 é aceitável em setores com crescimento moderado (ex.: varejo, indústria). Investidores devem verificar a tendência do EBITDA e o potencial de expansão.'},
        'bom': {'min': 3, 'max': 6, 'descrbom': 'Subvalorizada, oportunidade atrativa: Um P/EBITDA de 3-6 indica que a empresa está barata em relação à geração de caixa. Comum em setores cíclicos ou em recuperação. Investidores devem verificar riscos operacionais.'},
        'otimo': {'min': float('-inf'), 'max': 3, 'descotimo': 'Extremamente subvalorizada ou EBITDA negativo: Um P/EBITDA abaixo de 3 sugere subvalorização significativa, mas um EBITDA negativo indica prejuízo operacional. Investidores devem analisar as causas e o potencial de recuperação.'},
        'descricao': 'Relação entre preço de mercado e EBITDA. Quanto menor, mais barata a empresa.',
        'agrupador': 'Valuation'
    },
    'P/L': {
        'pessimo': {'min': -150, 'max': -10, 'descrpessimo': 'Empresa com prejuízo significativo, alto risco financeiro: Um P/L negativo entre -150 e -10 indica que a empresa está gerando prejuízo, tornando o múltiplo inválido para avaliação tradicional. Comum em empresas em crise ou setores em declínio (ex.: varejo físico, companhias aéreas em crise). Investidores devem evitar, salvo reestruturação robusta com plano claro de recuperação.'},
        'ruim': {'min': 0, 'max': 3, 'descrruim': 'Empresa extremamente barata, mas com alto risco de armadilha de valor: Um P/L entre 0 e 3 sugere que a ação está muito barata, mas pode indicar lucros voláteis ou riscos setoriais (ex.: siderurgia, construção). Investidores devem comparar com peers e analisar a sustentabilidade dos lucros.'},
        'moderado': {'min': 3, 'max': 6, 'descrmoderado': 'Avaliação atrativa, oportunidade potencial com riscos moderados: Um P/L de 3 a 6 indica que a ação está subvalorizada, comum em setores cíclicos em recuperação (ex.: indústria). Investidores devem avaliar a qualidade dos lucros e riscos macroeconômicos.'},
        'bom': {'min': 6, 'max': 10, 'descrbom': 'Avaliação justa, boa relação preço-lucro: Um P/L de 6 a 10 sugere que a ação está precificada de forma justa, comum em empresas maduras (ex.: bens de consumo, utilities). Investidores devem analisar a consistência dos lucros e o potencial de crescimento.'},
        'otimo': {'min': 10, 'max': float('inf'), 'descotimo': 'Empresa cara, mas potencialmente justificada por crescimento: Um P/L acima de 10 indica que a ação está cara, comum em setores de alto crescimento (ex.: tecnologia, saúde). Investidores devem comparar com a média setorial e avaliar o crescimento projetado.'},
        'descricao': 'Relação entre preço da ação e lucro por ação. Quanto menor, mais barata a ação.',
        'agrupador': 'Valuation'
    },
    'P/VPA': {
        'pessimo': {'min': 3, 'max': float('inf'), 'descrpessimo': 'Empresa extremamente cara, alto risco de sobrevalorização: O preço excede o triplo do valor patrimonial por ação, sugerindo expectativas irreais de crescimento. Comum em setores intangíveis (ex.: tecnologia em bolha). Investidores devem comparar com o ROE e avaliar o crescimento projetado.'},
        'ruim': {'min': 2, 'max': 3, 'descrruim': 'Empresa cara, possível sobrevalorização patrimonial: O preço é 2-3 vezes o valor patrimonial, indicando valuation alta. Investidores devem analisar a qualidade do patrimônio e o retorno sobre o mesmo.'},
        'moderado': {'min': 1, 'max': 2, 'descrmoderado': 'Avaliação equilibrada, razoável para o setor: O preço está alinhado com o valor patrimonial, aceitável em setores maduros (ex.: bancos, indústria). Investidores devem verificar a consistência do ROE.'},
        'bom': {'min': 0.5, 'max': 1, 'descrbom': 'Subvalorizada, boa oportunidade: O preço é inferior ao valor patrimonial, indicando possível subvalorização. Comum em setores cíclicos ou em recuperação. Investidores devem investigar riscos patrimoniais.'},
        'otimo': {'min': float('-inf'), 'max': 0.5, 'descotimo': 'Extremamente subvalorizada: Um P/VPA abaixo de 0,5 sugere subvalorização significativa, indicando que a empresa está precificada abaixo do valor patrimonial. Investidores devem analisar riscos operacionais.'},
        'descricao': 'Relação entre preço de mercado e valor patrimonial por ação. Quanto menor, mais barata a empresa.',
        'agrupador': 'Valuation'
    },
    'Patrimônio/Ativos': {
        'pessimo': {'min': float('-inf'), 'max': 0, 'descrpessimo': 'Patrimônio líquido negativo, risco crítico de insolvência: Os passivos superam os ativos, indicando uma situação financeira crítica. Comum em empresas em crise ou falência iminente. Investidores devem exigir um plano robusto de recuperação.'},
        'ruim': {'min': 0, 'max': 0.2, 'descrruim': 'Alta alavancagem, estrutura frágil: O patrimônio representa menos de 20% dos ativos, indicando forte dependência de dívidas. Comum em empresas com alavancagem elevada (ex.: construção). Investidores devem analisar a capacidade de pagamento de juros.'},
        'moderado': {'min': 0.2, 'max': 0.5, 'descrmoderado': 'Alavancagem moderada, estrutura razoável: O patrimônio financia 20-50% dos ativos, aceitável em setores com alavancagem moderada (ex.: indústria). Investidores devem monitorar a gestão da dívida e a geração de caixa.'},
        'bom': {'min': 0.5, 'max': 0.8, 'descrbom': 'Baixa alavancagem, estrutura sólida: O patrimônio representa 50-80% dos ativos, indicando baixa dependência de dívidas. Comum em setores estáveis (ex.: utilities). Investidores veem isso como sinal de segurança.'},
        'otimo': {'min': 0.8, 'max': float('inf'), 'descotimo': 'Alavancagem mínima, estrutura excepcional: O patrimônio financia mais de 80% dos ativos, indicando uma empresa extremamente bem capitalizada. Ideal para setores voláteis (ex.: commodities). Investidores devem verificar se o excesso de capital próprio limita crescimento.'},
        'descricao': 'Relação entre patrimônio líquido e ativos totais. Quanto maior, menor a dependência de dívidas.',
        'agrupador': 'Endividamento'
    },
    'PSR': {
        'pessimo': {'min': 4, 'max': float('inf'), 'descrpessimo': 'Empresa extremamente cara, alto risco de sobrevalorização: O preço é mais de 4 vezes a receita, sugerindo expectativas irreais de crescimento. Comum em setores de tecnologia em bolha. Investidores devem comparar com peers e avaliar o crescimento projetado.'},
        'ruim': {'min': 3, 'max': 4, 'descrruim': 'Empresa cara, possível sobrevalorização: Um PSR de 3-4 indica valuation alta, comum em setores de crescimento moderado (ex.: saúde). Investidores devem analisar a margem líquida e o potencial de crescimento da receita.'},
        'moderado': {'min': 1, 'max': 3, 'descrmoderado': 'Avaliação equilibrada, depende do setor: Um PSR de 1-3 é aceitável em setores com margens moderadas (ex.: varejo, indústria). Investidores devem verificar a tendência da receita e a margem líquida.'},
        'bom': {'min': 0.5, 'max': 1, 'descrbom': 'Subvalorizada, boa oportunidade: Um PSR de 0,5-1 indica que a empresa está barata em relação à receita. Comum em setores cíclicos ou em recuperação. Investidores devem investigar riscos operacionais.'},
        'otimo': {'min': float('-inf'), 'max': 0.5, 'descotimo': 'Extremamente subvalorizada: Um PSR abaixo de 0,5 sugere subvalorização significativa, indicando que a empresa está precificada muito abaixo da receita. Investidores devem analisar riscos setoriais.'},
        'descricao': 'Relação entre preço de mercado e receita. Quanto menor, mais barata a empresa.',
        'agrupador': 'Valuation'
    },
    'ROA': {
        'pessimo': {'min': -150, 'max': 0, 'descrpessimo': 'Retorno negativo, ativos gerando prejuízo: A empresa gera prejuízo em relação aos ativos, indicando ineficiência grave. Comum em empresas em crise ou com investimentos mal sucedidos. Investidores devem analisar as causas e o potencial de recuperação.'},
        'ruim': {'min': 0, 'max': 3, 'descrruim': 'Baixa eficiência, rentabilidade limitada: Um ROA abaixo de 3% sugere que os ativos geram retorno limitado, comum em setores intensivos em capital (ex.: infraestrutura). Investidores devem investigar o potencial de melhoria operacional.'},
        'moderado': {'min': 3, 'max': 8, 'descrmoderado': 'Eficiência moderada, desempenho razoável: Um ROA de 3-8% é aceitável em setores com margens moderadas (ex.: indústria). Investidores devem analisar a tendência do ROA e a composição dos ativos.'},
        'bom': {'min': 8, 'max': 15, 'descrbom': 'Boa eficiência, retorno sólido: Um ROA de 8-15% indica que a empresa utiliza seus ativos de forma eficiente, comum em setores com alta rotatividade (ex.: varejo). Investidores devem confirmar a sustentabilidade do ROA.'},
        'otimo': {'min': 15, 'max': float('inf'), 'descotimo': 'Alta eficiência, retorno excepcional: Um ROA acima de 15% reflete excelente utilização dos ativos, comum em setores de alta margem (ex.: tecnologia). Investidores devem verificar a consistência do ROA frente a riscos setoriais.'},
        'descricao': 'Retorno sobre ativos totais. Quanto maior, mais eficiente o uso dos ativos.',
        'agrupador': 'Rentabilidade'
    },
    'ROE': {
        'pessimo': {'min': -150, 'max': 0, 'descrpessimo': 'Retorno negativo, prejuízo sobre o patrimônio: A empresa gera prejuízo, indicando ineficiência grave na utilização do capital próprio. Comum em empresas em crise. Investidores devem analisar as causas e o potencial de recuperação.'},
        'ruim': {'min': 0, 'max': 5, 'descrruim': 'Baixa rentabilidade, retorno limitado: Um ROE abaixo de 5% sugere dificuldades em gerar retorno sobre o patrimônio, comum em setores competitivos ou com alta alavancagem. Investidores devem investigar o potencial de melhoria.'},
        'moderado': {'min': 5, 'max': 15, 'descrmoderado': 'Rentabilidade moderada, desempenho razoável: Um ROE de 5-15% é aceitável em setores maduros (ex.: indústria, varejo). Investidores devem analisar a alavancagem e a consistência dos lucros.'},
        'bom': {'min': 15, 'max': 25, 'descrbom': 'Boa rentabilidade, desempenho sólido: Um ROE de 15-25% indica eficiência sólida na utilização do patrimônio, comum em setores com margens moderadas (ex.: bancos). Investidores devem confirmar a sustentabilidade.'},
        'otimo': {'min': 25, 'max': float('inf'), 'descotimo': 'Alta rentabilidade, desempenho excepcional: Um ROE acima de 25% reflete eficiência excepcional, comum em setores de alta margem (ex.: tecnologia). Investidores devem verificar a sustentabilidade frente a riscos setoriais.'},
        'descricao': 'Retorno sobre o patrimônio líquido. Quanto maior, mais eficiente o uso do capital próprio.',
        'agrupador': 'Rentabilidade'
    },
    'CAGR Lucros 5 anos': {
        'pessimo': {'min': -150, 'max': -10, 'descrpessimo': 'Declínio grave, risco de deterioração financeira: Um CAGR abaixo de -10% indica queda acentuada nos lucros, sugerindo problemas estruturais graves. Comum em empresas em crise ou setores em declínio. Investidores devem evitar, salvo recuperação clara.'},
        'ruim': {'min': -10, 'max': 0, 'descrruim': 'Declínio leve ou estagnação, situação preocupante: Um CAGR negativo ou próximo de zero indica estagnação ou declínio nos lucros, comum em setores maduros ou em crise (ex.: varejo físico). Investidores devem investigar as causas e o potencial de recuperação.'},
        'moderado': {'min': 0, 'max': 5, 'descrmoderado': 'Crescimento moderado ou estagnado, depende do setor: Um CAGR de 0-5% é aceitável em setores maduros (ex.: utilities). Investidores devem analisar as causas da estagnação e o potencial de retomada do crescimento.'},
        'bom': {'min': 5, 'max': 15, 'descrbom': 'Crescimento sólido, desempenho atrativo: Um CAGR de 5-15% indica crescimento consistente, comum em setores com expansão moderada (ex.: indústria). Investidores devem verificar a sustentabilidade do crescimento.'},
        'otimo': {'min': 15, 'max': float('inf'), 'descotimo': 'Crescimento excepcional, empresa em forte expansão: Um CAGR acima de 15% reflete forte desempenho, comum em setores de tecnologia ou saúde. Investidores devem confirmar a sustentabilidade frente a concorrência.'},
        'descricao': 'Crescimento anual composto dos lucros nos últimos 5 anos. Quanto maior, melhor o desempenho.',
        'agrupador': 'Rentabilidade'
    },
    'ROIC': {
        'pessimo': {'min': -150, 'max': 0, 'descrpessimo': 'Retorno negativo, capital investido não rentável: A empresa gera prejuízo sobre o capital investido, indicando ineficiência grave. Comum em empresas em crise ou com investimentos mal sucedidos. Investidores devem analisar as causas e o potencial de recuperação.'},
        'ruim': {'min': 0, 'max': 5, 'descrruim': 'Baixo retorno, eficiência limitada: Um ROIC abaixo de 5% sugere que o capital investido gera retorno limitado, comum em setores intensivos em capital (ex.: infraestrutura). Investidores devem investigar o potencial de melhoria.'},
        'moderado': {'min': 5, 'max': 12, 'descrmoderado': 'Retorno moderado, desempenho razoável: Um ROIC de 5-12% é aceitável em setores com margens moderadas (ex.: varejo). Investidores devem analisar a tendência do ROIC e a eficiência dos investimentos.'},
        'bom': {'min': 12, 'max': 20, 'descrbom': 'Boa eficiência, retorno sólido: Um ROIC de 12-20% indica que a empresa utiliza o capital investido de forma eficiente, comum em setores com investimentos moderados (ex.: bens de consumo). Investidores devem confirmar a sustentabilidade.'},
        'otimo': {'min': 20, 'max': float('inf'), 'descotimo': 'Alta eficiência, retorno excepcional: Um ROIC acima de 20% reflete excelente alocação de capital, comum em setores de alta margem (ex.: tecnologia). Investidores devem verificar a consistência frente a riscos setoriais.'},
        'descricao': 'Retorno sobre o capital investido. Quanto maior, mais eficiente a alocação de capital.',
        'agrupador': 'Rentabilidade'
    },
    'VPA': {
        'pessimo': {'min': float('-inf'), 'max': 0, 'descrpessimo': 'Patrimônio negativo, risco crítico financeiro: O VPA negativo indica que os passivos superam os ativos, sugerindo dificuldades financeiras graves ou falência iminente. Investidores devem evitar, salvo reestruturação com forte potencial de recuperação.'},
        'ruim': {'min': 0, 'max': 0.1, 'descrruim': 'Patrimônio muito baixo, estrutura frágil: Um VPA próximo de zero sugere um patrimônio líquido extremamente baixo, indicando alta alavancagem ou dificuldades financeiras. Investidores devem investigar o risco de deterioração patrimonial.'},
        'moderado': {'min': 0.1, 'max': 0.5, 'descrmoderado': 'Patrimônio moderado, situação razoável: Um VPA de 0,1-0,5 reflete um patrimônio limitado, mas positivo, comum em empresas em crescimento ou com alavancagem moderada. Investidores devem analisar a tendência do VPA.'},
        'bom': {'min': 0.5, 'max': 1, 'descrbom': 'Patrimônio sólido, boa capitalização: Um VPA de 0,5-1 indica um patrimônio líquido robusto por ação, comum em setores maduros (ex.: bancos). Investidores devem comparar com o P/VPA.'},
        'otimo': {'min': 1, 'max': float('inf'), 'descotimo': 'Patrimônio excepcional, altíssima capitalização: Um VPA acima de 1 reflete uma empresa extremamente bem capitalizada, comum em setores estáveis (ex.: utilities). Investidores devem verificar se o alto VPA reflete subinvestimento.'},
        'descricao': 'Valor patrimonial por ação. Quanto maior, mais robusto o patrimônio por ação.',
        'agrupador': 'Valuation'
    },
    'PL/Ativos': {
        'pessimo': {'min': float('-inf'), 'max': 0, 'descrpessimo': 'Patrimônio líquido negativo, risco crítico de insolvência: Os passivos superam os ativos, indicando uma situação financeira crítica. Comum em empresas em crise ou falência iminente. Investidores devem exigir um plano robusto de recuperação.'},
        'ruim': {'min': 0, 'max': 0.2, 'descrruim': 'Alta alavancagem, estrutura frágil: O patrimônio representa menos de 20% dos ativos, indicando forte dependência de dívidas. Comum em empresas com alavancagem elevada (ex.: construção). Investidores devem analisar a capacidade de pagamento de juros.'},
        'moderado': {'min': 0.2, 'max': 0.5, 'descrmoderado': 'Alavancagem moderada, estrutura razoável: O patrimônio financia 20-50% dos ativos, aceitável em setores com alavancagem moderada (ex.: indústria). Investidores devem monitorar a gestão da dívida e a geração de caixa.'},
        'bom': {'min': 0.5, 'max': 0.8, 'descrbom': 'Baixa alavancagem, estrutura sólida: O patrimônio representa 50-80% dos ativos, indicando baixa dependência de dívidas. Comum em setores estáveis (ex.: utilities). Investidores veem isso como sinal de segurança.'},
        'otimo': {'min': 0.8, 'max': float('inf'), 'descotimo': 'Alavancagem mínima, estrutura excepcional: O patrimônio financia mais de 80% dos ativos, indicando uma empresa extremamente bem capitalizada. Ideal para setores voláteis (ex.: commodities). Investidores devem verificar se o excesso de capital próprio limita crescimento.'},
        'descricao': 'Relação entre patrimônio líquido e ativos totais. Quanto maior, menor a dependência de dívidas.',
        'agrupador': 'Endividamento'
    }
}

wbsaida = openpyxl.Workbook()


# define selenium webdriver options
options = webdriver.ChromeOptions()

# create selenium webdriver instance
driver = webdriver.Chrome(options=options)

def categorizar_valor(metrica, valor):
    try:
        if metrica not in MetricasStatus:
            return 'Métrica não reconhecida'
        valor2 = float(valor)
        #print("categoriza_valor", valor2)
        for categoria, limites in MetricasStatus[metrica].items():

            if categoria in ['descricao', 'agrupador','descrpessimo','descrruim','descrmoderado','descrbom','descotimo']:
                continue
            if limites['min'] <= valor2 < limites['max']:
                return categoria
        return 'Valor fora do alcance definido'
    except Exception as e:
        print("Debug - limites:", limites)
        print(f"Erro inesperado categorizar: {e}")
        print(f"Erro metrica: {e}")
        print("categorizar_valor - Erro" , metrica)
        print("categoriza_valor", valor2)
        print('categoriza_valor' ,stock)
    finally:
       # print("categorizar_valor - OK2")
       pass

def criaPlanilhaIndRentabilidade(wbsaida):
    wbsaida.create_sheet('IndiRentabilidade')
    IndiRentabilidade = wbsaida['IndiRentabilidade']
    IndiRentabilidade.append(
        ['Agrupador', 'Fonte', 'ATIVO', 'Indicador', 'Valor', 'Referencia', 'Pessimo','Ruim', 'Moderado', 'Bom', 'Otimo', 'Descrição'])

    for cell in IndiRentabilidade[1]:  # Apenas o cabeçalho
        cell.fill = filltitulo
        cell.font = font_branca

def tratamento(indicador):
    indicador2 = indicador

    try:
        if indicador2 in ["-", "--","--%"]:
            indicador2 = 0
        elif indicador2 is None or is_null_zero_or_spaces(indicador2):
            indicador2 = 0
        else:
            if indicador2 == "":  # Verificação adicionada para string vazia
                indicador2 = 0
            else:
                indicador2 = float(indicador2.strip('%')) / 100
        return indicador2

    except Exception as e:
        print(f"Erro inesperado tratamento : {e}", "metrica  ", metricasts, " indicador  ", indicador, " stock  ",stock)
        # print(metrica)  # Certifique-se de que metrica está definida
        # print(indicadortratado)  # Certifique-se de que indicadortratado está definida
        #print('tratamneto - erro', stock, "   ", metrica)


    finally:
       # print('tratamneto OK')
       pass
def tratamento3(indicador):
    indicador2 = indicador

    try:
        if indicador2 in ["-", "--","--%"]:
            indicador2 = 0
        elif indicador2 is None or is_null_zero_or_spaces(indicador2):
            indicador2 = 0
        else:
            if indicador2 == "":  # Verificação adicionada para string vazia
                indicador2 = 0
            else:
                indicador2 = float(indicador2)
        return indicador2

    except Exception as e:
        print(f"Erro inesperado tratamento 3 : {e}", " metrica  ", metricasts, " indicador  ", indicador, " stock  ",stock)
        # print(metrica)  # Certifique-se de que metrica está definida
        # print(indicadortratado)  # Certifique-se de que indicadortratado está definida
        #print('tratamneto3 - erro', stock, "   ", metrica)


    finally:
        #print('tratamneto3 OK')
        pass
# Certifique-se de que as variáveis `metrica` e `indicadortratado` estão definidas corretamente no contexto onde a função é chamada.

def tratamento2(indicador):
    indicador2 = indicador

    try:
        if indicador2 in ["-", "--","--%"]:
            indicador2 = 0
        elif indicador2 is None or is_null_zero_or_spaces(indicador2):
            indicador2 = 0
        elif indicador2 == "":  # Verificação adicionada para string vazia
            indicador2 = 0
        else:
            indicador2 = float(indicador2)
        return indicador2


        return indicador2
    except Exception as e:
      # print(f"Erro inesperado tratamento2: {e}", " metrica  ", metrica, " indicador  ", indicador ," stock ", stock)
        # print(metrica)  # Certifique-se de que metrica está definida
        # print(indicadortratado)  # Certifique-se de que indicadortratado está definida
        print('tratamneto2 - erro', stock)


    finally:
        #print('tratamneto2 OK', indicador)
        pass
def gravaIndiEficiênciaoStaus(wsIndiRentabilidade, dict_stocks, stock):
    # fontes ['StatusInvest', 'Fundamentus']



    global linha2
    global metricasts
    #linha2 = 1
    try:
        #print(dict_stocks)
        for metrica, detalhes in MetricasStatus.items():
    #        print(f'Métrica: {metrica}')
            linha2 += 1
            metricasts = metrica
            if metrica in ['Giro ativos', 'Div. liquida/PL','Div. liquida/EBITDA','Div. liquida/EBIT','PL/Ativos',
                           'Passivos/Ativos','Liq. corrente','P/L','PEG Ratio','P/VP','EV/EBITDA','EV/EBIT',
                            'P/EBITDA','P/EBIT','VPA','P/Ativo','LPA',
                            'P/SR','P/Ativo Circ. Liq.']:
             #   print('IF tratamneto2 ',metrica )

                indicadortratado = tratamento2(dict_stocks[stock].get(metrica))
                valor_pl = indicadortratado
                categoria_pl = categorizar_valor(metrica, (valor_pl))
              #  print("categoria " + categoria_pl)
               # print("categoria tratamento2 " + categoria_pl)
            elif metrica in ['Valor atual','LIQUIDEZ MEDIA DIARIA','Patrimonio liquido',
                             'Ativos','Ativo circulante','Divida bruta','Disponibilidade',
                             'Divida liquida','Valor de mercado','Valor de firma']:
               # print('IF tratamneto3 ', metrica)
               # print("categoria " + categoria_pl)
                indicadortratado = tratamento3(dict_stocks[stock].get(metrica))
                valor_pl = indicadortratado
                categoria_pl = categorizar_valor(metrica, (valor_pl))
               # print("categoria tratamento3" + categoria_pl)
            else:
               # print('IF tratamneto ', metrica)
                indicadortratado = tratamento(dict_stocks[stock].get(metrica))
                valor_pl = indicadortratado
                categoria_pl = categorizar_valor(metrica,(valor_pl * 100))
                #print("categoria tratamento" + categoria_pl)
                # Certifique-se de que 'ROE' é o valor correto para a métrica
   #         print(f'O índice P/L {valor_pl} é categorizado como: {categoria_pl}')
   #         print(f"  Agrupador: {detalhes['agrupador']}")


            wsIndiRentabilidade.cell(row=linha2, column=1, value=detalhes['agrupador'])
            wsIndiRentabilidade.cell(row=linha2, column=2, value='StausInvest')
            wsIndiRentabilidade.cell(row=linha2, column=3, value=stock)
           # print('celula - Indicador', metrica)
            wsIndiRentabilidade.cell(row=linha2, column=4, value=metrica)
            if metrica in ['Giro ativos', 'Div. liquida/PL','Div. liquida/EBITDA','Div. liquida/EBITDA',
                           'Div. liquida/EBIT','PL/Ativos','Passivos/Ativos','Liq. corrente',
                           'P/L','PEG Ratio','P/VP','EV/EBITDA','EV/EBIT',
                            'P/EBITDA','P/EBIT','VPA','P/Ativo','LPA',
                            'P/SR','P/Ativo Circ. Liq.']:
               # print('IF da celula - Indicador', metrica )
               # print('IF da celula - valor', valor_pl)
                wsIndiRentabilidade.cell(row=linha2, column=5, value=valor_pl).number_format = numbers.FORMAT_NUMBER_00
            elif  metrica in ['Valor atual','LIQUIDEZ MEDIA DIARIA','Patrimonio liquido',
                             'Ativos','Ativo circulante','Divida bruta','Disponibilidade',
                             'Divida liquida','Valor de mercado','Valor de firma']:
                  wsIndiRentabilidade.cell(row=linha2, column=5, value=valor_pl).number_format = 'R$ #,##0.00'
            else:
                wsIndiRentabilidade.cell(row=linha2, column=5, value=valor_pl).number_format = numbers.FORMAT_PERCENTAGE_00

            wsIndiRentabilidade.cell(row=linha2, column=7,
                             value=f"Mínimo = {detalhes['pessimo']['min']}, Máximo = {detalhes['pessimo']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=8,
                                     value=f"Mínimo = {detalhes['ruim']['min']}, Máximo = {detalhes['ruim']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=9,
                                     value=f"Mínimo = {detalhes['moderado']['min']}, Máximo = {detalhes['moderado']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=10,
                                     value=f"Mínimo = {detalhes['bom']['min']}, Máximo = {detalhes['bom']['max']}")
            wsIndiRentabilidade.cell(row=linha2, column=11,
                                     value=f"Mínimo = {detalhes['otimo']['min']}, Máximo = {detalhes['otimo']['max']}")
            #print(detalhes)

            #if metrica == 'Div. liquida/EBITDA':
            if categoria_pl == 'pessimo':
                wsIndiRentabilidade.cell(row=linha2, column=6, value=categoria_pl).fill = fillvermelho
                wsIndiRentabilidade.cell(row=linha2, column=11, value=f"{detalhes['descrpessimo']}")
            if categoria_pl == 'ruim':
                wsIndiRentabilidade.cell(row=linha2, column=6, value=categoria_pl).fill = fillvermelho
                wsIndiRentabilidade.cell(row=linha2, column=11, value=f"{detalhes['descrruim']}")
            if  categoria_pl == 'moderado':
                wsIndiRentabilidade.cell(row=linha2, column=6, value=categoria_pl).fill =fillamarelo
                wsIndiRentabilidade.cell(row=linha2, column=11, value=f"{detalhes['descrmoderado']}")
            if  categoria_pl == 'bom':
                wsIndiRentabilidade.cell(row=linha2, column=6, value=categoria_pl).fill = fillverde
                wsIndiRentabilidade.cell(row=linha2, column=11, value=f"{detalhes['descrbom']}")
            if categoria_pl == 'otimo':
                wsIndiRentabilidade.cell(row=linha2, column=6, value=categoria_pl).fill =fillazul
                wsIndiRentabilidade.cell(row=linha2, column=11, value=f"{detalhes['descotimo']}")


    except Exception as e:
        print(f"Erro inesperado grava planilha2: {e}")
        print(metrica)
        print(indicadortratado)

        print('gravaIndiEficiênciaoStaus - erro' ,  stock,"    ", metrica)
    finally:
        print('gravaIndiEficiênciaoStaus  OK''', stock)

def is_null_zero_or_spaces(variable):
    # Verifica se a variável é None
    if variable is None:
        return True
    # Verifica se a variável é zero (0)
    elif variable == 0:
        return True
    # Verifica se a variável é uma string e contém apenas espaços
    elif isinstance(variable, str) and variable.strip() == '':
        return True
    elif variable == '-%':
        return True
    else:
        return False


def get_stock_soup(stock):
    ''' Get raw html from a stock '''

    # access the stock urlww
    driver.get(f'https://statusinvest.com.br/acoes/{stock}')

    # get html from stock
    html = driver.find_element(By.ID, 'main-2').get_attribute('innerHTML')

    # remove accents from html and transform html into soup
    soup = BeautifulSoup(unidecode(html), 'html.parser')

    return soup


def soup_to_dict(soup):
    '''Get all data from stock soup and return as a dictionary '''
    keys, values = [], []

    # get divs from stock
    soup1 = soup.find('div', class_='pb-3 pb-md-5')
    soup2 = soup.find('div', class_='card rounded text-main-green-dark')
    soup3 = soup.find('div', class_='indicator-today-container')
    soup4 = soup.find(
        'div', class_='top-info info-3 sm d-flex justify-between mb-3')
    soups = [soup1, soup2, soup3, soup4]

    for s in soups:
        # get only titles from a div and append to keys
        titles = s.find_all('h3', re.compile('title m-0[^"]*'))
        titles = [t.get_text() for t in titles]
        keys += titles

        # get only numbers from a div and append to values
        numbers = s.find_all('strong', re.compile('value[^"]*'))
        numbers = [n.get_text()for n in numbers]
        values += numbers

    # remove unused key and insert needed keys
    keys.remove('PART. IBOV')
    keys.insert(6, 'TAG ALONG')
    keys.insert(7, 'LIQUIDEZ MEDIA DIARIA')

    # clean keys list
    keys = [k.replace('\nhelp_outline', '').strip() for k in keys]
    keys = [k for k in keys if k != '']

    # clean values list
    values = [v.replace('\nhelp_outline', '').strip() for v in values]
    values = [v.replace('.', '').replace(',', '.') for v in values]

    # create a dictionary using keys and values from indicators
    d = {k: v for k, v in zip(keys, values)}

    return d


if __name__ == "__main__":
    dict_stocks = {}
    criaPlanilhaIndRentabilidade(wbsaida)
    wsIndiRentabilidade = wbsaida['IndiRentabilidade']

    # start timer
    start = time.time()

    # read file with stocks codes to get stock information
    with open('stocks.txt', 'r') as f:
        stocks = f.read().splitlines()

        # get stock information and create excel sheet
        for stock in stocks:
            #print("stock :"  ,stock)
            try:
                # get data and transform into dictionary
                soup = get_stock_soup(stock)
                dict_stock = soup_to_dict(soup)
                dict_stocks[stock] = dict_stock
                gravaIndiEficiênciaoStaus(wsIndiRentabilidade, dict_stocks, stock)
            except:
                # if we not get the information... just skip it
                print(f'Could not get {stock} information', "    ", metricasts)

    # create dataframe using dictionary of stocks informations
    df = pd.DataFrame(dict_stocks)

    # replace missing values with NaN to facilitate processing
    df = df.replace(['', '-', '--', '-%', '--%'], np.nan)

    # write dataframe into csv file
    df.to_excel('stocks_data.xlsx', index_label='indicadores')

    # exit the driver
    driver.quit()

    # end timer
    end = time.time()
    wbsaida.save("StatusInvest.xlsx")
    print("teste")
    print(f'Brasilian stocks information got in {int(end-start)} s')
# silvio teste