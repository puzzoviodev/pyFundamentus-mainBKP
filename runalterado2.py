#!/usr/bin/env python
# encoding: utf-8

# ------------------------------------------------------------------------------
#  Name: run.py
#  Version: 0.0.1
#
#  Summary: Python Fundamentus
#           Python Fundamentus is a Python API that allows you to quickly
#           access the main fundamental indicators of the main stocks
#           in the Brazilian market.
#
#  Author: Alexsander Lopes Camargos
#  Author-email: alcamargos@vivaldi.net
#
#  License: MIT g
# ------------------------------------------------------------------------------

"""
Python Fundamentus API: Instant access to key financial indicators of
Brazilian stocks, empowering investors with comprehensive market analysis.
"""

import fundamentus
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
import requests
import re
import time
import numpy as np
import pandas as pd
from unidecode import unidecode
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By

TITLES = [
    'Identificação', 'Resumo Financeiro', 'Cotações', 'Informações Básicas',
    'Oscilações', 'Indicadores de Valuation', 'Indicadores de Rentabilidade',
    'Indicadores de Endividamento', 'Balanço Patrimonial', 'Demonstrativo de Resultados'
]

meu_dicionario = {}
wbsaida = openpyxl.Workbook()

#Silvio inicio  criaPlanilhaIndiEmpresa b
def criaPlanilhaIndiEmpresa(wbsaida):
    wbsaida.create_sheet('IndEmpresa')
    IndValuation = wbsaida['IndEmpresa']
    IndValuation.append(["ATIVO","Valor atual",
                         "Min. 52 semanas",
                          "Max. 52 semanas",
                          "dividend Yield",
                          "Valorizacao (12m)",
                          "Tipo",
                          "TAG ALONG",
                          "LIQUIDEZ MEDIA DIARIA",
                          "PARTICIPACAO NO IBOV",
                          "MERCADO DE OPCOES",
                          "Patrimonio liquido",
                          "Ativos",
                          "Ativo circulante",
                          "Divida bruta",
                          "Disponibilidade",
                          "Divida liquida",
                          "Valor de mercado",
                          "Valor de firma"])
    return

def criaPlanilaIndValuation(wbsaida):
    wbsaida.create_sheet('IndValuation')
    IndValuation = wbsaida['IndValuation']
    IndValuation.append(['FONTE','ATIVO','D.Y', 'P/L', ' PEG Ratio','P/VP','EV/EBITDA','EV/EBIT','P/EBITDA','P/EBIT','VPA','P/Ativo',
                         'LPA','P/SR','P/Ativo Circ. Liq.'])
    return
def criaPlanilhaIndEndividamento(IndEndividamento):
    wbsaida.create_sheet('IndEndividamento')
    IndEndividamento = wbsaida['IndEndividamento']
    IndEndividamento.append(['ATIVO','Dív. líquida/PL', 'Dív. líquida/EBITDA', 'Dív. líquida/EBIT','PL/Ativos','Passivos/Ativos','Liq. corrente'])
    return

def criaPlanilhaIndiEficiência(IndiEficiência):
    wbsaida.create_sheet('IndiEficiência')
    IndiEficiência = wbsaida['IndiEficiência']
    IndiEficiência.append(['FONTE','ATIVO','M. Bruta', 'M. EBITDA', 'M. EBIT', 'M. Líquida'])
    return

def criaPlanilhaIndRentabilidade(IndiRentabilidade):
    wbsaida.create_sheet('IndiRentabilidade')
    IndiRentabilidade = wbsaida['IndiRentabilidade']
    IndiRentabilidade.append(['FONTE','ATIVO','ROE', 'ROA', 'ROIC','Giro ativos',''])
    return
def criaPlanilhaIndiCrescimento(IndiCrescimento):
    wbsaida.create_sheet('IndiCrescimento')
    IndiCrescimento = wbsaida['IndiCrescimento']
    IndiCrescimento.append(['FONTE','ATIVO','CAGR Receitas 5 anos', 'CAGR Lucros 5 anos'])
    return
def criaPlanilaDiveros(wbsaida):
    wbsaida.create_sheet('IndDiversos')
    IndValuation = wbsaida['IndDiversos']
    IndValuation.append(['FONTE','ATIVO','D.Y', 'P/L', ' PEG Ratio','P/VP','EV/EBITDA','EV/EBIT','P/EBITDA','P/EBIT','VPA','P/Ativo',
                         'LPA','P/SR','P/Ativo Circ. Liq.'])
    return
def gravaIndiRentabilidade(wsIndiRentabilidade,linha,ATIVO,ROE,ROA,ROIC,Giroativos):
    # Condicional corrigida
    if is_null_zero_or_spaces(ROE):
        ROE = 0
    else:
        ROE = float(ROE.strip('%')) / 100



    if is_null_zero_or_spaces(ROA):
        ROA = 0
    else:

        ROA = float(ROA.strip('%')) / 100

    if is_null_zero_or_spaces(ROIC):
        ROIC = 0
    else:
        ROIC = float(ROIC.strip('%')) / 100


        wsIndiRentabilidade.cell(row=linha, column=2, value=ATIVO)
        wsIndiRentabilidade.cell(row=linha, column=3, value=ROE)
        wsIndiRentabilidade.cell(row=linha, column=4, value=ROA)
        wsIndiRentabilidade.cell(row=linha, column=5, value=ROIC)
        wsIndiRentabilidade.cell(row=linha, column=6, value=Giroativos)

Dicrentabilidade = {}
# pylint: disable=line-too-long
linha = 1

if __name__ == '__main__':
    criaPlanilaIndValuation(wbsaida)
    criaPlanilhaIndEndividamento(wbsaida)
    criaPlanilhaIndiEficiência(wbsaida)
    criaPlanilhaIndRentabilidade(wbsaida)
    criaPlanilhaIndiCrescimento(wbsaida)
    criaPlanilhaIndiEmpresa(wbsaida)
    IndiRentabilidade = wbsaida['IndiRentabilidade']
    with open('stocks.txt', 'r') as f:
        stocks = f.read().splitlines()
        for stock in stocks:
            try:
                linha = linha + 1  # silvio
                #main_pipeline = fundamentus.Pipeline('VALE3')
                main_pipeline = fundamentus.Pipeline(stock)
                response = main_pipeline.get_all_information()

                # Extract the information from the response.
                stock_identification = response.transformed_information['stock_identification']
                financial_summary = response.transformed_information['financial_summary']
                price_information = response.transformed_information['price_information']
                detailed_information = response.transformed_information[
                    'detailed_information']
                oscillations = response.transformed_information['oscillations']
                valuation_indicators = response.transformed_information[
                    'valuation_indicators']
                profitability_indicators = response.transformed_information[
                    'profitability_indicators']
                indebtedness_indicators = response.transformed_information[
                    'indebtedness_indicators']
                balance_sheet = response.transformed_information['balance_sheet']
                income_statement = response.transformed_information['income_statement']


                print(TITLES[0])
                print('-' * len(TITLES[0]))
                for information in stock_identification:
                    print(
                        f'{stock_identification[information].title}: {stock_identification[information].value}'
                    )



                # Adicionando dados ao dicionário vazio


                # Verificando o dicionário atualizado

                print(f'\n{TITLES[1]}')
                print('-' * len(TITLES[1]))
                for information in financial_summary:
                    print(
                        f'{financial_summary[information].title}: {financial_summary[information].value}'
                    )

                print(f'\n{TITLES[2]}')
                print('-' * len(TITLES[2]))
                for information in price_information:
                    print(
                        f'{price_information[information].title}: {price_information[information].value}'
                    )

                print(f'\n{TITLES[3]}')
                print('-' * len(TITLES[2]))
                for information in detailed_information:
                    if information != 'variation_52_weeks':
                        print(
                            f'{detailed_information[information].title}: {detailed_information[information].value}'
                        )
                    else:
                        for sub_information in detailed_information[information]:
                            print(
                                f'{detailed_information[information][sub_information].title}: {detailed_information[information][sub_information].value}'
                            )

                print(f'\n{TITLES[4]}')
                print('-' * len(TITLES[4]))
                for information in oscillations:
                    print(
                        f'{oscillations[information].title}: {oscillations[information].value}'
                    )

                print(f'\n{TITLES[5]}')
                print('-' * len(TITLES[5]))
                for information in valuation_indicators:
                    print(
                        f'{valuation_indicators[information].title}: {valuation_indicators[information].value}'
                    )

                print(f'\n{TITLES[6]}')
                print('-' * len(TITLES[6]))
                for information in profitability_indicators:
                    print(
                        f'{profitability_indicators[information].title}: {profitability_indicators[information].value}'
                    )
                    Dicrentabilidade[profitability_indicators[information].title] = profitability_indicators[
                        information].value
                    wsIndiRentabilidade = wbsaida['IndiRentabilidade']
                    gravaIndiRentabilidade(wsIndiRentabilidade, linha, stock, "ROE", "ROA","ROIC","Giroativos")
                print(f'\n{TITLES[7]}')
                print('-' * len(TITLES[7]))
                for information in indebtedness_indicators:
                    print(
                        f'{indebtedness_indicators[information].title}: {indebtedness_indicators[information].value}'
                    )

                print(f'\n{TITLES[8]}')
                print('-' * len(TITLES[8]))
                for information in balance_sheet:
                    print(
                        f'{balance_sheet[information].title}: {balance_sheet[information].value}'
                    )

                print(f'\n{TITLES[9]}')
                print('-' * len(TITLES[9]))
                print('Últimos 03 meses')
                for information in income_statement['three_months']:
                    print(
                        f"\t{income_statement['three_months'][information].title}: {income_statement['three_months'][information].value}"
                    )
                print('Últimos 12 meses')
                for information in income_statement['twelve_months']:
                    print(
                        f"\t{income_statement['twelve_months'][information].title}: {income_statement['twelve_months'][information].value}"
                    )
            except:
                # if we not get the information... just skip it
                print(f'Could not get {stock} information')



print(linha)

import warnings


# define selenium webdriver options
options = webdriver.ChromeOptions()

# create selenium webdriver instancee
driver = webdriver.Chrome(options=options)

#silvio 2
#wbsaida = openpyxl.Workbook()



def gravaIndiCrescimento(wsIndiCrescimento, linha, ATIVO, CAGRReceitas5, CAGRLucros5):
    if is_null_zero_or_spaces(CAGRReceitas5):
        CAGRReceitas5 = 0
    else:
        RCAGRReceitas5OE = float(CAGRReceitas5.strip('%')) / 100

    if is_null_zero_or_spaces(CAGRLucros5):
       CAGRLucros5 = 0
    else:
       CAGRLucros5 = float(CAGRLucros5.strip('%')) / 100

       wsIndiCrescimento.cell(row=linha, column=2, value=ATIVO)
       wsIndiCrescimento.cell(row=linha, column=3, value=CAGRReceitas5)
       wsIndiCrescimento.cell(row=linha, column=4, value=CAGRLucros5)


def gravaIndiEficiência(wsIndiEficiência, linha, ATIVO, MBruta, MEBITDA,MEBIT,MLiquida):
   # MBruta = float(MBruta.strip('%')) /100
   # MEBITDA = float(MEBITDA.strip('%')) /100
   # MEBIT = float(MEBIT.strip('%')) /100
   # MLiquida = float(MLiquida.strip('%')) /100


   # Condicional corrigida
    if is_null_zero_or_spaces(MBruta):
        MBruta   = 0
    else:
       MBruta = float(MBruta.strip('%')) /100
    if is_null_zero_or_spaces(MEBITDA):
       MEBITDA =0
    else:
        MEBITDA = float(MEBITDA.strip('%')) / 100


    if is_null_zero_or_spaces(MEBIT):
       MEBIT = 0
    else:

      MEBIT = float(MEBIT.strip('%')) / 100

    if is_null_zero_or_spaces(MLiquida):
       MLiquida =0
    else:
        MLiquida = float(MLiquida.strip('%')) / 100


    wsIndiEficiência.cell(row=linha, column=2, value=ATIVO)
    wsIndiEficiência.cell(row=linha, column=3, value=MBruta)
    wsIndiEficiência.cell(row=linha, column=4, value=MEBITDA)
    wsIndiEficiência.cell(row=linha, column=5, value=MEBIT)
    wsIndiEficiência.cell(row=linha, column=6, value=MLiquida)

def gravaIndEndividamento(wsIndEndividamento, linha, ATIVO, MivliquidaPL, DivliquidaEBITDA,
                                    DivliquidaEBIT, PLAtivos,PassivosAtivos,Liqcorrente):
    wsIndEndividamento.cell(row=linha, column=2, value=ATIVO)
    wsIndEndividamento.cell(row=linha, column=3, value=MivliquidaPL)
    wsIndEndividamento.cell(row=linha, column=4, value=DivliquidaEBITDA)
    wsIndEndividamento.cell(row=linha, column=5, value=DivliquidaEBIT)
    wsIndEndividamento.cell(row=linha, column=6, value=PLAtivos)
    wsIndEndividamento.cell(row=linha, column=7, value=PassivosAtivos)
    wsIndEndividamento.cell(row=linha, column=8, value=Liqcorrente)

def gravaIndValuation(wsIndValuation, linha, ATIVO, DY, PL,PEGRatio,
                                      PVP, EVEBITDA, EVEBIT, PEBITDA,PEBIT,VPA,
                                      PAtivo,LPA,PSR,PCapGiro,PAtivoCircLiq):

        wsIndValuation.cell(row=linha, column=2, value=ATIVO)
        wsIndValuation.cell(row=linha, column=3, value=DY)
        wsIndValuation.cell(row=linha, column=4, value=PL)
        wsIndValuation.cell(row=linha, column=5, value=PEGRatio)
        wsIndValuation.cell(row=linha, column=6, value=PVP)
        wsIndValuation.cell(row=linha, column=7, value=EVEBITDA)
        wsIndValuation.cell(row=linha, column=8, value=EVEBIT)
        wsIndValuation.cell(row=linha, column=9, value=PEBITDA)
        wsIndValuation.cell(row=linha, column=10, value=PEBIT)
        wsIndValuation.cell(row=linha, column=11, value=VPA)
        wsIndValuation.cell(row=linha, column=12, value=PAtivo)
        wsIndValuation.cell(row=linha, column=13, value=LPA)
        wsIndValuation.cell(row=linha, column=14, value=PSR)
        wsIndValuation.cell(row=linha, column=15, value=PCapGiro)
        wsIndValuation.cell(row=linha, column=16, value=PAtivoCircLiq)


#Silvio fim

def get_stock_soup(stock):
    ''' Get raw html from a stock '''

    # access the stock url
    driver.get(f'https://statusinvest.com.br/acoes/{stock}')
    #driver.get(f'https://statusinvest.com.br/acoes/eua/{stock}')

    # get html from stock
    html = driver.find_element(By.ID, 'main-2').get_attribute('innerHTML')

    # remove accents from html and transform html into soup
    soup = BeautifulSoup(unidecode(html), 'html.parser')

    return soup

def is_null_zero_or_spaces(variable):
       # Verifica se a variável é None
       if variable is None:
           return True
       # Verifica se a variável é zero (0)
       elif variable == 0:
           return True
       # Verifica se a variável é uma string e contém apenas espaços
       elif isinstance(variable, str) and variable.strip() == '':
           return True
       elif variable == '-%':
           return True
       else:
           return False

def soup_to_dict(soup):
    '''Get all data from stock soup and return as a dictionary '''
    keys, values = [], []

    # get divs from stock
    soup1 = soup.find('div', class_='pb-3 pb-md-5')
    soup2 = soup.find('div', class_='card rounded text-main-green-dark')
    soup3 = soup.find('div', class_='indicator-today-container')
    soup4 = soup.find(
        'div', class_='top-info info-3 sm d-flex justify-between mb-3')
    soups = [soup1, soup2, soup3, soup4]

    for s in soups:
        # get only titles from a div and append to keys
        titles = s.find_all('h3', re.compile('title m-0[^"]*'))

        titles = [t.get_text() for t in titles]
        keys += titles
        print(keys)
        # get only numbers from a div and append to values
        numbers = s.find_all('strong', re.compile('value[^"]*'))
        numbers = [n.get_text()for n in numbers]
        values += numbers
        print(keys)
        print(values)
    # remove unused key and insert needed keys
    keys.remove('PART. IBOV')
    keys.insert(6, 'TAG ALONG')
    keys.insert(7, 'LIQUIDEZ MEDIA DIARIA')
    print(keys)
    print(values)
    # clean keys list
    keys = [k.replace('\nhelp_outline', '').strip() for k in keys]
    keys = [k for k in keys if k != '']

    # clean values list
    values = [v.replace('\nhelp_outline', '').strip() for v in values]
    values = [v.replace('.', '').replace(',', '.') for v in values]

    # create a dictionary using keys and values from indicators
    d = {k: v for k, v in zip(keys, values)}

    return d


if __name__ == "__main__":
    dict_stocks = {}

    # start timer
    start = time.time()
    # Silvio Inicio

    # Silvio Fim
    # read file with stocks codes to get stock information
    with open('stocks.txt', 'r') as f:
        stocks = f.read().splitlines()
       # linha = 1 # silvio
        # get stock information and create excel sheet
        for stock in stocks:
            try:
                # get data and transform into dictionary
                soup = get_stock_soup(stock)
                dict_stock = soup_to_dict(soup)
                dict_stocks[stock] = dict_stock
                linha = linha + 1 #silvio


                #IndiRentabilidade
                print(dict_stocks[stock].get("ROE"))
                print(dict_stocks[stock].get("ROA"))
                print(dict_stocks[stock].get("ROIC"))
                print(dict_stocks[stock].get("Giro ativos"))
                ROE =dict_stocks[stock].get("ROE")
                ROA =dict_stocks[stock].get("ROA")
                ROIC =dict_stocks[stock].get("ROIC")
                Giroativos =dict_stocks[stock].get("Giro ativos")


                #wsIndiRentabilidade = wbsaida['IndiRentabilidade']
                #gravaIndiRentabilidade(wsIndiRentabilidade, linha,stock, ROE,ROA,ROIC,Giroativos)

                #IndiCrescimento

                print(dict_stocks[stock].get("CAGR Receitas 5 anos"))
                print(dict_stocks[stock].get("CAGR Lucros 5 anos"))
                CAGRReceitas5 = dict_stocks[stock].get("CAGR Receitas 5 anos")
                CAGRLucros5  = dict_stocks[stock].get("CAGR Lucros 5 anos")
                wsIndiCrescimento = wbsaida['IndiCrescimento']
                gravaIndiCrescimento(wsIndiCrescimento, linha, stock, CAGRReceitas5, CAGRLucros5)


                #IndiEficiência
                print(dict_stocks[stock].get("M. Bruta"))
                print(dict_stocks[stock].get("M. EBITDA"))
                print(dict_stocks[stock].get("M. EBIT"))
                print(dict_stocks[stock].get("M. Liquida"))

                MBruta =  dict_stocks[stock].get("M. Bruta")
                MEBITDA = dict_stocks[stock].get("M. EBITDA")
                MEBIT =   dict_stocks[stock].get("M. EBIT")
                MLiquida =dict_stocks[stock].get("M. Liquida")

                wsIndiEficiência = wbsaida['IndiEficiência']
                gravaIndiEficiência(wsIndiEficiência, linha, stock, MBruta, MEBITDA,MEBIT,MLiquida)

                #IndEndividamento
                print(dict_stocks[stock].get("Div. liquida/PL"))
                print(dict_stocks[stock].get("Div. liquida/EBITDA"))
                print(dict_stocks[stock].get("Div. liquida/EBIT"))
                print(dict_stocks[stock].get("PL/Ativos"))
                print(dict_stocks[stock].get("Passivos/Ativos"))
                print(dict_stocks[stock].get("Liq. corrente"))

                DivliquidaPL = dict_stocks[stock].get("Div. liquida/PL")
                DivliquidaEBITDA = dict_stocks[stock].get("Div. liquida/EBITDA")
                DivliquidaEBIT = dict_stocks[stock].get("Div. liquida/EBIT")
                PLAtivos = dict_stocks[stock].get("PL/Ativos")
                PassivosAtivos = dict_stocks[stock].get("Passivos/Ativos")
                Liqcorrente = dict_stocks[stock].get("Liq. corrente")

                wsIndEndividamento = wbsaida['IndEndividamento']
                gravaIndEndividamento(wsIndEndividamento, linha, stock, DivliquidaPL, DivliquidaEBITDA,
                                    DivliquidaEBIT, PLAtivos,PassivosAtivos,Liqcorrente)

               # IndValuation
                print(dict_stocks[stock].get("D.Y"))
                print(dict_stocks[stock].get("P/L"))
                print(dict_stocks[stock].get("PEG Ratio"))
                print(dict_stocks[stock].get("P/VP"))
                print(dict_stocks[stock].get("EV/EBITDA"))
                print(dict_stocks[stock].get("EV/EBIT"))
                print(dict_stocks[stock].get("P/EBITDA"))
                print(dict_stocks[stock].get("P/EBIT"))
                print(dict_stocks[stock].get("VPA"))
                print(dict_stocks[stock].get("P/Ativo"))
                print(dict_stocks[stock].get("LPA"))
                print(dict_stocks[stock].get("P/SR"))
                print(dict_stocks[stock].get("P/Cap. Giro"))
                print(dict_stocks[stock].get("P/Ativo Circ. Liq."))

                DY             = dict_stocks[stock].get("D.Y")
                PL             = dict_stocks[stock].get("P/L")
                PEGRatio       = dict_stocks[stock].get("PEG Ratio")
                PVP            = dict_stocks[stock].get("P/VP")
                EVEBITDA       = dict_stocks[stock].get("EV/EBITDA")
                EVEBIT         = dict_stocks[stock].get("EV/EBIT")
                PEBITDA        = dict_stocks[stock].get("P/EBITDA")
                PEBIT          = dict_stocks[stock].get("P/EBIT")
                VPA            = dict_stocks[stock].get("VPA")
                PAtivo         = dict_stocks[stock].get("P/Ativo")
                LPA            = dict_stocks[stock].get("LPA")
                PSR            = dict_stocks[stock].get("P/SR")
                PCapGiro       = dict_stocks[stock].get("P/Cap. Giro")
                PAtivoCircLiq  = dict_stocks[stock].get("P/Ativo Circ. Liq.")

                wsIndValuation = wbsaida['IndValuation']
                gravaIndValuation(wsIndValuation, linha, stock, DY, PL,PEGRatio,
                                      PVP, EVEBITDA, EVEBIT, PEBITDA,PEBIT,VPA,
                                      PAtivo,LPA,PSR,PCapGiro,PAtivoCircLiq)
                # Empresa

                print(dict_stocks[stock].get("Valor atual"))
                print(dict_stocks[stock].get("Min. 52 semanas"))
                print(dict_stocks[stock].get("Max. 52 semanas"))
                print(dict_stocks[stock].get("dividend Yield"))
                print(dict_stocks[stock].get("Valorizacao (12m)"))
                print(dict_stocks[stock].get("Tipo"))
                print(dict_stocks[stock].get("TAG ALONG"))
                print(dict_stocks[stock].get("LIQUIDEZ MEDIA DIARIA"))
                print(dict_stocks[stock].get("PARTICIPACAO NO IBOV"))
                print(dict_stocks[stock].get("MERCADO DE OPCOES"))
                print(dict_stocks[stock].get("Patrimonio liquido"))
                print(dict_stocks[stock].get("Ativos"))
                print(dict_stocks[stock].get("Ativo circulante"))
                print(dict_stocks[stock].get("Divida bruta"))
                print(dict_stocks[stock].get("Disponibilidade"))
                print(dict_stocks[stock].get("Divida liquida"))
                print(dict_stocks[stock].get("Valor de mercado"))
                print(dict_stocks[stock].get("Valor de firma"))

            except:
                # if we not get the information... just skip it
                print(f'Could not get {stock} information')

    # create dataframe using dictionary of stocks informations
    df = pd.DataFrame(dict_stocks)

    # replace missing values with NaN to facilitate processing
    df = df.replace(['', '-', '--', '-%', '--%'], np.nan)

    # write dataframe into csv file
    df.to_excel('stocks_data.xlsx', index_label='indicadores')

    # exit the driver
    driver.quit()

    # end timer
    end = time.time()
    wbsaida.save("StatusInvest.xlsx") # silvio
    print(f'Brasilian stocks information got in {int(end-start)} s')

   # print(meu_dicionario)
    #nome = meu_dicionario["ROE"]
    #print (nome)
