from openpyxl import Workbook
from openpyxl.styles import numbers


def criaPlanilhaIndEndividamento(wbsaida):
    ws = wbsaida.create_sheet('IndEndividamento')
    ws.append(['Fonte', 'ATIVO', 'Dív. líquida/PL', 'Dív. líquida/EBITDA', 'Dív. líquida/EBIT', 'PL/Ativos',
               'Passivos/Ativos', 'Liq. corrente'])

    # Definir a formatação de porcentagem para as colunas apropriadas
    columns_to_format = ['C', 'D', 'E', 'F', 'G', 'H']
    for col in columns_to_format:
        ws[f'{col}1'].number_format = numbers.FORMAT_PERCENTAGE_00  # Formato de porcentagem com duas casas decimais

    # Exemplo de como adicionar valores a essas colunas com porcentagem
    exemplo_valores = {
        'Fonte': 'ExemploFonte',
        'ATIVO': 'ExemploAtivo',
        'Dív. líquida/PL': 0.25,  # 25%
        'Dív. líquida/EBITDA': 0.40,  # 40%
        'Dív. líquida/EBIT': 0.30,  # 30%
        'PL/Ativos': 0.50,  # 50%
        'Passivos/Ativos': 0.60,  # 60%
        'Liq. corrente': 0.70  # 70%
    }

    # Adicionar linha de exemplo com valores em porcentagem
    ws.append([
        exemplo_valores['Fonte'],
        exemplo_valores['ATIVO'],
        exemplo_valores['Dív. líquida/PL'],
        exemplo_valores['Dív. líquida/EBITDA'],
        exemplo_valores['Dív. líquida/EBIT'],
        exemplo_valores['PL/Ativos'],
        exemplo_valores['Passivos/Ativos'],
        exemplo_valores['Liq. corrente']
    ])

    # Formatar a segunda linha com porcentagem
    for col in columns_to_format:
        ws[f'{col}2'].number_format = numbers.FORMAT_PERCENTAGE_00  # Formato de porcentagem com duas casas decimais

    return


# Exemplo de uso
wbsaida = Workbook()
criaPlanilhaIndEndividamento(wbsaida)
wbsaida.save('IndEndividamento.xlsx')
